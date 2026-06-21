import csv
import gc
import hashlib
import json
import math
import os
import pickle
import re
import shutil
import struct
import sys
import glob
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import datetime
from tkinter import simpledialog, messagebox
import traceback
from frame_compression import process_videos_in_folder
import joblib
import numpy as np
import pandas as pd
from PyQt5.QtCore import QDir, Qt, QEvent
from PyQt5.QtCore import QObject, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QColor, QFontMetrics, QTextOption, QTextCursor
from PyQt5.QtWidgets import QApplication, QWidget, QFileSystemModel, QMainWindow, QProgressBar, QDialog, QLabel, \
    QVBoxLayout, QTableWidgetItem, QMessageBox, QLineEdit, QPushButton, QTableWidget, QInputDialog, QFileDialog, \
    QListWidget, QAction, QMenu, QComboBox, QHBoxLayout, QCheckBox, QGroupBox, QSizePolicy, QPlainTextEdit, QFrame
from PyQt5 import uic, QtWidgets
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from simhash import Simhash
from Train_GRUprocess import twoTrainClass
from clustering1 import trainClustering
from Train_GRUprocess_multi import TrainClass
from extractframe_single import extractGOP
from extract_sps import parse_sps
from pps import analyzesps

os.environ["CUDA_VISIBLE_DEVICES"]="0"

# Detect: 케이스 내 모든 feature 상세 폴더를 쓸 때 _choose_detail_folder_for_case 반환값
_DETECT_DETAIL_ALL = object()

# VISION 데이터셋 정답지 (URL 목록) 기본 경로
_VISION_GROUND_TRUTH_DEFAULT = r"J:\vision_dataset\VISION_base_files.txt"
_VISION_VIDEO_EXTENSIONS = (".mp4", ".mov", ".3gp", ".avi", ".mkv")

# 병렬 처리를 위한 독립 함수들 (pickle 가능하도록 클래스 밖에 정의)
def _compute_file_hash(file_path):
    """파일 해시 계산 (캐시 없이, 병렬 처리용) - get_fast_file_hash와 동일한 로직"""
    hash_type = 'md5'
    sample_size = 8192  # 원래 get_fast_file_hash와 동일한 샘플 크기
    
    hash_func = getattr(hashlib, hash_type)()
    
    try:
        with open(file_path, 'rb') as f:
            # 파일의 처음 부분에서 sample_size 만큼 읽기
            start_chunk = f.read(sample_size)
            hash_func.update(start_chunk)
            
            # 파일의 마지막 부분에서 sample_size 만큼 읽기
            f.seek(0, 2)  # 파일 끝으로 이동
            file_size = f.tell()
            if file_size > sample_size * 2:
                f.seek(-sample_size, 2)  # 파일 끝에서 sample_size 전으로 이동
                end_chunk = f.read(sample_size)
                hash_func.update(end_chunk)
        
        return hash_func.hexdigest()
    except Exception as e:
        print(f"해시 계산 오류 ({file_path}): {e}")
        return ""


def _numeric_class_atoms_from_mapping_label_cell(label_cell):
    """숫자 클래스 복합이면 원자 문자열 리스트, 아니면 None. 매핑 JSON 로드 시 보조 키용."""
    s = str(label_cell or "").strip()
    if not s:
        return None
    if " / " in s:
        parts = [p.strip() for p in re.split(r"\s*/\s*", s) if p.strip()]
    else:
        parts = [s]
    norm = []
    for p in parts:
        try:
            v = float(p)
            if not np.isfinite(v):
                return None
            norm.append(str(int(v)))
        except (ValueError, TypeError, OverflowError):
            return None
    return norm


def _mapping_class_key_to_display_name(mapping_data):
    """
    매핑 JSON에서 (CSV 클래스 인덱스/키 문자열) -> 사람이 읽는 라벨명.
    병합 우선순위: label_to_group(최우선) -> groups_detail/groups(빈 키 보완) -> label_name_display_overrides(최우선 override).
    즉, label_to_group에 이미 값이 있으면 groups_detail의 label_name으로 덮어쓰지 않는다.
    """
    if not isinstance(mapping_data, dict):
        return {}
    out = {}
    raw = mapping_data.get('label_to_group') or {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            if k is None or str(k).strip() == '':
                continue
            if v is None or str(v).strip() == '':
                continue
            out[str(k).strip()] = str(v).strip()
    groups = mapping_data.get('groups_detail') or mapping_data.get('groups') or []
    for entry in groups:
        if not isinstance(entry, dict):
            continue
        lbl = str(entry.get('label', '')).strip()
        lnm = str(entry.get('label_name', '')).strip()
        if lbl and lnm and lbl not in out:
            out[lbl] = lnm
    for entry in groups:
        if not isinstance(entry, dict):
            continue
        lbl = str(entry.get('label', '')).strip()
        lnm = str(entry.get('label_name', '')).strip()
        if not lbl or not lnm:
            continue
        atoms = _numeric_class_atoms_from_mapping_label_cell(lbl)
        if atoms:
            for ak in atoms:
                if ak not in out:
                    out[ak] = lnm
    for entry in mapping_data.get('label_name_display_overrides') or []:
        if not isinstance(entry, dict):
            continue
        fl = str(entry.get('full_label', entry.get('label', ''))).strip()
        disp = str(entry.get('display', '')).strip()
        if fl and disp:
            out[fl] = disp
    return out


def _folder_path_between_case_and_csv(case_dir, csv_path):
    """케이스 루트에서 CSV가 있는 폴더까지의 상대 경로(중간 폴더 이름). CSV가 케이스 밖이면 CSV 상위 폴더 절대경로."""
    if not case_dir or not csv_path:
        return ""
    try:
        ca = os.path.abspath(os.path.normpath(str(case_dir).strip()))
        cs = os.path.abspath(os.path.normpath(str(csv_path).strip()))
        parent = os.path.dirname(cs)
        common = os.path.commonpath([ca, parent])
        if os.path.normcase(common) != os.path.normcase(ca):
            return parent
        rel = os.path.relpath(parent, ca)
        if rel in (".", ""):
            return "(케이스 폴더와 동일)"
        return rel
    except (ValueError, OSError):
        return ""


def _process_one_file(file_path, state_flags, seqdict):
    """
    단일 파일에서 박스/프레임/SPS 정보를 추출하는 독립 함수.
    pickle 가능하도록 클래스 밖에 정의하고 필요한 값들을 인자로 전달받습니다.
    
    Args:
        file_path: 처리할 파일 경로
        state_flags: dict with keys: structure_val_state, structure_seq_state, 
                    frame_gop_state, frame_ratio_state, frame_sps_state
        seqdict: 시퀀스 딕셔너리
    """
    try:
        results = []
        filename = os.path.basename(file_path)
        results.append(('name', filename))

        # MD5 / 해시 (캐시 없이 직접 계산)
        try:
            hashf = _compute_file_hash(file_path)
            results.append(('md5', hashf))
        except Exception as e:
            # 해시 계산 오류는 조용하게 처리
            # print(f"해시 계산 오류 ({filename}): {e}")
            results.append(('md5', ''))

        # 구조 값 / 시퀀스
        if state_flags.get('structure_val_state') == True or state_flags.get('structure_seq_state') == True:
            onesequence = []

            def parse_box(f, end_position, depth=0, max_depth=300):
                if depth > max_depth:
                    print("최대 재귀 깊이 도달 에러")
                    return

                while f.tell() < end_position:
                    box_start = f.tell()
                    box_header = f.read(8)
                    if len(box_header) < 8:
                        break

                    box_size, box_type = struct.unpack(">I4s", box_header)
                    # FourCC는 ASCII 기반이므로 utf-8 대신 안정적으로 처리
                    try:
                        # ©(0xA9) 등 확장 FourCC도 있으므로 latin-1로 보존
                        box_type_str = box_type.decode("latin-1", errors="replace")
                        # 4바이트 모두 printable(또는 0xA9)인 경우만 유효 처리
                        raw = box_type
                        ok = True
                        for b in raw:
                            if b == 0xA9:
                                continue
                            if b < 0x20 or b > 0x7E:
                                ok = False
                                break
                        if not ok or len(box_type_str) != 4:
                            raise ValueError("invalid fourcc")
                        box_type = box_type_str
                    except Exception:
                        # 동기화 복구: 1바이트 이동
                        try:
                            f.seek(box_start + 1)
                        except Exception:
                            pass
                        continue

                    if box_size == 0:
                        # box가 컨테이너 끝까지 확장되는 케이스(예: mdat/unknown) 방어
                        # 컨테이너 경계(end_position)까지로 처리하고 진행
                        actual_box_size = max(0, end_position - box_start)
                        header_len = 8
                    elif box_size == 1:
                        large_size = f.read(8)
                        if len(large_size) < 8:
                            break
                        actual_box_size = struct.unpack(">Q", large_size)[0]
                        header_len = 16
                    else:
                        actual_box_size = box_size
                        header_len = 8

                    # 박스 끝 위치 계산: box_start + actual_box_size
                    try:
                        box_end_position = box_start + int(actual_box_size)
                    except Exception:
                        break
                    # 비정상 사이즈 방어
                    if actual_box_size < header_len:
                        # 사이즈가 헤더보다 작으면 동기화가 깨진 것으로 보고 1바이트 이동
                        try:
                            f.seek(box_start + 1)
                        except Exception:
                            pass
                        continue
                    # 컨테이너 경계/파일 경계 초과 방지 (중간 동기화 깨짐 복구)
                    if box_end_position > end_position:
                        # 사이즈가 비정상(컨테이너 밖)일 때는 1바이트 이동 후 재동기화 시도
                        try:
                            f.seek(box_start + 1)
                        except Exception:
                            pass
                        continue

                    # 학습 CSV와 동일하게 "컨테이너 자체도 컬럼으로 남기는" 기존 방식을 유지하기 위해
                    # 컨테이너 재귀 파싱 범위를 최소화합니다.
                    # (meta/dinf/stsd/dref 등을 컨테이너로 확장하면 학습에 있던 meta/dinf 컬럼이 사라지고,
                    #  mdta/url/mp4a 같은 하위 박스가 '초과 컬럼'으로 생겨 피처 공간이 달라집니다.)
                    if box_type in ('moov', 'trak', 'mdia', 'minf', 'stbl', 'udta', 'edts', 'moof', 'traf'):
                        parse_box(f, box_end_position, depth + 1, max_depth)
                    else:
                        if box_type == 'mdat':
                            f.seek(box_end_position)
                            continue

                        data_size = int(actual_box_size) - header_len
                        box_data = f.read(data_size) if data_size > 0 else b""

                        if state_flags.get('structure_seq_state') == True:
                            # 시퀀스 사전에 없는 박스 타입도 일반화해서 반영
                            # (예: ctts/sdtp/sgpd/sbgp 등 누락 시에도 sequence 구분 가능)
                            token = seqdict.get(box_type)
                            if token is None:
                                token = f"__UNKBOX__:{box_type}"
                            onesequence.append(str(token))

                        if box_type == 'mvhd':
                            version = box_data[0]
                            if version == 0:
                                create_time, modify_time, timescale, duration = struct.unpack(">IIII", box_data[4:20])
                            else:
                                create_time, modify_time = struct.unpack(">QQ", box_data[4:20])
                                timescale, duration = struct.unpack(">II", box_data[20:28])
                            preferred_rate = struct.unpack(">I", box_data[28:32])[0]
                            preferred_volume = struct.unpack(">H", box_data[32:34])[0]
                            next_track_id = struct.unpack(">I", box_data[96:100])[0]
                            results.append((
                                box_type,
                                f"Create Time: {create_time}, Modify Time: {modify_time}, "
                                f"Timescale: {timescale}, Duration: {duration}, "
                                f"Preferred Rate: {preferred_rate}, Preferred Volume: {preferred_volume}, "
                                f"Next Track ID: {next_track_id}"
                            ))

                        elif box_type == 'tkhd':
                            version = box_data[0]
                            if version == 0:
                                create_time, modify_time, track_id, duration = struct.unpack(">IIII", box_data[4:20])
                            else:
                                create_time, modify_time = struct.unpack(">QQ", box_data[4:20])
                                track_id, duration = struct.unpack(">II", box_data[20:28])
                            width, height = struct.unpack(">II", box_data[76:84])
                            results.append((
                                box_type,
                                f"Track ID: {track_id}, Create Time: {create_time}, Modify Time: {modify_time}, "
                                f"Duration: {duration}, Width: {width}, Height: {height}"
                            ))

                        elif box_type == 'mdhd':
                            version = box_data[0]
                            if version == 0:
                                create_time, modify_time, timescale, duration = struct.unpack(">IIII", box_data[4:20])
                            else:
                                create_time, modify_time = struct.unpack(">QQ", box_data[4:20])
                                timescale, duration = struct.unpack(">II", box_data[20:28])
                            language_code = struct.unpack(">H", box_data[20:22])[0]
                            results.append((
                                box_type,
                                f"Create Time: {create_time}, Modify Time: {modify_time}, "
                                f"Timescale: {timescale}, Duration: {duration}, Language Code: {language_code}"
                            ))

                        elif box_type == 'elst':
                            version = box_data[0]
                            entry_count = struct.unpack(">I", box_data[4:8])[0]
                            entries = []
                            offset = 8
                            for _ in range(entry_count):
                                if version == 1:
                                    segment_duration, media_time, media_rate = struct.unpack(
                                        ">QqI", box_data[offset:offset + 16]
                                    )
                                    entries.append(
                                        f"Duration: {segment_duration}, Media Time: {media_time}, Rate: {media_rate}"
                                    )
                                    offset += 16
                                else:
                                    segment_duration, media_time, media_rate = struct.unpack(
                                        ">Iii", box_data[offset:offset + 12]
                                    )
                                    entries.append(
                                        f"Duration: {segment_duration}, Media Time: {media_time}, Rate: {media_rate}"
                                    )
                                    offset += 12
                            results.append((box_type, f"Entry Count: {entry_count}, Entries: {entries}"))

                        elif box_type == 'stsd':
                            if len(box_data) >= 8:
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                        elif box_type == 'stts':
                            if len(box_data) >= 8:
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                        elif box_type == 'stsc':
                            if len(box_data) >= 8:
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                        elif box_type == 'stsz':
                            if len(box_data) >= 12:
                                sample_size = struct.unpack(">I", box_data[4:8])[0]
                                sample_count = struct.unpack(">I", box_data[8:12])[0]
                                results.append((
                                    box_type,
                                    f"Sample Size: {sample_size}, Sample Count: {sample_count}"
                                ))

                        elif box_type == 'co64':
                            if len(box_data) >= 8:
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                        # 그 외 박스: MPEG/MP4 표준에 없는 타입은 structure value에 넣지 않음 (표준 구조만 사용)
                        # sequence에는 여전히 __UNKBOX__:타입 으로 반영됨

                        f.seek(box_end_position)

            # 파일 열어서 파싱
            try:
                with open(file_path, 'rb', buffering=1024 * 1024) as f:
                    file_size = f.seek(0, 2)
                    f.seek(0)
                    parse_box(f, file_size)
            except Exception as e:
                # 박스 파싱 오류는 조용하게 처리 (extract_sps.py에서 이미 출력됨)
                # print(f"박스 파싱 오류 ({filename}): {e}")
                pass

        # sequence 값 심해시화 (중복 코드 제거)
            if state_flags.get('structure_seq_state') == 1:
                try:
                    if isinstance(onesequence, list) and len(onesequence) == 0:
                        # sequence 정보가 전혀 없을 때는 명시적 결측 토큰 사용
                        onesequence = "__SEQ_MISSING__"
                    else:
                        onesequence = Simhash(onesequence).value
                except Exception:
                    onesequence = "__SEQ_MISSING__"

                results.append(('sequence', onesequence))

        # GOP 정보 (패턴만 추출하여 크기 축소)
        if state_flags.get('frame_gop_state') == 1:
            try:
                gop_string = extractGOP(file_path)
                # 전체 문자열 대신 패턴만 추출: "IPPPPP..." -> "30,30,30..." (각 GOP의 P 개수)
                if gop_string:
                    # 'I'로 분리하여 각 GOP의 P 개수 추출
                    gop_patterns = []
                    segments = gop_string.split('I')
                    for segment in segments:
                        if segment:  # 빈 문자열이 아닌 경우
                            p_count = segment.count('P')
                            if p_count > 0:
                                gop_patterns.append(str(p_count))
                    # 패턴을 쉼표로 구분하여 저장 (예: "30,30,30")
                    gop_pattern = ','.join(gop_patterns) if gop_patterns else ''
                else:
                    gop_pattern = ''
                results.append(('GOP', gop_pattern))
            except Exception as e:
                # GOP 추출 오류는 조용하게 처리
                # print(f"GOP 추출 오류 ({filename}): {e}")
                results.append(('GOP', ''))

        # GOP 압축비
        if state_flags.get('frame_ratio_state') == 1:
            try:
                ratio = process_videos_in_folder(file_path)
                results.append(('GOP compression', ratio))
            except Exception as e:
                # GOP 압축비 계산 오류는 조용하게 처리
                # print(f"GOP 압축비 계산 오류 ({filename}): {e}")
                results.append(('GOP compression', ''))

        # SPS 정보
        if state_flags.get('frame_sps_state') == 1:
            sps_filepath = None
            try:
                parse_sps(file_path)
                file_name = os.path.basename(file_path) + ".264"
                sps_filepath = file_name
                spsresult = analyzesps(sps_filepath)
                results.append(('SPS', spsresult))
            except Exception as e:
                # SPS 추출 오류는 조용하게 처리 (너무 많은 메시지 방지)
                # print(f"SPS 추출 오류 ({filename}): {e}")
                results.append(('SPS', ''))
            finally:
                if sps_filepath and os.path.exists(sps_filepath):
                    try:
                        os.remove(sps_filepath)
                    except Exception:
                        pass

        return results
    except Exception as e:
        # 전체 함수 레벨 예외 처리 - 최소한 파일명과 해시는 반환
        print(f"파일 처리 중 치명적 오류 ({os.path.basename(file_path)}): {e}")
        return [('name', os.path.basename(file_path)), ('md5', '')]

class ProgressWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("파일 처리 진행 상황")

        self.progress_bar = QProgressBar(self)
        self.progress_bar.setGeometry(100, 50, 300, 20)

        self.label = QLabel("파일 처리 중...", self)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.progress_bar)
        self.setLayout(layout)

    def set_progress(self, value):
        self.progress_bar.setValue(value)

    def set_label_text(self, text):
        self.label.setText(text)

class TrainingProgressDialog(QDialog):
    """학습 중 상태 표시용(무한 진행바 + 단계 + 경과시간)."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("학습 진행 중")
        self.setModal(False)
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)

        self.label = QLabel("학습을 준비 중...")
        self.elapsed_label = QLabel("경과: 0초")
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)  # indeterminate

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.elapsed_label)
        layout.addWidget(self.progress_bar)
        self.setLayout(layout)

        self._start_ts = None
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)

    def start(self):
        import time
        self._start_ts = time.time()
        self._timer.start(500)
        self.show()

    def stop(self):
        self._timer.stop()
        self.close()

    def set_stage(self, text: str):
        self.label.setText(text)

    def _tick(self):
        import time
        if not self._start_ts:
            return
        sec = int(time.time() - self._start_ts)
        self.elapsed_label.setText(f"경과: {sec}초")


class TrainingWorker(QObject):
    progress = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)
    notify = pyqtSignal(str)  # show_alert 대체용

    def __init__(self, train_obj, classmode, aimodel, trainindex, csv_path, train_jobs=None):
        super().__init__()
        self.train_obj = train_obj
        self.classmode = classmode
        self.aimodel = aimodel
        self.trainindex = trainindex
        self.csv_path = csv_path
        self.train_jobs = train_jobs or []
        self.completed_jobs = []
        self.batch_errors = []
        self.batch_notify_log = []

    def run(self):
        try:
            if self.train_jobs:
                import time
                total = len(self.train_jobs)
                for pos, job in enumerate(self.train_jobs, 1):
                    train_obj = job.get("train_obj")
                    classmode = job.get("classmode", self.classmode)
                    aimodel = job.get("aimodel", "")
                    trainindex = job.get("trainindex", 0)
                    csv_path = job.get("csv_path", self.csv_path)
                    csv_name = os.path.basename(str(csv_path)) if csv_path else "(CSV 없음)"
                    selected_feature_set = job.get("selected_feature_set", "")
                    start_ts = time.time()
                    try:
                        if hasattr(train_obj, "show_alert"):
                            def _proxy(msg, _log=self.batch_notify_log):
                                _log.append(str(msg))
                            train_obj.show_alert = _proxy
                        train_obj.progress_callback = lambda s, p=pos, t=total, m=aimodel, c=csv_name: self.progress.emit(
                            f"[{p}/{t}] {m} / {c}: {str(s)}"
                        )
                        self.progress.emit(f"[{pos}/{total}] {aimodel} / {csv_name} 학습 시작...")
                        train_obj.gotrain(classmode, aimodel, trainindex, csv_path)
                        self.completed_jobs.append({
                            "status": "ok",
                            "classmode": classmode,
                            "aimodel": aimodel,
                            "trainindex": trainindex,
                            "csv_path": csv_path,
                            "selected_feature_set": selected_feature_set,
                            "training_duration": time.time() - start_ts,
                        })
                    except Exception as e:
                        tb = traceback.format_exc()
                        msg = f"{aimodel}: {e}"
                        print(f"[ERROR] 전체 모델 학습 실패: {msg}")
                        print(tb)
                        self.batch_errors.append(f"{msg}\n{tb}")
                        self.completed_jobs.append({
                            "status": "error",
                            "classmode": classmode,
                            "aimodel": aimodel,
                            "trainindex": trainindex,
                            "csv_path": csv_path,
                            "selected_feature_set": selected_feature_set,
                            "training_duration": time.time() - start_ts,
                            "error": f"{e}\n\n{tb}",
                        })
                        self.progress.emit(f"[{pos}/{total}] {aimodel} 실패, 다음 모델로 진행...")
                self.finished.emit()
                return

            # 학습 코드 내부에서 UI 팝업을 띄우지 않도록 show_alert를 메인 스레드 신호로 우회
            if hasattr(self.train_obj, "show_alert"):
                def _proxy(msg):
                    self.notify.emit(str(msg))
                self.train_obj.show_alert = _proxy

            # 진행 단계 콜백이 있으면 사용
            self.progress.emit("학습 시작...")
            try:
                self.train_obj.progress_callback = lambda s: self.progress.emit(str(s))
            except Exception:
                pass

            self.train_obj.gotrain(self.classmode, self.aimodel, self.trainindex, self.csv_path)
            self.finished.emit()
        except Exception as e:
            tb = traceback.format_exc()
            self.error.emit(f"{e}\n\n{tb}")

def copyable_message_box(parent, icon, title, text, buttons=QMessageBox.Ok, default_btn=QMessageBox.NoButton):
    """텍스트 선택/복사 가능한 QMessageBox (에러문 복사용). 모듈 레벨 함수."""
    msg = QMessageBox(parent)
    msg.setIcon(icon)
    msg.setWindowTitle(title)
    msg.setText(text)
    msg.setStandardButtons(buttons)
    msg.setDefaultButton(default_btn)
    msg.setStyleSheet("""
        QMessageBox {
            background-color: #2e2e2e;
            color: #f5f5f5;
            font: 10pt "Helvetica";
        }
        QMessageBox QLabel {
            color: #f5f5f5;
            background-color: transparent;
            font: 10pt "Helvetica";
        }
        QMessageBox QPushButton {
            background-color: #444;
            color: #ffffff;
            border: 1px solid #777;
            border-radius: 5px;
            padding: 6px 14px;
            min-width: 72px;
        }
        QMessageBox QPushButton:hover {
            background-color: #555;
        }
        QMessageBox QPushButton:pressed {
            background-color: #666;
        }
    """)
    for label in msg.findChildren(QLabel):
        label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        label.setStyleSheet("color: #f5f5f5; background-color: transparent;")
    return msg.exec_()


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

form_path = resource_path(os.path.join("UI_Design", "new.ui"))
form_class = uic.loadUiType(form_path)[0]

class createtrainclass(QMainWindow, form_class):
    def __init__(self, case_direc, dataset_direc, skip_initial_csv_load=False):
        super(createtrainclass, self).__init__()
        self.choice = 0
        self.file_paths = []
        # 초기 CSV 자동 로드를 건너뛸지 여부
        self.skip_initial_csv_load = skip_initial_csv_load
        
        # 최적화: 캐싱을 위한 딕셔너리 추가
        self._file_hash_cache = {}
        self._simhash_cache = {}

        self.setupUi(self) # UI 요소 초기화
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.clustering = trainClustering()
        self.trainclass = twoTrainClass()
        self.existval = 0
        # 확장자 필터
        self.extension_list = ["확장자", ".mp4",  ".mov"]
        self.comboBox.addItems(self.extension_list)
        self.csv_file = ''
        self.tempcsv_file = ''
        self.comboBox.currentIndexChanged.connect(
            lambda index: self.filter_files_by_extension(self.comboBox.itemText(index)))

        # 매핑 Label List는 QListWidget 대신 Train 탭의 tableWidget_train 한 곳에만 표시 (update_label_list_from_json)
        self.label_list_widget = None

        self.dirModel = QFileSystemModel()
        self.dirModel.setRootPath(QDir.rootPath())

        self.exist_csv_but.clicked.connect(self.open_existcsv)
        self.treeView.setModel(self.dirModel)

        initialcode = 0
        self.detectmode = 0
        self.detect_all_detail_folders = False

        if initialcode == 0:
            print("=====================================")
            try:
                self.case_direc = case_direc
                self.dataset_direc = dataset_direc
                # 케이스 이름이 공백 또는 비어 있는 경우
                if not self.case_direc or self.case_direc.strip() == "":
                    print("기본 default_case를 생성합니다.")
                    self.case_direc = 'default_case'
                    raise ValueError("케이스명이 공백이라 default_case로 설정")
                else:
                    # 공백이 아닌 경우
                    if os.path.exists(self.case_direc):
                        print(f"{self.case_direc} 케이스가 이미 존재합니다.")
                    else:
                        print(f"{self.case_direc} 케이스가 존재하지 않아 생성합니다.")
                        os.makedirs(self.case_direc, exist_ok=True)
                        os.makedirs(os.path.join(self.case_direc, "config"), exist_ok=True)
                        raise ValueError("경로가 존재하지 않아 생성")

                # 데이터셋 경로 유효성 검사 (존재하는 경로인지 확인)
                if not os.path.isdir(self.dataset_direc):
                    raise ValueError(f"유효하지 않은 데이터셋 경로: {self.dataset_direc}")

                print("유효한 데이터셋 경로와 케이스 경로가 설정되었습니다.")

            except Exception as e:
                # 경로가 유효하지 않으면 기본 경로를 설정
                self.dataset_direc = "y:\\"
                print(f"데이터셋 경로가 {self.dataset_direc}로 설정되었습니다.")

            print("=====================================")
            try:
                print("케이스 이름: [", self.case_direc, "]")
                print("데이터셋 경로: [", self.dataset_direc, "]")
                print("=====================================")
            except:
                self.ask_input()

        self.default_states = {
            "structure_seq_state": 0,
            "frame_gop_state": 0,
            "frame_ratio_state": 0,
            "frame_sps_state": 0,
            "structure_val_state": 0,
            "train_structure_val_state": 1,
            "train_structure_seq_state": 1,
            "train_frame_sps_state": 1,
            "train_frame_gop_state": 1,
            "train_frame_ratio_state": 1,
            "train_execution_scope": "case_all",  # case_all | single
            "train_save_outputs_state": 1,        # 1: 저장, 0: 저장 안 함
        }
        # Data View 트리 루트·CSV 자동 로드·메뉴 등 초기화
        self._init_after_paths()

    def _config_dir(self):
        """케이스 내 config 폴더 경로 반환 (설정 파일 통합). 없으면 생성."""
        case = getattr(self, "case_direc", None)
        if not case or not str(case).strip():
            return None
        d = os.path.join(case, "config")
        os.makedirs(d, exist_ok=True)
        return d

    def _states_json_base_dir(self):
        """states.json을 둘 기준 디렉터리: 현재 선택된 CSV와 같은 폴더, 없으면 케이스 루트."""
        for key in ("csv_path", "csv_file"):
            p = getattr(self, key, None)
            if not p or not str(p).strip():
                continue
            ap = os.path.abspath(os.path.normpath(str(p).strip()))
            parent = os.path.dirname(ap)
            if parent and parent != ap:
                return parent
        case = getattr(self, "case_direc", None)
        if case and str(case).strip():
            return os.path.abspath(os.path.normpath(case))
        return None

    def _states_json_path(self):
        """`<CSV와 같은 폴더>/config/states.json` (또는 CSV 없을 때만 케이스/config/states.json)."""
        base = self._states_json_base_dir()
        if not base:
            return None
        cfg = os.path.join(base, "config")
        try:
            os.makedirs(cfg, exist_ok=True)
        except OSError as e:
            print(f"[WARN] config 폴더를 만들 수 없습니다: {e}")
        return os.path.join(cfg, "states.json")

    def _legacy_states_json_path(self):
        """하위 호환: 같은 기준 폴더 루트의 states.json (config 밖)"""
        base = self._states_json_base_dir()
        if not base:
            return None
        return os.path.join(base, "states.json")

    def get_selected_feature_set_string(self):
        """Create 탭 기준 선택된 피처를 val_seq_sps_gop 형식 문자열로 반환."""
        parts = []
        if getattr(self, 'structure_val_state', 0):
            parts.append('val')
        if getattr(self, 'structure_seq_state', 0):
            parts.append('seq')
        if getattr(self, 'frame_sps_state', 0):
            parts.append('sps')
        if getattr(self, 'frame_gop_state', 0):
            parts.append('gop')
        if getattr(self, 'frame_ratio_state', 0):
            parts.append('ratio')
        return '_'.join(parts) if parts else ''

    def _layout_train_tab_target_section(self):
        """학습 대상 그룹 너비·높이를 텍스트에 맞추고, 그 아래 피처·Label list·테이블 위치를 맞춘다."""
        train_tab = getattr(self, "tab", None)
        ib = getattr(self, "train_info_box", None)
        fb = getattr(self, "train_feature_box", None)
        te_case = getattr(self, "train_case_info_text", None)
        lbl_csv = getattr(self, "train_csv_info_label", None)
        rel_blk = getattr(self, "train_csv_rel_block", None)
        if not train_tab or not ib or not fb or not te_case or not lbl_csv or not rel_blk:
            return
        tab_inner_w = max(360, train_tab.width() - 32)
        # 가로: 탭에서 여유만 두고 거의 전체 사용 (긴 경로 줄바꿈)
        inner_min = 260
        content_w = max(inner_min, tab_inner_w - 56)
        lay = ib.layout()
        m = lay.contentsMargins()
        sp = lay.spacing()
        fm = QFontMetrics(lbl_csv.font())
        path_row = 0
        te_case.setFixedWidth(content_w)
        doc = te_case.document()
        doc.setTextWidth(content_w)
        dl = doc.documentLayout()
        if dl is not None:
            try:
                dl.update()
            except Exception:
                pass
        h_raw = 0.0
        try:
            if dl is not None:
                h_raw = float(dl.documentSize().height())
            else:
                h_raw = float(doc.size().height())
        except Exception:
            h_raw = 0.0
        plain = te_case.toPlainText() or ""
        char_w = max(1.0, float(fm.horizontalAdvance("M")))
        chars_per_line = max(8.0, float(content_w) / char_w)
        lines_by_chars = max(1.0, (len(plain) + chars_per_line - 1.0) / chars_per_line)
        h_fallback = int(math.ceil(lines_by_chars * float(fm.lineSpacing()))) + 2
        if h_raw < 4.0:
            h_case = h_fallback
        else:
            h_case = max(int(math.ceil(h_raw)) + 2, h_fallback)
        h_case = max(24, h_case)
        rel_blk.setFixedWidth(content_w)
        lbl_csv.setFixedWidth(content_w)
        if hasattr(rel_blk, "document"):
            rdoc = rel_blk.document()
            rdoc.setTextWidth(content_w)
            rdl = rdoc.documentLayout()
            if rdl is not None:
                try:
                    rdl.update()
                except Exception:
                    pass
            rh_raw = 0.0
            try:
                if rdl is not None:
                    rh_raw = float(rdl.documentSize().height())
                else:
                    rh_raw = float(rdoc.size().height())
            except Exception:
                rh_raw = 0.0
            rplain = rel_blk.toPlainText() or ""
            rlines = max(1.0, (len(rplain) + chars_per_line - 1.0) / chars_per_line)
            rh_fallback = int(math.ceil(rlines * float(fm.lineSpacing()))) + 2
            if rh_raw < 4.0:
                h_rel = rh_fallback
            else:
                h_rel = max(int(math.ceil(rh_raw)) + 2, rh_fallback)
            h_rel = max(fm.height() + 2, h_rel)
            h_rel = min(h_rel, 120)
            rel_blk.setFixedHeight(h_rel)
        else:
            h_rel = max(rel_blk.heightForWidth(content_w), fm.height() + 2)
        h2 = max(lbl_csv.heightForWidth(content_w), fm.height() + 2)
        # 피처·라벨 제목·테이블 최소만 남기고 케이스 경로에 나머지 세로 공간 할당(테이블은 남는 높이 전부 사용)
        reserve_min = 118 + 10 + 24 + 20 + 72
        extra_mid = h_rel
        max_case_h = max(
            100,
            train_tab.height()
            - 208
            - reserve_min
            - m.top()
            - m.bottom()
            - h2
            - extra_mid
            - 28,
        )
        h_case = min(h_case, max_case_h)
        te_case.setFixedHeight(h_case)
        try:
            lay.invalidate()
            lay.activate()
            inner_body_h = lay.sizeHint().height()
        except Exception:
            inner_body_h = h_case + h_rel + h2
        group_chrome = 8
        box_h = max(56, inner_body_h + m.top() + m.bottom() + group_chrome)
        box_w = content_w + m.left() + m.right() + 4
        info_y = 208
        ib.setGeometry(120, info_y, box_w, box_h)
        gap = 8
        fb.setGeometry(120, ib.y() + ib.height() + gap, box_w, 118)
        ly = fb.y() + fb.height() + 10
        lbl_w = min(693, tab_inner_w - 8)
        if hasattr(self, "label_7") and self.label_7 is not None:
            self.label_7.setGeometry(120, ly, min(480, lbl_w), 20)
        tbl_y = ly + 24
        bottom_pad = 36
        remain_h = train_tab.height() - tbl_y - bottom_pad
        remain_h = max(72, remain_h)
        if hasattr(self, "tableWidget_train") and self.tableWidget_train is not None:
            self.tableWidget_train.setGeometry(120, tbl_y, lbl_w, remain_h)

    def _refresh_train_tab_case_csv_info(self):
        """Train 탭에 표시하는 케이스 경로·학습 CSV 정보 갱신."""
        try:
            if not getattr(self, 'train_case_info_text', None):
                return
            case = (getattr(self, 'case_direc', None) or '').strip()
            csvp = (getattr(self, 'csv_path', None) or '').strip()
            self.train_case_info_text.setPlainText(
                f"케이스 경로\n{case or '(없음)'}"
            )
            self.train_case_info_text.setToolTip(case if case else '')
            base = os.path.basename(csvp) if csvp else ''
            self.train_csv_info_label.setText(f"학습 CSV: {base or '(선택 전 — 분류 모델 학습 시 선택)'}")
            self.train_csv_info_label.setToolTip(os.path.abspath(csvp) if csvp else '')
            mappingp = (getattr(self, 'mapping_json_path', None) or '').strip()
            mapping_abs = os.path.abspath(mappingp) if mappingp else ''
            mapping_line = f"매핑 JSON 경로\n{(mapping_abs if mapping_abs else '(없음)')}"
            rel = _folder_path_between_case_and_csv(case, csvp) if (case and csvp) else ""
            cap_line = "CSV까지(케이스 내부 폴더)"
            if getattr(self, "train_csv_rel_block", None) is not None:
                if not csvp:
                    body = "(학습 CSV를 선택하면 표시됩니다)"
                    self.train_csv_rel_block.setToolTip("")
                elif rel == "(케이스 폴더와 동일)":
                    body = rel
                    self.train_csv_rel_block.setToolTip(os.path.dirname(os.path.abspath(csvp)))
                elif rel and os.path.isabs(rel):
                    body = f"(케이스 밖) {rel}"
                    self.train_csv_rel_block.setToolTip(rel)
                elif rel:
                    body = str(rel)
                    self.train_csv_rel_block.setToolTip(os.path.dirname(os.path.abspath(csvp)))
                else:
                    body = "(상대 경로를 계산할 수 없음)"
                    self.train_csv_rel_block.setToolTip(
                        os.path.dirname(os.path.abspath(csvp)) if csvp else ""
                    )
                block_text = f"{cap_line}\n{body}\n\n{mapping_line}"
                self.train_csv_rel_block.setPlainText(block_text)
                tip_parts = []
                try:
                    tip = self.train_csv_rel_block.toolTip()
                    if tip:
                        tip_parts.append(tip)
                except Exception:
                    pass
                if mapping_abs:
                    tip_parts.append(f"매핑 JSON: {mapping_abs}")
                self.train_csv_rel_block.setToolTip("\n".join(tip_parts))
            self._layout_train_tab_target_section()
            try:
                tc = self.train_case_info_text
                cur = tc.textCursor()
                cur.movePosition(QTextCursor.End)
                tc.setTextCursor(cur)
                tc.ensureCursorVisible()
            except Exception:
                pass
        except Exception:
            pass

    def _mark_train_feature_selection_synced(self):
        """Train 탭 피처 선택이 저장·로드된 상태와 일치함을 표시 (분류 학습 시 경고 생략)."""
        self._train_feature_selection_dirty = False

    def _on_train_feature_checkbox_changed(self, *_args):
        if getattr(self, 'train_structure_val_but', None) is None:
            return
        self._train_feature_selection_dirty = True

    def _setup_train_feature_checkboxes(self):
        """Train 탭에 학습용 피처 선택 체크박스 추가 (Create 탭과 별도)."""
        try:
            train_tab = getattr(self, 'tab', None)
            if not train_tab:
                return
            # Train 탭 본문 라벨(학습 CSV 등)과 동일: UI의 label_7은 기본 16pt Arial이라 여기서 통일
            lbl7 = getattr(self, "label_7", None)
            if lbl7 is not None:
                lbl7.setStyleSheet(
                    'color: rgb(230, 230, 230); font: 9pt "Helvetica"; margin: 0; padding: 0;'
                )
            # Train 탭 고정 위젯과 겹치지 않게; 매핑은 tableWidget_train만 사용하므로 가로 전체 사용 가능
            info_box = QGroupBox("학습 대상 (케이스 / CSV)", train_tab)
            info_box.setObjectName("trainInfoBox")
            info_box.setFlat(True)
            info_box.setStyleSheet(
                "QGroupBox#trainInfoBox { font: bold 10pt \"Helvetica\"; color: rgb(255, 255, 255); "
                "border: 1px solid #666; margin-top: 4px; padding: 0px 2px 1px 2px; }"
                "QGroupBox#trainInfoBox::title { subcontrol-origin: margin; left: 8px; padding: 0px; }"
            )
            info_box.setGeometry(120, 208, 400, 96)
            self.train_info_box = info_box
            info_layout = QVBoxLayout()
            info_layout.setContentsMargins(4, 2, 4, 2)
            info_layout.setSpacing(0)
            info_box.setLayout(info_layout)
            self.train_case_info_text = QPlainTextEdit(info_box)
            self.train_case_info_text.setReadOnly(True)
            self.train_case_info_text.setFrameShape(QFrame.NoFrame)
            self.train_case_info_text.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.train_case_info_text.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.train_case_info_text.setLineWrapMode(QPlainTextEdit.WidgetWidth)
            self.train_case_info_text.setTabChangesFocus(False)
            try:
                _doc = self.train_case_info_text.document()
                _doc.setDocumentMargin(0)
                _rf = _doc.rootFrame()
                _ff = _rf.frameFormat()
                _ff.setTopMargin(0)
                _ff.setBottomMargin(0)
                _ff.setLeftMargin(0)
                _ff.setRightMargin(0)
                _ff.setPadding(0)
                _rf.setFrameFormat(_ff)
            except Exception:
                pass
            _wrap_opt = QTextOption()
            _wrap_opt.setWrapMode(QTextOption.WrapAnywhere)
            self.train_case_info_text.document().setDefaultTextOption(_wrap_opt)
            self.train_case_info_text.setStyleSheet(
                "QPlainTextEdit { background-color: transparent; color: rgb(230, 230, 230); "
                "font: 9pt \"Helvetica\"; margin: 0; padding: 0; border: none; }"
            )
            self.train_case_info_text.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            try:
                self.train_case_info_text.setViewportMargins(0, 0, 0, 0)
            except Exception:
                pass
            lab_style = "color: rgb(230, 230, 230); font: 9pt \"Helvetica\"; margin: 0; padding: 0;"
            self.train_csv_rel_block = QPlainTextEdit(info_box)
            self.train_csv_rel_block.setReadOnly(True)
            self.train_csv_rel_block.setFrameShape(QFrame.NoFrame)
            self.train_csv_rel_block.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.train_csv_rel_block.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.train_csv_rel_block.setLineWrapMode(QPlainTextEdit.WidgetWidth)
            self.train_csv_rel_block.setPlainText(
                "CSV까지(케이스 내부 폴더)\n(학습 CSV를 선택하면 표시됩니다)"
            )
            self.train_csv_rel_block.setStyleSheet(
                "QPlainTextEdit { background-color: transparent; color: rgb(210, 210, 210); "
                "font: 9pt \"Helvetica\"; margin: 0; padding: 0; border: none; }"
            )
            self.train_csv_rel_block.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            try:
                self.train_csv_rel_block.setViewportMargins(0, 0, 0, 0)
                _rd = self.train_csv_rel_block.document()
                _rd.setDocumentMargin(0)
                _rw = QTextOption()
                _rw.setWrapMode(QTextOption.WrapAnywhere)
                _rd.setDefaultTextOption(_rw)
            except Exception:
                pass
            self.train_csv_info_label = QLabel("학습 CSV: (선택 전)", info_box)
            self.train_csv_info_label.setStyleSheet(lab_style)
            self.train_csv_info_label.setWordWrap(True)
            self.train_csv_info_label.setAlignment(Qt.AlignTop | Qt.AlignLeft)
            self.train_csv_info_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
            info_layout.addWidget(self.train_case_info_text)
            info_layout.addWidget(self.train_csv_rel_block)
            info_layout.addWidget(self.train_csv_info_label)

            box = QGroupBox("학습에 사용할 피처 (Train 탭)", train_tab)
            box.setStyleSheet("font: bold 10pt \"Helvetica\"; color: rgb(255, 255, 255);")
            box.setGeometry(120, 312, 400, 118)
            self.train_feature_box = box
            main_layout = QVBoxLayout()
            main_layout.setSpacing(10)
            main_layout.setContentsMargins(12, 14, 12, 10)
            box.setLayout(main_layout)
            cb_style = "color: rgb(255, 255, 255);"
            self.train_structure_val_but = QCheckBox("Structure value", box)
            self.train_structure_val_but.setChecked(True)
            self.train_structure_val_but.setStyleSheet(cb_style)
            self.train_structure_seq_but = QCheckBox("Structure Sequence", box)
            self.train_structure_seq_but.setChecked(True)
            self.train_structure_seq_but.setStyleSheet(cb_style)
            self.train_frame_sps_but = QCheckBox("SPS/PPS", box)
            self.train_frame_sps_but.setChecked(True)
            self.train_frame_sps_but.setStyleSheet(cb_style)
            self.train_frame_gop_but = QCheckBox("GOP frame", box)
            self.train_frame_gop_but.setChecked(True)
            self.train_frame_gop_but.setStyleSheet(cb_style)
            self.train_frame_ratio_but = QCheckBox("Encode ratio", box)
            self.train_frame_ratio_but.setChecked(True)
            self.train_frame_ratio_but.setStyleSheet(cb_style)
            # 위쪽 행: 2개
            row1 = QHBoxLayout()
            row1.setSpacing(24)
            row1.addWidget(self.train_structure_val_but)
            row1.addWidget(self.train_structure_seq_but)
            row1.addStretch()
            main_layout.addLayout(row1)
            # 아래쪽 행: 3개 + 저장 버튼
            row2 = QHBoxLayout()
            row2.setSpacing(24)
            row2.addWidget(self.train_frame_sps_but)
            row2.addWidget(self.train_frame_gop_but)
            row2.addWidget(self.train_frame_ratio_but)
            row2.addStretch()
            self.train_save_state_but = QPushButton("저장", box)
            self.train_save_state_but.setStyleSheet(
                "QPushButton { background-color: rgb(212, 175, 55); color: rgb(35, 37, 42); "
                "font: bold 10pt \"Playfair Display\"; border-radius: 6px; padding: 6px 12px; }"
                "QPushButton:hover { background-color: rgb(183, 149, 11); }"
            )
            self.train_save_state_but.setToolTip("선택한 피처 설정을 config에 저장합니다.")
            self.train_save_state_but.clicked.connect(self._save_train_feature_states)
            row2.addWidget(self.train_save_state_but)
            main_layout.addLayout(row2)
            row3 = QHBoxLayout()
            self.train_show_fi_plot_but = QCheckBox("학습 후 피처 중요도 차트(브라우저) 표시", box)
            self.train_show_fi_plot_but.setChecked(False)
            self.train_show_fi_plot_but.setStyleSheet(cb_style)
            self.train_show_fi_plot_but.setToolTip(
                "체크한 경우에만 학습 직후 Plotly 차트(HTML/브라우저)가 열립니다.\n"
                "feature_importance.csv 저장은 체크 여부와 관계없이 수행됩니다."
            )
            row3.addWidget(self.train_show_fi_plot_but)
            row3.addStretch()
            main_layout.addLayout(row3)

            # 실행 옵션: 케이스 전체/단일 + 결과 저장 여부
            row4 = QHBoxLayout()
            row4.setSpacing(10)
            opt_label = QLabel("실행 옵션", box)
            opt_label.setStyleSheet(cb_style)
            self.train_execution_scope_combo = QComboBox(box)
            self.train_execution_scope_combo.addItems(["케이스 전체 자동", "단일 CSV"])
            self.train_execution_scope_combo.setToolTip(
                "케이스 전체 자동: training_results 없는 폴더의 CSV를 자동 일괄 학습\n"
                "단일 CSV: 현재 선택된 CSV 1개만 학습"
            )
            self.train_save_outputs_but = QCheckBox("결과 저장", box)
            self.train_save_outputs_but.setChecked(True)
            self.train_save_outputs_but.setStyleSheet(cb_style)
            self.train_save_outputs_but.setToolTip(
                "해제하면 model/scaler, training_results, training_history 기록을 저장하지 않습니다."
            )
            row4.addWidget(opt_label)
            row4.addWidget(self.train_execution_scope_combo, 1)
            row4.addWidget(self.train_save_outputs_but)
            main_layout.addLayout(row4)

            for _cb in (
                self.train_structure_val_but,
                self.train_structure_seq_but,
                self.train_frame_sps_but,
                self.train_frame_gop_but,
                self.train_frame_ratio_but,
            ):
                _cb.stateChanged.connect(self._on_train_feature_checkbox_changed)

            self._refresh_train_tab_case_csv_info()
        except Exception as e:
            print(f"[WARN] Train 탭 피처 체크박스 생성 실패: {e}")
            self.train_structure_val_but = getattr(self, 'train_structure_val_but', None)
            self.train_structure_seq_but = getattr(self, 'train_structure_seq_but', None)
            self.train_frame_sps_but = getattr(self, 'train_frame_sps_but', None)
            self.train_frame_gop_but = getattr(self, 'train_frame_gop_but', None)
            self.train_frame_ratio_but = getattr(self, 'train_frame_ratio_but', None)
            self.train_show_fi_plot_but = getattr(self, 'train_show_fi_plot_but', None)
            self.train_execution_scope_combo = getattr(self, 'train_execution_scope_combo', None)
            self.train_save_outputs_but = getattr(self, 'train_save_outputs_but', None)
            self.train_case_info_text = getattr(self, 'train_case_info_text', None)
            self.train_csv_info_label = getattr(self, 'train_csv_info_label', None)
            self.train_csv_rel_block = getattr(self, 'train_csv_rel_block', None)
            self.train_info_box = getattr(self, 'train_info_box', None)
            self.train_feature_box = getattr(self, 'train_feature_box', None)

    def eventFilter(self, obj, event):
        if obj is getattr(self, "tab", None) and event.type() == QEvent.Resize:
            try:
                self._layout_train_tab_target_section()
            except Exception:
                pass
        return super(createtrainclass, self).eventFilter(obj, event)

    def _save_train_feature_states(self):
        """Train 탭 피처 선택을 states에 반영 후 config에 저장 (저장 버튼 클릭 시)."""
        try:
            if not hasattr(self, "states"):
                return
            if getattr(self, 'train_structure_val_but', None) is None:
                return
            self.states["train_structure_val_state"] = 1 if self.train_structure_val_but.isChecked() else 0
            self.states["train_structure_seq_state"] = 1 if self.train_structure_seq_but.isChecked() else 0
            self.states["train_frame_sps_state"] = 1 if self.train_frame_sps_but.isChecked() else 0
            self.states["train_frame_gop_state"] = 1 if self.train_frame_gop_but.isChecked() else 0
            self.states["train_frame_ratio_state"] = 1 if self.train_frame_ratio_but.isChecked() else 0
            if getattr(self, "train_execution_scope_combo", None) is not None:
                self.states["train_execution_scope"] = "single" if self.train_execution_scope_combo.currentIndex() == 1 else "case_all"
            if getattr(self, "train_save_outputs_but", None) is not None:
                self.states["train_save_outputs_state"] = 1 if self.train_save_outputs_but.isChecked() else 0
            self.save_states()
            self._mark_train_feature_selection_synced()
            self.statusBar().showMessage("Train 피처/옵션 설정이 저장되었습니다.", 3000)
        except Exception as e:
            print(f"[WARN] Train 피처 설정 저장 실패: {e}")

    def get_train_selected_feature_set_string(self):
        """Train 탭에서 선택된 피처만 val_seq_sps_gop 형식으로 반환. 학습 시 이 값 사용."""
        if not getattr(self, 'train_structure_val_but', None):
            return self.get_selected_feature_set_string()
        parts = []
        if self.train_structure_val_but.isChecked():
            parts.append('val')
        if self.train_structure_seq_but.isChecked():
            parts.append('seq')
        if self.train_frame_sps_but.isChecked():
            parts.append('sps')
        if self.train_frame_gop_but.isChecked():
            parts.append('gop')
        if self.train_frame_ratio_but.isChecked():
            parts.append('ratio')
        return '_'.join(parts) if parts else ''

    def _init_after_paths(self):
        """경로 설정 직후 Data View·트리·CSV 자동 로드 등 초기화 (원래 __init__에 있던 블록)."""
        self.treeView.setRootIndex(self.dirModel.index(self.dataset_direc))
        self.treeView.clicked.connect(self.file_selected)
        # 케이스 디렉토리에서 .csv 찾아서 csv_files 리스트로 반환
        all_files = glob.glob(os.path.join(self.case_direc, "*.csv"))
        print("All CSV files:", all_files)
        # _train_이 포함된 파일만 필터링 (feature_importance.csv, _processed 제외)
        csv_files = [file for file in all_files 
                     if '_train_' in os.path.basename(file) 
                     and 'feature_importance.csv' not in os.path.basename(file)
                     and '_processed' not in os.path.basename(file)]
        # 최신 mtime 순으로 정렬해 가장 최근 CSV를 우선 사용
        csv_files = sorted(csv_files, key=lambda p: os.path.getmtime(p), reverse=True)
        self.csv_path = ""

        # .csv 파일이 하나 이상 있을 때 일단은 첫 번째 파일을 열기
        if csv_files:
            if not getattr(self, "skip_initial_csv_load", False):
                self.csv_path = csv_files[0]  # 첫 번째 _train_ CSV (states.json 위치는 이 경로 폴더 기준)
                # 피처 옵션은 체크박스·states.json만 사용 (CSV 파일명으로 덮어쓰지 않음)
                # 매핑 JSON 자동 찾기
                mapping_json_path = self.find_mapping_json(self.csv_path)
                if mapping_json_path:
                    self.set_mapping_json_path(mapping_json_path)
                    print(f"[INFO] 매핑 JSON 파일 자동 로드: {os.path.basename(mapping_json_path)}")
                else:
                    self.set_mapping_json_path(None)
                self.open_csv2(self.csv_path, self.tableWidget)
                try:
                    if hasattr(self, "csv_info_label"):
                        self.csv_info_label.setVisible(True)
                except Exception:
                    pass

        try :
            file_name = os.path.basename(self.csv_path)
            self.csvlabel.setText(file_name)
            self.caselabel.setText(self.case_direc[-20:])
        except Exception as e:
            pass
        try:
            self._refresh_train_tab_case_csv_info()
        except Exception:
            pass
        # Data View 탭 하단 CSV 정보 라벨 초기화
        try:
            if hasattr(self, "csv_info_label"):
                self.csv_info_label.setText("CSV 정보를 표시합니다.")
        except Exception:
            pass
        
        # 전체 보기 체크박스 (Data View 탭 tab1 안, 테이블 위)
        try:
            data_tab = getattr(self, "tab1", None)
            _cb_parent = data_tab if data_tab is not None else self
            self.show_all_checkbox = QCheckBox("전체 보기", _cb_parent)
            self.show_all_checkbox.setGeometry(520, 278, 140, 24)
            self.show_all_checkbox.setStyleSheet(
                "QCheckBox {color: white; font: 10pt 'Helvetica';}"
                "QCheckBox::indicator {width: 18px; height: 18px; border: 2px solid #888; border-radius: 3px; background-color: #2e2e2e;}"
                "QCheckBox::indicator:checked {background-color: #4a9eff; border-color: #4a9eff;}"
            )
            self.show_all_checkbox.stateChanged.connect(self.on_show_all_changed)
            self.current_csv_path = None  # 현재 열린 CSV 경로 저장용
        except Exception as e:
            print(f"[WARN] 전체 보기 체크박스 생성 실패: {e}")
            self.show_all_checkbox = None
        
        # 헤더 설정
        header = self.treeView.header()
        header.setSectionResizeMode(0, header.Interactive)
        header.resizeSection(0, 400)
        self.create_value2.clicked.connect(lambda: setattr(self, 'choice', 2))

        self.structure_val_state = False
        self.structure_seq_state = False
        self.frame_sps_state = False
        self.frame_gop_state = False
        self.frame_ratio_state = False

        self.tabWidget.setCurrentIndex(0)

        self.structure_val_but.stateChanged.connect(self.on_structure_val_changed)
        self.structure_seq_but.stateChanged.connect(self.on_structure_seq_changed)
        self.frame_sps_but.stateChanged.connect(self.on_frame_sps_changed)
        self.frame_gop_but.stateChanged.connect(self.on_frame_gop_changed)
        self.frame_ratio_but.stateChanged.connect(self.on_frame_ratio_changed)
        self.stateButton.clicked.connect(self.save_states)

        self.LoadButton.clicked.connect(self.main) # Load 버튼 클릭 시 self.main() 호출
        self.cluster_train.clicked.connect(self.clustermain)
        self.class_train.clicked.connect(self.classmain)

        # Train 탭 전용 피처 선택 (Create 탭과 별도)
        self._train_feature_selection_dirty = False
        self._setup_train_feature_checkboxes()
        # setupUi 직후 load에서는 Train 탭 체크박스가 없었으므로, 생성 후 states.json 다시 반영
        self.load_or_initialize_states()
        self._refresh_train_tab_case_csv_info()
        try:
            if getattr(self, "tab", None) is not None:
                self.tab.installEventFilter(self)
        except Exception:
            pass

        self.class_detect.clicked.connect(self.load_file_for_prediction)
        self._setup_ground_truth_compare_button()
        # 파일 목록에서 아이템을 더블 클릭할 때 호출되는 슬롯을 연결합니다.
        self.listWidget.itemDoubleClicked.connect(self.remove_selected_file)
        self.list_del.clicked.connect(self.remove_all_file)

        self.label_info.clicked.connect(self.open_data_entry_window)
        if self.binButton.isChecked():
            self.label_datacsv = 'labeldata_bin.csv'
        elif self.mulButton.isChecked():
            self.label_datacsv = 'labeldata_mul.csv'
        self.labelinfofile = ""

        try :
            pass
            self.load_excel_data()
        except:
            pass
        self.label_input_but.clicked.connect(self.input_label)
        self.model_combo.currentTextChanged.connect(self.on_modelcombobox_select)
        self._ensure_train_all_models_option()
        self._ensure_detect_all_models_option()
        self.trainindex = self.comboBox.currentIndex()
        
        # 메뉴바에 데이터셋 경로 설정 및 CSV 파일 선택 기능 추가
        try:
            menu_bar = self.menuBar()
            if menu_bar:
                file_menu = None
                # 기존 File 메뉴 찾기
                for action in menu_bar.actions():
                    if action.menu() and action.menu().title() == "File":
                        file_menu = action.menu()
                        break
                
                # File 메뉴가 없으면 생성
                if not file_menu:
                    file_menu = menu_bar.addMenu("File")
                
                # 케이스 전환 (다른 Cases / 임의 폴더)
                case_switch_action = QAction("케이스 변경…", self)
                case_switch_action.triggered.connect(self.switch_case_from_menu)
                file_menu.addAction(case_switch_action)
                
                # 데이터셋 경로 설정 액션 추가
                dataset_dir_action = QAction("데이터셋 경로 설정", self)
                dataset_dir_action.triggered.connect(self.change_dataset_directory)
                file_menu.addAction(dataset_dir_action)
                
                # 구분선 추가
                file_menu.addSeparator()
                
                # CSV 파일 선택 액션 추가 (Train/Detect 공통)
                csv_select_action = QAction("CSV 파일 선택 (Train/Detect)", self)
                csv_select_action.triggered.connect(self.select_csv_file_from_menu)
                file_menu.addAction(csv_select_action)

                train_all_models_action = QAction("전체 모델 학습 실행 (Train)", self)
                train_all_models_action.triggered.connect(self.train_all_models_from_menu)
                file_menu.addAction(train_all_models_action)

                train_feature_summary_action = QAction("피처 조합 성능 요약 (케이스)", self)
                train_feature_summary_action.triggered.connect(self.show_case_feature_combo_summary)
                file_menu.addAction(train_feature_summary_action)
                
                # 구분선 추가
                file_menu.addSeparator()
                
                # Model/Scaler 선택 액션 추가 (Detect용)
                model_scaler_action = QAction("Model/Scaler 선택 (Detect)", self)
                model_scaler_action.triggered.connect(self.select_model_scaler_from_menu)
                file_menu.addAction(model_scaler_action)

                detect_all_models_action = QAction("전체 상세폴더 Detect 실행", self)
                detect_all_models_action.triggered.connect(self.load_file_for_prediction_all_models)
                file_menu.addAction(detect_all_models_action)

                gt_compare_action = QAction("정답지 비교 (VISION)…", self)
                gt_compare_action.triggered.connect(self.compare_detect_with_vision_ground_truth)
                file_menu.addAction(gt_compare_action)

                # 모드 변경 액션 추가 (Train/Detect 전환)
                mode_change_action = QAction("모드 변경 (Train/Detect)", self)
                mode_change_action.triggered.connect(self.prompt_mode_change)
                file_menu.addAction(mode_change_action)
        except Exception as e:
            print(f"메뉴 추가 중 오류 (무시하고 계속): {e}")
    
    # 상태값을 불러오거나 초기화하는 함수
    def load_or_initialize_states(self):
        # `<선택 CSV 폴더>/config/states.json` 우선, 없으면 같은 폴더 루트의 states.json (하위 호환).
        primary = self._states_json_path()
        legacy = self._legacy_states_json_path()
        statepath = None
        if primary and os.path.isfile(primary):
            statepath = primary
        elif legacy and os.path.isfile(legacy):
            statepath = legacy
        if statepath:
            with open(statepath, "r", encoding="utf-8") as file:
                self.states = json.load(file)
            for k, v in self.default_states.items():
                if k not in self.states:
                    self.states[k] = v
            print("기존 상태값을 불러왔습니다:", self.states)
        else:
            self.states = self.default_states.copy()
            if primary:
                try:
                    with open(primary, "w", encoding="utf-8") as file:
                        json.dump(self.states, file, indent=4, ensure_ascii=False)
                    print("파일이 없어 기본 상태값으로 초기화하고 저장했습니다:", primary)
                except OSError as e:
                    print(f"[WARN] states.json 최초 저장 실패: {e}")
            else:
                print("파일이 없어 기본 상태값으로 초기화했습니다 (case_direc 없음, 저장 생략):", self.states)


        try:
            self.structure_seq_state = self.states["structure_seq_state"]
            self.frame_gop_state = self.states["frame_gop_state"]
            self.frame_ratio_state = self.states["frame_ratio_state"]
            self.frame_sps_state = self.states["frame_sps_state"]
            self.structure_val_state = self.states["structure_val_state"]
        except Exception as e:
            pass
        # Create 탭 체크박스를 states.json과 동기화 (파일명과 무관)
        try:
            if getattr(self, "structure_val_but", None) is not None:
                _ccb = (
                    self.structure_val_but,
                    self.structure_seq_but,
                    self.frame_sps_but,
                    self.frame_gop_but,
                    self.frame_ratio_but,
                )
                for _c in _ccb:
                    _c.blockSignals(True)
                self.structure_val_but.setChecked(bool(self.states.get("structure_val_state", 0)))
                self.structure_seq_but.setChecked(bool(self.states.get("structure_seq_state", 0)))
                self.frame_sps_but.setChecked(bool(self.states.get("frame_sps_state", 0)))
                self.frame_gop_but.setChecked(bool(self.states.get("frame_gop_state", 0)))
                self.frame_ratio_but.setChecked(bool(self.states.get("frame_ratio_state", 0)))
                for _c in _ccb:
                    _c.blockSignals(False)
        except Exception:
            pass
        # Train 탭 전용 피처 선택 복원
        try:
            if getattr(self, 'train_structure_val_but', None) is not None:
                _tcb = (
                    self.train_structure_val_but,
                    self.train_structure_seq_but,
                    self.train_frame_sps_but,
                    self.train_frame_gop_but,
                    self.train_frame_ratio_but,
                )
                for _c in _tcb:
                    _c.blockSignals(True)
                self.train_structure_val_but.setChecked(bool(self.states.get("train_structure_val_state", 1)))
                self.train_structure_seq_but.setChecked(bool(self.states.get("train_structure_seq_state", 1)))
                self.train_frame_sps_but.setChecked(bool(self.states.get("train_frame_sps_state", 1)))
                self.train_frame_gop_but.setChecked(bool(self.states.get("train_frame_gop_state", 1)))
                self.train_frame_ratio_but.setChecked(bool(self.states.get("train_frame_ratio_state", 1)))
                for _c in _tcb:
                    _c.blockSignals(False)
                if getattr(self, "train_execution_scope_combo", None) is not None:
                    scope = str(self.states.get("train_execution_scope", "case_all")).strip()
                    self.train_execution_scope_combo.setCurrentIndex(1 if scope == "single" else 0)
                if getattr(self, "train_save_outputs_but", None) is not None:
                    self.train_save_outputs_but.setChecked(bool(self.states.get("train_save_outputs_state", 1)))
                self._mark_train_feature_selection_synced()
        except Exception as e:
            pass
    
    def prompt_mode_change(self):
        """수동으로 모드를 변경 (Create / Train / Detect)"""
        try:
            mode_items = ["Create (데이터 생성)", "Train (학습)", "Detect"]
            mode_selected, ok = QInputDialog.getItem(
                self,
                "모드 변경",
                "모드를 선택하세요:",
                mode_items,
                0,
                False
            )
            if not ok:
                return

            if mode_selected.startswith("Create"):
                self.detectmode = 0
                self.choice = 2
                self.existval = 0
                # 새로 만들기 시 초기 CSV 자동 로드 방지
                self.skip_initial_csv_load = True

                create_mode_items = ["새로 만들기", "기존 데이터에 추가"]
                create_mode_selected, ok = QInputDialog.getItem(
                    self,
                    "Create 모드 선택",
                    "새로 만들기 또는 기존 데이터에 추가를 선택하세요:",
                    create_mode_items,
                    0,
                    False
                )
                if not ok:
                    return

                if create_mode_selected == "새로 만들기":
                    self.csv_file = ''
                    try:
                        self.ensure_new_csv_path()
                    except Exception as e:
                        print(f"[WARN] 신규 CSV 경로 설정 실패(무시): {e}")
                    print("[INFO] 모드 변경: Create > 새로 만들기")
                elif create_mode_selected == "기존 데이터에 추가":
                    csv_file_path = (
                        getattr(self, "csv_file", "")
                        or getattr(self, "csv_path", "")
                    )
                    if not csv_file_path:
                        csv_file_path, _ = QFileDialog.getOpenFileName(
                            self,
                            "기존 CSV 파일 선택",
                            self.case_direc,
                            "CSV Files (*.csv);;All Files (*)"
                        )
                        if not csv_file_path:
                            try:
                                self.show_alert("CSV 파일이 선택되지 않았습니다.")
                            except Exception:
                                pass
                            self.choice = 0
                            return

                    self.csv_file = csv_file_path
                    self.csv_path = csv_file_path
                    self.existval = 1
                    self.skip_initial_csv_load = False

                    try:
                        mapping_json_path = self.find_mapping_json(csv_file_path)
                        self.set_mapping_json_path(mapping_json_path)
                    except Exception as e:
                        print(f"[WARN] 매핑 JSON 설정 실패(무시): {e}")
                    try:
                        self.open_csv2(csv_file_path, self.tableWidget)
                        self.current_csv_path = csv_file_path
                    except Exception as e:
                        print(f"[WARN] CSV 표시 실패(무시): {e}")
                    try:
                        if hasattr(self, 'csvlabel'):
                            self.csvlabel.setText(os.path.basename(csv_file_path))
                    except Exception:
                        pass
                    try:
                        self.load_or_initialize_states()
                    except Exception:
                        pass
                    print("[INFO] 모드 변경: Create > 기존 데이터에 추가")

            elif mode_selected.startswith("Train"):
                # Train 모드: 기존 CSV 선택 필수
                self.detectmode = 0
                self.choice = 0
                self.existval = 1
                self.skip_initial_csv_load = False

                csv_file_path, _ = QFileDialog.getOpenFileName(
                    self,
                    "학습용 CSV 파일 선택",
                    self.case_direc,
                    "CSV Files (*.csv);;All Files (*)"
                )
                if not csv_file_path:
                    try:
                        self.show_alert("CSV 파일이 선택되지 않았습니다.")
                    except Exception:
                        pass
                    self.choice = 0
                    return

                self.csv_file = csv_file_path
                self.csv_path = csv_file_path

                try:
                    mapping_json_path = self.find_mapping_json(csv_file_path)
                    self.set_mapping_json_path(mapping_json_path)
                except Exception as e:
                    print(f"[WARN] 매핑 JSON 설정 실패(무시): {e}")
                try:
                    self.open_csv2(csv_file_path, self.tableWidget)
                    self.current_csv_path = csv_file_path
                except Exception as e:
                    print(f"[WARN] CSV 표시 실패(무시): {e}")
                try:
                    if hasattr(self, 'csvlabel'):
                        self.csvlabel.setText(os.path.basename(csv_file_path))
                except Exception:
                    pass
                try:
                    self.load_or_initialize_states()
                except Exception:
                    pass
                print("[INFO] 모드 변경: Train (학습)")

            else:
                # Detect 모드
                self.detectmode = 1
                self.choice = 0
                self.skip_initial_csv_load = False
                print("[INFO] 모드 변경: Detect (예측)")
        except Exception as e:
            print(f"[WARN] 모드 변경 중 오류: {e}")

    def set_state(self, state_name, value):
        self.states[state_name] = value

    def _train_all_models_label(self):
        return "전체 모델"

    def _ensure_train_all_models_option(self):
        """Train 모델 콤보에 전체 학습 옵션을 추가 (기본 선택은 유지)."""
        try:
            combo = getattr(self, "model_combo", None)
            if combo is None:
                return
            prev_idx = combo.currentIndex()
            prev_text = combo.currentText().strip() if prev_idx >= 0 else ""
            label = self._train_all_models_label()
            if combo.findText(label) < 0:
                combo.addItem(label)
            # 기존 선택을 유지하고, 선택이 비어 있을 때만 첫 모델로 맞춘다.
            if prev_text:
                keep_idx = combo.findText(prev_text)
                if keep_idx >= 0:
                    combo.setCurrentIndex(keep_idx)
                    return
            if combo.currentIndex() < 0 and combo.count() > 0:
                combo.setCurrentIndex(0)
        except Exception as e:
            print(f"[WARN] Train 전체 모델 옵션 추가 실패: {e}")

    def _train_run_all_models_selected(self):
        try:
            return self.model_combo.currentText().strip() == self._train_all_models_label()
        except Exception:
            return False

    def _train_model_items(self):
        """Train 전체 학습에 사용할 실제 모델 항목(index, name)을 반환."""
        items = []
        seen = set()
        excluded = {"knn", "k-nearest neighbors", "knearestneighbors", "lstm"}
        try:
            combo = self.model_combo
            for idx in range(combo.count()):
                name = str(combo.itemText(idx)).strip()
                if not name or name == self._train_all_models_label():
                    continue
                key = name.lower()
                key_compact = re.sub(r"[\s_\-]+", "", key)
                if key in excluded or key_compact in excluded:
                    continue
                if key in seen:
                    continue
                seen.add(key)
                items.append((idx, name))
        except Exception as e:
            print(f"[WARN] Train 모델 목록 수집 실패: {e}")
        return items

    def _new_train_obj_for_current_mode(self):
        if getattr(self, "classmode", "") == "bin_":
            return twoTrainClass()
        return TrainClass()

    def _begin_batch_run(self):
        """전체 모델 실행 중 개별 팝업을 막고 메시지를 모아 둔다."""
        self._batch_run_quiet = True
        self._batch_run_log = []

    def _end_batch_run(self):
        """전체 모델 실행 종료 후 수집된 메시지를 반환."""
        log = list(getattr(self, "_batch_run_log", []) or [])
        self._batch_run_quiet = False
        self._batch_run_log = []
        return log

    def _feature_set_from_folder_name(self, csv_path):
        """
        CSV 상위 폴더명(및 케이스 기준 상대 경로)에서 피처 토큰을 추출해
        val_seq_sps_gop_ratio 형식 문자열로 반환.
        """
        tokens_in_order = ("val", "seq", "sps", "gop", "ratio")
        text_parts = []
        try:
            if csv_path:
                csv_dir = os.path.dirname(os.path.abspath(csv_path))
                text_parts.append(os.path.basename(csv_dir))
                case_dir = getattr(self, "case_direc", None) or ""
                if case_dir and os.path.isdir(case_dir):
                    try:
                        rel_dir = os.path.relpath(csv_dir, os.path.abspath(case_dir))
                        if rel_dir and rel_dir != ".":
                            text_parts.append(rel_dir)
                    except Exception:
                        pass
        except Exception:
            pass

        raw = " ".join(text_parts).lower()
        found = [tok for tok in tokens_in_order if re.search(rf"(?<![a-z0-9]){tok}(?![a-z0-9])", raw)]
        return "_".join(found)

    def _resolve_train_mapping_json_path(self, selected_csv_path=None):
        """Train에서 공통으로 사용할 단일 매핑 JSON 경로를 결정."""
        try:
            cur = (getattr(self, "mapping_json_path", "") or "").strip()
            if cur and os.path.exists(cur):
                return cur
        except Exception:
            pass

        try:
            case_dir = (getattr(self, "case_direc", "") or "").strip()
            if case_dir:
                case_map = os.path.join(os.path.abspath(case_dir), "config", "label_mapping.json")
                if os.path.exists(case_map):
                    return case_map
        except Exception:
            pass

        try:
            if selected_csv_path:
                return self.find_mapping_json(selected_csv_path)
        except Exception:
            pass
        return None

    def _configure_train_obj(self, train_obj, csv_path, selected_feature_set=None, save_outputs=True, mapping_json_path=None):
        train_obj.csv_path = csv_path
        train_obj.comboBox = getattr(self, "model_combo_2", None)
        train_obj.case_direc = getattr(self, 'case_direc', None) or ''
        train_obj.save_training_outputs = bool(save_outputs)
        folder_feature_set = self._feature_set_from_folder_name(csv_path)
        if selected_feature_set is None:
            selected_feature_set = folder_feature_set or self.get_train_selected_feature_set_string()
        train_obj.selected_feature_set = selected_feature_set
        print(
            f"[INFO] Train 피처셋 적용: csv={os.path.basename(str(csv_path) or '')}, "
            f"folder={folder_feature_set or '(없음)'}, applied={selected_feature_set or '(전체/기본)'}"
        )
        train_obj.show_feature_importance_plot = bool(
            getattr(self, "train_show_fi_plot_but", None) and self.train_show_fi_plot_but.isChecked()
        )
        mapping_path = (mapping_json_path or "").strip() if mapping_json_path else None
        if mapping_path:
            train_obj.mapping_json_path = mapping_path
            print(f"[INFO] Train 공통 매핑 JSON 경로 전달: {os.path.basename(mapping_path)}")
        else:
            train_obj.mapping_json_path = None
        return train_obj

    def _train_execution_scope_mode(self):
        """Train 실행 범위 모드: case_all | single"""
        try:
            combo = getattr(self, "train_execution_scope_combo", None)
            if combo is not None:
                return "single" if combo.currentIndex() == 1 else "case_all"
        except Exception:
            pass
        scope = str(getattr(self, "states", {}).get("train_execution_scope", "case_all")).strip()
        return "single" if scope == "single" else "case_all"

    def _train_save_outputs_enabled(self):
        """Train 결과 저장 여부."""
        try:
            cb = getattr(self, "train_save_outputs_but", None)
            if cb is not None:
                return bool(cb.isChecked())
        except Exception:
            pass
        try:
            return bool(int(getattr(self, "states", {}).get("train_save_outputs_state", 1) or 0))
        except Exception:
            return True

    def _collect_train_targets_by_mode(self):
        """실행 옵션에 따라 학습 대상 CSV 경로 목록을 반환."""
        mode = self._train_execution_scope_mode()
        if mode == "single":
            current_csv = (getattr(self, "csv_path", "") or "").strip()
            if current_csv and os.path.isfile(current_csv):
                return [os.path.abspath(current_csv)]
            targets = self._collect_untrained_case_csv_targets()
            return targets[:1]
        return self._collect_untrained_case_csv_targets()

    def _detect_all_models_label(self):
        return "전체 상세폴더"

    def _ensure_detect_all_models_option(self):
        """Detect 모델 콤보에 전체 실행 옵션을 추가."""
        try:
            combo = getattr(self, "model_combo_2", None)
            if combo is None:
                return
            label = self._detect_all_models_label()
            if combo.findText(label) < 0:
                combo.addItem(label)
            idx = combo.findText(label)
            if idx >= 0 and combo.currentIndex() < 0:
                combo.setCurrentIndex(idx)
        except Exception as e:
            print(f"[WARN] Detect 전체 모델 옵션 추가 실패: {e}")

    def _detect_run_all_models_selected(self):
        try:
            if getattr(self, "detect_all_detail_folders", False):
                return True
            return self.model_combo_2.currentText().strip() == self._detect_all_models_label()
        except Exception:
            return bool(getattr(self, "detect_all_detail_folders", False))

    def save_states(self):
        # Create·Train 탭 체크 상태를 states에 반영 후 저장 (파일명과 무관)
        try:
            if getattr(self, "structure_val_but", None) is not None:
                self.states["structure_val_state"] = 1 if self.structure_val_but.isChecked() else 0
                self.states["structure_seq_state"] = 1 if self.structure_seq_but.isChecked() else 0
                self.states["frame_sps_state"] = 1 if self.frame_sps_but.isChecked() else 0
                self.states["frame_gop_state"] = 1 if self.frame_gop_but.isChecked() else 0
                self.states["frame_ratio_state"] = 1 if self.frame_ratio_but.isChecked() else 0
        except Exception:
            pass
        try:
            if getattr(self, 'train_structure_val_but', None) is not None:
                self.states["train_structure_val_state"] = 1 if self.train_structure_val_but.isChecked() else 0
                self.states["train_structure_seq_state"] = 1 if self.train_structure_seq_but.isChecked() else 0
                self.states["train_frame_sps_state"] = 1 if self.train_frame_sps_but.isChecked() else 0
                self.states["train_frame_gop_state"] = 1 if self.train_frame_gop_but.isChecked() else 0
                self.states["train_frame_ratio_state"] = 1 if self.train_frame_ratio_but.isChecked() else 0
                if getattr(self, "train_execution_scope_combo", None) is not None:
                    self.states["train_execution_scope"] = "single" if self.train_execution_scope_combo.currentIndex() == 1 else "case_all"
                if getattr(self, "train_save_outputs_but", None) is not None:
                    self.states["train_save_outputs_state"] = 1 if self.train_save_outputs_but.isChecked() else 0
        except Exception:
            pass
        statepath = self._states_json_path()
        if not statepath:
            print("[WARN] states.json 저장 생략: CSV·케이스 경로가 없습니다.")
            return
        try:
            with open(statepath, "w", encoding="utf-8") as file:
                json.dump(self.states, file, indent=4, ensure_ascii=False)
            print(f"상태값이 JSON 파일에 저장되었습니다: {statepath}")
        except OSError as e:
            print(f"[ERROR] states.json 저장 실패: {e}")

    def on_combobox_select(self, index):
        self.trainclass.index = index
    def on_modelcombobox_select(self):
        self.aimodel = self.model_combo.currentText()

    def train_all_models_from_menu(self):
        """File 메뉴에서 Train 전체 모델 학습을 실행."""
        try:
            idx = self.model_combo.findText(self._train_all_models_label())
            if idx >= 0:
                self.model_combo.setCurrentIndex(idx)
        except Exception:
            pass
        self.classmain()

    def clustermain(self):
        binstat = self.binButton_3.isChecked()
        mulstat = self.mulButton_3.isChecked()
        if binstat:
            cluster_mode = 'bin_'
        elif mulstat:
            cluster_mode = 'mul_'
        else:
            try:
                self.show_alert("바이너리/멀티 모드를 선택하세요.")
            except Exception:
                messagebox.showerror("에러", "바이너리/멀티 모드를 선택")
            return
        if not getattr(self, "csv_path", ""):
            self.show_alert("클러스터링할 CSV 파일을 먼저 선택하세요.")
            return

        output_paths = self.clustering.gotrain(self.csv_path, cluster_mode)
        output_dir = ""
        if isinstance(output_paths, dict):
            output_dir = output_paths.get("output_dir", "")
        if output_dir:
            msg = f"클러스터링 결과 저장 완료: {output_dir}"
            try:
                self.statusBar().showMessage(msg, 5000)
            except Exception:
                pass
            self.show_alert(msg)

    def showFileDialog(self):
        # 파일 다이얼로그를 띄워서 파일 선택
        self.csv_path, _ = QFileDialog.getOpenFileName(self, '파일 선택', '',
                                                   '모든 파일 (*);;텍스트 파일 (*.csv)')

        # 선택한 파일 경로를 라벨에 표시
        if self.csv_path:
            print(f'선택된 파일 경로: {self.csv_path}')
        else:
            print('파일이 선택되지 않았습니다.')

    def classmain(self):
        # 버튼 클릭 즉시 피드백(학습이 오래 걸릴 수 있어 "눌림" 확인용)
        try:
            self.statusBar().showMessage("학습 시작... (완료 시 정확도 팝업이 표시됩니다)")
        except Exception:
            pass
        binstat = self.binButton_3.isChecked()
        mulstat = self.mulButton_3.isChecked()
        if binstat:
            self.trainclass = twoTrainClass() ##### 이진으로 설정
            self.classmode = 'bin_'
        elif mulstat:
            self.trainclass = TrainClass() ##### 다중으로 설정
            self.classmode = 'mul_'
        else :
            try:
                self.show_alert("바이너리/멀티 모드를 선택하세요.")
            except Exception:
                messagebox.showerror("에러", "바이너리/멀티 모드를 선택")
            return
        if getattr(self, '_train_feature_selection_dirty', False) and getattr(self, 'train_structure_val_but', None) is not None:
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Train 피처 설정 미저장")
            msg.setText(
                "학습에 사용할 피처(Train 탭 체크박스)가 변경되었습니다.\n"
                "config에 반영하려면 Train 탭의 [저장]을 누른 뒤 학습하는 것이 좋습니다.\n\n"
                "「저장 후 학습」을 누르면 지금 체크 상태를 저장하고 학습을 진행합니다.\n"
                "「취소」는 학습을 중단합니다."
            )
            btn_save_go = msg.addButton("저장 후 학습", QMessageBox.AcceptRole)
            btn_cancel = msg.addButton("취소", QMessageBox.RejectRole)
            msg.setDefaultButton(btn_save_go)
            msg.exec_()
            if msg.clickedButton() != btn_save_go:
                return
            self._save_train_feature_states()
        # 실행 옵션(케이스 전체 자동 / 단일 CSV)에 맞춰 대상 CSV 결정
        selected_csv_paths = self._collect_train_targets_by_mode()
        if not selected_csv_paths:
            self.show_alert(
                "학습 대상을 찾지 못했습니다.\n"
                "- 케이스 내부 폴더에 CSV가 있어야 합니다.\n"
                "- 해당 폴더에 training_results가 있으면 학습 완료로 간주해 제외합니다."
            )
            return
        print(f"[INFO] 학습 대상 CSV 수: {len(selected_csv_paths)} (mode={self._train_execution_scope_mode()})")
        selected_csv_path = selected_csv_paths[0]
        
        # 선택한 CSV 파일 설정
        self.csv_path = selected_csv_path
        self.csv_file = selected_csv_path
        # Train/Create 피처는 체크·states.json 기준만 사용 (CSV 파일명과 무관)

        # CSV 파일을 테이블에 표시
        try:
            self.open_csv2(selected_csv_path, self.tableWidget)
        except Exception as e:
            print(f"[WARN] CSV 파일 열기 오류: {e}")
        
        # csvlabel 업데이트
        try:
            if hasattr(self, 'csvlabel'):
                self.csvlabel.setText(os.path.basename(selected_csv_path))
        except Exception:
            pass
        try:
            self._refresh_train_tab_case_csv_info()
        except Exception:
            pass
        
        # Train은 CSV별이 아닌 공통 매핑 JSON 1개만 사용
        train_mapping_json_path = self._resolve_train_mapping_json_path(selected_csv_path)
        if train_mapping_json_path and os.path.exists(train_mapping_json_path):
            self.set_mapping_json_path(train_mapping_json_path)
            print(f"[INFO] Train 공통 매핑 JSON 사용: {os.path.basename(train_mapping_json_path)}")
        else:
            self.set_mapping_json_path(None)
            print(f"[INFO] Train 공통 매핑 JSON 파일을 찾을 수 없습니다.")

        # 선택한 CSV 폴더의 states.json 반영
        try:
            self.load_or_initialize_states()
        except Exception:
            pass

        self.trainindex = self.model_combo.currentIndex()
        self.aimodel = self.model_combo.currentText()
        train_all_models = self._train_run_all_models_selected()
        save_outputs_enabled = self._train_save_outputs_enabled()
        train_jobs = None

        if train_all_models or len(selected_csv_paths) > 1:
            train_jobs = []
            if train_all_models:
                for csv_path in selected_csv_paths:
                    for model_index, model_name in self._train_model_items():
                        train_obj = self._new_train_obj_for_current_mode()
                        self._configure_train_obj(
                            train_obj,
                            csv_path,
                            save_outputs=save_outputs_enabled,
                            mapping_json_path=train_mapping_json_path
                        )
                        train_jobs.append({
                            "train_obj": train_obj,
                            "classmode": self.classmode,
                            "aimodel": model_name,
                            "trainindex": model_index,
                            "csv_path": csv_path,
                            "selected_feature_set": getattr(train_obj, "selected_feature_set", ""),
                        })
            else:
                for csv_path in selected_csv_paths:
                    train_obj = self._new_train_obj_for_current_mode()
                    self._configure_train_obj(
                        train_obj,
                        csv_path,
                        save_outputs=save_outputs_enabled,
                        mapping_json_path=train_mapping_json_path
                    )
                    train_jobs.append({
                        "train_obj": train_obj,
                        "classmode": self.classmode,
                        "aimodel": self.aimodel,
                        "trainindex": self.trainindex,
                        "csv_path": csv_path,
                        "selected_feature_set": getattr(train_obj, "selected_feature_set", ""),
                    })
            if not train_jobs:
                self.show_alert("학습할 모델 목록을 찾지 못했습니다.")
                return
            self.trainclass = train_jobs[0]["train_obj"]
            if train_all_models:
                self.aimodel = self._train_all_models_label()
            else:
                self.aimodel = f"{self.aimodel} (다중 CSV)"
            self.trainindex = -1
            print(f"[INFO] 일괄 학습 대상 작업 수: {len(train_jobs)}")
        else:
            self._configure_train_obj(
                self.trainclass,
                self.csv_path,
                save_outputs=save_outputs_enabled,
                mapping_json_path=train_mapping_json_path
            )
        
        # 학습 시작 시간 기록
        import time
        self._training_start_time = time.time()
        
        # UI가 멈추지 않도록 학습은 백그라운드 스레드에서 실행 + 진행창 표시
        self._train_progress_dialog = TrainingProgressDialog(self)
        self._train_progress_dialog.set_stage("학습 준비 중...")
        self._train_progress_dialog.start()

        self._train_thread = QThread(self)
        self._train_worker = TrainingWorker(
            train_obj=self.trainclass,
            classmode=self.classmode,
            aimodel=self.aimodel,
            trainindex=self.trainindex,
            csv_path=self.csv_path,
            train_jobs=train_jobs
        )
        self._train_worker.moveToThread(self._train_thread)

        # signals
        self._train_thread.started.connect(self._train_worker.run)
        self._train_worker.progress.connect(lambda s: self._train_progress_dialog.set_stage(str(s)))
        if train_jobs:
            self._begin_batch_run()
        else:
            self._train_worker.notify.connect(lambda msg: self.show_alert(str(msg)))

        def _safe_quit_train_thread():
            try:
                th = getattr(self, "_train_thread", None)
                if th is not None:
                    th.quit()
            except RuntimeError as e:
                print(f"[WARN] train_thread 종료 중 RuntimeError 무시: {e}")
            except Exception as e:
                print(f"[WARN] train_thread 종료 실패: {e}")

        def _on_train_done():
            try:
                self.statusBar().showMessage("학습 완료")
            except Exception:
                pass
            try:
                self._train_progress_dialog.stop()
            except Exception:
                pass
            # 학습 결과 기록
            try:
                import time
                completed_jobs = getattr(self._train_worker, "completed_jobs", None) or []
                if completed_jobs:
                    ok_jobs = [job for job in completed_jobs if job.get("status") == "ok"]
                    failed_jobs = [job for job in completed_jobs if job.get("status") != "ok"]
                    if save_outputs_enabled:
                        for job in ok_jobs:
                            self.save_training_result(
                                classmode=job.get("classmode", self.classmode),
                                aimodel=job.get("aimodel", ""),
                                trainindex=job.get("trainindex", -1),
                                csv_path=job.get("csv_path", self.csv_path),
                                selected_features=job.get("selected_feature_set", ""),
                                training_duration=job.get("training_duration"),
                            )
                    summary_lines = [
                        "일괄 학습 완료",
                        f"- 성공: {len(ok_jobs)}개",
                        f"- 실패: {len(failed_jobs)}개",
                        f"- 결과 저장: {'ON' if save_outputs_enabled else 'OFF'}",
                    ]
                    if ok_jobs:
                        success_texts = []
                        for job in ok_jobs:
                            model_name = job.get("aimodel", "")
                            csv_name = os.path.basename(str(job.get("csv_path", "") or ""))
                            success_texts.append(f"{model_name} / {csv_name}")
                        summary_lines.append("성공 작업: " + ", ".join(success_texts))
                    if failed_jobs:
                        summary_lines.append("실패 작업:")
                        for job in failed_jobs:
                            model_name = job.get("aimodel", "")
                            csv_name = os.path.basename(str(job.get("csv_path", "") or ""))
                            summary_lines.append(f"- {model_name} / {csv_name}: {str(job.get('error', '')).splitlines()[0]}")
                    notify_log = getattr(self._train_worker, "batch_notify_log", []) or []
                    if notify_log:
                        summary_lines.append("")
                        summary_lines.append("[모델별 학습 결과]")
                        for entry in notify_log:
                            summary_lines.append(f"- {entry}")
                    self._end_batch_run()
                    self._copyable_msg(QMessageBox.Information, "일괄 학습 결과", "\n".join(summary_lines))
                else:
                    training_duration = None
                    if hasattr(self, '_training_start_time'):
                        training_duration = time.time() - self._training_start_time

                    if save_outputs_enabled:
                        self.save_training_result(
                            classmode=self.classmode,
                            aimodel=self.aimodel,
                            trainindex=self.trainindex,
                            csv_path=self.csv_path,
                            selected_features=getattr(self.trainclass, "selected_feature_set", ""),
                            training_duration=training_duration
                        )
            except Exception as e:
                print(f"[WARN] 학습 결과 기록 실패: {e}")
            _safe_quit_train_thread()

        def _on_train_error(err_text):
            try:
                self.statusBar().showMessage("학습 실패")
            except Exception:
                pass
            try:
                self._train_progress_dialog.stop()
            except Exception:
                pass
            if getattr(self, "_batch_run_quiet", False):
                self._end_batch_run()
            _safe_quit_train_thread()
            try:
                self.show_alert(f"학습 중 오류가 발생했습니다.\n\n{err_text}")
            except Exception:
                pass

        self._train_worker.finished.connect(_on_train_done)
        self._train_worker.error.connect(_on_train_error)

        # cleanup: 다음 이벤트 루프에서 실행해 스레드 정리 중 크래시/강제종료 방지
        def _cleanup_thread():
            try:
                self._train_worker.deleteLater()
            except Exception:
                pass
            try:
                self._train_thread.deleteLater()
            except Exception:
                pass
        self._train_thread.finished.connect(lambda: QTimer.singleShot(0, _cleanup_thread))

        self._train_thread.start()


    def classdetect(self):
        self.detectclass.predict(file_path=self.file_paths[0])

    def filter_files_by_extension(self, extension):  # 선택된 확장자에 따라 필터링
        if extension and extension != "확장자":  # Ensure it's not the placeholder text
            self.extension = extension  # Directly assign the selected extension
            self.dirModel.setNameFilters([f"*{extension}"])
            self.dirModel.setNameFilterDisables(False)
        else:
            self.dirModel.setNameFilters([])

    def file_selected(self, index): # 파일 또는 디렉토리 선택 시 호출
        file_info = self.dirModel.fileInfo(index)
        try:
            if file_info.isDir():  # If a directory is selected
                if self.extension :
                    self.select_all_files_in_directory(file_info.absoluteFilePath())
            else:
                file_path = file_info.absoluteFilePath()
                extension = os.path.splitext(file_path)[1]
                self.filter_files_by_extension(extension)
                if file_path not in self.file_paths:
                    if extension.lower() == self.extension.lower():
                        self.listWidget.addItem(file_path)
                        self.file_paths.append(file_path)
                if extension == '.csv':
                    self.csv_path = file_path
                    try:
                        self._refresh_train_tab_case_csv_info()
                    except Exception:
                        pass
                    # 매핑 JSON 자동 찾기
                    mapping_json_path = self.find_mapping_json(self.csv_path)
                    if mapping_json_path:
                        self.set_mapping_json_path(mapping_json_path)
                        print(f"[INFO] 매핑 JSON 파일 자동 로드: {os.path.basename(mapping_json_path)}")
                    else:
                        self.set_mapping_json_path(None)
                    try:
                        self.load_or_initialize_states()
                    except Exception:
                        pass

        except Exception as e:
            self.show_alert(str(e))

    def ask_input(self):
            try:
                self.dataset_direc = input("데이터셋 경로를 입력하세요: ")
            except:
                pass


    def display_dataframe(self, df, widgettype):
        """읽은 데이터의 전체 행 표시 (open_csv2에서 이미 체크박스 상태에 따라 읽음)"""
        try:
            df = self.move_label_to_second_column(df)
        except:
            pass
        
        # 읽은 데이터의 전체 행 표시 (open_csv2에서 이미 체크박스 상태에 따라 데이터를 읽었음)
        display_rows = df.shape[0]
        
        widgettype.setRowCount(display_rows)
        widgettype.setColumnCount(df.shape[1])
        widgettype.setHorizontalHeaderLabels(df.columns)

        # 최적화: 벌크 업데이트를 위해 blockSignals 사용
        widgettype.blockSignals(True)
        for i in range(display_rows):
            for j in range(df.shape[1]):
                item = QTableWidgetItem(str(df.iat[i, j]))
                widgettype.setItem(i, j, item)
        widgettype.blockSignals(False)


    def open_csv2(self, csvfile, widgett):
        file_name = csvfile
        if file_name:
            try:
                # 전체 보기 체크박스 확인 (tableWidget인 경우에만)
                nrows_limit = None
                show_all_mode = False
                if widgett is getattr(self, "tableWidget", None):
                    try:
                        if hasattr(self, "show_all_checkbox") and self.show_all_checkbox:
                            show_all_mode = self.show_all_checkbox.isChecked()
                            if not show_all_mode:
                                nrows_limit = 100
                            # 현재 CSV 경로 저장 (체크박스 변경 시 다시 로드용)
                            self.current_csv_path = file_name
                            # 현재 선택 CSV를 명시적으로 상태에 반영 (추가 시 새 파일 생성 방지)
                            try:
                                self.csv_path = file_name
                                self.csv_file = file_name
                                self.existval = 1
                            except Exception:
                                pass
                    except Exception as e:
                        print(f"[WARN] 체크박스 상태 확인 실패: {e}")
                        nrows_limit = 100
                else:
                    nrows_limit = 100
                
                print(f"[DEBUG] CSV 읽기: 파일={os.path.basename(file_name)}, 전체보기={show_all_mode}, nrows_limit={nrows_limit}")
                
                # 최적화: 한 번에 읽고 헤더 확인 (pandas 버전 호환성 처리)
                header_used = 0
                # 토크나이즈 오류 방지: python 엔진 + on_bad_lines='skip'를 기본 사용
                try:
                    if nrows_limit is not None:
                        df = pd.read_csv(file_name, nrows=nrows_limit, engine='python', on_bad_lines='skip')
                    else:
                        df = pd.read_csv(file_name, engine='python', on_bad_lines='skip')
                except Exception as e2:
                    print(f"[ERROR] CSV 읽기 실패 (python 엔진): {e2}")
                    raise
                
                print(f"[DEBUG] CSV 읽기 완료: 실제 읽은 행 수={df.shape[0]}")
                
                # 첫 번째 컬럼명이 'name'이 아니면 header=1로 다시 읽기
                if df.columns[0] != 'name':
                    try:
                        if nrows_limit is not None:
                            df = pd.read_csv(file_name, header=1, nrows=nrows_limit, engine='python', on_bad_lines='skip')
                        else:
                            df = pd.read_csv(file_name, header=1, engine='python', on_bad_lines='skip')
                    except Exception as e2:
                        print(f"[ERROR] header=1 CSV 읽기 실패 (python 엔진): {e2}")
                        raise
                    header_used = 1
                    print(f"[DEBUG] header=1로 재읽기 완료: 실제 읽은 행 수={df.shape[0]}")
                
                self.display_dataframe(df, widgettype=widgett)

                # 현재 CSV 파일명 라벨도 함께 갱신
                try:
                    if hasattr(self, "csvlabel"):
                        self.csvlabel.setText(os.path.basename(file_name))
                except Exception:
                    pass
                try:
                    self._refresh_train_tab_case_csv_info()
                except Exception:
                    pass

                # Data View 탭: 현재 보고 있는 CSV 정보 표시(테이블 아래)
                try:
                    if hasattr(self, "csv_info_label") and widgett is getattr(self, "tableWidget", None):
                        abs_path = os.path.abspath(file_name)
                        stat = os.stat(file_name)
                        size_mb = stat.st_size / (1024 * 1024)
                        mtime_str = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                        # 전체 보기 체크박스 상태 확인
                        show_all = False
                        try:
                            if hasattr(self, "show_all_checkbox") and self.show_all_checkbox:
                                show_all = self.show_all_checkbox.isChecked()
                        except Exception:
                            pass
                        
                        if show_all:
                            preview_rows = df.shape[0]
                            display_text = "전체"
                        else:
                            preview_rows = min(100, df.shape[0])
                            display_text = "미리보기"
                        
                        cols = df.shape[1]
                        header_hint = "header=1(2번째 줄을 헤더로 사용)" if header_used == 1 else "header=0(첫 줄을 헤더로 사용)"
                        self.csv_info_label.setText(
                            f"파일: {os.path.basename(file_name)}\n"
                            f"경로: {abs_path}\n"
                            f"{display_text}: {preview_rows}행 / {cols}열, {header_hint}, 크기: {size_mb:.2f}MB, 수정: {mtime_str}"
                        )
                        self.csv_info_label.setVisible(True)
                        self.csv_info_label.raise_()
                except Exception:
                    pass

                # 보조: 상태바에도 한 줄 요약 표시(탭/해상도 상관 없이 보이게)
                try:
                    abs_path = os.path.abspath(file_name)
                    stat = os.stat(file_name)
                    size_mb = stat.st_size / (1024 * 1024)
                    mtime_str = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                    header_hint = "header=1" if header_used == 1 else "header=0"
                    # 전체 보기 체크박스 상태 확인
                    show_all = False
                    try:
                        if hasattr(self, "show_all_checkbox") and self.show_all_checkbox:
                            show_all = self.show_all_checkbox.isChecked()
                    except Exception:
                        pass
                    
                    if show_all:
                        display_text = f"전체 {df.shape[0]}"
                    else:
                        display_text = f"미리보기 {min(100, df.shape[0])}"
                    
                    self.statusBar().showMessage(
                        f"CSV: {os.path.basename(file_name)} | {header_hint} | {display_text}x{df.shape[1]} | {size_mb:.2f}MB | {mtime_str} | {abs_path}"
                    )
                except Exception:
                    pass
            except Exception as e:
                self.show_alert("CSV 파일을 읽는 중 오류가 발생했습니다: " + str(e))

        return


    def select_all_files_in_directory(self, directory_path):
        """최적화: 세트를 사용하여 중복 확인 속도 향상"""
        try:
            # 최적화: 리스트 대신 세트로 중복 확인
            existing_paths = set(self.file_paths)
            new_paths = []
            
            for root, _, files in os.walk(directory_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    extension = os.path.splitext(file_path)[1]

                    # 같은 확장자만 담기
                    if extension.lower() == self.extension and file_path not in existing_paths:
                        new_paths.append(file_path)
                        existing_paths.add(file_path)
            
            # 최적화: 벌크로 UI 업데이트
            if new_paths:
                self.listWidget.addItems(new_paths)
                self.file_paths.extend(new_paths)
        except Exception as e:
            self.show_alert(str(e))

    def remove_selected_file(self, item): # 선택한 파일을 목록에서 제거
        # 더블 클릭한 파일 아이템을 목록에서 제거합니다.
        file_path = item.text()
        self.listWidget.takeItem(self.listWidget.row(item))
        self.file_paths.remove(file_path)

    def remove_all_file(self):
        self.listWidget.clear()
        self.file_paths = []

    def process_files(self):
        progress_window = ProgressWindow()
        progress_window.setModal(True)
        progress_window.show()

        total_files = len(self.file_paths)

        for i, fname in enumerate(self.file_paths):
            full_path = fname
            if os.path.isfile(full_path):
                self.extract_value(full_path)
                self.all_result.append(self.reres)

            # 진행 상황 업데이트
            progress_percentage = (i + 1) / total_files * 100
            self.progress_bar.setValue(progress_percentage)
            QApplication.processEvents()

        progress_window.set_label_text("작업 완료")
        progress_window.exec_()

    def get_files_value(self): # self.file_paths의 모든 파일 처리하고 self.all_result 저장 및 반환
        self.all_result = []
        self.count = -1
        # 폴더 내 모든 파일에 대해 수행
        for i, fname in enumerate(self.file_paths):
            self.count += 1
            full_path = fname

            if os.path.isfile(full_path):
                self.extract_value(full_path)
                self.all_result.append(self.reres)

            # 진행 상황 업데이트
            progress_percentage = (i + 1) / len(self.file_paths) * 100
            self.progress_bar.setValue(progress_percentage)
            QApplication.processEvents()

        # 파일 처리가 완료되면 최장 리스트를 찾음
        longest_list = max(self.all_result, key=len)

        # 모든 리스트를 가장 긴 리스트와 비교하며 첫 번째 요소가 없으면 추가하고 두 번째 요소는 0으로 설정
        for lst in self.all_result:
            for i in range(len(longest_list)):
                if len(lst) <= i or lst[i][0] != longest_list[i][0]:
                    # 첫 번째 요소가 없으면 추가하고 두 번째 요소를 0으로 설정
                    if len(lst) <= i:
                        lst.insert(i, [longest_list[i][0], '0'])
                    else:
                        lst[i] = [longest_list[i][0], '0']

        return self.all_result

    def merge_lists2(self, ngram): # LCS 받아서 연속되거나 유사한 n-gram 병합하여 하나의 긴 패턴으로 생성
        count, count2, onecount = 0,0,0
        new_list = []
        merged_list=[]
        one_merged_list = []
        previous_gram = ''
        for onegram1 in ngram :

            if previous_gram == onegram1 and onegram1 != '00000000': # 00000000 이면 제외
                one_merged_list.append(onegram1)
                onecount += 1
                pass

            else : # 현재 n-gram과 이전 n-gram 마지막 부분이 일치하는지 확인
                if count == 0:
                    previous_gram = onegram1
                    count += 1
                else :
                    lengh = (len(onegram1) - 1)
                    previous_gram_2=previous_gram[-lengh:]
                    onegram_2 = onegram1[:-1]

                    if previous_gram_2 == onegram_2 and onegram1 != '00000000':
                        previous_gram = previous_gram+onegram1[-1]
                        #count2 += 1
                    else :
                        previous_gram_2 = ''
                        if count2 != 0 and onegram1 != '00000000':
                            one_merged_list.append(previous_gram)
                            count = 0
                        else :
                            if previous_gram != '00000000':
                                one_merged_list.append(previous_gram)

                        previous_gram = onegram1
                if onecount == len(ngram)-1: # 리스트의 마지막 n-gram 처리
                    if previous_gram != '00000000':
                        one_merged_list.append(previous_gram)
                    break

                onecount += 1

        merged_list.append(one_merged_list)
        return merged_list

    def add_numbers_to_duplicates(self, input_list):
        counts = {}  # 요소별로 카운트를 저장할 딕셔너리

        for i in range(len(input_list[0])):
            item = input_list[0][i]

            # 이미 등장한 요소인 경우
            if item in counts:
                counts[item] += 1
                input_list[0][i] = f"{item}_{counts[item]}"
            else:
                counts[item] = 0

        return input_list

    def extract_value(self, fpath): # 파일에서 n-gram 추출하고 병합된 n-gram 리스트와 비교하여 일치하는 패턴 찾아 res 리스트에 저장
        file_type = ""
        res = []
        self.reres = []

        # 파일 경로에서 파일 이름 추출
        file_name = os.path.basename(fpath)

        for file_info in self.ngrams_list:
            _, ngrams = file_info
            _ = os.path.basename(_)
            if file_name == _:  # 파일 이름과 fpath의 파일 이름을 비교
                res.append(('name', os.path.basename(fpath)))

                with open(fpath, 'rb') as fp:
                    check_opcode = self.mergelist
                    check_opcode = self.add_numbers_to_duplicates(check_opcode)

                    # 데이터 추출
                    count = 0
                    tempvalue = ''
                    mvalue = 0
                    for i in range(len(check_opcode[0]) + 1):
                        for j in range(mvalue, len(ngrams)):
                            try:
                                nowvalue = ngrams[j]

                                headerfeat = check_opcode[0][i]

                                if '_' in headerfeat :
                                    headerfeatemp = headerfeat[:-2]
                                else :
                                    headerfeatemp = headerfeat
                                k = 8
                                m = j
                                if nowvalue in headerfeatemp:

                                    while len(nowvalue) < len(headerfeatemp) and (j + k) < len(ngrams):
                                        temppp = ngrams[m + k]
                                        nowvalue += temppp[0]
                                        k += 1

                                if nowvalue == headerfeatemp:
                                    count += 1
                                elif count != 0:
                                    lennowvalue = len(nowvalue)
                                    testvalue = j + lennowvalue
                                    for kn in range(testvalue, testvalue + (lennowvalue * 2), 8):
                                        temppppp = ngrams[kn]  # 수정된 부분
                                        tempvalue += temppppp

                                    res.append((headerfeat, tempvalue))
                                    tempvalue = ''
                                    count = 0
                                    mvalue = j + 1
                                    break

                                if j == len(ngrams) - 1 and tempvalue == '':
                                    res.append((headerfeat, '0'))

                                    break

                            except Exception as e:
                                pass

        self.reres = res
        return res

    def extract_rengram(self, result): # self.ngrams_list와 result 간 교집합을 계산해 각 파일별로 공통된 n-gram들을 찾아냄
        self.intersection_lists = []

        result_set = set(result)
        for name, ngram in self.ngrams_list:
            intersection_list = [onegram for onegram in ngram if onegram in result_set]
            self.intersection_lists.append(intersection_list)

        return self.intersection_lists

    def find_duplicates_count(self): # ngrams_list에서 공통적으로 출현하는 요소 찾기

        self.data_list = []
        self.newlist = []
        duplicates = []

        element_count = {} # n-gram 요소 출현 횟수 저장 딕셔너리, 키: 요소, 값: 출현 횟수

        # 모든 리스트에서 요소의 출현 횟수를 카운트
        for k in range(len(self.ngrams_list)):
            for lst in self.ngrams_list[k][1]:
                    if lst in element_count:
                        element_count[lst] += 1
                    else:
                        element_count[lst] = 1

        # 출현 횟수가 2번 이상인 요소만 self.newlist에 저장
        basenum = int(len(self.ngrams_list)*0.7)
        self.newlist = [key for key, value in element_count.items() if value >= basenum]

        #중복이 없는 교집합 리스트를 commonlist.pkl에 저장
        commonlistpkl = str(self.extension + '\\' + "commonlist.pkl")
        with open(commonlistpkl, "wb") as fw: #
            pickle.dump(self.newlist, fw)

    def add_string_if_not_exists(self, filename, target_string):
        try:
            with open(filename, 'r', encoding='utf-8') as file:
                content = file.read()

            # 없으면 추가
            if target_string not in content:
                with open(filename, 'a', encoding='utf-8') as file:
                    file.write(target_string)

        except Exception as e :
            with open(filename, 'a', encoding='utf-8') as file:
                file.write(target_string)

    def feature_dictionary(self, hexa):

        array10 = []
        # 엑셀 파일에서 데이터 읽어오기 (엑셀 파일 경로를 설정해 주세요)
        excel_file = str(self.extension + '\\' + '_dict.xlsx')  # 엑셀 파일 경로
        self.resource_path(excel_file)
        df = pd.read_excel(excel_file)  # 엑셀 파일 읽기

        # 엑셀 데이터를 딕셔너리로 변환 (엑셀 파일의 첫 번째 열을 key로, 두 번째 열을 value로)
        newdict = dict(zip(df.iloc[:, 0], df.iloc[:, 1]))
        result = hexa
        # newdict의 value가 result[1]에 있으면 key를 array10에 추가
        for key, value in newdict.items():
            if value in result[1]:
                array10.append(str(key))

        # array10을 ","로 구분된 문자열로 만들어서 simhash 계산
        sequencedem = ", ".join(array10)
        sequencedem = self.simhash(sequencedem)


        # sequencedem과 hexa[0]을 함께 저장
        self.sequencedem.append((hexa[0], sequencedem))

    def save_lists_of_10_to_csv(self, data_list, file_name): # 더 긴 패턴(mergelist 리스트) 주어진 데이터를 CSV로 저장
        with open(file_name, 'w', newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)

            row = [j for j in range(1, len(data_list[0])+1)] # 헤더 작성
            csv_writer.writerow(row)

            for row in data_list:
                csv_writer.writerow(row)

    # value로 key찾기
    def find_key_by_value(self, dictionary, value):
        for key, val in dictionary.items():
            if val == value:
                return key
        return None  # 해당 값과 일치하는 키가 없을 경우 None을 반환

    def make_features(self, input_str):
        length = 3
        input_str = input_str.lower()
        out_str = re.sub(r'[^\w]+', '', input_str)
        return [out_str[i:i + length] for i in range(max(len(out_str) - length + 1, 1))]
    def simhash(self, input_str):
        features = self.make_features(input_str)
        return Simhash(features).value
    def headersimhash(self, input_list):
        string_result = ''.join(map(str, input_list))
        self.simhash(string_result)
    def makearray(self, featurelist, newdict):
        newlist2 = []
        newlist = []
        # 최적화: values를 세트로 변환하여 조회 속도 향상
        dict_values_set = set(newdict.values())
        # 피쳐를 딕셔너리 사전의 10진수값에 매핑
        for item in featurelist:
            if item in dict_values_set:
                newlist.append(self.find_key_by_value(newdict, item))

        newlist2 =[self.file_paths]
        newlist2.append(newlist)

        pklname = str(self.extension + '\\' + "vectordb.pkl")
        with open(pklname, "wb") as fw:
            pickle.dump(newlist2, fw)

    # 헤더를 csv에 저장
    def save_list_of_indivi_to_csv(self, data_list, file_name):

        with open(file_name, 'w', newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            sub_strings = data_list.split(',')
            csv_writer.writerow(sub_strings)
    def file_exists(self, folder_path, filename):
        file_path = os.path.join(folder_path, filename)
        return os.path.isfile(file_path)

    def center_window(self, root, width=300, height=200):
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)

        root.geometry('%dx%d+%d+%d' % (width, height, x, y))

    def merge_and_save_pkl(self, data, pkl_path):
        if os.path.exists(pkl_path) and os.path.getsize(pkl_path) > 0:
            with open(pkl_path, 'rb') as f:
                existing_data = pickle.load(f)
        else:
            existing_data = []

        # 기존 데이터와 새로운 데이터 병합
        existing_file_names = {item[0] for item in existing_data}
        new_items = [item for item in data if item[0] not in existing_file_names]
        combined_data = existing_data + new_items

        with open(pkl_path, 'wb') as f:
            pickle.dump(combined_data, f)

        return combined_data

    def get_fast_file_hash(self, filepath, hash_type='md5', sample_size=65536):
        """최적화: 캐싱을 사용하여 중복 계산 방지"""
        # 캐시 키: (파일경로, 파일 수정시간, 파일크기)
        try:
            stat_info = os.stat(filepath)
            cache_key = (filepath, stat_info.st_mtime, stat_info.st_size)
            
            # 캐시에서 확인
            if cache_key in self._file_hash_cache:
                return self._file_hash_cache[cache_key]
        except:
            cache_key = None
        
        hash_func = getattr(hashlib, hash_type)()

        with open(filepath, 'rb') as f:
            # 파일의 처음 부분에서 sample_size 만큼 읽기
            start_chunk = f.read(sample_size)
            hash_func.update(start_chunk)

            # 파일의 마지막 부분에서 sample_size 만큼 읽기
            f.seek(0, 2)  # 파일 끝으로 이동
            file_size = f.tell()
            if file_size > sample_size * 2:
                f.seek(-sample_size, 2)  # 파일 끝에서 sample_size 전으로 이동
                end_chunk = f.read(sample_size)
                hash_func.update(end_chunk)

        result = hash_func.hexdigest()
        
        # 캐시에 저장
        if cache_key:
            self._file_hash_cache[cache_key] = result
            
        return result


    def extract_box_feature(self, file_paths):
        """
        기존 개념(무엇을 추출하는지, CSV 포맷)은 그대로 두고
        파일 단위 처리를 병렬화한 버전입니다.
        """
        # 새 데이터셋을 만들 때는 처리 시작 전에 최신 규칙으로 CSV 경로를 만들어 둔다
        if self.detectmode == 0:
            self.ensure_new_csv_path()

        from joblib import Parallel, delayed

        # 엑셀에서 seqdict 로딩 (공통자원)
        excel_file = os.path.join('mp4', '_dict.xlsx')
        excel_file = self.resource_path(excel_file)
        df = pd.read_excel(excel_file)
        self.seqdict = dict(zip(df.iloc[:, 0], df.iloc[:, 1]))

        if isinstance(file_paths, str):
            file_paths = [file_paths]

        total_files = len(file_paths)
        
        # 진행상황 파일: config 폴더 우선, 없으면 케이스 루트 (불러올 때도 동일, 하위 호환)
        config_d = self._config_dir()
        progress_file = os.path.join(config_d, "progress_current.txt") if (config_d and os.path.isdir(config_d)) else os.path.join(self.case_direc, "progress_current.txt")
        legacy_progress = os.path.join(self.case_direc, "progress_current.txt")
        processed_files = set()
        
        for pf in (progress_file, legacy_progress):
            if pf and os.path.exists(pf):
                try:
                    with open(pf, 'r', encoding='utf-8') as f:
                        processed_files = set(line.strip() for line in f if line.strip())
                    print(f"기존 진행상황 로드: {len(processed_files)}개 파일 이미 처리됨 ({os.path.basename(os.path.dirname(pf))})")
                    progress_file = pf  # 삭제 시 사용할 경로
                    break
                except Exception as e:
                    print(f"진행상황 파일 읽기 오류: {e}")
        
        # 이미 처리된 파일은 제외
        files_to_process = [
            fp for fp in file_paths
            if os.path.basename(fp) not in processed_files
        ]

        print(f"처리 대상 파일 수: {len(files_to_process)}/{total_files}")

        if not files_to_process:
            print("새로 처리할 파일이 없습니다.")
            return []

        # CPU 코어 수 기준 병렬 작업 개수
        try:
            n_jobs = max(1, (os.cpu_count() or 2) - 1)
        except Exception:
            n_jobs = 1

        print(f"병렬 처리 작업 수(n_jobs): {n_jobs}")
        print(f"처리 시작: {len(files_to_process)}개 파일")
        print("※ 일부 파일에서 '박스 타입 디코딩 오류' 또는 'h264_analyze 오류'가 발생할 수 있습니다.")
        print("  이는 정상적인 예외 처리이며, 작업은 계속 진행됩니다. 걱정하지 마세요.")

        # 상태값들을 딕셔너리로 준비 (pickle 가능하도록)
        state_flags = {
            'structure_val_state': self.structure_val_state,
            'structure_seq_state': self.structure_seq_state,
            'frame_gop_state': self.frame_gop_state,
            'frame_ratio_state': self.frame_ratio_state,
            'frame_sps_state': self.frame_sps_state
        }
        
        # 디버깅: detect 시 상태값 확인
        if self.detectmode == 1:
            print(f"[DEBUG] extract_box_feature 상태값 (detectmode=1): {state_flags}")
            active_options = [k for k, v in state_flags.items() if v]
            print(f"[DEBUG] 활성화된 옵션: {active_options}")
        
        # seqdict도 복사 (pickle 가능하도록)
        seqdict_copy = self.seqdict.copy()

        # 병렬 처리 실행 (독립 함수 사용, verbose=1로 진행 상황 표시)
        # verbose=1은 각 작업이 시작/완료될 때마다 출력 (너무 많을 수 있음)
        # verbose=0은 출력 없음, verbose=10은 10개마다 출력
        import sys
        import time
        start_time = time.time()
        
        print("병렬 처리 중... (진행 상황은 작업 완료 시 표시됩니다)")
        results_list = Parallel(n_jobs=n_jobs, backend="loky", verbose=10)(
            delayed(_process_one_file)(fp, state_flags, seqdict_copy) for fp in files_to_process
        )
        
        elapsed_time = time.time() - start_time
        print(f"\n처리 완료: {len(results_list)}개 파일 처리됨 (소요 시간: {elapsed_time:.1f}초)")
        
        # 디버깅: detect 시 결과 확인
        if self.detectmode == 1 and len(results_list) > 0:
            first_result = results_list[0]
            result_keys = [k for k, v in first_result] if isinstance(first_result, list) else []
            print(f"[DEBUG] _process_one_file 반환값: {len(result_keys)}개 키")
            print(f"[DEBUG] 주요 키 (처음 30개): {result_keys[:30]}")
            # SPS, sequence, GOP 확인
            has_sps = any(k == 'SPS' for k, v in first_result) if isinstance(first_result, list) else False
            has_sequence = any(k == 'sequence' for k, v in first_result) if isinstance(first_result, list) else False
            has_gop = any(k == 'GOP' for k, v in first_result) if isinstance(first_result, list) else False
            print(f"[DEBUG] SPS 포함: {has_sps}, sequence 포함: {has_sequence}, GOP 포함: {has_gop}")

        # CSV 저장 (append 모드, 한 번에)
        if self.detectmode == 0 and len(results_list) > 0:
            self.save_to_csv(results_list, append_mode=True, show_alert=True)

        # 진행상황 파일은 작업 완료 후 정리
        try:
            if os.path.exists(progress_file):
                os.remove(progress_file)
                print("진행상황 파일 정리 완료")
        except Exception as e:
            print(f"진행상황 파일 정리 오류: {e}")
            
        return results_list

    def on_structure_val_changed(self, state):
        if state == Qt.Checked:
            print('structure_val Box is checked')
            self.structure_val_state = 1
            self.set_state("structure_val_state", 1)
            if '_val' not in self.tempcsv_file:
                self.tempcsv_file += '_val'
        else:
            print('structure_val Box is unchecked')
            self.structure_val_state = 1
            self.set_state("structure_val_state", 0)
            self.tempcsv_file = self.tempcsv_file.replace('_val', '')

    def on_structure_seq_changed(self, state):
        if state == Qt.Checked:
            print('structure_seq Box is checked')
            self.structure_seq_state = 1
            self.set_state("structure_seq_state", 1)
            if '_seq' not in self.tempcsv_file:
                self.tempcsv_file += '_seq'
        else:
            print('structure_seq Box is unchecked')
            self.structure_seq_state = 0
            self.set_state("structure_seq_state", 0)
            self.tempcsv_file = self.tempcsv_file.replace('_seq', '')

    def on_frame_sps_changed(self, state):
        if state == Qt.Checked:
            print('frame_sps Box is checked')
            self.set_state("frame_sps_state", 1)
            self.frame_sps_state = 1
            if '_sps' not in self.tempcsv_file:
                self.tempcsv_file += '_sps'
        else:
            print('frame_sps Box is unchecked')
            self.frame_sps_state = 0
            self.set_state("frame_sps_state", 0)
            self.tempcsv_file = self.tempcsv_file.replace('_sps', '')

    def on_frame_gop_changed(self, state):
        if state == Qt.Checked:
            print('frame_gop Box is checked')
            self.frame_gop_state = 1
            self.set_state("frame_gop_state", 1)
            if '_gop' not in self.tempcsv_file:
                self.tempcsv_file += '_gop'
        else:
            print('frame_gop Box is unchecked')
            self.set_state("frame_gop_state", 0)
            self.frame_gop_state = 0
            self.tempcsv_file = self.tempcsv_file.replace('_gop', '')

    def on_frame_ratio_changed(self, state):
        if state == Qt.Checked:
            print('frame_ratio Box is checked')
            self.set_state("frame_ratio_state", 1)
            self.frame_ratio_state = 1
            if '_ratio' not in self.tempcsv_file:
                self.tempcsv_file += '_ratio'
        else:
            print('frame_ratio Box is unchecked')
            self.set_state("frame_ratio_state", 0)
            self.frame_ratio_state = 0
            self.tempcsv_file = self.tempcsv_file.replace('_ratio', '')

    def on_show_all_changed(self, state):
        """전체 보기 체크박스 변경 시 CSV 다시 로드"""
        try:
            # current_csv_path가 없으면 csv_path 사용
            csv_path_to_load = None
            if hasattr(self, "current_csv_path") and self.current_csv_path:
                csv_path_to_load = self.current_csv_path
            elif hasattr(self, "csv_path") and self.csv_path:
                csv_path_to_load = self.csv_path
            
            if csv_path_to_load:
                print(f"[DEBUG] 전체 보기 체크박스 변경: 체크={state == 2}, CSV 다시 로드: {os.path.basename(csv_path_to_load)}")
                # 체크박스 상태 변경 후 CSV 다시 로드
                self.open_csv2(csv_path_to_load, self.tableWidget)
            else:
                print(f"[WARN] 전체 보기 변경: CSV 경로가 없습니다.")
        except Exception as e:
            import traceback
            print(f"[WARN] 전체 보기 변경 시 CSV 다시 로드 실패: {e}")
            traceback.print_exc()

    def ensure_new_csv_path(self):
        """새 데이터셋 생성 시 최신 규칙으로 csv_path/csvlabel을 설정"""
        if self.existval != 1 and not self.csv_file:
            timestamp = datetime.now().strftime("%y%m%d%H%M")
            base_name = self.tempcsv_file.strip("_") if self.tempcsv_file else ""
            # base_name이 비어있거나 "train"만 있는 경우 중복 방지
            if not base_name or base_name == "train":
                filename = f"_train_{timestamp}.csv"
            else:
                filename = f"_{base_name}_train_{timestamp}.csv"
            self.csv_path = os.path.join(self.case_direc, filename)
            try:
                if hasattr(self, 'csvlabel'):
                    self.csvlabel.setText(os.path.basename(self.csv_path))
            except Exception:
                pass

    # 결과를 CSV로 저장 (메모리 최적화)
    def save_to_csv(self, all_data, append_mode=False, show_alert=True):
        # 신규 데이터셋(기존 CSV 추가 모드가 아닐 때)에는 항상 최신 파일명으로 생성
        # 선택된 CSV가 이미 있는 경우 그것을 우선 사용하도록 보강
        try:
            if not self.csv_file and hasattr(self, "current_csv_path") and self.current_csv_path:
                self.csv_file = self.current_csv_path
                self.csv_path = self.current_csv_path
                self.existval = 1
        except Exception:
            pass

        if self.existval != 1 and not self.csv_file:
            self.ensure_new_csv_path()

        # 기본 경로 설정 (선택된 CSV가 있으면 우선 사용)
        csv_path = self.csv_path

        if self.csv_file!='':
            csv_path = self.csv_file

        # 최적화: append_mode일 때는 기존 데이터를 읽지 않음 (메모리 절약)
        existing_data = []
        existing_fieldnames = []
        if os.path.exists(csv_path) and not append_mode:
            try:
                # error_bad_lines=False (pandas 구버전) 또는 on_bad_lines='skip' (신버전)
                try:
                    existing_df = pd.read_csv(csv_path, encoding='utf-8', on_bad_lines='skip')
                except TypeError:
                    # pandas 구버전
                    existing_df = pd.read_csv(csv_path, encoding='utf-8', error_bad_lines=False, warn_bad_lines=True)
                existing_fieldnames = existing_df.columns.tolist()
                existing_data = existing_df.to_dict('records')
            except Exception as e:
                print(f"기존 CSV 읽기 오류: {e}")
                existing_fieldnames = []
        elif os.path.exists(csv_path) and append_mode:
            # append 모드: 헤더만 읽기
            try:
                # pandas 버전 호환성 처리
                try:
                    existing_df = pd.read_csv(csv_path, encoding='utf-8', nrows=0, on_bad_lines='skip')
                except TypeError:
                    existing_df = pd.read_csv(csv_path, encoding='utf-8', nrows=0, error_bad_lines=False)
                existing_fieldnames = existing_df.columns.tolist()
                
                # 기존 데이터가 있는지 확인 (안전하게 처리)
                try:
                    try:
                        existing_df_check = pd.read_csv(csv_path, encoding='utf-8', nrows=1, on_bad_lines='skip')
                    except TypeError:
                        existing_df_check = pd.read_csv(csv_path, encoding='utf-8', nrows=1, error_bad_lines=False)
                    if len(existing_df_check) > 0:
                        # 전체 행 수를 세는 대신 wc를 사용하거나 간단히 표시
                        print(f"기존 CSV 파일에서 데이터 발견 (append 모드)")
                except Exception as count_error:
                    print(f"기존 데이터 확인 중 오류 (무시하고 계속): {count_error}")
            except Exception as e:
                print(f"헤더 읽기 오류: {e}")
                existing_fieldnames = []

        # 메모리 최적화: append 모드일 때는 'a' 모드 사용, 아니면 'w' 모드
        # 단, append_mode라도 파일이 없거나 비어 있으면 헤더를 쓰도록 'w'로 전환
        file_exists = os.path.exists(csv_path)
        file_empty = False
        if file_exists:
            try:
                file_empty = os.path.getsize(csv_path) == 0
            except Exception:
                file_empty = False

        # 필드명 추출 - 모든 파일의 필드를 확인하여 중복 필드 처리
        fieldnames = existing_fieldnames.copy()
        key_count_global = {}  # 전체 파일에서의 중복 key 카운트 딕셔너리

        # 모든 파일의 데이터를 순회하여 필드 추출
        for file_data in all_data:
            key_count_local = {}  # 각 파일 내에서의 중복 key 카운트
            for key, value in file_data:
                if key!='GOP':
                    # 중복 필드 처리 (필드 이름 중복 시 숫자를 붙임)
                    if key in key_count_local:
                        key_count_local[key] += 1
                        key_with_count = f"{key}({key_count_local[key]})"
                    else:
                        key_count_local[key] = 1
                        key_with_count = key

                    # 콜론이 있는 경우 자식 속성 분리
                    if isinstance(value, str) and ":" in value:
                        attributes = [attr.strip() for attr in value.split(",")]
                        for attr in attributes:
                            if ":" in attr:
                                attr_name = f"{key_with_count}_{attr.split(':')[0].strip()}"
                                attr_value = attr.split(":")[1].strip()

                                # 필드가 fieldnames에 없고 값이 있는 경우만 필드를 추가
                                if attr_name not in fieldnames and attr_value:
                                    fieldnames.append(attr_name)
                    else:
                        if key_with_count not in fieldnames:
                            fieldnames.append(key_with_count)
                else:
                    # GOP는 모든 파일에 공통으로 1개 컬럼만 있으면 됩니다.
                    # (중복 추가되면 pandas가 GOP.1/GOP.2...로 자동 변경되어 학습/예측 정합성이 깨질 수 있음)
                    if key not in fieldnames:
                        fieldnames.append(key)

        ##1025 레이블 추가
        if 'label' not in fieldnames:
            fieldnames.append('label')

        print('최종 필드명 확인: ', fieldnames)

        # append 모드에서 새 컬럼이 등장하면 헤더를 다시 써서 정합성 유지
        # (이전 저장 단계에서 NaN 컬럼 제거로 헤더가 줄어든 경우, 새 컬럼을 append만 하면 필드 불일치 발생)
        try:
            new_columns = [c for c in fieldnames if c not in existing_fieldnames]
            if append_mode and file_exists and not file_empty and new_columns:
                print(f"[INFO] append 모드에서 새 컬럼 {len(new_columns)}개 발견 -> 전체 재작성으로 전환: {new_columns[:5]}{'...' if len(new_columns) > 5 else ''}")
                try:
                    # 전체 데이터를 다시 읽어서 기존 행을 유지한 채 새 컬럼을 포함하는 헤더로 재작성
                    try:
                        existing_df_full = pd.read_csv(csv_path, encoding='utf-8', engine='python', on_bad_lines='skip')
                    except TypeError:
                        existing_df_full = pd.read_csv(csv_path, encoding='utf-8', error_bad_lines=False, warn_bad_lines=True)
                    existing_data = existing_df_full.to_dict('records')
                    # 기존 헤더를 새 필드명 리스트로 갱신
                    existing_fieldnames = existing_df_full.columns.tolist()
                    # fieldnames를 기존+신규 컬럼의 합집합으로 재정렬
                    field_set = list(existing_fieldnames)
                    for col in fieldnames:
                        if col not in field_set:
                            field_set.append(col)
                    fieldnames = field_set
                    # 기존 행에도 신규 컬럼 키를 채워 넣도록 append_mode를 False로 전환
                    append_mode = False
                    write_mode = 'w'
                    write_header = True
                except Exception as e:
                    print(f"[WARN] 새 컬럼 반영을 위한 재작성 실패(계속 append 시도): {e}")
        except Exception as e:
            print(f"[WARN] 새 컬럼 검출 중 오류(무시하고 진행): {e}")

        # CSV에 기존 데이터와 함께 쓰기
        try:
            all_data = self.move_label_to_second_column(all_data)
        except Exception as e:
            strtem = f'{e}, " -- label 컬럼이 존재하지 않습니다."'
            self.show_alert(strtem)

        if 'write_mode' not in locals():
            if append_mode and file_exists and not file_empty:
                write_mode = 'a'
                write_header = False
            else:
                write_mode = 'w'
                write_header = True
        
        with open(csv_path, write_mode, newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            # 헤더 쓰기 (append 모드가 아닐 때만)
            if write_header:
                writer.writeheader()

            # 기존 데이터 쓰기 (append 모드가 아닐 때만)
            if not append_mode:
                for row in existing_data:
                    writer.writerow(row)

            # 새로운 데이터 쓰기
            rows_written = 0
            for file_data in all_data:
                row_data = {}
                key_count_local = {}

                for key, value in file_data:
                    # GOP는 별도 처리: 헤더는 1개만, 각 행의 값은 그대로 저장
                    if key == 'GOP':
                        if 'GOP' in fieldnames:
                            row_data['GOP'] = value if value else ''
                        continue
                    
                    if key in key_count_local:
                        key_count_local[key] += 1
                        key_with_count = f"{key}({key_count_local[key]})"
                    else:
                        key_count_local[key] = 1
                        key_with_count = key

                    if isinstance(value, str):
                        attributes = [attr.strip() for attr in value.split(",")]
                        for attr in attributes:
                            if ":" in attr:
                                attr_name, attr_value = attr.split(":", 1)
                                attr_name_clean = attr_name.strip()
                                attr_value_clean = attr_value.strip()
                                
                                # Entries가 리스트 문자열인 경우 추가 파싱 (예: "Entries: ['Duration: 123, Media Time: 456']")
                                if attr_name_clean == "Entries" and attr_value_clean.startswith("['") and ":" in attr_value_clean:
                                    # 리스트 내의 각 entry를 파싱
                                    entries_str = attr_value_clean.strip("[]'\"")
                                    if "Duration:" in entries_str or "Media Time:" in entries_str or "Rate:" in entries_str:
                                        # 이미 파싱된 형식이므로 추가 파싱
                                        entry_parts = entries_str.split(",")
                                        for entry_part in entry_parts:
                                            if ":" in entry_part:
                                                entry_key, entry_val = entry_part.split(":", 1)
                                                entry_key_clean = entry_key.strip()
                                                entry_val_clean = entry_val.strip()
                                                # elst_Media Time, elst_Rate 등으로 저장
                                                if f"{key_with_count}_{entry_key_clean}" in fieldnames:
                                                    row_data[f"{key_with_count}_{entry_key_clean}"] = entry_val_clean
                                                elif f"{key_with_count}_'Duration" in fieldnames and entry_key_clean == "Duration":
                                                    row_data[f"{key_with_count}_'Duration"] = entry_val_clean
                                else:
                                    # 일반적인 속성 처리
                                    if f"{key_with_count}_{attr_name_clean}" in fieldnames:
                                        row_data[f"{key_with_count}_{attr_name_clean}"] = attr_value_clean
                            else:
                                if key_with_count in fieldnames:
                                    row_data[key_with_count] = value
                    elif isinstance(value, list):
                        # 리스트 타입 처리 (elst의 Entries 같은 경우)
                        for item in value:
                            if isinstance(item, str) and ":" in item:
                                item_attrs = [ia.strip() for ia in item.split(",")]
                                for item_attr in item_attrs:
                                    if ":" in item_attr:
                                        item_key, item_val = item_attr.split(":", 1)
                                        item_key_clean = item_key.strip()
                                        item_val_clean = item_val.strip()
                                        if f"{key_with_count}_{item_key_clean}" in fieldnames:
                                            row_data[f"{key_with_count}_{item_key_clean}"] = item_val_clean
                            else:
                                if key_with_count in fieldnames:
                                    row_data[key_with_count] = str(item)
                    else:
                        if key_with_count in fieldnames:
                            row_data[key_with_count] = value

                try:
                    if hasattr(self, 'label_data') and self.label_data:
                        row_data['label'] = self.label_data
                    else:
                        print("Warning: 'label_data' is not set or is empty.")
                except Exception as e:
                    pass

                writer.writerow({key: row_data.get(key, "") for key in fieldnames})
                rows_written += 1

        print(f"Results saved to {csv_path} (mode={write_mode}, rows_written={rows_written})")
        
        # 방금 생성/추가한 CSV를 기본 선택 상태로 반영 (Data View 업데이트)
        try:
            self.csv_path = csv_path
            self.csv_file = csv_path
            self.current_csv_path = csv_path
            # 이후 추가 작업에서 새 파일 생성하지 않도록 append 모드 상태로 간주
            self.existval = 1
            self.skip_initial_csv_load = False
            # Data View 테이블에 즉시 반영
            if hasattr(self, "tableWidget"):
                self.open_csv2(csv_path, self.tableWidget)
            if hasattr(self, "csvlabel"):
                self.csvlabel.setText(os.path.basename(csv_path))
        except Exception as e:
            print(f"[WARN] CSV 선택/표시 자동 반영 실패: {e}")
        
        # 모든 행이 NaN인 컬럼 제거 전처리
        try:
            # 원본 라인 수(헤더 제외)를 먼저 계산해 파싱 시 스킵된 행이 있으면 재저장을 건너뜀
            raw_rows = 0
            try:
                with open(csv_path, "r", encoding="utf-8", errors="ignore") as f:
                    raw_rows = sum(1 for _ in f) - 1  # 헤더 제외
            except Exception:
                raw_rows = 0

            # CSV 파일을 DataFrame으로 읽기
            try:
                df = pd.read_csv(csv_path, encoding='utf-8', on_bad_lines='skip')
            except TypeError:
                df = pd.read_csv(csv_path, encoding='utf-8', error_bad_lines=False, warn_bad_lines=True)

            # 파싱 과정에서 행이 스킵된 경우(열 불일치 등)에는 재저장으로 기존 행을 잃을 수 있으므로 컬럼 제거를 건너뜀
            if raw_rows > 0 and df.shape[0] < raw_rows:
                print(f"[WARN] CSV 파싱 시 {raw_rows - df.shape[0]}개 행이 스킵되어 NaN 컬럼 제거/재저장을 건너뜁니다.")
                df = None
            
            if df is not None:
                # NaN/빈 문자열/문자열/bytes 리터럴 등을 NaN으로 통일 후 컬럼 단위로 제거
                def _normalize_nan(v):
                    # bytes -> str 디코드 시도
                    try:
                        if isinstance(v, (bytes, bytearray)):
                            v = v.decode("utf-8", errors="ignore")
                    except Exception:
                        pass
                    # 이미 NaN이면 그대로
                    if pd.isna(v):
                        return np.nan
                    s = str(v)
                    s_strip = s.strip()
                    s_lower = s_strip.lower()
                    # 빈 문자열, 제어문자만 있는 경우, nan/none/null 계열, bytes 빈 리터럴 등 처리
                    if s_strip == "":
                        return np.nan
                    if re.fullmatch(r"(nan|none|null|n/?a|na|nul)", s_lower):
                        return np.nan
                    if re.fullmatch(r"b['\"]\s*['\"]", s_lower):
                        return np.nan
                    # 공백/개행만 있는 경우
                    if re.fullmatch(r"\s*", s):
                        return np.nan
                    return v

                df = df.applymap(_normalize_nan)
                # 모든 행이 NaN인 컬럼 제거 (세로 방향)
                before_cols = set(df.columns)
                df = df.dropna(axis=1, how="all")
                cols_to_drop = list(before_cols - set(df.columns))
                
                # NaN 컬럼 제거
                if cols_to_drop:
                    print(f"[INFO] 모든 행이 NaN인 컬럼 {len(cols_to_drop)}개 제거: {cols_to_drop[:10]}{'...' if len(cols_to_drop) > 10 else ''}")
                    # errors='ignore'로 존재하지 않는 컬럼 drop 시 예외 방지
                    df = df.drop(columns=cols_to_drop, errors='ignore')
                    # 원본 파일에 다시 저장 (인덱스 없이)
                    df.to_csv(csv_path, index=False, encoding='utf-8')
                    print(f"[INFO] NaN 컬럼 제거 후 CSV 저장 완료: {csv_path}")
        except Exception as e:
            print(f"[WARN] NaN 컬럼 제거 중 오류 발생 (계속 진행): {e}")
            import traceback
            traceback.print_exc()
        
        # show_alert가 True일 때만 알림창 표시
        if show_alert:
            savemassage = f"학습데이터셋이 파일 {csv_path} 에 저장되었습니다."
            self.show_file_alert(csv_path, savemassage, self.tableWidget_Create)
            # 후처리 CSV(_processed) 자동 생성은 중복 파일을 늘리고 혼동을 줄 수 있어 비활성화
            # self.post_process_csv(csv_path)

        self.save_states()

    def get_fieldnames(self, all_data, existing_data):
        fieldnames = set()
        for row in existing_data:
            fieldnames.update(row.keys())
        for file_data in all_data:
            fieldnames.update(file_data.keys())
        if 'label' not in fieldnames:
            fieldnames.add('label')
        return list(fieldnames)

    def prepare_row_data(self, file_data, fieldnames):
        row_data = {}
        for key, value in file_data.items():
            row_data[key] = value
        if 'label' in fieldnames and self.label_data:
            row_data['label'] = self.label_data
        return row_data

    # Function to post-process the saved CSV
    def post_process_csv(self, csv_path):
        result_path = csv_path.replace(".csv", "_processed.csv")
        
        # pandas 버전 호환성 처리
        try:
            try:
                df = pd.read_csv(csv_path, on_bad_lines='skip')
            except TypeError:
                df = pd.read_csv(csv_path, error_bad_lines=False, warn_bad_lines=True)
        except Exception as e:
            print(f"CSV 읽기 오류 (post_process): {e}")
            print("후처리를 건너뜁니다.")
            return

        # Apply transformations for specific columns
        # 학습 CSV의 원본 값을 그대로 맞추기 위해 시간/길이/크기 변환을 적용하지 않는다.
        # (과거에는 adjust_time_columns/adjust_duration_columns/adjust_dimensions에서 1/3/0 등으로 축약하여 값 불일치가 발생)
        # self.adjust_time_columns(df)
        # self.adjust_duration_columns(df)
        # self.adjust_dimensions(df)

        # Save processed data
        df.to_csv(result_path, index=False)
        print(f"Processed data saved to {result_path}")

    def post_process(self, df):
        oridf = df
        # Apply transformations for specific columns
        self.adjust_time_columns(df)
        self.adjust_duration_columns(df)
        self.adjust_dimensions(df)

        return df


    # Functions for column-specific adjustments (최적화)
    def adjust_time_columns(self, df):
        pattern = re.compile(r'.*(Create Time|Modify Time)', re.IGNORECASE)
        # 최적화: 컬럼 리스트를 미리 필터링
        time_cols = [col for col in df.columns if pattern.search(col)]
        for col in time_cols:
            df[col] = df[col].apply(self.transform_time)

    def transform_time(self, value):
        if pd.notna(value) and len(str(value)) >= 4:
            if str(value).startswith('1'):
                return 1
            elif str(value).startswith('3'):
                return 3
        return value

    def adjust_duration_columns(self, df):
        pattern = re.compile(r'.*(duration|Entry|Entries|Sample Count)\b', re.IGNORECASE)
        # 최적화: 컬럼 리스트를 미리 필터링
        duration_cols = [col for col in df.columns if pattern.search(col)]
        for col in duration_cols:
            df[col] = df[col].apply(self.transform_duration)

    def transform_duration(self, value):
        if pd.notna(value):
            return 1
        return -1 if pd.isna(value) or value == '' else value

    def adjust_dimensions(self, df):
        pattern = re.compile(r'.*(width|height)\b$', re.IGNORECASE)
        # 최적화: 컬럼 리스트를 미리 필터링
        dimension_cols = [col for col in df.columns if pattern.search(col)]
        for col in dimension_cols:
            df[col] = df[col].apply(self.transform_dimension)

    def transform_dimension(self, value):
        if type(value) is str :
            value = int(value)

        if pd.notna(value):
            if value > 1:
                return 1
            elif value == 0:
                return 0
        return -1 if pd.isna(value) or value == '' else value

    @staticmethod
    def calculate_simhash_lib(value):
        try:
            value = str(value)
        except:
            pass

        try:
            if value in [0, None, ""] or (isinstance(value, float) and math.isnan(value)):
                return -99999999
        except Exception as e:
            pass
        try:
            try:
                simval = Simhash(str(value)).value
            except:
                simval = Simhash(str(value[:200])).value
        except Exception as e:
            print(e)
            simval = -99999999
        return simval

    def calculate_simhash_lib(self, value, zero_as_missing=True, missing_sentinel=-99999999):
        """Train 클래스와 동일한 calculate_simhash_lib 구현 (multi-class 기준).
        학습 시 pandas가 CSV에서 1->1.0으로 읽으므로, 1.0과 1을 동일한 simhash로 맞춤."""
        try:
            if value in [None, ""] or (isinstance(value, float) and math.isnan(float(value))):
                return missing_sentinel
            if zero_as_missing:
                try:
                    if float(value) == 0.0:
                        return missing_sentinel
                except Exception:
                    pass
        except Exception:
            pass
        # 숫자 정규화: 1.0, "1.0", 1, "1" -> "1" (학습/탐지 simhash 일치)
        try:
            v = float(value)
            if not math.isfinite(v):  # inf, -inf, NaN 처리 (int 변환 시 OverflowError 방지)
                return missing_sentinel
            if v == int(v):
                value = str(int(v))
            else:
                value = str(v)
        except (ValueError, TypeError, OverflowError):
            value = str(value)
        try:
            try:
                simval = Simhash(str(value)).value
            except Exception:
                simval = Simhash(str(value)[:200]).value
        except Exception as e:
            print(f"[WARN] Simhash 계산 오류: {e}")
            simval = missing_sentinel
        return simval

    def apply_simhash(self, df):
        """Simhash 적용 (Train 클래스와 동일한 방식)"""
        df.columns = df.columns.astype(str)
        columns_to_process = [col for col in df.columns if col not in ['name', 'label']]
        
        # Train 클래스와 동일하게 apply 사용
        for column in columns_to_process:
            if column == 'sequence':
                # sequence는 0도 유효값으로 취급 (결측과 구분)
                df[column] = df[column].apply(
                    lambda v: self.calculate_simhash_lib(v, zero_as_missing=False, missing_sentinel=-99999998)
                )
            else:
                df[column] = df[column].apply(self.calculate_simhash_lib)

        return df

    def _detect_results_base_dir(self):
        """Detect 결과 저장 기준 폴더 (상세 폴더 → CSV 폴더 → 케이스)."""
        detail = getattr(self, "detect_detail_dir", None)
        if detail and os.path.isdir(detail):
            return os.path.abspath(detail)
        csv_path = getattr(self, "csv_path", None)
        if csv_path and os.path.exists(csv_path):
            return os.path.dirname(os.path.abspath(csv_path))
        case = getattr(self, "case_direc", None)
        if case and os.path.isdir(case):
            return os.path.abspath(case)
        return os.path.abspath(".")

    def _detect_run_metadata(self, file_paths, run_all_models=False, detail_folders=None):
        """논문/검증용 Detect 실행 메타데이터."""
        binstat = getattr(self, "binButton_2", None) and self.binButton_2.isChecked()
        mulstat = getattr(self, "mulButton_2", None) and self.mulButton_2.isChecked()
        if binstat:
            class_mode = "Binary"
        elif mulstat:
            class_mode = "Multi-class"
        else:
            class_mode = "Unknown"
        case = getattr(self, "case_direc", "") or ""
        detail_list = detail_folders or []
        if not detail_list and getattr(self, "detect_detail_dir", None):
            detail_list = [getattr(self, "detect_detail_dir")]
        detail_labels = []
        for d in detail_list:
            if not d:
                continue
            try:
                detail_labels.append(os.path.relpath(d, case) if case else d)
            except Exception:
                detail_labels.append(d)
        return {
            "run_mode": "all_detail_folders" if run_all_models else "single_model",
            "model_selection": (
                self._detect_all_models_label()
                if run_all_models or self._detect_run_all_models_selected()
                else getattr(self, "aimodel", "")
            ),
            "detail_folders": detail_labels,
            "classification_mode": class_mode,
            "num_files": len(file_paths or []),
            "file_paths": [os.path.abspath(p) for p in (file_paths or [])],
            "case_directory": getattr(self, "case_direc", "") or "",
            "detail_directory": getattr(self, "detect_detail_dir", "") or "",
            "csv_path": getattr(self, "csv_path", "") or "",
            "mapping_json_path": getattr(self, "mapping_json_path", "") or "",
            "feature_states": {
                "structure_val": getattr(self, "structure_val_state", None),
                "structure_seq": getattr(self, "structure_seq_state", None),
                "frame_sps": getattr(self, "frame_sps_state", None),
                "frame_gop": getattr(self, "frame_gop_state", None),
                "frame_ratio": getattr(self, "frame_ratio_state", None),
            },
        }

    def _detect_prediction_record(self, df, file_path, model_name=None, model_pkl=None, detail_folder=None):
        """Detect 예측 1건을 논문/검증용 구조화 레코드로 변환."""
        detail_folder = detail_folder or getattr(self, "detect_detail_dir", "") or ""
        case = getattr(self, "case_direc", "") or ""
        detail_label = detail_folder
        if case and detail_folder:
            try:
                detail_label = os.path.relpath(detail_folder, case)
            except Exception:
                pass
        pkl_path = model_pkl or getattr(self, "pklpath", "") or ""
        resolved_name = (model_name or "").strip()
        if pkl_path and (not resolved_name or re.match(r"^\d+$", resolved_name)):
            resolved_name = self._detect_record_model_name(pkl_path)
        if not resolved_name:
            combo_name = getattr(self, "aimodel", "") or ""
            if combo_name and combo_name != self._detect_all_models_label():
                resolved_name = combo_name
        return {
            "detail_folder": detail_label,
            "detail_folder_path": detail_folder,
            "file_path": os.path.abspath(file_path) if file_path else "",
            "file_name": os.path.basename(file_path) if file_path else "",
            "model_name": resolved_name,
            "model_algorithm": self._detect_model_family_from_path(pkl_path) if pkl_path else "",
            "predicted_label": str(df.attrs.get("detect_predicted_label", "")),
            "predicted_label_name": str(df.attrs.get("detect_predicted_label_name", "")),
            "probability_percent": str(df.attrs.get("detect_probability_percent", "")),
            "summary": str(df.attrs.get("detect_result_summary", "")),
            "model_pkl": pkl_path,
            "scaler_pkl": getattr(self, "scalerpath", "") or "",
            "label_encoder_pkl": getattr(self, "label_encoder_path", "") or "",
            "feature_json": getattr(self, "detect_feature_json_path", "") or "",
            "label_mapping_json": (
                getattr(self, "detect_mapping_json_path", "")
                or getattr(self, "mapping_json_path", "")
                or ""
            ),
        }

    def _save_detect_run(self, records, run_meta, summary_text, save_base_dir=None):
        """Detect 실행 결과를 detect_results/ 및 detect_history.json에 저장."""
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_dir = save_base_dir or self._detect_results_base_dir()
            run_dir = os.path.join(base_dir, "detect_results", ts)
            os.makedirs(run_dir, exist_ok=True)

            if records:
                pd.DataFrame(records).to_csv(
                    os.path.join(run_dir, "results.csv"),
                    index=False,
                    encoding="utf-8-sig",
                )

            with open(os.path.join(run_dir, "summary.txt"), "w", encoding="utf-8") as f:
                f.write(summary_text or "")

            run_info = {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "run_id": ts,
                "output_dir": run_dir,
                "num_records": len(records or []),
                **(run_meta or {}),
            }
            with open(os.path.join(run_dir, "run_info.json"), "w", encoding="utf-8") as f:
                json.dump(run_info, f, ensure_ascii=False, indent=2)

            history_file = os.path.join(base_dir, "detect_history.json")
            history = []
            if os.path.exists(history_file):
                try:
                    with open(history_file, "r", encoding="utf-8") as f:
                        history = json.load(f)
                    if not isinstance(history, list):
                        history = []
                except Exception:
                    history = []
            history.append({
                "timestamp": run_info["timestamp"],
                "run_id": ts,
                "output_dir": run_dir,
                "run_mode": run_meta.get("run_mode", ""),
                "model_selection": run_meta.get("model_selection", ""),
                "num_files": run_meta.get("num_files", 0),
                "num_records": len(records or []),
                "case_directory": run_meta.get("case_directory", ""),
                "detail_directory": run_meta.get("detail_directory", ""),
                "csv_path": run_meta.get("csv_path", ""),
                "results_csv": os.path.join(run_dir, "results.csv"),
                "summary_txt": os.path.join(run_dir, "summary.txt"),
            })
            with open(history_file, "w", encoding="utf-8") as f:
                json.dump(history, f, ensure_ascii=False, indent=2)

            print(f"[INFO] Detect 결과 저장: {run_dir}")
            return run_dir
        except Exception as e:
            print(f"[ERROR] Detect 결과 저장 실패: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _setup_ground_truth_compare_button(self):
        """Detect 버튼 아래에 VISION 정답지 비교 버튼 추가."""
        try:
            ref = getattr(self, "class_detect", None)
            if ref is None:
                return
            parent = ref.parent() or self
            btn = QPushButton("정답지 비교 (VISION)", parent)
            btn.setToolTip(
                "VISION_base_files.txt 정답지와 Detect results.csv를 비교해\n"
                "정답/오탐(iOS·Android) 및 모델별 정확도를 분석합니다."
            )
            btn.setGeometry(ref.x(), ref.y() + ref.height() + 6, ref.width(), 36)
            try:
                btn.setStyleSheet(ref.styleSheet())
            except Exception:
                pass
            btn.clicked.connect(self.compare_detect_with_vision_ground_truth)
            btn.show()
            self.btn_compare_vision_gt = btn
        except Exception as e:
            print(f"[WARN] 정답지 비교 버튼 추가 실패: {e}")

    def _default_vision_ground_truth_path(self):
        path = _VISION_GROUND_TRUTH_DEFAULT
        if os.path.isfile(path):
            return path
        ds = getattr(self, "dataset_direc", "") or ""
        if ds:
            alt = os.path.join(ds, "VISION_base_files.txt")
            if os.path.isfile(alt):
                return alt
            alt2 = os.path.join(os.path.dirname(ds), "VISION_base_files.txt")
            if os.path.isfile(alt2):
                return alt2
        return path

    def _vision_gt_label_from_device_folder(self, device_folder_name):
        """
        URL 경로의 D##_기기명 폴더 기준 OS 라벨.
        예: .../D01_Samsung_GalaxyS3Mini/... → Android
            .../D02_Apple_iPhone4s/...       → iOS
        """
        name = (device_folder_name or "").strip()
        if not name:
            return -1, "Unknown"
        if name.startswith("Apple_"):
            return 1, "iOS"
        if name.startswith("Microsoft_") or name.startswith("BlackBerry_"):
            return -1, "Other"
        # VISION 이진 분류: Apple 외 스마트폰/태블릿 → Android
        return 0, "Android"

    def _parse_vision_ground_truth_line(self, line):
        """VISION_base_files.txt 한 줄 → URL 경로 전체 파싱."""
        line = (line or "").strip()
        if not line:
            return None
        m = re.search(
            r"/VISION/dataset/(D\d+)_([^/]+)/(images|videos)/([^/]+)/([^/?#\s]+)\s*$",
            line,
            re.IGNORECASE,
        )
        if not m:
            return None
        dcode, dname, media, scene, fname = m.groups()
        gt_label, gt_name = self._vision_gt_label_from_device_folder(dname)
        rel_path = f"{media}/{scene}/{fname}"
        return {
            "device_code": dcode,
            "device_name": dname,
            "media_type": media,
            "scene": scene,
            "file_name": fname,
            "relative_path": rel_path,
            "dataset_key": rel_path.replace("\\", "/").lower(),
            "source_url": line,
            "gt_label": gt_label,
            "gt_label_name": gt_name,
        }

    def _load_vision_ground_truth_index(self, gt_path):
        """
        VISION_base_files.txt 전체 파싱.
        - device_registry: URL 경로 D##_기기명 (images+videos 전체)
        - by_name: 동영상 파일명 → 경로 메타
        - by_relpath: videos/장면/파일명 → 경로 메타
        """
        if not gt_path or not os.path.isfile(gt_path):
            raise FileNotFoundError(f"정답지 파일을 찾을 수 없습니다: {gt_path}")

        by_name = {}
        by_relpath = {}
        device_registry = {}
        stats = {
            "total_lines": 0,
            "parsed_lines": 0,
            "video_entries": 0,
            "image_entries": 0,
            "unparsed_lines": 0,
        }

        with open(gt_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                stats["total_lines"] += 1
                entry = self._parse_vision_ground_truth_line(line)
                if not entry:
                    stats["unparsed_lines"] += 1
                    continue
                stats["parsed_lines"] += 1
                dcode = entry["device_code"]
                device_registry[dcode] = {
                    "device_code": dcode,
                    "device_name": entry["device_name"],
                    "gt_label": entry["gt_label"],
                    "gt_label_name": entry["gt_label_name"],
                }
                if entry["media_type"].lower() == "videos":
                    stats["video_entries"] += 1
                    by_name[entry["file_name"].lower()] = entry
                    by_relpath[entry["dataset_key"]] = entry
                else:
                    stats["image_entries"] += 1

        return {
            "by_name": by_name,
            "by_relpath": by_relpath,
            "device_registry": device_registry,
            "stats": stats,
        }

    def _lookup_vision_ground_truth(self, gt_bundle, file_name="", file_path=""):
        """Detect 결과 행 → 정답지 매칭 (파일명 → 상대경로 → D코드 순)."""
        by_name = gt_bundle.get("by_name", {})
        by_relpath = gt_bundle.get("by_relpath", {})
        device_registry = gt_bundle.get("device_registry", {})

        fname = (file_name or "").strip()
        if not fname and file_path:
            fname = os.path.basename(str(file_path))
        fp_norm = str(file_path or "").replace("\\", "/")

        if fname:
            hit = by_name.get(fname.lower())
            if hit:
                return hit, "file_name"

        if fp_norm:
            m = re.search(r"((?:videos|images)/[^/]+/[^/]+)$", fp_norm, re.IGNORECASE)
            if m:
                rel_key = m.group(1).lower()
                hit = by_relpath.get(rel_key)
                if hit:
                    return hit, "relative_path"

        m = re.match(r"(D\d+)_", fname, re.IGNORECASE)
        if m:
            dcode = m.group(1).upper()
            if dcode in device_registry:
                reg = device_registry[dcode]
                return {
                    "device_code": dcode,
                    "device_name": reg["device_name"],
                    "media_type": "",
                    "scene": "",
                    "file_name": fname,
                    "relative_path": "",
                    "dataset_key": "",
                    "source_url": "",
                    "gt_label": reg["gt_label"],
                    "gt_label_name": reg["gt_label_name"],
                }, "device_code_from_filename"

        return None, "no_match"

    def _normalize_detect_label_int(self, label_val, label_name=""):
        """예측 라벨을 0(Android)/1(iOS) 정수로 통일."""
        if label_name:
            ln = str(label_name).strip().lower()
            if ln in ("ios", "iphone", "apple", "1"):
                return 1
            if ln in ("android", "0"):
                return 0
        if label_val is None or (isinstance(label_val, float) and pd.isna(label_val)):
            return None
        s = str(label_val).strip()
        if not s:
            return None
        if s.lower() in ("ios", "iphone", "apple"):
            return 1
        if s.lower() == "android":
            return 0
        try:
            v = int(float(s))
            if v in (0, 1):
                return v
        except (TypeError, ValueError):
            pass
        return None

    def _find_latest_detect_results_csv(self):
        """가장 최근 Detect results.csv 경로."""
        bases = []
        case = getattr(self, "case_direc", None)
        if case:
            bases.append(case)
        try:
            bases.append(self._detect_results_base_dir())
        except Exception:
            pass
        seen = set()
        for base in bases:
            if not base:
                continue
            base = os.path.abspath(base)
            if base in seen:
                continue
            seen.add(base)
            history_file = os.path.join(base, "detect_history.json")
            if os.path.isfile(history_file):
                try:
                    with open(history_file, "r", encoding="utf-8") as f:
                        history = json.load(f)
                    if isinstance(history, list) and history:
                        latest = max(history, key=lambda x: str(x.get("run_id", "")))
                        csv_path = latest.get("results_csv")
                        if csv_path and os.path.isfile(csv_path):
                            return csv_path
                except Exception:
                    pass
            pattern = os.path.join(base, "detect_results", "*", "results.csv")
            found = glob.glob(pattern)
            if found:
                return max(found, key=os.path.getmtime)
        return None

    def _classify_detect_verdict(self, pred_int, gt_int):
        """정답/오탐/미검 분류."""
        if gt_int is None or gt_int < 0:
            return "GT미분류"
        if pred_int is None:
            return "예측없음"
        if pred_int == gt_int:
            return "정답"
        if pred_int == 1 and gt_int == 0:
            return "오탐(iOS)"  # Android인데 iOS로 예측
        if pred_int == 0 and gt_int == 1:
            return "미검(iOS)"  # iOS인데 Android로 예측
        return "불일치"

    def _compare_detect_results_with_vision_gt(self, results_csv, gt_path):
        """results.csv와 VISION 정답지 비교 후 enriched CSV·요약 저장."""
        gt_bundle = self._load_vision_ground_truth_index(gt_path)
        gt_stats = gt_bundle.get("stats", {})
        df = pd.read_csv(results_csv, encoding="utf-8-sig")
        if df.empty:
            raise ValueError("results.csv가 비어 있습니다.")

        rows = []
        match_methods = Counter()
        for _, row in df.iterrows():
            fname = str(row.get("file_name", "") or "").strip()
            fpath = str(row.get("file_path", "") or "").strip()
            gt, match_method = self._lookup_vision_ground_truth(gt_bundle, fname, fpath)
            match_methods[match_method] += 1

            pred_int = self._normalize_detect_label_int(
                row.get("predicted_label"),
                row.get("predicted_label_name", ""),
            )
            if gt:
                gt_int = gt["gt_label"]
                gt_name = gt["gt_label_name"]
                gt_device = gt["device_code"]
                gt_device_name = gt["device_name"]
                gt_url = gt.get("source_url", "")
                gt_media = gt.get("media_type", "")
                gt_scene = gt.get("scene", "")
                gt_rel_path = gt.get("relative_path", "")
            else:
                gt_int = None
                gt_name = ""
                gt_device = ""
                gt_device_name = ""
                gt_url = ""
                gt_media = ""
                gt_scene = ""
                gt_rel_path = ""

            verdict = self._classify_detect_verdict(pred_int, gt_int)
            is_correct = verdict == "정답"
            enriched = {k: row.get(k, "") for k in df.columns}
            enriched.update({
                "gt_device_code": gt_device,
                "gt_device_name": gt_device_name,
                "gt_media_type": gt_media,
                "gt_scene": gt_scene,
                "gt_relative_path": gt_rel_path,
                "gt_label": "" if gt_int is None else gt_int,
                "gt_label_name": gt_name,
                "pred_label_int": "" if pred_int is None else pred_int,
                "is_correct": is_correct,
                "verdict": verdict,
                "ground_truth_url": gt_url,
                "gt_match_method": match_method,
                "in_ground_truth_file": match_method in ("file_name", "relative_path"),
            })
            rows.append(enriched)

        out_df = pd.DataFrame(rows)
        run_dir = os.path.dirname(os.path.abspath(results_csv))
        out_csv = os.path.join(run_dir, "comparison_with_ground_truth.csv")
        out_df.to_csv(out_csv, index=False, encoding="utf-8-sig")

        registry_path = os.path.join(run_dir, "vision_device_registry.json")
        with open(registry_path, "w", encoding="utf-8") as f:
            json.dump(
                gt_bundle.get("device_registry", {}),
                f,
                ensure_ascii=False,
                indent=2,
            )

        eval_df = out_df[
            out_df["verdict"].isin(["정답", "오탐(iOS)", "미검(iOS)"])
        ].copy()
        total_eval = len(eval_df)
        correct = int((eval_df["verdict"] == "정답").sum())
        fp = int((eval_df["verdict"] == "오탐(iOS)").sum())
        fn = int((eval_df["verdict"] == "미검(iOS)").sum())
        accuracy = (correct / total_eval * 100.0) if total_eval else 0.0

        no_gt = int((out_df["verdict"] == "GT미분류").sum())
        no_match = int((out_df["gt_match_method"] == "no_match").sum())
        no_pred = int((out_df["verdict"] == "예측없음").sum())

        summary_lines = [
            "VISION 정답지 vs Detect 결과 비교",
            "=" * 50,
            f"정답지: {gt_path}",
            f"결과 CSV: {results_csv}",
            "",
            "[정답지 파싱 (URL 경로 기준)]",
            f"  전체 줄: {gt_stats.get('total_lines', 0)}",
            f"  파싱 성공: {gt_stats.get('parsed_lines', 0)}",
            f"  동영상 entries: {gt_stats.get('video_entries', 0)}",
            f"  이미지 entries: {gt_stats.get('image_entries', 0)} (기기 레지스트리용)",
            f"  등록 기기(D01~): {len(gt_bundle.get('device_registry', {}))}종",
            f"  파싱 실패 줄: {gt_stats.get('unparsed_lines', 0)}",
            "",
            "[매칭 방식]",
        ]
        for method, cnt in match_methods.most_common():
            summary_lines.append(f"  {method}: {cnt}")
        summary_lines.extend([
            "",
            f"총 비교 행: {len(out_df)}",
            f"비교 가능(정답·오탐·미검): {total_eval}",
            f"  정답: {correct}",
            f"  오탐(iOS, Android→iOS): {fp}",
            f"  미검(iOS, iOS→Android): {fn}",
            f"정확도: {accuracy:.2f}%",
            f"정답지 미매칭: {no_match}",
            f"GT 미분류(Other): {no_gt}",
            f"예측 라벨 없음: {no_pred}",
            "",
            "※ 정답 OS = URL 경로의 D##_기기명 폴더 기준",
            "  (Apple_* → iOS, 그 외 VISION 기기 → Android, Microsoft_* → 제외)",
            "",
            "[모델별 정확도]",
        ])
        if "model_name" in eval_df.columns and total_eval:
            for model_name, grp in eval_df.groupby("model_name", dropna=False):
                n = len(grp)
                c = int((grp["verdict"] == "정답").sum())
                fpi = int((grp["verdict"] == "오탐(iOS)").sum())
                fni = int((grp["verdict"] == "미검(iOS)").sum())
                acc = c / n * 100.0 if n else 0.0
                summary_lines.append(
                    f"  {model_name}: {acc:.1f}% ({c}/{n}) | 오탐 {fpi} | 미검 {fni}"
                )
        if "detail_folder" in eval_df.columns and total_eval:
            summary_lines.append("")
            summary_lines.append("[상세폴더별 정확도]")
            for detail, grp in eval_df.groupby("detail_folder", dropna=False):
                n = len(grp)
                c = int((grp["verdict"] == "정답").sum())
                acc = c / n * 100.0 if n else 0.0
                summary_lines.append(f"  {detail}: {acc:.1f}% ({c}/{n})")

        summary_text = "\n".join(summary_lines)
        summary_path = os.path.join(run_dir, "comparison_summary.txt")
        with open(summary_path, "w", encoding="utf-8") as f:
            f.write(summary_text)

        meta = {
            "ground_truth_path": gt_path,
            "results_csv": results_csv,
            "comparison_csv": out_csv,
            "comparison_summary": summary_path,
            "device_registry_json": registry_path,
            "ground_truth_stats": gt_stats,
            "match_methods": dict(match_methods),
            "accuracy_percent": round(accuracy, 4),
            "total_rows": len(out_df),
            "evaluated_rows": total_eval,
            "correct": correct,
            "false_positive_ios": fp,
            "false_negative_ios": fn,
        }
        meta_path = os.path.join(run_dir, "comparison_meta.json")
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)

        return {
            "comparison_csv": out_csv,
            "summary_path": summary_path,
            "summary_text": summary_text,
            "meta": meta,
        }

    def compare_detect_with_vision_ground_truth(self):
        """UI: Detect results.csv와 VISION 정답지 비교."""
        default_gt = self._default_vision_ground_truth_path()
        gt_path, _ = QFileDialog.getOpenFileName(
            self,
            "VISION 정답지 파일 선택",
            os.path.dirname(default_gt) if default_gt else "",
            "Text Files (*.txt);;All Files (*)",
        )
        if not gt_path:
            gt_path = default_gt
        if not os.path.isfile(gt_path):
            self._copyable_msg(
                QMessageBox.Warning,
                "정답지 없음",
                f"정답지 파일을 찾을 수 없습니다.\n{gt_path}",
            )
            return

        latest = self._find_latest_detect_results_csv()
        start_dir = os.path.dirname(latest) if latest else (
            os.path.join(getattr(self, "case_direc", "") or "", "detect_results")
        )
        results_csv, _ = QFileDialog.getOpenFileName(
            self,
            "비교할 Detect results.csv 선택",
            latest or start_dir,
            "CSV Files (*.csv);;All Files (*)",
        )
        if not results_csv:
            if latest:
                results_csv = latest
            else:
                self._copyable_msg(
                    QMessageBox.Warning,
                    "결과 없음",
                    "비교할 results.csv를 선택하세요.\n먼저 Detect를 실행해 결과를 저장해야 합니다.",
                )
                return
        if not os.path.isfile(results_csv):
            self._copyable_msg(QMessageBox.Warning, "오류", "선택한 results.csv가 없습니다.")
            return

        try:
            out = self._compare_detect_results_with_vision_gt(results_csv, gt_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._copyable_msg(
                QMessageBox.Critical,
                "비교 실패",
                f"정답지 비교 중 오류가 발생했습니다.\n{e}",
            )
            return

        msg = (
            f"{out['summary_text']}\n\n"
            f"[저장 파일]\n"
            f"- {out['comparison_csv']}\n"
            f"- {out['summary_path']}"
        )
        self._copyable_msg(QMessageBox.Information, "정답지 비교 완료", msg)
        try:
            self.statusBar().showMessage(
                f"정답지 비교 완료 — 정확도 {out['meta']['accuracy_percent']:.2f}%",
                10000,
            )
        except Exception:
            pass

    def _detect_candidate_file_paths(self):
        """Detect 대상 파일 목록(CSV 제외, 존재하는 파일만)."""
        paths = []
        seen = set()
        for raw in getattr(self, "file_paths", []) or []:
            if not raw:
                continue
            path = os.path.abspath(os.path.normpath(str(raw)))
            if path in seen or not os.path.isfile(path):
                continue
            if os.path.splitext(path)[1].lower() == ".csv":
                continue
            seen.add(path)
            paths.append(path)
        return paths

    def _run_detect_on_files(self, file_paths, run_all_models=False):
        """선택된 여러 파일에 대해 Detect 실행 후 결과를 한 창에 표시."""
        paths = list(file_paths or [])
        if not paths:
            try:
                self.show_alert("먼저 좌측 트리/파일 목록에서 동영상 파일을 선택해 주세요.")
            except Exception:
                self._copyable_msg(
                    QMessageBox.Warning,
                    "오류",
                    "먼저 좌측 트리/파일 목록에서 동영상 파일을 선택해 주세요.",
                )
            return

        self.detectmode = 1
        run_all = bool(run_all_models or self._detect_run_all_models_selected())
        result_lines = []
        error_lines = []
        all_records = []
        processed_detail_dirs = []

        if run_all:
            detail_dirs = self._case_feature_detail_dirs()
            if not detail_dirs:
                self.detectmode = 0
                self._copyable_msg(
                    QMessageBox.Warning,
                    "Detect 결과",
                    "현재 케이스에서 Detect 가능한 feature 상세 폴더를 찾지 못했습니다.\n"
                    "(training_results 안에 model.pkl/scaler.pkl 쌍이 있어야 합니다.)",
                )
                return
            case = getattr(self, "case_direc", "") or ""
            title = "전체 상세폴더 Detect 결과"
            result_lines.append(f"{title} ({len(paths)}개 파일 × {len(detail_dirs)}개 상세폴더)")
            result_lines.append("")
        else:
            try:
                if getattr(self, "structure_val_but", None) is not None:
                    self.structure_val_state = 1 if self.structure_val_but.isChecked() else 0
                    self.structure_seq_state = 1 if self.structure_seq_but.isChecked() else 0
                    self.frame_sps_state = 1 if self.frame_sps_but.isChecked() else 0
                    self.frame_gop_state = 1 if self.frame_gop_but.isChecked() else 0
                    self.frame_ratio_state = 1 if self.frame_ratio_but.isChecked() else 0
            except Exception:
                pass
            try:
                self.load_or_initialize_states()
            except Exception:
                pass
            if getattr(self, "detect_detail_dir", None):
                self._apply_detect_detail_context(self.detect_detail_dir)
            result_lines.append(f"Detect 결과 ({len(paths)}개 파일)")
            result_lines.append("")

        self._begin_batch_run()
        try:
            if run_all:
                for detail_dir in detail_dirs:
                    processed_detail_dirs.append(detail_dir)
                    case = getattr(self, "case_direc", "") or ""
                    try:
                        rel_detail = os.path.relpath(detail_dir, case) if case else os.path.basename(detail_dir)
                    except Exception:
                        rel_detail = os.path.basename(detail_dir)
                    result_lines.append(f"######## [{rel_detail}] ########")
                    result_lines.append("")
                    try:
                        self._apply_detect_detail_context(detail_dir)
                    except Exception as e:
                        error_lines.append(f"{rel_detail}: 상세 폴더 설정 실패 - {e}")
                        result_lines.append("")
                        continue

                    for idx, file_path in enumerate(paths, 1):
                        name = os.path.basename(file_path)
                        result_lines.append(f"=== [{idx}/{len(paths)}] {name} ===")
                        try:
                            feature_data = self.extract_features_from_file(file_path)
                            structured_data = self.flatten_features_for_prediction(feature_data)
                            out = self.predict_all_detect_models(
                                structured_data,
                                file_path,
                                show_message=False,
                                compact=True,
                            )
                            block = ""
                            if isinstance(out, dict):
                                block = str(out.get("message", "")).strip()
                                all_records.extend(out.get("records", []) or [])
                            elif isinstance(out, str) and out.strip():
                                block = out.strip()
                            if block:
                                result_lines.append(block)
                            else:
                                result_lines.append("예측 결과를 만들지 못했습니다.")
                        except ValueError as ve:
                            if "Found array with 0 sample(s)" in str(ve):
                                err = (
                                    f"{rel_detail} / {name}: 피처 상태가 비어 있습니다. "
                                    "해당 상세 폴더 config/states.json을 확인하세요."
                                )
                            else:
                                err = f"{rel_detail} / {name}: {ve}"
                            error_lines.append(err)
                        except Exception as e:
                            error_lines.append(f"{rel_detail} / {name}: {e}")
                        result_lines.append("")
            else:
                self.load_model_and_scaler()
                for idx, file_path in enumerate(paths, 1):
                    name = os.path.basename(file_path)
                    result_lines.append(f"=== [{idx}/{len(paths)}] {name} ===")
                    try:
                        feature_data = self.extract_features_from_file(file_path)
                        structured_data = self.flatten_features_for_prediction(feature_data)
                        df = self.predict_data1(
                            structured_data,
                            file_path,
                            skip_simhash=False,
                            show_message=False,
                            show_warnings=False,
                        )
                        all_records.append(self._detect_prediction_record(
                            df,
                            file_path,
                            model_pkl=getattr(self, "pklpath", ""),
                            detail_folder=getattr(self, "detect_detail_dir", ""),
                        ))
                        summary = df.attrs.get("detect_result_summary", "")
                        paths_text = df.attrs.get("detect_used_paths_text", "")
                        if summary:
                            result_lines.append(summary)
                        else:
                            result_lines.append("예측 결과를 만들지 못했습니다.")
                        if paths_text:
                            result_lines.append(paths_text)
                    except ValueError as ve:
                        if "Found array with 0 sample(s)" in str(ve):
                            err = (
                                "피처 상태가 비어 있습니다. Create 탭에서 피처를 선택하고 "
                                "'save state'로 저장하세요."
                            )
                        else:
                            err = str(ve)
                        error_lines.append(f"{name}: {err}")
                    except Exception as e:
                        error_lines.append(f"{name}: {e}")
                    result_lines.append("")
        finally:
            batch_log = self._end_batch_run()
            self.detectmode = 0

        if error_lines:
            result_lines.append("[실패한 파일]")
            result_lines.extend(error_lines)
        if batch_log:
            result_lines.append("")
            result_lines.append("[참고 메시지]")
            result_lines.extend(batch_log)

        message = "\n".join(result_lines).strip()
        print(message)
        run_meta = self._detect_run_metadata(
            paths,
            run_all_models=run_all,
            detail_folders=processed_detail_dirs if run_all else None,
        )
        save_base = getattr(self, "case_direc", None) if run_all else None
        saved_dir = self._save_detect_run(all_records, run_meta, message, save_base_dir=save_base)
        if saved_dir:
            message += f"\n\n[기록 저장]\n{saved_dir}\n- results.csv (표 형식)\n- summary.txt\n- run_info.json\n- detect_history.json (누적 이력)"
        dialog_title = "전체 상세폴더 Detect 결과" if run_all else "Detect 결과"
        self._copyable_msg(QMessageBox.Information, dialog_title, message)

    def load_file_for_prediction(self):
        """목록에 선택된 파일(들)에 대해 Detect 실행."""
        paths = self._detect_candidate_file_paths()
        self._run_detect_on_files(paths, run_all_models=False)

    def load_file_for_prediction_all_models(self):
        """목록에 선택된 파일(들)을 현재 상세 폴더의 전체 Detect 모델로 예측."""
        paths = self._detect_candidate_file_paths()
        self._run_detect_on_files(paths, run_all_models=True)

    def move_label_to_second_column(self, df):
        # 'label' 컬럼을 분리합니다.
        label_column = df.pop('label')

        # 'label' 컬럼을 두 번째 위치로 삽입합니다.
        df.insert(1, 'label', label_column)

        return df

    def predict_on_file(self, file_path, run_all_models=False):
        """단일 파일 Detect (저장·결과 창 표시는 _run_detect_on_files와 동일)."""
        if not file_path:
            return
        self._run_detect_on_files([file_path], run_all_models=run_all_models)

    def list_to_dict(self, data):
        """Convert a list of strings into a dictionary by identifying 'key: value' pairs."""
        data_dict = {}
        for index, item in enumerate(data):
            # If the item is a string with colon-separated pairs
            if isinstance(item, str) and ":" in item:
                attributes = [attr.strip() for attr in item.split(",")]
                for attr in attributes:
                    if ":" in attr:
                        key, value = attr.split(":", 1)
                        data_dict[key.strip()] = value.strip()
                    else:
                        # If parsing fails, add the entire item as a single entry
                        data_dict[key] = item
            else:
                # Add non-string items or items without ':' as is
                data_dict[key] = item
        return data_dict

    def flatten_features_for_prediction(self, data):
        flattened_rows = []

        global_fieldnames = set()

        for file_data in data:
            flattened = {}
            key_count_local = {}  # Local key duplication count for each file
            for key, value in file_data:
                if key != 'GOP':
                    if key in key_count_local:
                        key_count_local[key] += 1
                        key_with_count = f"{key}({key_count_local[key]})"
                    else:
                        key_count_local[key] = 1
                        key_with_count = key

                    if isinstance(value, str) and ":" in value:
                        # SPS는 analyzesps가 반환한 "profile_idc:66, constraint_set0_flag:0, ..." 형식
                        # 이 경우 key='SPS', value='profile_idc:66, constraint_set0_flag:0, ...'
                        if key == 'SPS' and ',' in value:
                            # SPS 결과를 파싱: "profile_idc:66, constraint_set0_flag:0, ..."
                            sps_attrs = [attr.strip() for attr in value.split(",")]
                            for sps_attr in sps_attrs:
                                if ":" in sps_attr:
                                    sps_name, sps_val = sps_attr.split(":", 1)
                                    sps_name_clean = sps_name.strip()
                                    sps_val_clean = sps_val.strip()
                                    # SPS_profile_idc, SPS_constraint_set0_flag 등으로 저장
                                    sps_field_name = f"{key}_{sps_name_clean}"
                                    flattened[sps_field_name] = sps_val_clean
                                    global_fieldnames.add(sps_field_name)
                        else:
                            # 일반적인 박스 속성 파싱: "Entry Count: 1, Entries: [...]" 형식
                            attributes = [attr.strip() for attr in value.split(",")]
                            for attr in attributes:
                                if ":" in attr:
                                    attr_name, attr_value = attr.split(":", 1)
                                    attr_name_clean = attr_name.strip()
                                    attr_value_clean = attr_value.strip()
                                    
                                    # Entries가 리스트 문자열인 경우 추가 파싱 (학습 시와 동일하게)
                                    # 예: "Entries: ['Duration: 123, Media Time: 456, Rate: 789']"
                                    if attr_name_clean == "Entries" and attr_value_clean.startswith("['") and "Duration:" in attr_value_clean:
                                        # 리스트 내의 각 entry를 파싱
                                        entries_str = attr_value_clean.strip("[]'\"")
                                        entry_parts = entries_str.split(",")
                                        for entry_part in entry_parts:
                                            if ":" in entry_part:
                                                entry_key, entry_val = entry_part.split(":", 1)
                                                entry_key_clean = entry_key.strip()
                                                entry_val_clean = entry_val.strip()
                                                # 학습 CSV에는 Duration이 별도 컬럼으로 없으므로 Media Time과 Rate만 저장
                                                # Duration은 Entries 안에만 있고 별도 컬럼으로 분리되지 않음
                                                if entry_key_clean in ["Media Time", "Rate"]:
                                                    entry_field_name = f"{key_with_count}_{entry_key_clean}"
                                                    flattened[entry_field_name] = entry_val_clean
                                                    global_fieldnames.add(entry_field_name)
                                        # Entry Count는 이미 위에서 파싱되었으므로 덮어쓰지 않음
                                        # (Entry Count는 "Entry Count: 1" 형식으로 이미 flattened에 저장됨)
                                    else:
                                        # 일반적인 속성 처리 (Entry Count, Media Time, Rate 등)
                                        attr_field_name = f"{key_with_count}_{attr_name_clean}"
                                        flattened[attr_field_name] = attr_value_clean
                                        global_fieldnames.add(attr_field_name)
                                else:
                                    # ":"가 없는 경우 원본 키로 저장
                                    flattened[key_with_count] = value
                                    global_fieldnames.add(key_with_count)
                    elif isinstance(value, list):
                        # 리스트 타입 처리 (학습 시와 동일)
                        for item in value:
                            if isinstance(item, str) and ":" in item:
                                # "Duration: 123, Media Time: 456" 형식 파싱
                                item_attrs = [ia.strip() for ia in item.split(",")]
                                for item_attr in item_attrs:
                                    if ":" in item_attr:
                                        item_key, item_val = item_attr.split(":", 1)
                                        item_key_clean = item_key.strip()
                                        item_val_clean = item_val.strip()
                                        item_field_name = f"{key_with_count}_{item_key_clean}"
                                        flattened[item_field_name] = item_val_clean
                                        global_fieldnames.add(item_field_name)
                            else:
                                flattened[key_with_count] = str(item)
                                global_fieldnames.add(key_with_count)
                    else:
                        flattened[key_with_count] = value
                        global_fieldnames.add(key_with_count)
                else:
                    # Handle the 'GOP' key separately
                    # GOP는 이미 "30,30,30" 형식의 쉼표로 구분된 문자열이므로 그대로 사용
                    flattened[key] = value if value else ''
                    global_fieldnames.add(key)

            flattened_rows.append(flattened)

        df = pd.DataFrame(flattened_rows)

        # 학습 CSV(save_to_csv)는 raw 값 그대로 저장하므로 변환 적용하지 않음
        # (adjust_* 적용 시 학습값 87, 3805416987 등이 1로 바뀌어 불일치 발생)
        # self.adjust_time_columns(df)
        # self.adjust_duration_columns(df)
        # self.adjust_dimensions(df)

        for field in global_fieldnames:
            if field not in df.columns:
                df[field] = None
        
        # 디버깅: flatten 후 피처 개수와 주요 키 확인
        print(f"[DEBUG] flatten_features_for_prediction 완료: {len(df.columns)}개 피처 생성")
        if len(df.columns) > 0:
            print(f"[DEBUG] 생성된 피처 키 (처음 30개): {list(df.columns[:30])}")
            # SPS 관련 피처 확인
            sps_features = [c for c in df.columns if 'SPS' in str(c)]
            print(f"[DEBUG] SPS 관련 피처: {len(sps_features)}개 - {sps_features[:10] if sps_features else '없음'}")
            # sequence 관련 피처 확인
            seq_features = [c for c in df.columns if 'sequence' in str(c).lower()]
            print(f"[DEBUG] sequence 관련 피처: {len(seq_features)}개 - {seq_features}")
            # elst 관련 피처 확인
            elst_features = [c for c in df.columns if 'elst' in str(c).lower()]
            print(f"[DEBUG] elst 관련 피처: {len(elst_features)}개 - {elst_features[:10] if elst_features else '없음'}")
            
            # Detect에서는 학습 CSV 데이터와 직접 비교하지 않는다.
            if False and hasattr(self, 'csv_path') and self.csv_path and os.path.exists(self.csv_path):
                try:
                    try:
                        train_df = pd.read_csv(self.csv_path, nrows=0, on_bad_lines='skip')
                    except:
                        try:
                            train_df = pd.read_csv(self.csv_path, nrows=0, error_bad_lines=False)
                        except:
                            train_df = None
                    
                    if train_df is not None:
                        train_cols = set([str(c) for c in train_df.columns if c not in ['name', 'md5', 'label']])
                        current_cols = set([str(c) for c in df.columns])
                        
                        missing_in_current = train_cols - current_cols
                        extra_in_current = current_cols - train_cols
                        
                        print(f"[WARN] 피처 개수 불일치 진단:")
                        print(f"  - 학습 CSV 피처: {len(train_cols)}개 (name, md5, label 제외)")
                        print(f"  - 현재 생성 피처: {len(current_cols)}개")
                        print(f"  - 누락 피처: {len(missing_in_current)}개")
                        print(f"  - 초과 피처: {len(extra_in_current)}개")
                        
                        if missing_in_current:
                            print(f"  - 누락 피처 예시 (처음 20개): {list(missing_in_current)[:20]}")
                            # 누락 피처 카테고리별 분석
                            missing_sps = [c for c in missing_in_current if 'SPS' in str(c)]
                            missing_elst = [c for c in missing_in_current if 'elst' in str(c).lower()]
                            missing_seq = [c for c in missing_in_current if 'sequence' in str(c).lower()]
                            missing_co64 = [c for c in missing_in_current if 'co64' in str(c).lower()]
                            missing_box = [c for c in missing_in_current if any(box in str(c).lower() for box in ['stco', 'ctts', 'sdtp', 'sgpd', 'sbgp', 'free'])]
                            
                            print(f"    - SPS 관련 누락: {len(missing_sps)}개 - {missing_sps[:10] if missing_sps else '없음'}")
                            print(f"    - elst 관련 누락: {len(missing_elst)}개 - {missing_elst[:5] if missing_elst else '없음'}")
                            print(f"    - sequence 관련 누락: {len(missing_seq)}개 - {missing_seq}")
                            print(f"    - co64 관련 누락: {len(missing_co64)}개 - {missing_co64[:5] if missing_co64 else '없음'}")
                            print(f"    - 기타 박스 누락: {len(missing_box)}개 - {missing_box[:10] if missing_box else '없음'}")
                except Exception as e:
                    print(f"[WARN] 학습 CSV와 피처 비교 실패: {e}")

        return df

        global_fieldnames = list(global_fieldnames)
        for row in flattened_rows:
            for field in global_fieldnames:
                if field not in row:
                    row[field] = None

        return flattened_rows

    def transform_in_flattened(self, flattened, pattern, transform_func):
        for key, value in flattened.items():
            if pattern.search(key):
                flattened[key] = transform_func(value)
        return flattened

    def transform_duration_entry(self, x):
        """Transformation logic for duration, entry, and entries columns."""
        if pd.notna(x):  # If value is not NaN, set it to 1
            return 1
        elif pd.isna(x) or x == '':  # If value is NaN or empty, set it to -1
            return -1
        return x  # Otherwise, keep the original value

    def transform_time_value(self, x):
        """Transformation logic for time-related columns (Create Time, Modify Time)."""
        if pd.notna(x) and len(str(x)) >= 4:
            x_str = str(x)
            if x_str.startswith('1'):
                return 1
            elif x_str.startswith('3'):
                return 3
            else:
                return 0
        return x  # Keep original value if it doesn't match

    def transform_width_height(self, x):
        x = int(x)
        if pd.notna(x):
            if x > 1:
                return 1
            elif x == 0:
                return 0
        elif pd.isna(x) or x == '':
            return -1
        return x  # keep the original value


    def extract_features_from_file(self, file_path):
        """
        예측 파이프라인(flatten_features_for_prediction)이 기대하는 형태로 반환합니다.
        - 반환 형태: [ [ (key, value), (key, value), ... ] ]  (파일 1개 기준)
        - 학습 시와 동일하게 extract_box_feature를 사용하여 모든 옵션(structure_val, structure_seq, frame_sps, frame_gop 등)이 조합된 피처를 추출합니다.
        - extract_box_feature()는 내부적으로 _process_one_file을 호출하여 모든 state_flags를 조합합니다.
        """
        # Detect 모드에서는 predict_on_file이 이미 325피처용 플래그를 강제 설정했으므로
        # load_or_initialize_states()를 호출하면 states.json이 그 값을 덮어써 103피처만 추출되는 버그 발생.
        # 따라서 detect 모드에서는 상태 로드를 건너뜁니다.
        if getattr(self, "detectmode", 0) != 1:
            self.load_or_initialize_states()

        # 학습 시와 동일하게 extract_box_feature를 사용
        # extract_box_feature는 내부적으로 _process_one_file을 호출하여
        # 모든 옵션(structure_val, structure_seq, frame_sps, frame_gop, frame_ratio)을 조합합니다
        print(f"[DEBUG] extract_features_from_file: extract_box_feature 호출 시작 (파일: {os.path.basename(file_path)})")
        print(f"[DEBUG] 현재 상태값: structure_val={self.structure_val_state}, structure_seq={self.structure_seq_state}, frame_sps={self.frame_sps_state}, frame_gop={self.frame_gop_state}, frame_ratio={self.frame_ratio_state}")
        
        box_features_list = self.extract_box_feature(file_path)  # e.g. [ [('name',..), ('md5',..), ...] ]
        
        # 디버깅: 추출된 피처 수 확인
        if isinstance(box_features_list, list) and len(box_features_list) > 0:
            num_features = len(box_features_list[0]) if box_features_list[0] else 0
            print(f"[DEBUG] extract_box_feature 완료: {len(box_features_list)}개 파일, 첫 번째 파일 피처 수: {num_features}개")
            if num_features > 0:
                feature_keys = [k for k, v in box_features_list[0]]
                print(f"[DEBUG] 추출된 주요 피처 키 (처음 20개): {feature_keys[:20]}")
                return box_features_list
        
        # 추출 실패 시 최소한 파일명만 반환
        print(f"[WARN] extract_box_feature 결과가 비어있음, 기본값 반환")
        return [[('name', os.path.basename(file_path)), ('md5', '')]]

    def select_model_scaler_from_menu(self):
        """File 메뉴에서 Model/Scaler를 선택하는 함수 (detect용)"""
        try:
            # detect 모드 확인
            binstat = self.binButton_2.isChecked()
            mulstat = self.mulButton_2.isChecked()
            if not binstat and not mulstat:
                self._copyable_msg(QMessageBox.Warning, "오류", "먼저 Detect 탭에서 바이너리/멀티 모드를 선택해주세요.")
                return
            
            # Model/Scaler 선택 (load_model_and_scaler의 로직 재사용)
            self._select_model_scaler_files()
            
        except Exception as e:
            self._copyable_msg(QMessageBox.Warning, "오류", f"Model/Scaler 선택 중 오류가 발생했습니다:\n{str(e)}")
            print(f"[ERROR] Model/Scaler 선택 오류: {e}")
    
    def _select_model_scaler_files(self):
        """Model과 Scaler 파일을 선택하는 내부 함수 (detect용)"""
        # 케이스 디렉토리에서 모든 model.pkl과 scaler.pkl 파일 찾기 (training_results 하위 포함)
        search_dirs = []
        if self.csv_path:
            search_dirs.append(os.path.dirname(os.path.abspath(self.csv_path)))
        if self.case_direc:
            search_dirs.append(os.path.abspath(self.case_direc))
        search_dirs.append(os.path.abspath("."))
        
        # 중복 제거
        search_dirs = list(set(search_dirs))
        
        model_files = []
        scaler_files = []
        
        for search_dir in search_dirs:
            if os.path.isdir(search_dir):
                model_pattern = os.path.join(search_dir, "*model.pkl")
                model_files.extend(glob.glob(model_pattern))
                scaler_pattern = os.path.join(search_dir, "*scaler.pkl")
                scaler_files.extend(glob.glob(scaler_pattern))
                # training_results/{run}/ 내 model.pkl, scaler.pkl
                tr_dir = os.path.join(search_dir, "training_results")
                if os.path.isdir(tr_dir):
                    for run_name in os.listdir(tr_dir):
                        run_path = os.path.join(tr_dir, run_name)
                        if os.path.isdir(run_path):
                            m, s = os.path.join(run_path, "model.pkl"), os.path.join(run_path, "scaler.pkl")
                            if os.path.isfile(m):
                                model_files.append(m)
                            if os.path.isfile(s):
                                scaler_files.append(s)
        
        model_files = list(set(model_files))
        scaler_files = list(set(scaler_files))
        
        if not model_files:
            self._copyable_msg(QMessageBox.Warning, "오류", "Model 파일을 찾을 수 없습니다.")
            return
        
        # Model 파일 선택
        selected_model_path = None
        if len(model_files) == 1:
            selected_model_path = model_files[0]
            print(f"[INFO] Model 파일 자동 선택: {os.path.basename(selected_model_path)}")
        else:
            model_names = [os.path.basename(f) for f in model_files]
            model_name, ok = QInputDialog.getItem(
                self, 
                "Model 선택", 
                "사용할 Model 파일을 선택하세요:",
                model_names, 
                0, 
                False
            )
            if not ok:
                return
            
            selected_model_path = next((f for f in model_files if os.path.basename(f) == model_name), None)
            if not selected_model_path:
                self._copyable_msg(QMessageBox.Warning, "오류", f"선택한 Model 파일을 찾을 수 없습니다: {model_name}")
                return
        
        # Model 파일명의 앞부분 추출 (scaler 매칭용 - 모델명 포함)
        model_basename = os.path.basename(selected_model_path)
        model_dir = os.path.dirname(selected_model_path)
        # training_results run 폴더 안의 model.pkl이면 같은 폴더의 scaler.pkl / label_encoder.pkl 사용
        is_run_folder_model = ("training_results" in selected_model_path and model_basename == "model.pkl")
        
        model_prefix_for_scaler = model_basename.replace("model.pkl", "").replace(".pkl", "")
        
        # 매핑 JSON용 prefix (모델명과 .csv 제거)
        model_prefix_for_json = model_prefix_for_scaler
        # 모델명 제거 (Xgboost, RandomForest 등)
        for model_name in ["Xgboost", "RandomForest", "LGBM", "LogisticRegression", "LSTM"]:
            if model_prefix_for_json.endswith(model_name):
                model_prefix_for_json = model_prefix_for_json[:-len(model_name)]
                break
        # 끝에 _ 가 남아있으면 먼저 제거 (모델명 제거 후 남은 _)
        if model_prefix_for_json.endswith("_"):
            model_prefix_for_json = model_prefix_for_json[:-1]
        # .csv 또는 .csv_ 제거 (Model 파일명에 .csv가 포함된 경우)
        if model_prefix_for_json.endswith(".csv_"):
            model_prefix_for_json = model_prefix_for_json[:-5]
        elif model_prefix_for_json.endswith(".csv"):
            model_prefix_for_json = model_prefix_for_json[:-4]
        # 끝에 _ 가 남아있으면 다시 제거
        if model_prefix_for_json.endswith("_"):
            model_prefix_for_json = model_prefix_for_json[:-1]
        
        matching_scaler = self._find_matching_scaler_for_model(selected_model_path, scaler_files)
        if not matching_scaler:
            same_dir_scalers = [
                f for f in scaler_files
                if os.path.dirname(os.path.abspath(f)) == os.path.abspath(model_dir)
            ]
            if same_dir_scalers:
                scaler_names = [os.path.basename(f) for f in same_dir_scalers]
                scaler_name, ok = QInputDialog.getItem(
                    self,
                    "Scaler 선택",
                    f"Model과 같은 폴더의 Scaler만 선택할 수 있습니다:\n{model_dir}",
                    scaler_names,
                    0,
                    False
                )
                if not ok:
                    return
                selected = next((f for f in same_dir_scalers if os.path.basename(f) == scaler_name), None)
                if selected and self._model_scaler_paths_compatible(selected_model_path, selected):
                    matching_scaler = selected
            if not matching_scaler:
                self._copyable_msg(
                    QMessageBox.Warning,
                    "오류",
                    "Model과 같은 폴더에서 호환되는 Scaler 파일을 찾을 수 없습니다.\n"
                    f"Model: {selected_model_path}"
                )
                return
        
        # 선택한 파일 경로 저장 (다음 detect 시 사용)
        self.selected_model_path = selected_model_path
        self.selected_scaler_path = matching_scaler
        
        # 매핑 JSON 파일 찾기 (config 폴더 우선)
        # 1. Model 파일과 같은 디렉토리에서 찾기
        model_dir = os.path.dirname(os.path.abspath(selected_model_path))
        csv_dir = None
        if self.csv_path:
            csv_dir = os.path.dirname(os.path.abspath(self.csv_path))
        
        mapping_json_path = None
        found_path_info = []
        
        csv_prefix = None
        if self.csv_path:
            csv_basename = os.path.splitext(os.path.basename(self.csv_path))[0]
            csv_prefix = csv_basename
        
        possible_mapping_paths = []
        config_dir = self._config_dir()
        if config_dir:
            if csv_prefix:
                possible_mapping_paths.append(os.path.join(config_dir, f"{csv_prefix}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(config_dir, "label_mapping.json"))
        if csv_prefix:
            if csv_dir:
                possible_mapping_paths.append(os.path.join(csv_dir, f"{csv_prefix}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(model_dir, f"{csv_prefix}_label_mapping.json"))
            if csv_dir:
                possible_mapping_paths.append(os.path.join(os.path.dirname(csv_dir), f"{csv_prefix}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(os.path.dirname(model_dir), f"{csv_prefix}_label_mapping.json"))
        
        # Model prefix 기반 (모델명 제거한 버전 사용)
        if model_prefix_for_json:
            possible_mapping_paths.append(os.path.join(model_dir, f"{model_prefix_for_json}_label_mapping.json"))
            if csv_dir:
                possible_mapping_paths.append(os.path.join(csv_dir, f"{model_prefix_for_json}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(os.path.dirname(model_dir), f"{model_prefix_for_json}_label_mapping.json"))
        
        # 기본 경로들
        possible_mapping_paths.extend([
            os.path.join(model_dir, "label_mapping.json"),
            os.path.join(os.path.dirname(model_dir), "label_mapping.json"),
        ])
        if csv_dir:
            possible_mapping_paths.append(os.path.join(csv_dir, "label_mapping.json"))
            possible_mapping_paths.append(os.path.join(os.path.dirname(csv_dir), "label_mapping.json"))
        
        # CSV 기반 find_mapping_json도 시도 (중복 제거)
        if self.csv_path:
            csv_mapping = self.find_mapping_json(self.csv_path)
            if csv_mapping and csv_mapping not in possible_mapping_paths:
                possible_mapping_paths.insert(0, csv_mapping)
        
        # 각 경로 확인 및 로그 출력
        print(f"[DEBUG] 매핑 JSON 검색 시작 (Model: {os.path.basename(selected_model_path)})")
        if self.csv_path:
            print(f"[DEBUG] 현재 CSV 경로: {self.csv_path}")
            print(f"[DEBUG] 추출된 CSV Prefix: {csv_prefix}")
        print(f"[DEBUG] Model Prefix for Scaler: {model_prefix_for_scaler}")
        print(f"[DEBUG] Model Prefix for JSON: {model_prefix_for_json}")
        for i, path in enumerate(possible_mapping_paths):
            exists = os.path.exists(path)
            print(f"[DEBUG] 경로 {i+1}: {path} - {'존재함' if exists else '없음'}")
            if exists and not mapping_json_path:
                mapping_json_path = path
                found_path_info.append(f"발견: {path}")
        
        # 매핑 JSON 설정
        if mapping_json_path:
            self.set_mapping_json_path(mapping_json_path)
            print(f"[INFO] 매핑 JSON 파일 자동 로드: {mapping_json_path}")
            info_text = f"Model/Scaler가 선택되었습니다:\n\n"
            info_text += f"Model: {os.path.basename(selected_model_path)}\n"
            info_text += f"Scaler: {os.path.basename(matching_scaler)}\n\n"
            info_text += f"매핑 JSON: {os.path.basename(mapping_json_path)}\n"
            info_text += f"전체 경로: {mapping_json_path}"
            self._copyable_msg(QMessageBox.Information, "선택 완료", info_text)
        else:
            self.set_mapping_json_path(None)
            print(f"[WARN] 매핑 JSON 파일을 찾을 수 없습니다.")
            print(f"[DEBUG] 검색한 경로들:")
            for path in possible_mapping_paths:
                print(f"  - {path}")
            info_text = f"Model/Scaler가 선택되었습니다:\n\n"
            info_text += f"Model: {os.path.basename(selected_model_path)}\n"
            info_text += f"Scaler: {os.path.basename(matching_scaler)}\n\n"
            info_text += f"매핑 JSON: 없음\n\n"
            info_text += f"검색한 경로:\n"
            for path in possible_mapping_paths[:3]:  # 처음 3개만 표시
                info_text += f"  - {os.path.basename(os.path.dirname(path))}/{os.path.basename(path)}\n"
            self._copyable_msg(QMessageBox.Warning, "선택 완료 (매핑 JSON 없음)", info_text)
        
        print(f"[INFO] Model/Scaler 선택 완료:\n  Model: {selected_model_path}\n  Scaler: {matching_scaler}")
    
    def load_model_and_scaler(self):
        """Load the trained model and scaler from disk."""
        self.label_encoder_path = ""
        binstat = self.binButton_2.isChecked()
        mulstat = self.mulButton_2.isChecked()
        self.aimodel = 'Xgboost'
        self.model_combo_2.activated.connect(self.on_combobox_select)
        self.aimodel = self.model_combo_2.currentText()
        if binstat:
            self.classmode = 'bin_'
        elif mulstat:
            self.classmode = 'mul_'
        else:
            messagebox.showerror("에러", "바이너리/멀티 모드를 선택하세요")
            return

        selected_model_ok = (
            hasattr(self, 'selected_model_path')
            and hasattr(self, 'selected_scaler_path')
            and self.selected_model_path
            and self.selected_scaler_path
            and os.path.exists(self.selected_model_path)
            and os.path.exists(self.selected_scaler_path)
            and self._model_scaler_paths_compatible(self.selected_model_path, self.selected_scaler_path)
            and self._detect_paths_in_current_scope(self.selected_model_path, self.selected_scaler_path)
        )
        if not selected_model_ok:
            try:
                if getattr(self, 'selected_model_path', None) or getattr(self, 'selected_scaler_path', None):
                    print("[INFO] 현재 Detect 상세 폴더 밖의 기존 Model/Scaler 선택을 초기화합니다.")
                self.selected_model_path = ""
                self.selected_scaler_path = ""
            except Exception:
                pass
            self._auto_select_detect_model_for_current_csv(show_message=False)
        
        # 이미 선택된 model/scaler가 있으면 사용
        if hasattr(self, 'selected_model_path') and hasattr(self, 'selected_scaler_path'):
            if os.path.exists(self.selected_model_path) and os.path.exists(self.selected_scaler_path):
                self.pklpath = self.selected_model_path
                self.scalerpath = self.selected_scaler_path
                self.model = joblib.load(self.pklpath)
                self.scaler = joblib.load(self.scalerpath)
                print(f"[INFO] 선택된 Model 로드: {os.path.basename(self.pklpath)}")
                print(f"[INFO] 선택된 Scaler 로드: {os.path.basename(self.scalerpath)}")
                
                # LabelEncoder 로드 (XGBoost 등 인코딩 사용 시)
                model_dir = os.path.dirname(os.path.abspath(self.selected_model_path))
                model_basename = os.path.basename(self.selected_model_path)
                # encoder 파일명: model 파일명에서 "model.pkl" -> "label_encoder.pkl"
                # 예: "csv_path_Xgboostmodel.pkl" -> "csv_path_Xgboostlabel_encoder.pkl"
                enc_filename = model_basename.replace("model.pkl", "label_encoder.pkl")
                enc_path = os.path.join(model_dir, enc_filename)
                # 대체 패턴: csv_path + aimodel + label_encoder.pkl
                if not os.path.exists(enc_path) and self.csv_path:
                    alt_enc_path = str(self.csv_path + "_" + self.aimodel + "label_encoder.pkl")
                    if os.path.exists(alt_enc_path):
                        enc_path = alt_enc_path
                
                if os.path.exists(enc_path):
                    try:
                        self.label_encoder = joblib.load(enc_path)
                        self.label_encoder_path = enc_path
                        print(f"[INFO] LabelEncoder 로드: {os.path.basename(enc_path)}")
                    except Exception as e:
                        print(f"[WARN] LabelEncoder 로드 실패 (인코딩 없이 계속): {e}")
                        self.label_encoder = None
                else:
                    self.label_encoder = None
                    print(f"[WARN] LabelEncoder 없음 (인코딩 없이 예측 진행)")
                    print(f"[DEBUG] LabelEncoder 검색 경로: {enc_path}")
                    if self.csv_path:
                        print(f"[DEBUG] 대체 경로 시도: {str(self.csv_path + '_' + self.aimodel + 'label_encoder.pkl')}")
                    # 모델이 XGBoost인 경우 경고 강화
                    if self.aimodel == 'Xgboost' or 'Xgboost' in model_basename:
                        print(f"[WARN] ⚠️ XGBoost 모델인데 LabelEncoder를 찾을 수 없습니다!")
                        print(f"[WARN] 예상 파일명: {enc_filename}")
                        print(f"[WARN] 학습 시 저장 패턴: csv_path + '_' + aimodel + 'label_encoder.pkl'")
                        print(f"[WARN] 모델이 학습될 때 LabelEncoder가 저장되었는지 확인하세요.")
                
                # feature.json 유무 확인 (모델 run 폴더 → config 폴더 순으로 검색)
                feature_json_path = os.path.join(model_dir, "feature.json")
                if not os.path.exists(feature_json_path) and "training_results" in os.path.normpath(model_dir):
                    case_root = os.path.dirname(os.path.dirname(model_dir))
                    config_feature = os.path.join(case_root, "config", "feature.json")
                    if os.path.exists(config_feature):
                        feature_json_path = config_feature
                if not os.path.exists(feature_json_path) and self._config_dir():
                    config_feature = os.path.join(self._config_dir(), "feature.json")
                    if os.path.exists(config_feature):
                        feature_json_path = config_feature
                if not os.path.exists(feature_json_path) and not getattr(self, "_batch_run_quiet", False):
                    warn_msg = (
                        "feature.json 파일을 찾을 수 없습니다.\n"
                        "모델이 학습될 때의 feature.json을 모델/스케일러 폴더에 두면 피처 정합성 확인이 더 정확해집니다."
                    )
                    try:
                        self.show_alert(warn_msg)
                    except Exception:
                        print(f"[WARN] {warn_msg}")
                elif not os.path.exists(feature_json_path):
                    print("[WARN] feature.json 파일을 찾을 수 없습니다. (전체 모델 실행 중 팝업 생략)")

                # 매핑 JSON이 이미 설정되어 있지 않으면 찾기 (config 폴더 우선)
                if not hasattr(self, 'mapping_json_path') or not self.mapping_json_path:
                    model_dir = os.path.dirname(os.path.abspath(self.selected_model_path))
                    model_basename = os.path.basename(self.selected_model_path)
                    model_prefix = model_basename.replace("model.pkl", "").replace(".pkl", "")
                    possible_mapping_paths = []
                    if self._config_dir():
                        config_dir = self._config_dir()
                        if self.csv_path:
                            csv_prefix = os.path.splitext(os.path.basename(self.csv_path))[0]
                            possible_mapping_paths.append(os.path.join(config_dir, f"{csv_prefix}_label_mapping.json"))
                        possible_mapping_paths.append(os.path.join(config_dir, "label_mapping.json"))
                    possible_mapping_paths.extend([
                        os.path.join(model_dir, "label_mapping.json"),
                        os.path.join(model_dir, f"{model_prefix}_label_mapping.json"),
                        os.path.join(os.path.dirname(model_dir), "label_mapping.json"),
                    ])
                    if self.csv_path:
                        csv_mapping = self.find_mapping_json(self.csv_path)
                        if csv_mapping and csv_mapping not in possible_mapping_paths:
                            possible_mapping_paths.insert(0, csv_mapping)
                    
                    for path in possible_mapping_paths:
                        if os.path.exists(path):
                            self.set_mapping_json_path(path)
                            print(f"[INFO] 매핑 JSON 파일 자동 로드: {os.path.basename(path)}")
                            break
                
                return

        # 케이스 디렉토리에서 모든 model.pkl과 scaler.pkl 파일 찾기 (training_results 하위 포함)
        search_dirs = self._detect_model_search_roots()
        
        model_files = []
        scaler_files = []
        for search_dir in search_dirs:
            if os.path.isdir(search_dir):
                model_files.extend(glob.glob(os.path.join(search_dir, "*model.pkl")))
                scaler_files.extend(glob.glob(os.path.join(search_dir, "*scaler.pkl")))
                tr_dir = os.path.join(search_dir, "training_results")
                if os.path.isdir(tr_dir):
                    for run_name in os.listdir(tr_dir):
                        run_path = os.path.join(tr_dir, run_name)
                        if os.path.isdir(run_path):
                            m = os.path.join(run_path, "model.pkl")
                            s = os.path.join(run_path, "scaler.pkl")
                            if os.path.isfile(m):
                                model_files.append(m)
                            if os.path.isfile(s):
                                scaler_files.append(s)
        model_files = list(set(model_files))
        scaler_files = list(set(scaler_files))
        
        # 기존 방식으로 파일 찾기 (하위 호환성)
        old_pklname = str(self.csv_path + "_" + self.aimodel + "model.pkl")
        old_pklpath = self.resource_path(old_pklname)
        old_scalername = str(self.csv_path + "_" + self.aimodel + "scaler.pkl")
        old_scalerpath = self.resource_path(old_scalername)
        
        # 기존 파일이 있으면 우선 추가
        if os.path.exists(old_pklpath) and old_pklpath not in model_files:
            model_files.append(old_pklpath)
        if os.path.exists(old_scalerpath) and old_scalerpath not in scaler_files:
            scaler_files.append(old_scalerpath)
        
        if not model_files:
            model_name = self._detect_model_name()
            searched = "\n".join(f"- {p}" for p in search_dirs[:20])
            raise FileNotFoundError(
                f"Model 파일을 찾을 수 없습니다.\n"
                f"Detect 모델 선택: {model_name or '(비어 있음)'}\n"
                f"검색 경로:\n{searched}"
            )
        
        # Model 파일 선택
        selected_model_path = None
        if len(model_files) == 1:
            # 파일이 하나면 자동 선택
            selected_model_path = model_files[0]
            print(f"[INFO] Model 파일 자동 선택: {os.path.basename(selected_model_path)}")
        else:
            # 여러 개면 사용자에게 선택 요청
            model_names = [os.path.basename(f) for f in model_files]
            model_name, ok = QInputDialog.getItem(
                self, 
                "Model 선택", 
                "사용할 Model 파일을 선택하세요:",
                model_names, 
                0, 
                False
            )
            if not ok:
                raise FileNotFoundError("Model 파일 선택이 취소되었습니다.")
            
            # 선택한 파일명으로 경로 찾기
            selected_model_path = next((f for f in model_files if os.path.basename(f) == model_name), None)
            if not selected_model_path:
                raise FileNotFoundError(f"선택한 Model 파일을 찾을 수 없습니다: {model_name}")
        
        # Model 파일명의 앞부분 추출 (scaler 매칭용)
        model_basename = os.path.basename(selected_model_path)
        model_dir_here = os.path.dirname(selected_model_path)
        is_run_folder_model = ("training_results" in selected_model_path and model_basename == "model.pkl")
        model_prefix_for_scaler = model_basename.replace("model.pkl", "").replace(".pkl", "")
        
        model_prefix_for_json = model_prefix_for_scaler
        for model_name in ["Xgboost", "RandomForest", "LGBM", "LogisticRegression", "LSTM"]:
            if model_prefix_for_json.endswith(model_name):
                model_prefix_for_json = model_prefix_for_json[:-len(model_name)]
                break
        if model_prefix_for_json.endswith("_"):
            model_prefix_for_json = model_prefix_for_json[:-1]
        if model_prefix_for_json.endswith(".csv_"):
            model_prefix_for_json = model_prefix_for_json[:-5]
        elif model_prefix_for_json.endswith(".csv"):
            model_prefix_for_json = model_prefix_for_json[:-4]
        if model_prefix_for_json.endswith("_"):
            model_prefix_for_json = model_prefix_for_json[:-1]
        
        matching_scaler = self._find_matching_scaler_for_model(selected_model_path, scaler_files)
        if not matching_scaler:
            same_dir_scalers = [
                f for f in scaler_files
                if os.path.dirname(os.path.abspath(f)) == os.path.abspath(model_dir_here)
            ]
            if same_dir_scalers:
                scaler_names = [os.path.basename(f) for f in same_dir_scalers]
                scaler_name, ok = QInputDialog.getItem(
                    self,
                    "Scaler 선택",
                    f"Model과 같은 폴더의 Scaler만 선택할 수 있습니다:\n{model_dir_here}",
                    scaler_names,
                    0,
                    False
                )
                if not ok:
                    raise FileNotFoundError("Scaler 파일 선택이 취소되었습니다.")
                selected = next((f for f in same_dir_scalers if os.path.basename(f) == scaler_name), None)
                if selected and self._model_scaler_paths_compatible(selected_model_path, selected):
                    matching_scaler = selected
            if not matching_scaler:
                raise FileNotFoundError(
                    "Model과 같은 폴더에서 호환되는 Scaler 파일을 찾을 수 없습니다.\n"
                    f"Model: {selected_model_path}"
                )
        
        # 이미 선택된 model/scaler가 있으면 사용
        if hasattr(self, 'selected_model_path') and hasattr(self, 'selected_scaler_path'):
            if os.path.exists(self.selected_model_path) and os.path.exists(self.selected_scaler_path):
                self.pklpath = self.selected_model_path
                self.scalerpath = self.selected_scaler_path
                self.model = joblib.load(self.pklpath)
                self.scaler = joblib.load(self.scalerpath)
                print(f"[INFO] 선택된 Model 로드: {os.path.basename(self.pklpath)}")
                print(f"[INFO] 선택된 Scaler 로드: {os.path.basename(self.scalerpath)}")
                
                # LabelEncoder 로드 (XGBoost 등 인코딩 사용 시)
                model_dir = os.path.dirname(os.path.abspath(self.selected_model_path))
                model_basename = os.path.basename(self.selected_model_path)
                # encoder 파일명: model 파일명에서 "model.pkl" -> "label_encoder.pkl"
                # 예: "csv_path_Xgboostmodel.pkl" -> "csv_path_Xgboostlabel_encoder.pkl"
                enc_filename = model_basename.replace("model.pkl", "label_encoder.pkl")
                enc_path = os.path.join(model_dir, enc_filename)
                # 대체 패턴: csv_path + aimodel + label_encoder.pkl
                if not os.path.exists(enc_path) and self.csv_path:
                    alt_enc_path = str(self.csv_path + "_" + self.aimodel + "label_encoder.pkl")
                    if os.path.exists(alt_enc_path):
                        enc_path = alt_enc_path
                
                if os.path.exists(enc_path):
                    try:
                        self.label_encoder = joblib.load(enc_path)
                        self.label_encoder_path = enc_path
                        print(f"[INFO] LabelEncoder 로드: {os.path.basename(enc_path)}")
                    except Exception as e:
                        print(f"[WARN] LabelEncoder 로드 실패 (인코딩 없이 계속): {e}")
                        self.label_encoder = None
                else:
                    self.label_encoder = None
                    print(f"[WARN] LabelEncoder 없음 (인코딩 없이 예측 진행)")
                    print(f"[DEBUG] LabelEncoder 검색 경로: {enc_path}")
                    if self.csv_path:
                        print(f"[DEBUG] 대체 경로 시도: {str(self.csv_path + '_' + self.aimodel + 'label_encoder.pkl')}")
                    # 모델이 XGBoost인 경우 경고 강화
                    if self.aimodel == 'Xgboost' or 'Xgboost' in model_basename:
                        print(f"[WARN] ⚠️ XGBoost 모델인데 LabelEncoder를 찾을 수 없습니다!")
                        print(f"[WARN] 예상 파일명: {enc_filename}")
                        print(f"[WARN] 학습 시 저장 패턴: csv_path + '_' + aimodel + 'label_encoder.pkl'")
                        print(f"[WARN] 모델이 학습될 때 LabelEncoder가 저장되었는지 확인하세요.")
                
                # 매핑 JSON이 이미 설정되어 있지 않으면 찾기
                if not hasattr(self, 'mapping_json_path') or not self.mapping_json_path:
                    model_dir = os.path.dirname(os.path.abspath(self.selected_model_path))
                    csv_dir = None
                    if self.csv_path:
                        csv_dir = os.path.dirname(os.path.abspath(self.csv_path))
                    
                    model_basename = os.path.basename(self.selected_model_path)
                    model_prefix_for_scaler = model_basename.replace("model.pkl", "").replace(".pkl", "")
                    
                    # 매핑 JSON용 prefix (모델명과 .csv 제거)
                    model_prefix_for_json = model_prefix_for_scaler
                    # 모델명 제거 (Xgboost, RandomForest 등)
                    for model_name in ["Xgboost", "RandomForest", "LGBM", "LogisticRegression", "LSTM"]:
                        if model_prefix_for_json.endswith(model_name):
                            model_prefix_for_json = model_prefix_for_json[:-len(model_name)]
                            break
                    # 끝에 _ 가 남아있으면 먼저 제거 (모델명 제거 후 남은 _)
                    if model_prefix_for_json.endswith("_"):
                        model_prefix_for_json = model_prefix_for_json[:-1]
                    # .csv 또는 .csv_ 제거 (Model 파일명에 .csv가 포함된 경우)
                    if model_prefix_for_json.endswith(".csv_"):
                        model_prefix_for_json = model_prefix_for_json[:-5]
                    elif model_prefix_for_json.endswith(".csv"):
                        model_prefix_for_json = model_prefix_for_json[:-4]
                    # 끝에 _ 가 남아있으면 다시 제거
                    if model_prefix_for_json.endswith("_"):
                        model_prefix_for_json = model_prefix_for_json[:-1]
                    
                    # CSV 파일명 기반 prefix 추출 (예: _train_2512241542_processed_labeled200)
                    # CSV 파일명 전체를 prefix로 사용 (확장자만 제거)
                    csv_prefix = None
                    if self.csv_path:
                        csv_basename = os.path.splitext(os.path.basename(self.csv_path))[0]
                        # CSV 파일명 전체를 prefix로 사용 (예: _train_2512241542_processed_labeled200)
                        csv_prefix = csv_basename
                        print(f"[DEBUG] CSV 파일명: {os.path.basename(self.csv_path)}")
                        print(f"[DEBUG] 추출된 CSV prefix: {csv_prefix}")
                    
                    possible_mapping_paths = []
                    config_dir = self._config_dir()
                    if config_dir:
                        if csv_prefix:
                            possible_mapping_paths.append(os.path.join(config_dir, f"{csv_prefix}_label_mapping.json"))
                        possible_mapping_paths.append(os.path.join(config_dir, "label_mapping.json"))
                    if csv_prefix:
                        if csv_dir:
                            possible_mapping_paths.append(os.path.join(csv_dir, f"{csv_prefix}_label_mapping.json"))
                        possible_mapping_paths.append(os.path.join(model_dir, f"{csv_prefix}_label_mapping.json"))
                        if csv_dir:
                            possible_mapping_paths.append(os.path.join(os.path.dirname(csv_dir), f"{csv_prefix}_label_mapping.json"))
                        possible_mapping_paths.append(os.path.join(os.path.dirname(model_dir), f"{csv_prefix}_label_mapping.json"))
                    if model_prefix_for_json:
                        possible_mapping_paths.append(os.path.join(model_dir, f"{model_prefix_for_json}_label_mapping.json"))
                        if csv_dir:
                            possible_mapping_paths.append(os.path.join(csv_dir, f"{model_prefix_for_json}_label_mapping.json"))
                        possible_mapping_paths.append(os.path.join(os.path.dirname(model_dir), f"{model_prefix_for_json}_label_mapping.json"))
                    
                    # 기본 경로들
                    possible_mapping_paths.extend([
                        os.path.join(model_dir, "label_mapping.json"),
                        os.path.join(os.path.dirname(model_dir), "label_mapping.json"),
                    ])
                    if csv_dir:
                        possible_mapping_paths.append(os.path.join(csv_dir, "label_mapping.json"))
                        possible_mapping_paths.append(os.path.join(os.path.dirname(csv_dir), "label_mapping.json"))
                    
                    # CSV 기반 find_mapping_json도 시도 (중복 제거)
                    if self.csv_path:
                        csv_mapping = self.find_mapping_json(self.csv_path)
                        if csv_mapping and csv_mapping not in possible_mapping_paths:
                            possible_mapping_paths.insert(0, csv_mapping)
                    
                    # 각 경로 확인 및 로그 출력
                    print(f"[DEBUG] 매핑 JSON 검색 시작 (이미 선택된 Model: {os.path.basename(self.selected_model_path)})")
                    for i, path in enumerate(possible_mapping_paths):
                        exists = os.path.exists(path)
                        print(f"[DEBUG] 경로 {i+1}: {path} - {'존재함' if exists else '없음'}")
                        if exists and not self.mapping_json_path:
                            self.set_mapping_json_path(path)
                            print(f"[INFO] 매핑 JSON 파일 자동 로드: {path}")
                            # 매핑 JSON 정보를 메시지 박스로 표시
                            info_text = f"Model/Scaler 로드 완료:\n\n"
                            info_text += f"Model: {os.path.basename(self.selected_model_path)}\n"
                            info_text += f"Scaler: {os.path.basename(self.selected_scaler_path)}\n\n"
                            info_text += f"매핑 JSON: {os.path.basename(path)}\n"
                            info_text += f"전체 경로: {path}"
                            self._copyable_msg(QMessageBox.Information, "Model/Scaler 로드 완료", info_text)
                            break
                    
                    if not self.mapping_json_path:
                        print(f"[WARN] 매핑 JSON 파일을 찾을 수 없습니다.")
                        print(f"[DEBUG] 검색한 경로들:")
                        for path in possible_mapping_paths:
                            print(f"  - {path}")
                        # 매핑 JSON 없음 정보를 메시지 박스로 표시
                        info_text = f"Model/Scaler 로드 완료:\n\n"
                        info_text += f"Model: {os.path.basename(self.selected_model_path)}\n"
                        info_text += f"Scaler: {os.path.basename(self.selected_scaler_path)}\n\n"
                        info_text += f"매핑 JSON: 없음\n\n"
                        info_text += f"검색한 경로:\n"
                        for path in possible_mapping_paths[:3]:  # 처음 3개만 표시
                            info_text += f"  - {os.path.basename(os.path.dirname(path))}/{os.path.basename(path)}\n"
                        self._copyable_msg(QMessageBox.Warning, "Model/Scaler 로드 완료 (매핑 JSON 없음)", info_text)
                
                return

        # 파일 로드
        self.pklpath = selected_model_path
        self.scalerpath = matching_scaler
        
        # 선택한 파일 경로 저장 (다음 detect 시 사용)
        self.selected_model_path = selected_model_path
        self.selected_scaler_path = matching_scaler
        
        # 매핑 JSON 파일 찾기
        model_dir = os.path.dirname(os.path.abspath(selected_model_path))
        csv_dir = None
        if self.csv_path:
            csv_dir = os.path.dirname(os.path.abspath(self.csv_path))
        
        mapping_json_path = None
        
        csv_prefix = None
        if self.csv_path:
            csv_basename = os.path.splitext(os.path.basename(self.csv_path))[0]
            csv_prefix = csv_basename
        
        possible_mapping_paths = []
        config_dir = self._config_dir()
        if config_dir:
            if csv_prefix:
                possible_mapping_paths.append(os.path.join(config_dir, f"{csv_prefix}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(config_dir, "label_mapping.json"))
        if csv_prefix:
            if csv_dir:
                possible_mapping_paths.append(os.path.join(csv_dir, f"{csv_prefix}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(model_dir, f"{csv_prefix}_label_mapping.json"))
            if csv_dir:
                possible_mapping_paths.append(os.path.join(os.path.dirname(csv_dir), f"{csv_prefix}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(os.path.dirname(model_dir), f"{csv_prefix}_label_mapping.json"))
        if model_prefix_for_json:
            possible_mapping_paths.append(os.path.join(model_dir, f"{model_prefix_for_json}_label_mapping.json"))
            if csv_dir:
                possible_mapping_paths.append(os.path.join(csv_dir, f"{model_prefix_for_json}_label_mapping.json"))
            possible_mapping_paths.append(os.path.join(os.path.dirname(model_dir), f"{model_prefix_for_json}_label_mapping.json"))
        possible_mapping_paths.extend([
            os.path.join(model_dir, "label_mapping.json"),
            os.path.join(os.path.dirname(model_dir), "label_mapping.json"),
        ])
        if csv_dir:
            possible_mapping_paths.append(os.path.join(csv_dir, "label_mapping.json"))
            possible_mapping_paths.append(os.path.join(os.path.dirname(csv_dir), "label_mapping.json"))
        if self.csv_path:
            csv_mapping = self.find_mapping_json(self.csv_path)
            if csv_mapping and csv_mapping not in possible_mapping_paths:
                possible_mapping_paths.insert(0, csv_mapping)
        
        print(f"[DEBUG] 매핑 JSON 검색 시작 (Model: {os.path.basename(selected_model_path)})")
        for i, path in enumerate(possible_mapping_paths):
            exists = os.path.exists(path)
            print(f"[DEBUG] 경로 {i+1}: {path} - {'존재함' if exists else '없음'}")
            if exists and not mapping_json_path:
                mapping_json_path = path
        
        # 매핑 JSON 설정
        if mapping_json_path:
            self.set_mapping_json_path(mapping_json_path)
            print(f"[INFO] 매핑 JSON 파일 자동 로드: {mapping_json_path}")
            print(f"[INFO] 매핑 JSON 전체 경로: {mapping_json_path}")
            # 매핑 JSON 정보를 메시지 박스로 표시
            info_text = f"Model/Scaler 로드 완료:\n\n"
            info_text += f"Model: {os.path.basename(selected_model_path)}\n"
            info_text += f"Scaler: {os.path.basename(matching_scaler)}\n\n"
            info_text += f"매핑 JSON: {os.path.basename(mapping_json_path)}\n"
            info_text += f"전체 경로: {mapping_json_path}"
            self._copyable_msg(QMessageBox.Information, "Model/Scaler 로드 완료", info_text)
        else:
            self.set_mapping_json_path(None)
            print(f"[WARN] 매핑 JSON 파일을 찾을 수 없습니다.")
            print(f"[DEBUG] 검색한 경로들:")
            for path in possible_mapping_paths:
                print(f"  - {path}")
            # 매핑 JSON 없음 정보를 메시지 박스로 표시
            info_text = f"Model/Scaler 로드 완료:\n\n"
            info_text += f"Model: {os.path.basename(selected_model_path)}\n"
            info_text += f"Scaler: {os.path.basename(matching_scaler)}\n\n"
            info_text += f"매핑 JSON: 없음\n\n"
            info_text += f"검색한 경로:\n"
            for path in possible_mapping_paths[:3]:  # 처음 3개만 표시
                info_text += f"  - {os.path.basename(os.path.dirname(path))}/{os.path.basename(path)}\n"
            self._copyable_msg(QMessageBox.Warning, "Model/Scaler 로드 완료 (매핑 JSON 없음)", info_text)
        
        if os.path.exists(self.pklpath) and os.path.exists(self.scalerpath):
            self.model = joblib.load(self.pklpath)
            self.scaler = joblib.load(self.scalerpath)
            print(f"[INFO] Model 로드 완료: {os.path.basename(self.pklpath)}")
            print(f"[INFO] Scaler 로드 완료: {os.path.basename(self.scalerpath)}")
            
            # LabelEncoder 로드 (XGBoost 등 인코딩 사용 시)
            model_dir = os.path.dirname(os.path.abspath(self.pklpath))
            model_basename = os.path.basename(self.pklpath)
            # encoder 파일명: model 파일명에서 "model.pkl" -> "label_encoder.pkl"
            # 예: "csv_path_Xgboostmodel.pkl" -> "csv_path_Xgboostlabel_encoder.pkl"
            enc_filename = model_basename.replace("model.pkl", "label_encoder.pkl")
            enc_path = os.path.join(model_dir, enc_filename)
            # 대체 패턴: csv_path + aimodel + label_encoder.pkl
            if not os.path.exists(enc_path) and self.csv_path:
                alt_enc_path = str(self.csv_path + "_" + self.aimodel + "label_encoder.pkl")
                if os.path.exists(alt_enc_path):
                    enc_path = alt_enc_path
            
            if os.path.exists(enc_path):
                try:
                    self.label_encoder = joblib.load(enc_path)
                    self.label_encoder_path = enc_path
                    print(f"[INFO] LabelEncoder 로드: {os.path.basename(enc_path)}")
                except Exception as e:
                    print(f"[WARN] LabelEncoder 로드 실패 (인코딩 없이 계속): {e}")
                    self.label_encoder = None
            else:
                self.label_encoder = None
                print(f"[INFO] LabelEncoder 없음 (인코딩 없이 예측 진행)")
        else:
            raise FileNotFoundError(f"Model 또는 Scaler 파일을 찾을 수 없습니다.\nModel: {self.pklpath}\nScaler: {self.scalerpath}")

    def predict_all_detect_models(self, structured_data, file_path=None, show_message=True, compact=False):
        """현재 선택된 feature 상세 폴더 안의 모든 model.pkl/scaler.pkl 쌍으로 예측."""
        candidates = self._detect_model_scaler_candidates(latest_per_family=False)
        if not candidates:
            warn = "현재 선택된 상세 폴더에서 호환되는 model.pkl/scaler.pkl 쌍을 찾지 못했습니다."
            if show_message:
                self._copyable_msg(QMessageBox.Warning, "Detect 결과", warn)
            return warn if compact else pd.DataFrame(structured_data)

        original_state = {}
        for name in (
            "selected_model_path", "selected_scaler_path", "pklpath", "scalerpath",
            "model", "scaler", "label_encoder", "label_encoder_path",
            "detect_feature_json_path", "detect_mapping_json_path", "mapping_json_path",
            "aimodel",
        ):
            original_state[name] = getattr(self, name, None)

        result_frames = []
        result_lines = []
        error_lines = []
        records = []
        filename = os.path.basename(file_path) if file_path else "선택 파일"
        if not compact:
            result_lines.append(f"{filename} 모델별 결과 ({len(candidates)}개 모델)")
            result_lines.append("")
        if show_message:
            self._begin_batch_run()

        try:
            for idx, cand in enumerate(candidates, 1):
                family = cand["family"]
                model_path = cand["model_path"]
                model_label = self._detect_record_model_name(model_path) or family
                scaler_path = cand["scaler_path"]
                try:
                    self.selected_model_path = model_path
                    self.selected_scaler_path = scaler_path
                    self.aimodel = model_label
                    self.label_encoder = None
                    self.label_encoder_path = ""
                    self.detect_feature_json_path = ""
                    self.detect_mapping_json_path = ""
                    self.load_model_and_scaler()
                    df = self.predict_data1(
                        structured_data,
                        file_path,
                        skip_simhash=False,
                        show_message=False,
                        show_warnings=False,
                    )
                    result_frames.append(df)
                    records.append(self._detect_prediction_record(
                        df,
                        file_path,
                        model_name=model_label,
                        model_pkl=model_path,
                        detail_folder=getattr(self, "detect_detail_dir", ""),
                    ))
                    summary = df.attrs.get("detect_result_summary", "")
                    paths_text = df.attrs.get("detect_used_paths_text", "")
                    result_lines.append(f"[{idx}] {model_label}")
                    result_lines.append(summary if summary else "예측 결과를 만들지 못했습니다.")
                    if paths_text:
                        result_lines.append(paths_text)
                    result_lines.append("")
                except Exception as e:
                    err = f"[{idx}] {model_label}: 실패 - {e}"
                    print(f"[WARN] 전체 모델 Detect 실패: {err}")
                    error_lines.append(err)
        finally:
            batch_log = self._end_batch_run() if show_message else []
            for name, value in original_state.items():
                try:
                    setattr(self, name, value)
                except Exception:
                    pass

        if error_lines:
            result_lines.append("[실패한 모델]")
            result_lines.extend(error_lines)
        if batch_log and show_message:
            result_lines.append("")
            result_lines.append("[참고 메시지]")
            for line in batch_log:
                result_lines.append(line)

        message = "\n".join(result_lines).strip()
        print(message)
        if show_message:
            run_meta = self._detect_run_metadata(
                [file_path] if file_path else [],
                run_all_models=True,
            )
            saved_dir = self._save_detect_run(records, run_meta, message)
            if saved_dir:
                message += (
                    f"\n\n[기록 저장]\n{saved_dir}\n"
                    "- results.csv (표 형식)\n- summary.txt\n- run_info.json"
                )
            self._copyable_msg(QMessageBox.Information, "전체 상세폴더 Detect 결과", message)
            if result_frames:
                return pd.concat(result_frames, ignore_index=True, sort=False)
            return pd.DataFrame(structured_data)
        return {"message": message, "records": records}

    def predict_data1(self, structured_data, file_path=None, skip_simhash=False, show_message=True, show_warnings=True):
        """Scale the features and predict the label.
        skip_simhash=True 인 경우(이미 가공된 CSV 행 사용) simhash를 다시 적용하지 않습니다.
        """
        import json as json_module  # 안전하게 모듈 별칭 사용 (json 변수 오염 방지)
        df = pd.DataFrame(structured_data)
        
        # 파일명 추출 (file_path가 제공된 경우)
        filename = None
        if file_path:
            filename = os.path.basename(file_path)

        # feature.json은 학습 파이프라인에서 생성되는 것이 정상인데,
        # 케이스/환경에 따라 누락될 수 있으므로 예측 단계에서 자동 복구합니다.
        # 우선순위: 모델 폴더 > CSV 폴더 > 케이스 폴더 > 현재 폴더
        base_dir = None
        try:
            if hasattr(self, "pklpath") and self.pklpath:
                base_dir = os.path.dirname(os.path.abspath(self.pklpath))
        except Exception:
            base_dir = None
        if not base_dir:
            try:
                if self.csv_path:
                    base_dir = os.path.dirname(os.path.abspath(self.csv_path))
            except Exception:
                base_dir = None
        if not base_dir:
            try:
                base_dir = os.path.abspath(self.case_direc)
            except Exception:
                base_dir = os.path.abspath(".")

        jsonpath = os.path.join(base_dir, "feature.json")
        if not os.path.isfile(jsonpath) and base_dir and "training_results" in os.path.normpath(base_dir):
            case_root = os.path.dirname(os.path.dirname(base_dir))
            config_feature = os.path.join(case_root, "config", "feature.json")
            if os.path.isfile(config_feature):
                jsonpath = config_feature
        config_dir = self._config_dir() if self.case_direc else None
        if config_dir and not os.path.isfile(jsonpath):
            config_json = os.path.join(config_dir, "feature.json")
            if os.path.isfile(config_json):
                jsonpath = config_json
        model_features = None
        try:
            with open(jsonpath, 'r', encoding='utf-8') as f:
                model_features = json_module.load(f)
            self.detect_feature_json_path = jsonpath
            try:
                print(f"[INFO] feature.json 사용: {jsonpath} (피처 {len(model_features)}개)")
            except Exception:
                pass
        except FileNotFoundError:
            try:
                model_features = list(getattr(self.scaler, "feature_names_in_", []))
            except Exception:
                model_features = []
            if not model_features:
                model_features = [str(c) for c in df.columns]
            try:
                write_path = os.path.join(config_dir, "feature.json") if config_dir else jsonpath
                with open(write_path, 'w', encoding='utf-8') as f:
                    json_module.dump(model_features, f, ensure_ascii=False, indent=2)
                self.detect_feature_json_path = write_path
                print(f"[INFO] feature.json이 없어 자동 생성했습니다: {write_path}")
            except Exception as e:
                print(f"[WARN] feature.json 자동 생성 실패(무시하고 계속): {e}")
        except Exception as e:
            # 형식 오류 등은 scaler 피처로 대체
            try:
                model_features = list(getattr(self.scaler, "feature_names_in_", []))
            except Exception:
                model_features = [str(c) for c in df.columns]
            print(f"[WARN] feature.json 읽기 실패(대체 사용): {e}")
        try:
            scaler_feat_len = len(getattr(self.scaler, "feature_names_in_", []))
            model_feat_len = len(model_features) if model_features is not None else 0
            if scaler_feat_len and model_feat_len and scaler_feat_len != model_feat_len:
                print(f"[WARN] feature.json 피처수({model_feat_len})와 스케일러 피처수({scaler_feat_len})가 다릅니다.")
        except Exception:
            pass

        # 학습 시와 동일한 전처리 순서 유지
        # 1. md5 제거 (학습 시와 동일)
        try:
            df = df.drop(columns='md5')
        except Exception as e:
            pass

        # 2. name 제거 (학습 시와 동일)
        df = df.drop(columns=[col for col in df.columns if col == 'name'], errors='ignore')
        # 3. 컬럼명을 문자열로 변환 (학습 시 apply_simhash 전에 수행)
        df.columns = df.columns.astype(str)
        
        # 4. apply_simhash 적용 (학습 시와 동일)
        # skip_simhash=True이면 이미 가공된 CSV 행을 그대로 사용
        if not skip_simhash:
            df = self.apply_simhash(df)
        
        # 5. label 제거 (학습 시와 동일 - apply_simhash 후에 제거)
        df = df.drop(columns=[col for col in df.columns if col == 'label'], errors='ignore')

        # 6. feature.json 기준으로 컬럼 정렬/정합화 (누락은 0으로 채우고, 초과는 드롭)
        if model_features:
            model_features = [str(c) for c in model_features]
            current_cols = [str(c) for c in df.columns]
            missing_from_df = [c for c in model_features if c not in current_cols]
            extra_in_df = [c for c in current_cols if c not in model_features]
            if missing_from_df or extra_in_df:
                print(f"[WARN] feature.json 대비 정렬 수행: missing={len(missing_from_df)}, extra={len(extra_in_df)}")
            # 누락은 0으로 채워넣고, 초과는 제거
            for c in missing_from_df:
                df[c] = 0
            df = df.reindex(columns=model_features, fill_value=0)
            try:
                print(f"[INFO] feature.json 정렬 후 컬럼 수: {df.shape[1]} (누락 채움 {len(missing_from_df)}, 초과 제거 {len(extra_in_df)})")
            except Exception:
                pass
        
        # feature.json vs 현재 컬럼 헤더 비교 (학습 피처 정합성 확인용)
        try:
            if model_features:
                missing_from_df = [c for c in model_features if c not in df.columns]
                extra_in_df = [c for c in df.columns if c not in model_features]
                if missing_from_df or extra_in_df:
                    print(f"[WARN] feature.json 헤더와 입력 컬럼 불일치: feature.json {len(model_features)}개, 입력 {df.shape[1]}개")
                    if missing_from_df:
                        print(f"[WARN] 입력에 없는 피처({len(missing_from_df)}): {missing_from_df[:20]}{' ...' if len(missing_from_df)>20 else ''}")
                    if extra_in_df:
                        print(f"[WARN] feature.json에 없는 추가 컬럼({len(extra_in_df)}): {extra_in_df[:20]}{' ...' if len(extra_in_df)>20 else ''}")
        except Exception as e:
            print(f"[WARN] feature.json 대비 헤더 비교 실패: {e}")
        
        # Detect에서는 학습 CSV의 실제 행/값과 비교하지 않는다. 모델 피처 정합성은 feature.json/scaler 기준으로만 본다.
        try:
            ref_cols = []
            ref_row_series = None
            ref_name = None
            if False and self.csv_path and os.path.exists(self.csv_path):
                print(f"[DEBUG] 학습 CSV와 입력 값 비교 시작: 학습CSV={os.path.basename(self.csv_path)}, 입력파일={filename if filename else 'N/A'}")
                try:
                    ref_df_full = pd.read_csv(self.csv_path, on_bad_lines='skip')
                except TypeError:
                    ref_df_full = pd.read_csv(self.csv_path, error_bad_lines=False)
                ref_cols = [str(c) for c in ref_df_full.columns]
                print(f"[DEBUG] 학습 CSV 로드: {len(ref_df_full)}행, {len(ref_cols)}컬럼")
                
                match_series = None
                matched_by_name = False
                if file_path and 'name' in ref_df_full.columns:
                    fname = os.path.basename(file_path)
                    try:
                        match_series = ref_df_full.loc[ref_df_full['name'] == fname]
                        if match_series is not None and len(match_series) > 0:
                            matched_by_name = True
                            print(f"[DEBUG] 파일명으로 매칭 성공: {fname} (학습 CSV에서 {len(match_series)}개 행 발견)")
                    except Exception as e:
                        print(f"[DEBUG] 파일명 매칭 실패: {e}")
                        match_series = None
                
                if match_series is not None and len(match_series) > 0:
                    ref_row_series = match_series.iloc[0]
                    ref_name = fname if matched_by_name else match_series.iloc[0].get('name', 'matched_row')
                    print(f"[DEBUG] 비교 대상 행 선택: {ref_name}")
                else:
                    ref_row_series = ref_df_full.iloc[0]
                    ref_name = ref_row_series.get('name', 'first_row')
                    print(f"[DEBUG] 파일명 매칭 실패 -> 첫 번째 행 사용: {ref_name}")

                # 학습 CSV에 동일 파일이 있으면 값 자체를 학습 CSV 값으로 맞춰 변환/스케일 전에 정렬
                if ref_row_series is not None and df_raw_before_hash is not None:
                    overlap_cols = [c for c in df_raw_before_hash.columns if c in ref_row_series.index and c not in ['name', 'md5', 'label']]
                    if overlap_cols:
                        try:
                            df_raw_before_hash.loc[:, overlap_cols] = ref_row_series[overlap_cols].values
                            print(f"[INFO] 학습 CSV 매칭 행을 사용해 {len(overlap_cols)}개 컬럼 값을 정렬했습니다.")
                        except Exception as e:
                            print(f"[WARN] 학습 CSV 값 정렬 실패(무시): {e}")

                # 컬럼 존재 여부 비교 (simhash 적용 전 스냅샷 vs 학습 CSV)
                missing_vs_ref = [c for c in ref_cols if c not in df_raw_before_hash.columns]
                extra_vs_ref = [c for c in df_raw_before_hash.columns if c not in ref_cols]
                if missing_vs_ref or extra_vs_ref:
                    print(f"[WARN] 학습 CSV 헤더와 입력 컬럼 불일치: 학습CSV {len(ref_cols)}개, 입력 {df_raw_before_hash.shape[1]}개")
                    if missing_vs_ref:
                        print(f"[WARN] 입력에 없는 학습CSV 컬럼({len(missing_vs_ref)}): {missing_vs_ref[:20]}{' ...' if len(missing_vs_ref)>20 else ''}")
                    if extra_vs_ref:
                        print(f"[WARN] 학습CSV에 없는 추가 컬럼({len(extra_vs_ref)}): {extra_vs_ref[:20]}{' ...' if len(extra_vs_ref)>20 else ''}")

                # 값 비교 (공통 컬럼) - UI에 표시하기 위해 모든 비교 결과를 저장
                common_cols = [c for c in df_raw_before_hash.columns if c in ref_cols]
                print(f"[DEBUG] 공통 컬럼 수: {len(common_cols)}개 (비교 시작)")
                comparison_data = []  # UI 표시용: [(컬럼명, 입력값, 학습값, 상태, 이유), ...]
                diffs = []
                matches = []
                if ref_row_series is not None and len(common_cols) > 0:
                    for c in common_cols:
                        try:
                            v_in = df_raw_before_hash.iloc[0][c]
                            v_ref = ref_row_series[c]
                            status = "일치"
                            reason = ""
                            is_match = False
                            
                            # 둘 다 NaN이면 같은 것으로 간주
                            if (pd.isna(v_in) and pd.isna(v_ref)):
                                matches.append(c)
                                comparison_data.append((c, "NaN", "NaN", "일치", "둘 다 없음"))
                                continue
                            
                            # 빈 문자열과 NaN을 같은 것으로 간주 ("없음" 의미)
                            v_in_str = str(v_in).strip() if not pd.isna(v_in) else ""
                            v_ref_str = str(v_ref).strip() if not pd.isna(v_ref) else ""
                            # 빈 문자열, "/", "nan", "None" 등을 모두 "없음"으로 간주
                            empty_values = ["", "/", "nan", "None", "none", "null", "NULL"]
                            v_in_empty = v_in_str in empty_values or (pd.isna(v_in) if not isinstance(v_in, str) else False)
                            v_ref_empty = v_ref_str in empty_values or (pd.isna(v_ref) if not isinstance(v_ref, str) else False)
                            if v_in_empty and v_ref_empty:
                                matches.append(c)
                                comparison_data.append((c, v_in_str or "없음", v_ref_str or "없음", "일치", "둘 다 없음"))
                                continue  # 둘 다 "없음"이면 같은 것으로 간주
                            
                            # 하나만 "없음"이면 다름
                            if v_in_empty or v_ref_empty:
                                diffs.append(c)
                                status = "불일치"
                                reason = "한쪽만 없음"
                                comparison_data.append((c, v_in_str or "없음" if v_in_empty else str(v_in), 
                                                       v_ref_str or "없음" if v_ref_empty else str(v_ref), status, reason))
                                if len(diffs) <= 20:
                                    print(f"[WARN] 값 불일치: {c} | 입력:{v_in} / 학습:{v_ref} (행:{ref_name})")
                                continue
                            
                            # 숫자 값 비교 (정수/부동소수점 모두 처리)
                            try:
                                v_in_num = float(v_in)
                                v_ref_num = float(v_ref)
                                if abs(v_in_num - v_ref_num) < 1e-9:  # 부동소수점 오차 고려
                                    matches.append(c)
                                    is_match = True
                                    comparison_data.append((c, str(v_in), str(v_ref), "일치", "숫자 값 동일"))
                                else:
                                    status = "불일치"
                                    reason = f"숫자 값 차이: {abs(v_in_num - v_ref_num):.6f}"
                            except (ValueError, TypeError):
                                # 숫자로 변환 실패 시 문자열로 비교
                                is_match = False
                            
                            # 문자열로 비교 (숫자 변환 실패하거나 문자열인 경우)
                            if not is_match:
                                if v_in_str != v_ref_str:
                                    diffs.append(c)
                                    if status == "일치":  # 아직 상태가 설정되지 않았으면
                                        status = "불일치"
                                        reason = "문자열 값 다름"
                                    comparison_data.append((c, v_in_str, v_ref_str, status, reason))
                                    if len(diffs) <= 20:
                                        print(f"[WARN] 값 불일치: {c} | 입력:{v_in} / 학습:{v_ref} (행:{ref_name})")
                                else:
                                    matches.append(c)
                                    comparison_data.append((c, v_in_str, v_ref_str, "일치", "문자열 값 동일"))
                        except Exception as e:
                            print(f"[DEBUG] 컬럼 {c} 비교 중 오류: {e}")
                            try:
                                v_in_val = str(df_raw_before_hash.iloc[0][c]) if 'df_raw_before_hash' in locals() else "오류"
                            except:
                                v_in_val = "오류"
                            try:
                                v_ref_val = str(ref_row_series[c]) if 'ref_row_series' in locals() else "오류"
                            except:
                                v_ref_val = "오류"
                            comparison_data.append((c, v_in_val, v_ref_val, "오류", str(e)))
                            continue
                    
                    print(f"[DEBUG] 값 비교 완료: 일치={len(matches)}개, 불일치={len(diffs)}개")
                    if matches and len(matches) <= 10:
                        print(f"[DEBUG] 일치하는 컬럼 예시: {matches[:10]}")
                    if len(diffs) > 20:
                        print(f"[WARN] 값 불일치 컬럼 추가 {len(diffs)-20}개 더 있음...")
                    
                    # 로그에 모든 비교 결과 출력
                    if comparison_data:
                        print(f"\n{'='*80}")
                        print(f"[INFO] 피처별 비교 결과 상세 (총 {len(comparison_data)}개)")
                        print(f"{'='*80}")
                        print(f"{'컬럼명':<40} {'입력 값':<30} {'학습 CSV 값':<30} {'상태':<10} {'비고'}")
                        print(f"{'-'*80}")
                        
                        # 불일치 항목을 먼저 출력 (중요하므로)
                        for col_name, input_val, ref_val, status, reason in comparison_data:
                            if status == "불일치":
                                input_str = str(input_val)[:28] + "..." if len(str(input_val)) > 28 else str(input_val)
                                ref_str = str(ref_val)[:28] + "..." if len(str(ref_val)) > 28 else str(ref_val)
                                col_str = str(col_name)[:38] + "..." if len(str(col_name)) > 38 else str(col_name)
                                print(f"{col_str:<40} {input_str:<30} {ref_str:<30} {status:<10} {reason}")
                        
                        # 일치 항목 출력 (상위 50개만, 너무 많으면 생략)
                        match_count = 0
                        for col_name, input_val, ref_val, status, reason in comparison_data:
                            if status == "일치":
                                if match_count < 50:
                                    input_str = str(input_val)[:28] + "..." if len(str(input_val)) > 28 else str(input_val)
                                    ref_str = str(ref_val)[:28] + "..." if len(str(ref_val)) > 28 else str(ref_val)
                                    col_str = str(col_name)[:38] + "..." if len(str(col_name)) > 38 else str(col_name)
                                    print(f"{col_str:<40} {input_str:<30} {ref_str:<30} {status:<10} {reason}")
                                    match_count += 1
                                elif match_count == 50:
                                    remaining = len(matches) - 50
                                    print(f"... 일치 항목 {remaining}개 더 있음 (생략)")
                                    match_count += 1
                        
                        print(f"{'='*80}")
                        print(f"[INFO] 요약: 일치={len(matches)}개, 불일치={len(diffs)}개, 전체={len(comparison_data)}개")
                        print(f"{'='*80}\n")
                    
                    # UI에 비교 결과 표시 (불일치가 있거나 모든 결과를 보고 싶을 때)
                    if comparison_data:
                        try:
                            self.show_feature_comparison_dialog(comparison_data, filename if filename else "파일", ref_name, len(matches), len(diffs), len(common_cols))
                        except Exception as e:
                            print(f"[WARN] 피처 비교 UI 표시 실패: {e}")
                            import traceback
                            traceback.print_exc()
                else:
                    print(f"[WARN] 비교할 수 없음: ref_row_series={ref_row_series is not None}, common_cols={len(common_cols)}")
        except Exception as e:
            print(f"[WARN] 학습 CSV 헤더/행 비교 실패: {e}")
            import traceback
            traceback.print_exc()
        
        # Detect에서는 학습 CSV 헤더와도 직접 비교하지 않는다.
        try:
            if False and self.csv_path and os.path.exists(self.csv_path):
                try:
                    ref_df = pd.read_csv(self.csv_path, nrows=0, on_bad_lines='skip')
                except TypeError:
                    ref_df = pd.read_csv(self.csv_path, nrows=0, error_bad_lines=False)
                ref_cols = [str(c) for c in ref_df.columns]
                missing_from_df_csv = [c for c in ref_cols if c not in df.columns]
                extra_in_df_csv = [c for c in df.columns if c not in ref_cols]
                if missing_from_df_csv or extra_in_df_csv:
                    print(f"[WARN] 학습 CSV 헤더와 입력 컬럼 불일치: 학습CSV {len(ref_cols)}개, 입력 {df.shape[1]}개")
                    if missing_from_df_csv:
                        print(f"[WARN] 입력에 없는 학습CSV 컬럼({len(missing_from_df_csv)}): {missing_from_df_csv[:20]}{' ...' if len(missing_from_df_csv)>20 else ''}")
                    if extra_in_df_csv:
                        print(f"[WARN] 학습CSV에 없는 추가 컬럼({len(extra_in_df_csv)}): {extra_in_df_csv[:20]}{' ...' if len(extra_in_df_csv)>20 else ''}")
        except Exception as e:
            print(f"[WARN] 학습 CSV 헤더 대비 비교 실패: {e}")
        
        # 6. scaler_features에 맞춰 컬럼 순서 정렬 및 결측값 처리
        # Train에서는 apply_simhash 후에 name과 label을 제거한 후 scaler에 맞춰 정렬
        scaler_features = self.scaler.feature_names_in_
        # 디버그: 스케일러와 입력 컬럼 비교 + 정합성 검사 (있어도 계속 진행; 누락은 0으로 대체)
        missing = []
        extra = []
        try:
            missing = [c for c in scaler_features if c not in df.columns]
            extra = [c for c in df.columns if c not in scaler_features]
            print(f"[DEBUG] 스케일러 피처 수: {len(scaler_features)}, 입력 피처 수: {df.shape[1]}")
            if missing:
                print(f"[DEBUG] 누락 피처({len(missing)}): {missing[:10]}{' ...' if len(missing)>10 else ''}")
            if extra:
                print(f"[DEBUG] 초과 피처({len(extra)}): {extra[:10]}{' ...' if len(extra)>10 else ''}")
        except Exception as e:
            print(f"[WARN] 스케일러/입력 피처 비교 실패: {e}")

        # 피처 누락 처리 정책:
        # - 누락이 있어도 "되도록 예측은 진행" (0으로 채워진 피처가 많으면 정확도가 급락할 수 있으므로 경고는 반드시 표시)
        if missing:
            missing_ratio = (len(missing) / float(len(scaler_features))) if scaler_features is not None and len(scaler_features) else 1.0

            msg = (
                "입력 피처가 학습 피처와 완전히 일치하지 않습니다.\n\n"
                f"- 스케일러 피처: {len(scaler_features)}개\n"
                f"- 입력 피처: {df.shape[1]}개\n"
                f"- 누락 피처: {len(missing)}개 (비율 {missing_ratio:.1%})\n"
                f"- 누락 예시: {missing[:20]}{' ...' if len(missing)>20 else ''}\n\n"
                "결론: 누락된 피처는 0으로 채워 예측을 계속합니다.\n"
                "※ 누락이 클수록 예측 정확도가 크게 떨어질 수 있습니다. (추출 파이프라인/입력을 학습과 맞추는 것이 근본 해결)"
            )

            try:
                print(f"[WARN] {msg}")
            except Exception:
                pass
            if show_warnings:
                try:
                    self.show_alert(msg)
                except Exception:
                    pass

        df = df.reindex(columns=scaler_features, fill_value=0)  # 결측값을 0으로 채움 (scaler가 숫자 기대)
        X_new_scaled = self.scaler.transform(df)
        y_pred = self.model.predict(X_new_scaled)
        y_pred_probs = self.model.predict_proba(X_new_scaled)
        
        # y_pred를 항상 배열로 변환 (스칼라인 경우 처리)
        y_pred = np.atleast_1d(y_pred)
        
        # LabelEncoder가 있으면 inverse_transform으로 원래 라벨로 복원 (XGBoost 등 인코딩 사용 시)
        if hasattr(self, 'label_encoder') and self.label_encoder is not None:
            try:
                y_pred = self.label_encoder.inverse_transform(y_pred.astype(int))
                print(f"[INFO] LabelEncoder로 라벨 복원: {y_pred}")
            except Exception as e:
                print(f"[WARN] LabelEncoder inverse_transform 실패 (인코딩된 라벨 그대로 사용): {e}")
        
        # 첫 번째 샘플의 클래스 확률 벡터
        predicted_class_probs = y_pred_probs[0]

        # 확률 벡터 인덱스(0..K-1)와 실제 클래스 라벨 매핑을 안전하게 복원
        class_labels_for_prob = None
        try:
            if hasattr(self, 'label_encoder') and self.label_encoder is not None:
                class_labels_for_prob = self.label_encoder.inverse_transform(
                    np.arange(predicted_class_probs.shape[0], dtype=int)
                )
            elif hasattr(self, 'model') and hasattr(self.model, 'classes_'):
                class_labels_for_prob = np.asarray(self.model.classes_)
        except Exception as e:
            print(f"[WARN] 확률 클래스 라벨 복원 실패: {e}")
            class_labels_for_prob = None

        # 예측 라벨의 확률을 직접 매칭해서 가져온다.
        # (기존처럼 라벨 값을 확률 인덱스로 간주하면 잘못된 확률이 나올 수 있음)
        predicted_class_prob = None
        try:
            pred_label = y_pred[0]
            if class_labels_for_prob is not None and len(class_labels_for_prob) == predicted_class_probs.shape[0]:
                hit = np.where(class_labels_for_prob == pred_label)[0]
                if hit.size > 0:
                    predicted_class_prob = float(predicted_class_probs[int(hit[0])])
        except Exception as e:
            print(f"[WARN] 예측 라벨-확률 매칭 실패: {e}")

        # 매칭 실패 시 top1 확률로 대체 + 예측 라벨도 top1 기준으로 동기화
        try:
            top_idx = int(np.argmax(predicted_class_probs))
            top_prob = float(predicted_class_probs[top_idx])
            if predicted_class_prob is None:
                predicted_class_prob = top_prob
                if class_labels_for_prob is not None and len(class_labels_for_prob) > top_idx:
                    y_pred[0] = class_labels_for_prob[top_idx]
                else:
                    y_pred[0] = top_idx
                print("[WARN] 라벨-확률 매칭 실패로 top1 기준 동기화 적용")
            elif top_prob > 0 and abs(predicted_class_prob - top_prob) > 1e-9:
                print(
                    f"[WARN] 예측 라벨 확률({predicted_class_prob:.6f})과 top1 확률({top_prob:.6f}) 불일치 - "
                    "라벨 인코딩/클래스 매핑 확인 필요"
                )
        except Exception as e:
            print(f"[WARN] top1 확률 계산 실패: {e}")
            predicted_class_prob = float(np.max(predicted_class_probs))

        # Add predictions to DataFrame
        df['predicted_label'] = y_pred
        # 사람이 읽을 수 있는 레이블 이름 컬럼도 추가 (매핑 성공 시)
        df['predicted_label_name'] = ""

        # 먼저 CSV 에디터에서 저장한 레이블 매핑 JSON 파일을 찾아서 사용
        label_name = None
        temppred = int(y_pred[0])  # inverse_transform 후 원래 라벨 값
        
        # 이미 찾아진 매핑 JSON 경로가 있으면 사용, 없으면 찾기
        mapping_json_path = None
        mapping_json_path = self._find_mapping_json_for_detect_model(getattr(self, "pklpath", None))
        if mapping_json_path:
            self.detect_mapping_json_path = mapping_json_path
            self.set_mapping_json_path(mapping_json_path)
            print(f"[DEBUG] 사용된 매핑 JSON 경로: {mapping_json_path}")
        else:
            # 1. CSV 파일과 같은 디렉토리에서 label_mapping.json 찾기
            if self.csv_path:
                mapping_json_path = self.find_mapping_json(self.csv_path)
                if mapping_json_path:
                    self.detect_mapping_json_path = mapping_json_path
                    self.set_mapping_json_path(mapping_json_path)
            
            # 2. Model 파일과 같은 디렉토리에서 label_mapping.json 찾기
            if not mapping_json_path and hasattr(self, 'pklpath') and self.pklpath:
                model_dir = os.path.dirname(os.path.abspath(self.pklpath))
                model_basename = os.path.basename(self.pklpath)
                model_prefix = model_basename.replace("model.pkl", "").replace(".pkl", "")
                
                possible_mapping_paths = [
                    os.path.join(model_dir, "label_mapping.json"),
                    os.path.join(model_dir, f"{model_prefix}_label_mapping.json"),
                    os.path.join(os.path.dirname(model_dir), "label_mapping.json"),
                ]
                
                for path in possible_mapping_paths:
                    if os.path.exists(path):
                        mapping_json_path = path
                        self.detect_mapping_json_path = path
                        self.set_mapping_json_path(path)
                        print(f"[INFO] Model 경로에서 매핑 JSON 찾음: {os.path.basename(path)}")
                        break
            print(f"[DEBUG] 사용된 매핑 JSON 경로: {mapping_json_path}")
        
        # JSON 매핑 파일이 있으면 사용
        if mapping_json_path:
            try:
                import json
                with open(mapping_json_path, 'r', encoding='utf-8') as f:
                    mapping_data = json.load(f)
                print(f"[DEBUG] 매핑 JSON 로드: {mapping_json_path}")
                
                # 1) 클래스 키 -> 표시 라벨명 (그룹 패턴이 아닌 이름 우선)
                label_to_group = _mapping_class_key_to_display_name(mapping_data)
                # 2) 정보가 비었을 때만 구형: groups에서 label -> group(패턴)
                if not label_to_group:
                    for entry in mapping_data.get('groups_detail', []) or mapping_data.get('groups', []):
                        try:
                            lbl = str(entry.get('label', '')).strip()
                            grp = str(entry.get('group', '')).strip()
                            if lbl and grp and lbl not in label_to_group:
                                label_to_group[lbl] = grp
                        except Exception:
                            continue
                print(f"[DEBUG] 매핑 딕셔너리 크기: {len(label_to_group)}")
                
                # 매핑 정보가 전혀 없으면 경고 후 중단
                if not label_to_group and not (mapping_data.get('groups_detail') or mapping_data.get('groups')):
                    warn_msg = "매핑 JSON에 레이블 정보가 없습니다.\nlabel_to_group/groups_detail/groups가 모두 비어 있습니다."
                    print(f"[WARN] {warn_msg}")
                    if show_warnings:
                        try:
                            self.show_alert(warn_msg)
                        except Exception:
                            pass
                    return df

                label_str = str(temppred)
                
                # 정확한 매칭 시도
                if label_str in label_to_group:
                    label_name = label_to_group[label_str]
                else:
                    # 정규화된 레이블로도 시도 (1.0 -> 1 등)
                    for key, value in label_to_group.items():
                        try:
                            if int(float(key)) == temppred:
                                label_name = value
                                break
                        except (ValueError, TypeError):
                            continue
                
                # 3) 라벨 딕셔너리가 여전히 비어 있고, groups_detail 순서가 클래스 인덱스와 동일하다고 가정할 수 있을 때 위치 기반 fallback
                if not label_name:
                    groups_seq = mapping_data.get('groups_detail') or mapping_data.get('groups')
                    if isinstance(groups_seq, list) and temppred < len(groups_seq):
                        try:
                            ent = groups_seq[temppred]
                            candidate = (
                                str(ent.get('label_name', '')).strip()
                                or str(ent.get('label', '')).strip()
                                or str(ent.get('group', '')).strip()
                            )
                            if candidate:
                                label_name = candidate
                                print(f"[INFO] groups_detail 순서 기반 매핑 사용: idx={temppred} -> {label_name}")
                        except Exception:
                            pass
                
                if label_name:
                    print(f"[INFO] CSV 에디터 레이블 매핑 사용: {mapping_json_path}")
                    df['predicted_label_name'] = label_name
            except Exception as e:
                print(f"[WARN] 레이블 매핑 JSON 읽기 실패: {e}")
        
        # JSON 매핑이 없거나 찾지 못한 경우 기존 방식 사용
        if not label_name:
            try:
                # Load label information from CSV (케이스 디렉토리에서 찾기)
                labelname = "labeldata_mul.csv"
                labelname = os.path.join(self.case_direc, labelname)
                labelpath = self.resource_path(labelname)
                
                # 파일이 없으면 경고만 출력하고 계속 진행
                if not os.path.exists(labelpath):
                    print(f"[WARN] labeldata_mul.csv 파일을 찾을 수 없습니다: {labelpath}")
                    label_name = None
                else:
                    labeltransferdf = pd.read_csv(labelpath)

                    # Ensure label columns are integers
                    try:
                        labeltransferdf.columns = [int(col) for col in labeltransferdf.columns]
                    except (ValueError, TypeError) as e:
                        print(f"[WARN] labeldata_mul.csv 컬럼 변환 실패: {e}")
                        label_name = None
                        raise

                    # Filter the relevant label
                    if temppred in labeltransferdf.columns:
                        filtered_df = labeltransferdf[temppred]
                        label_name = filtered_df[0] if len(filtered_df) > 0 else None
                    else:
                        print(f"[WARN] 레이블 {temppred}가 labeldata_mul.csv에 없습니다.")
                        label_name = None
            except KeyError:
                message = f"해당 라벨({temppred})이 존재하지 않습니다. 라벨을 업데이트하세요."
                if show_warnings:
                    self.show_alert(message)
                label_name = None
            except Exception as e:
                print(f"[WARN] labeldata_mul.csv 읽기 실패: {e}")
                label_name = None

        # 매핑이 성공한 경우 DataFrame에도 기록
        if label_name:
            df['predicted_label_name'] = label_name

        # 예측 결과 메시지 표시
        try:
            # Format probability and show message
            # predicted_class_prob는 예측된 클래스의 확률 값 (스칼라)
            fileaccuracy = "{:.3f}".format(float(predicted_class_prob) * 100)
            
            # 파일명이 있으면 메시지 앞에 추가
            filename_prefix = f"{filename} 파일은 " if filename else ""
            
            if label_name:
                summary = f"{filename_prefix}{fileaccuracy}% 확률로 {label_name}({y_pred[0]}) 입니다"
            else:
                summary = f"{filename_prefix}{fileaccuracy}% 확률로 레이블 {y_pred[0]} 입니다"
            paths_text = self._detect_used_paths_text(jsonpath, mapping_json_path)
            message = summary + paths_text
            df.attrs["detect_result_summary"] = summary
            df.attrs["detect_used_paths_text"] = paths_text
            df.attrs["detect_probability_percent"] = fileaccuracy
            df.attrs["detect_predicted_label"] = str(y_pred[0])
            df.attrs["detect_predicted_label_name"] = str(label_name or "")
            print(message)
            if show_message:
                self._copyable_msg(QMessageBox.Information, "Detect 결과", message)
        except Exception as e:
            # 예외 처리: predicted_class_prob 사용
            fileaccuracy = "{:.3f}".format(float(predicted_class_prob) * 100)
            
            # 파일명이 있으면 메시지 앞에 추가
            filename_prefix = f"{filename} 파일은 " if filename else ""
            
            if label_name:
                summary = f"{filename_prefix}{fileaccuracy}% 확률로 {label_name}({y_pred[0]}) 입니다"
            else:
                summary = f"{filename_prefix}{fileaccuracy}% 확률로 레이블 {y_pred[0]} 입니다"
            paths_text = self._detect_used_paths_text(jsonpath, mapping_json_path)
            message = summary + paths_text
            df.attrs["detect_result_summary"] = summary
            df.attrs["detect_used_paths_text"] = paths_text
            df.attrs["detect_probability_percent"] = fileaccuracy
            df.attrs["detect_predicted_label"] = str(y_pred[0])
            df.attrs["detect_predicted_label_name"] = str(label_name or "")
            print(message)
            if show_message:
                self._copyable_msg(QMessageBox.Information, "Detect 결과", message)

        return df

    def show_select_file(self):
        """현재 케이스의 CSV에 추가할지, 다른 CSV 파일을 선택할지 묻는 다이얼로그"""
        app = QApplication.instance()
        if not app:
            app = QApplication(sys.argv)

        dialog = QDialog(self)
        dialog.setWindowTitle("CSV 파일 선택")
        dialog.setWindowFlags(Qt.WindowStaysOnTopHint)
        dialog.setModal(True)

        # 다크 모드 스타일 적용
        dialog.setStyleSheet("""
                    QDialog {
                        background-color: #2e2e2e;
                        border: 2px solid #444;
                        border-radius: 15px;
                        padding: 20px;
                        font: bold 10pt "Playfair Display";
                    }
                    QLabel {
                        color: #f5f5f5;
                        font-size: 16px;
                        font-weight: bold;
                        margin-bottom: 10px;
                    }
                    QPushButton {
                        background-color: #444;
                        color: white;
                        border: 1px solid #777;
                        border-radius: 5px;
                        padding: 8px 15px;
                        margin-top: 10px;
                        min-width: 100px;
                    }
                    QPushButton:hover {
                        background-color: #555;
                    }
                """)

        layout = QVBoxLayout()
        
        # 현재 선택된 CSV 파일명 표시
        csv_filename = os.path.basename(self.csv_path) if self.csv_path else "없음"
        messages = f"현재 케이스의 CSV 파일에 추가하시겠습니까?\n\n선택된 파일: {csv_filename}\n\n'아니오'를 선택하면 다른 CSV 파일을 직접 선택할 수 있습니다."
        message_label = QLabel(messages)
        message_label.setWordWrap(True)
        layout.addWidget(message_label)

        # 버튼 레이아웃
        button_layout = QVBoxLayout()
        
        # '예' 버튼 - 현재 CSV 사용
        yes_button = QPushButton("예 (현재 CSV 사용)")
        yes_button.clicked.connect(dialog.accept)
        button_layout.addWidget(yes_button)

        # '아니오' 버튼 - 다른 파일 선택
        no_button = QPushButton("아니오 (다른 파일 선택)")
        def select_other_file():
            self.filedialog()
            dialog.accept()
        no_button.clicked.connect(select_other_file)
        button_layout.addWidget(no_button)
        
        # '취소' 버튼
        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(dialog.reject)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        dialog.setLayout(layout)
        dialog.setFixedSize(450, 250)

        # 다이얼로그 실행 및 결과 반환
        result = dialog.exec_()
        return result == QDialog.Accepted


    def filedialog(self):
        """다른 CSV 파일을 선택하는 다이얼로그"""
        # 기본 경로를 현재 케이스 디렉토리로 설정
        default_path = self.case_direc if hasattr(self, 'case_direc') and self.case_direc else ""
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "CSV 파일 선택", 
            default_path, 
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            self.csv_file = file_path
            self.csv_path = file_path  # csv_path도 업데이트
            # 매핑 JSON 자동 찾기
            mapping_json_path = self.find_mapping_json(self.csv_path)
            if mapping_json_path:
                self.set_mapping_json_path(mapping_json_path)
                print(f"[INFO] 매핑 JSON 파일 자동 로드: {os.path.basename(mapping_json_path)}")
            else:
                self.set_mapping_json_path(None)
            print(f"선택된 CSV 파일: {file_path}")
            try:
                if hasattr(self, "csv_info_label"):
                    self.csv_info_label.setVisible(True)
            except Exception:
                pass
            try:
                self.load_or_initialize_states()
            except Exception:
                pass

            return file_path
        return None

    def show_file_alert(self, file_path, messagea, widgett):
        """파일 경로를 받아 사용자에게 알림을 표시하고 파일을 여는 함수."""
        app = QApplication.instance()  # 이미 실행 중인 QApplication 인스턴스 확인
        if not app:
            app = QApplication(sys.argv)

        # QDialog를 사용해 타이틀 없는 커스텀 알림창 생성
        dialog = QDialog()
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)  # 타이틀 바 제거 및 최상단 설정

        # 다크 모드 스타일 적용
        dialog.setStyleSheet("""
                    QDialog {
                        background-color: #2e2e2e;
                        border: 2px solid #444;
                        border-radius: 15px;
                        padding: 20px;
                        font: bold 10pt "Playfair Display";
                    }
                    QLabel {
                        color: #f5f5f5;
                        font-size: 20px;
                        font-weight: bold;
                        margin-bottom: 10px;
                        font: bold 10pt "Playfair Display";
                    }
                    QPushButton {
                        background-color: #444;
                        color: white;
                        border: 1px solid #777;
                        border-radius: 5px;
                        padding: 8px 15px;
                        margin-top: 10px;
                        font: bold 10pt "Playfair Display";
                    }
                    QPushButton:hover {
                        background-color: #555;
                    }
                """)

        # 레이아웃 생성 및 위젯 추가
        layout = QVBoxLayout()
        messages = messagea + "바로 확인하시겠습니까?"
        message_label = QLabel(messages)
        message_label.setWordWrap(True)
        message_label.setTextInteractionFlags(Qt.TextSelectableByMouse)  # 텍스트 선택/복사 가능
        layout.addWidget(message_label)

        # '확인' 버튼 추가
        open_button = QPushButton("확인")
        open_button.clicked.connect(lambda: self.open_csv2(file_path, widgett))  # 파일 열기 함수 호출
        open_button.clicked.connect(dialog.accept)
        layout.addWidget(open_button)

        # '취소' 버튼 추가
        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(dialog.reject)  # 창 닫기
        layout.addWidget(cancel_button)
        dialog.setFixedSize(400, 200)
        dialog.setLayout(layout)

        # 창 크기 조정 및 화면 중앙 배치
        dialog.adjustSize()

        # 알림 창 표시
        dialog.exec_()
        return


    def show_alert(self, message):
        if getattr(self, "_batch_run_quiet", False):
            log = getattr(self, "_batch_run_log", None)
            if log is not None:
                log.append(str(message))
            else:
                print(f"[BATCH] {message}")
            return
        title = "알림"
        app = QApplication.instance()  # 이미 실행 중인 QApplication 인스턴스 확인
        if not app:
            app = QApplication(sys.argv)

        # QDialog를 사용해 타이틀 없는 커스텀 알림창 생성
        dialog = QDialog()
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)  # 타이틀 바 제거 및 최상단 설정

        # 다크 모드 스타일 적용
        dialog.setStyleSheet("""
            QDialog {
                background-color: #2e2e2e;
                border: 2px solid #444;
                border-radius: 15px;
                padding: 20px;
                font: bold 10pt "Playfair Display";
            }
            QLabel {
                color: #f5f5f5;
                font-size: 20px;
                font-weight: bold;
                margin-bottom: 10px;
                font: bold 10pt "Playfair Display";
            }
            QLabel#alertTitleLabel {
                color: #ffffff;
                font: bold 11pt "Helvetica";
            }
            QLabel#alertMessageLabel {
                color: #f5f5f5;
                font: 10pt "Helvetica";
                background-color: transparent;
            }
            QPushButton {
                background-color: #444;
                color: white;
                border: 1px solid #777;
                border-radius: 5px;
                padding: 8px 15px;
                margin-top: 10px;
                font: bold 10pt "Playfair Display";
            }
            QPushButton:hover {
                background-color: #555;
            }
        """)

        # 레이아웃 생성 및 위젯 추가
        layout = QVBoxLayout()
        title_label = QLabel(title)
        title_label.setObjectName("alertTitleLabel")
        message_label = QLabel(message)
        message_label.setObjectName("alertMessageLabel")
        message_label.setStyleSheet("color: #f5f5f5; background-color: transparent; font: 10pt \"Helvetica\";")
        message_label.setTextInteractionFlags(Qt.TextSelectableByMouse)  # 텍스트 선택/복사 가능
        message_label.setWordWrap(True)
        layout.addWidget(title_label)
        layout.addWidget(message_label)

        # 확인 버튼 추가
        button = QPushButton("확인")
        button.clicked.connect(dialog.accept)  # 버튼 클릭 시 창 닫기
        layout.addWidget(button)

        dialog.setLayout(layout)

        # 창 크기 조정 및 화면 중앙 배치
        dialog.adjustSize()


        # 알림 창 표시
        dialog.exec_()

    def _copyable_msg(self, icon, title, text, buttons=QMessageBox.Ok, default_btn=QMessageBox.NoButton):
        """텍스트 선택/복사 가능한 QMessageBox (에러문 복사용)"""
        if getattr(self, "_batch_run_quiet", False):
            log = getattr(self, "_batch_run_log", None)
            if log is not None:
                log.append(f"[{title}] {text}")
            else:
                print(f"[BATCH] {title}: {text}")
            return None
        return copyable_message_box(self, icon, title, text, buttons, default_btn)

    def show_feature_comparison_dialog(self, comparison_data, input_filename, ref_filename, match_count, diff_count, total_count):
        """피처별 비교 결과를 UI 테이블로 표시하는 함수"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle(f"피처 비교 결과: {input_filename}")
            dialog.resize(1200, 700)
            
            # 레이아웃 생성
            layout = QVBoxLayout()
            
            # 요약 정보 표시
            summary_label = QLabel(
                f"비교 대상: {input_filename} ↔ 학습 CSV ({ref_filename})\n"
                f"총 비교 컬럼: {total_count}개 | 일치: {match_count}개 | 불일치: {diff_count}개"
            )
            summary_label.setStyleSheet("font-weight: bold; font-size: 12pt; padding: 10px;")
            layout.addWidget(summary_label)
            
            # 필터 버튼 (일치/불일치 필터링)
            filter_layout = QHBoxLayout()
            filter_label = QLabel("필터:")
            filter_combo = QComboBox()
            filter_combo.addItems(["전체", "일치만", "불일치만"])
            filter_layout.addWidget(filter_label)
            filter_layout.addWidget(filter_combo)
            filter_layout.addStretch()
            filter_widget = QWidget()
            filter_widget.setLayout(filter_layout)
            layout.addWidget(filter_widget)
            
            # 테이블 생성
            table = QTableWidget()
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels(["컬럼명", "입력 값", "학습 CSV 값", "상태", "비고"])
            
            # 테이블 데이터 채우기
            all_comparison_data = comparison_data.copy()  # 필터링을 위한 복사본
            
            def update_table(filter_text="전체"):
                table.setRowCount(0)
                filtered_data = all_comparison_data
                if filter_text == "일치만":
                    filtered_data = [d for d in all_comparison_data if d[3] == "일치"]
                elif filter_text == "불일치만":
                    filtered_data = [d for d in all_comparison_data if d[3] == "불일치"]
                
                table.setRowCount(len(filtered_data))
                for row_idx, (col_name, input_val, ref_val, status, reason) in enumerate(filtered_data):
                    # 값이 너무 길면 축약
                    input_display = str(input_val)[:100] + "..." if len(str(input_val)) > 100 else str(input_val)
                    ref_display = str(ref_val)[:100] + "..." if len(str(ref_val)) > 100 else str(ref_val)
                    
                    table.setItem(row_idx, 0, QTableWidgetItem(str(col_name)))
                    table.setItem(row_idx, 1, QTableWidgetItem(input_display))
                    table.setItem(row_idx, 2, QTableWidgetItem(ref_display))
                    table.setItem(row_idx, 3, QTableWidgetItem(status))
                    table.setItem(row_idx, 4, QTableWidgetItem(reason))
                    
                    # 불일치 행은 배경색 변경
                    if status == "불일치":
                        for col in range(5):
                            if table.item(row_idx, col):
                                table.item(row_idx, col).setBackground(QColor(255, 200, 200))
                    elif status == "일치":
                        for col in range(5):
                            if table.item(row_idx, col):
                                table.item(row_idx, col).setBackground(QColor(200, 255, 200))
            
            # 초기 데이터 로드
            update_table("전체")
            
            # 필터 변경 시 테이블 업데이트
            filter_combo.currentTextChanged.connect(update_table)
            
            # 테이블 설정
            table.setAlternatingRowColors(True)
            table.setSelectionBehavior(QTableWidget.SelectRows)
            table.horizontalHeader().setStretchLastSection(True)
            table.resizeColumnsToContents()
            
            layout.addWidget(table)
            
            # 닫기 버튼
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            close_button = QPushButton("닫기")
            close_button.clicked.connect(dialog.accept)
            button_layout.addWidget(close_button)
            button_widget = QWidget()
            button_widget.setLayout(button_layout)
            layout.addWidget(button_widget)
            
            dialog.setLayout(layout)
            dialog.exec_()
        except Exception as e:
            print(f"[WARN] 피처 비교 UI 생성 실패: {e}")
            import traceback
            traceback.print_exc()

    def extract_sps_features(self, file_path):
        """Extract SPS features."""
        try:
            parse_sps(file_path)
            file_name = os.path.basename(file_path) + ".264"
            sps_result = analyzesps(file_name)
            return sps_result
        finally:
            if os.path.exists(file_name):
                os.remove(file_name)


    def main(self):
        self.ngrams = []

        while True:
            choice = self.choice

            folder_path = os.getcwd()  # 폴더 경로
            filename = 'lcsdata.pkl'  # 확인하고 싶은 파일 이름

            a = self.file_exists(folder_path, filename) # lcsdata.pkl 확인

            if choice == 1: #기준 피처를 만들기 위함, 10개 이내의 파일로 파일형식의 피처 생성
                print("1클릭")
                self.extension = (self.file_paths[0].split('.'))[1]

                #헤더딕셔너리(기존딕셔너리에 없으면 추가하기 위함)
                """header = self.extract_value_tocsv(choice) # 헤더 추출해서 문자열로 반환
                headersave = header.replace('name,', '') # header에서 name 문자열 제거한 결과 저장
                filename = str(self.extension+ 'header.txt') # 헤더 정보 저장할 파일 경로, 이름
                self.add_string_if_not_exists(filename, headersave)
                messagebox.showinfo("Notification", "Learning data extraction has been completed")"""

                break

            elif choice == 2:
                # 기존 CSV 데이터셋에 새로운 데이터 추가 모드
                print("기존 CSV에 데이터 추가 모드")
                print("선택한 파일:", self.file_paths)
                
                # 파일이 선택되지 않았을 경우
                if not self.file_paths:
                    self.show_alert("추가할 파일을 먼저 선택해주세요.")
                    break
                
                # CSV 파일 경로 확인
                if self.csv_file != '' or self.existval == 1:
                    file_names = [os.path.basename(path) for path in self.file_paths]
                    
                    # CSV 파일 경로 결정
                    if self.csv_file != '':
                        csv_file_path = self.csv_file
                    else:
                        csv_file_path = self.csv_path
                    
                    # CSV 파일 존재 확인
                    if not csv_file_path or not os.path.exists(csv_file_path):
                        self.show_alert(f"CSV 파일을 찾을 수 없습니다: {csv_file_path}\n'기존 학습데이터에 추가' 버튼을 클릭하여 CSV 파일을 선택해주세요.")
                        break
                    
                    try:
                        # 기존 CSV 파일 읽기 (tokenizing 오류 방지를 위해 python 엔진 + on_bad_lines='skip' 기본 사용)
                        try:
                            df = pd.read_csv(csv_file_path, engine="python", on_bad_lines="skip")
                        except TypeError:
                            # pandas 구버전 호환
                            df = pd.read_csv(csv_file_path, error_bad_lines=False, warn_bad_lines=True)
                        
                        # 'md5' 컬럼이 없을 경우
                        if 'md5' not in df.columns:
                            self.show_alert("CSV 파일에 'md5' 컬럼이 없습니다. 유효한 데이터셋 파일이 아닙니다.")
                            break
                        
                        # 추가할 파일들의 해시 계산
                        file_info = {}
                        for path in self.file_paths:
                            try:
                                file_hash = self.get_fast_file_hash(path)
                                if file_hash:  # 해시가 성공적으로 계산된 경우만
                                    file_info[file_hash] = path
                            except Exception as e:
                                print(f"파일 해시 계산 오류 ({path}): {e}")
                                continue
                        
                        existing_hashes = set(df['md5'].dropna())  # NaN 제거

                        # CSV의 hash 컬럼과 file_info 해시 값을 비교
                        # 해시가 같으면서 name 컬럼의 파일명이 다른 경우 name 값을 업데이트
                        for i, row in df.iterrows():
                            file_hash = row['md5']
                            if pd.notna(file_hash) and file_hash in file_info:
                                csv_name = row['name']
                                file_name = os.path.basename(file_info[file_hash])
                                if csv_name != file_name:
                                    df.at[i, 'name'] = file_name
                        df.to_csv(csv_file_path, index=False)

                        # 기존 CSV에 없는 새로운 해시만 추출하여 new_entries에 저장
                        new_entries = [file_info[hash_val] for hash_val in file_info if hash_val not in existing_hashes]
                        
                        if new_entries:
                            print(f"새로 추가할 파일 수: {len(new_entries)}개 (전체: {len(self.file_paths)}개)")
                            self.extract_box_feature(new_entries)
                        else:
                            self.show_alert("모든 파일이 이미 CSV에 존재합니다. 새로 추가할 파일이 없습니다.")
                    except Exception as e:
                        self.show_alert(f"CSV 파일 처리 중 오류가 발생했습니다:\n{str(e)}")
                        print(f"CSV 파일 처리 오류: {e}")
                        import traceback
                        traceback.print_exc()

                else:
                    # CSV 파일이 선택되지 않은 경우 새로 생성
                    self.show_alert("CSV 파일이 선택되지 않았습니다. 새 CSV 파일을 생성합니다.")
                    self.extract_box_feature(self.file_paths)



                break

            elif choice == 3:
                self.file_paths = self.listWidget
                # 파일을 순서딕셔너리와 비교하여 1-2-3-4, 2-3-1-4 등 순서 리스트를 만들고, 이를 심해시화
                self.extension = 'mp4'

                self.sequencedem = []
                '''                hexlist = str(self.extension + '\\' +  "hexlist.pkl")
                                with open(hexlist, 'rb') as f:
                                    hexvalues = pickle.load(f)
                                for h in range(len(hexvalues)):
                                    self.feature_dictionary(hexvalues[h])'''


                self.feature_dictionary()

                filename_to_sequence = {}
                for path, value in self.sequencedem:
                    filename = os.path.basename(path)
                    filename_to_sequence[filename] = value

                extractvalue = str(self.extension + '\\' +  "extractvalues.csv")
                df = pd.read_csv(extractvalue)

                for index, row in df.iterrows():
                    if row['name'] in filename_to_sequence:
                        df.at[index, 'sequence'] = filename_to_sequence[row['name']]

                df.to_csv(extractvalue, index=False)
                messagebox.showinfo("Notification", "Learning data extraction has been completed")
                break

            else :
                self.show_alert("학습버튼을 선택하세요")
                break


    def ask_overwrite_labels(self):
        """덮어쓰기 여부를 묻는 팝업."""
        reply = QMessageBox.question(
            self, "Overwrite Labels", "Do you want to overwrite existing labels?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        return reply == QMessageBox.Yes

    def open_data_entry_window(self):
        """데이터 입력 창 열기."""
        overwrite_labels = self.ask_overwrite_labels()
        self.data_entry_window = DataEntryWindow(overwrite_labels, self.case_direc)
        self.data_entry_window.show()
        self.load_excel_data()

    def change_dataset_directory(self):
        """데이터셋 경로를 변경하는 함수"""
        # 폴더 선택 다이얼로그 사용
        dataset_direc = QFileDialog.getExistingDirectory(
            self, 
            "데이터셋 디렉터리 선택",
            self.dataset_direc if self.dataset_direc else os.getcwd()
        )
        
        if dataset_direc:
            # 경로 유효성 검사
            if not os.path.isdir(dataset_direc):
                self._copyable_msg(QMessageBox.Warning, "오류", "유효하지 않은 디렉터리입니다.")
                return
            
            # config/base_directory.xml에 저장
            xml_path = os.path.join(self._config_dir(), "base_directory.xml")
            root = ET.Element("settings")
            ET.SubElement(root, "dataset_directory").text = dataset_direc
            tree = ET.ElementTree(root)
            try:
                with open(xml_path, "wb") as file:
                    tree.write(file, encoding="utf-8", xml_declaration=True)
                self.dataset_direc = dataset_direc
                
                # treeView의 루트 경로 업데이트
                self.treeView.setRootIndex(self.dirModel.index(self.dataset_direc))
                
                self._copyable_msg(QMessageBox.Information, "성공", f"데이터셋 경로가 변경되었습니다:\n{dataset_direc}")
                print(f"Dataset directory changed to: {dataset_direc}")
            except IOError as e:
                self._copyable_msg(QMessageBox.Warning, "오류", f"데이터셋 경로 저장 실패: {str(e)}")

    def _createtraining_cases_root(self):
        """CaseSelectorApp과 동일: 스크립트 기준 Cases 폴더."""
        try:
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), "Cases")
        except Exception:
            return os.path.join(os.getcwd(), "Cases")

    def _read_dataset_dir_from_case_config(self, case_root):
        """케이스의 config/base_directory.xml(또는 루트)에서 dataset_directory 읽기. 없거나 무효면 None."""
        case_root = os.path.abspath(os.path.normpath(case_root))
        xml_candidates = [
            os.path.join(case_root, "config", "base_directory.xml"),
            os.path.join(case_root, "base_directory.xml"),
        ]
        for xml_path in xml_candidates:
            if not os.path.isfile(xml_path):
                continue
            try:
                tree = ET.parse(xml_path)
                root = tree.getroot()
                ds = root.findtext("dataset_directory")
                if ds:
                    ds = str(ds).strip()
                if ds and os.path.isdir(ds):
                    return ds
            except Exception as e:
                print(f"[WARN] base_directory.xml 읽기 실패 ({xml_path}): {e}")
        return None

    def _write_dataset_dir_to_case_config(self, case_root, dataset_direc):
        """케이스 config/base_directory.xml에 데이터셋 경로 저장."""
        case_root = os.path.abspath(os.path.normpath(case_root))
        config_d = os.path.join(case_root, "config")
        os.makedirs(config_d, exist_ok=True)
        xml_path = os.path.join(config_d, "base_directory.xml")
        root = ET.Element("settings")
        ET.SubElement(root, "dataset_directory").text = dataset_direc
        tree = ET.ElementTree(root)
        with open(xml_path, "wb") as file:
            tree.write(file, encoding="utf-8", xml_declaration=True)

    def _clear_data_view_table(self):
        try:
            tw = getattr(self, "tableWidget", None)
            if tw is not None:
                tw.clear()
                tw.setRowCount(0)
                tw.setColumnCount(0)
        except Exception:
            pass

    def _train_csv_files_in_dir(self, folder):
        """해당 폴더 바로 아래의 학습 CSV 후보를 최신순으로 반환."""
        if not folder or not os.path.isdir(folder):
            return []
        files = glob.glob(os.path.join(folder, "*.csv"))
        files = [
            f for f in files
            if "_train_" in os.path.basename(f)
            and "feature_importance.csv" not in os.path.basename(f)
            and "_processed" not in os.path.basename(f)
        ]
        return sorted(list(set(files)), key=lambda p: os.path.getmtime(p), reverse=True)

    def _case_detail_dirs_with_csv(self, case_root):
        """케이스 루트/하위 상세 폴더 중 _train_ CSV가 있는 폴더 목록."""
        if not case_root or not os.path.isdir(case_root):
            return []
        skip_names = {
            "config", "training_results", "clustering_results", "detect_results", "__pycache__",
        }
        dirs = []
        if self._train_csv_files_in_dir(case_root):
            dirs.append(os.path.abspath(case_root))
        for root, subdirs, _files in os.walk(case_root):
            subdirs[:] = [d for d in subdirs if d not in skip_names and not d.startswith(".")]
            if os.path.abspath(root) == os.path.abspath(case_root):
                pass
            elif self._train_csv_files_in_dir(root):
                dirs.append(os.path.abspath(root))
        return sorted(list(set(dirs)), key=lambda p: os.path.relpath(p, case_root).lower())

    def _detail_dir_has_detect_models(self, detail_dir):
        """상세 폴더에 호환 model/scaler 쌍이 있는지 확인."""
        if not detail_dir or not os.path.isdir(detail_dir):
            return False
        prev_detail = getattr(self, "detect_detail_dir", None)
        self.detect_detail_dir = os.path.abspath(detail_dir)
        try:
            return len(self._detect_model_scaler_candidates(latest_per_family=False)) > 0
        finally:
            self.detect_detail_dir = prev_detail

    def _case_feature_detail_dirs(self, case_root=None):
        """케이스 내 feature별 상세 폴더 중 Detect 가능한 폴더 목록."""
        case_root = case_root or getattr(self, "case_direc", None)
        if not case_root or not os.path.isdir(case_root):
            return []
        dirs = self._case_detail_dirs_with_csv(case_root)
        feature_dirs = [d for d in dirs if self._detail_dir_has_detect_models(d)]
        return feature_dirs

    def _apply_detect_detail_context(self, detail_dir):
        """Detect를 특정 feature 상세 폴더 기준으로 전환 (CSV·states·매핑)."""
        detail_dir = os.path.abspath(detail_dir)
        self.detect_detail_dir = detail_dir
        csv_files = self._train_csv_files_in_dir(detail_dir)
        if csv_files:
            self._apply_csv_for_current_case(csv_files[0])
        else:
            self.set_mapping_json_path(None)
            try:
                self.load_or_initialize_states()
            except Exception:
                pass

    def _resolve_case_and_detail_folder(self, selected_folder, cases_root):
        """직접 선택한 폴더가 상세 폴더이면 부모를 케이스로, 선택 폴더를 상세로 해석."""
        folder = os.path.abspath(os.path.normpath(selected_folder))
        parent = os.path.abspath(os.path.dirname(folder))
        try:
            cases_root_abs = os.path.abspath(os.path.normpath(cases_root))
        except Exception:
            cases_root_abs = ""
        if (
            self._train_csv_files_in_dir(folder)
            and cases_root_abs
            and os.path.dirname(parent) == cases_root_abs
        ):
            return parent, folder
        return folder, None

    def _choose_detail_folder_for_case(self, case_root, preselected_detail=None):
        """케이스 안에서 사용할 상세 폴더를 선택. 후보가 하나면 자동 선택."""
        detail_dirs = self._case_detail_dirs_with_csv(case_root)
        if not detail_dirs:
            return None
        if preselected_detail:
            pre = os.path.abspath(os.path.normpath(preselected_detail))
            if pre in detail_dirs:
                return pre
        if self._detect_run_all_models_selected() and not preselected_detail:
            return _DETECT_DETAIL_ALL
        if len(detail_dirs) == 1 and not getattr(self, "detect_all_detail_folders", False):
            return detail_dirs[0]

        all_label = self._detect_all_models_label()
        folder_labels = []
        for d in detail_dirs:
            rel = os.path.relpath(d, case_root)
            folder_labels.append("(케이스 루트)" if rel == "." else rel)
        labels = [all_label] + folder_labels
        current_index = 0
        if getattr(self, "detect_all_detail_folders", False) or self._detect_run_all_models_selected():
            current_index = 0
        else:
            cur = getattr(self, "detect_detail_dir", None)
            if cur:
                cur = os.path.abspath(os.path.normpath(cur))
                if cur in detail_dirs:
                    current_index = detail_dirs.index(cur) + 1
        selected_label, ok = QInputDialog.getItem(
            self,
            "상세 폴더 선택",
            "Detect에 사용할 케이스 상세 폴더를 선택하세요:\n"
            f"(「{all_label}」: Detect 실행 시 케이스 내 feature 상세 폴더를 모두 사용)",
            labels,
            current_index,
            False
        )
        if not ok:
            return None
        if selected_label == all_label:
            return _DETECT_DETAIL_ALL
        return detail_dirs[folder_labels.index(selected_label)]

    def _apply_detect_detail_choice(self, choice):
        """상세 폴더 선택 결과를 Detect 상태·CSV 표시에 반영."""
        if choice is _DETECT_DETAIL_ALL:
            self.detect_all_detail_folders = True
            self.detect_detail_dir = None
            self._reload_primary_csv_for_current_case(all_detail_folders=True)
            try:
                combo = self.model_combo_2
                idx = combo.findText(self._detect_all_models_label())
                if idx >= 0:
                    combo.setCurrentIndex(idx)
            except Exception:
                pass
            return
        self.detect_all_detail_folders = False
        detail_dir = choice if choice else None
        self._reload_primary_csv_for_current_case(detail_dir)
        try:
            combo = self.model_combo_2
            if combo.currentText().strip() == self._detect_all_models_label():
                for i in range(combo.count()):
                    text = combo.itemText(i).strip()
                    if text and text != self._detect_all_models_label():
                        combo.setCurrentIndex(i)
                        break
        except Exception:
            pass

    def _apply_csv_for_current_case(self, csv_path):
        """CSV 경로를 Detect/Train 공통 상태와 Data View에 반영."""
        self.csv_path = csv_path
        self.csv_file = csv_path
        self.current_csv_path = csv_path
        self.existval = 1
        try:
            mj = self.find_mapping_json(csv_path)
            self.set_mapping_json_path(mj)
        except Exception as e:
            print(f"[WARN] 매핑 JSON: {e}")
            self.set_mapping_json_path(None)
        try:
            self.open_csv2(csv_path, self.tableWidget)
        except Exception as e:
            print(f"[WARN] open_csv2: {e}")
        try:
            if hasattr(self, "csvlabel"):
                self.csvlabel.setText(os.path.basename(csv_path))
        except Exception:
            pass
        try:
            self.load_or_initialize_states()
        except Exception:
            pass
        self._refresh_train_tab_case_csv_info()

    def _reload_primary_csv_for_current_case(self, detail_dir=None, all_detail_folders=False):
        """케이스/상세 폴더의 _train_ CSV를 다시 찾아 연다. 없으면 테이블·매핑을 비운다."""
        self.csv_path = ""
        self.csv_file = ""
        self.current_csv_path = None
        case = getattr(self, "case_direc", None)
        if all_detail_folders:
            self.detect_all_detail_folders = True
            self.detect_detail_dir = None
        if not case or not os.path.isdir(case):
            self.set_mapping_json_path(None)
            self._clear_data_view_table()
            try:
                if hasattr(self, "csvlabel"):
                    self.csvlabel.setText("")
            except Exception:
                pass
            try:
                if hasattr(self, "csv_info_label"):
                    self.csv_info_label.setText("CSV 정보를 표시합니다.")
            except Exception:
                pass
            self._refresh_train_tab_case_csv_info()
            return

        csv_files = []
        if all_detail_folders:
            for d in self._case_detail_dirs_with_csv(case):
                csv_files.extend(self._train_csv_files_in_dir(d))
            csv_files = sorted(list(set(csv_files)), key=lambda p: os.path.getmtime(p), reverse=True)
            if not csv_files:
                csv_files = self._train_csv_files_in_dir(case)
        elif detail_dir and os.path.isdir(detail_dir):
            self.detect_detail_dir = os.path.abspath(detail_dir)
            csv_files = self._train_csv_files_in_dir(detail_dir)
        if not csv_files and not all_detail_folders:
            self.detect_detail_dir = None
            csv_files = self._train_csv_files_in_dir(case)
        if not csv_files and not all_detail_folders:
            for d in self._case_detail_dirs_with_csv(case):
                csv_files.extend(self._train_csv_files_in_dir(d))
            csv_files = sorted(list(set(csv_files)), key=lambda p: os.path.getmtime(p), reverse=True)

        if csv_files:
            if all_detail_folders:
                self.detect_detail_dir = None
                self._apply_csv_for_current_case(csv_files[0])
                try:
                    if hasattr(self, "csvlabel"):
                        self.csvlabel.setText(
                            f"(전체 상세폴더 · 참고용 {os.path.basename(csv_files[0])})"
                        )
                except Exception:
                    pass
            else:
                csv_dir = os.path.dirname(os.path.abspath(csv_files[0]))
                self.detect_detail_dir = csv_dir if os.path.abspath(csv_dir) != os.path.abspath(case) else None
                self._apply_csv_for_current_case(csv_files[0])
        else:
            self.set_mapping_json_path(None)
            self._clear_data_view_table()
            try:
                if hasattr(self, "csvlabel"):
                    self.csvlabel.setText("(케이스에 _train_ CSV 없음)")
            except Exception:
                pass
            try:
                if hasattr(self, "csv_info_label"):
                    self.csv_info_label.setText("CSV 정보를 표시합니다.")
            except Exception:
                pass

        try:
            if hasattr(self, "caselabel") and self.case_direc:
                self.caselabel.setText(self.case_direc[-20:])
        except Exception:
            pass
        self._refresh_train_tab_case_csv_info()

    def _detect_model_name(self):
        try:
            name = self.model_combo_2.currentText()
            name = str(name).strip() if name else ""
            if name == self._detect_all_models_label():
                return ""
            return name
        except Exception:
            return ""

    def _model_feature_count(self, model_obj):
        for attr in ("n_features_in_", "n_features_"):
            try:
                value = getattr(model_obj, attr, None)
                if value:
                    return int(value)
            except Exception:
                pass
        try:
            booster = model_obj.get_booster()
            value = booster.num_features()
            if value:
                return int(value)
        except Exception:
            pass
        return None

    def _scaler_feature_count(self, scaler_obj):
        try:
            names = getattr(scaler_obj, "feature_names_in_", None)
            if names is not None:
                return len(names)
        except Exception:
            pass
        try:
            value = getattr(scaler_obj, "n_features_in_", None)
            if value:
                return int(value)
        except Exception:
            pass
        return None

    def _model_scaler_paths_compatible(self, model_path, scaler_path):
        """model.pkl과 scaler.pkl의 입력 피처 수가 맞는지 확인."""
        if not model_path or not scaler_path:
            return False
        if not os.path.exists(model_path) or not os.path.exists(scaler_path):
            return False
        try:
            model_obj = joblib.load(model_path)
            scaler_obj = joblib.load(scaler_path)
            model_count = self._model_feature_count(model_obj)
            scaler_count = self._scaler_feature_count(scaler_obj)
            if model_count and scaler_count and model_count != scaler_count:
                print(
                    "[WARN] Model/Scaler 피처 수 불일치로 제외: "
                    f"model={model_count}, scaler={scaler_count}\n"
                    f"  Model: {model_path}\n"
                    f"  Scaler: {scaler_path}"
                )
                return False
            return True
        except Exception as e:
            print(f"[WARN] Model/Scaler 호환성 확인 실패: {e}")
            return False

    def _detect_used_paths_text(self, feature_json_path=None, mapping_json_path=None):
        """Detect에 실제 사용된 주요 파일 경로를 복사 가능한 텍스트로 반환."""
        paths = [
            ("Model PKL", getattr(self, "pklpath", "")),
            ("Scaler PKL", getattr(self, "scalerpath", "")),
            ("LabelEncoder PKL", getattr(self, "label_encoder_path", "")),
            ("feature.json", feature_json_path or getattr(self, "detect_feature_json_path", "")),
            ("label_mapping.json", mapping_json_path or getattr(self, "detect_mapping_json_path", "")),
        ]
        lines = ["", "[Detect 사용 파일]"]
        for title, path in paths:
            lines.append(f"{title}: {path if path else '(없음)'}")
        return "\n".join(lines)

    def _find_matching_scaler_for_model(self, selected_model_path, scaler_files):
        model_basename = os.path.basename(selected_model_path)
        model_dir = os.path.dirname(selected_model_path)
        candidates = []
        if "training_results" in selected_model_path and model_basename == "model.pkl":
            same_dir_scaler = os.path.join(model_dir, "scaler.pkl")
            if os.path.isfile(same_dir_scaler):
                candidates.append(same_dir_scaler)
        same_name_scaler = os.path.join(model_dir, model_basename.replace("model.pkl", "scaler.pkl"))
        if os.path.isfile(same_name_scaler):
            candidates.append(same_name_scaler)
        model_prefix = model_basename.replace("model.pkl", "").replace(".pkl", "")
        for scaler_file in scaler_files:
            if os.path.dirname(os.path.abspath(scaler_file)) != os.path.abspath(model_dir):
                continue
            scaler_basename = os.path.basename(scaler_file)
            scaler_prefix = scaler_basename.replace("scaler.pkl", "").replace(".pkl", "")
            if scaler_prefix == model_prefix:
                candidates.append(scaler_file)

        seen = set()
        for scaler_file in candidates:
            if not scaler_file or scaler_file in seen:
                continue
            seen.add(scaler_file)
            if self._model_scaler_paths_compatible(selected_model_path, scaler_file):
                return scaler_file
        print(f"[WARN] 같은 폴더에서 호환되는 scaler.pkl을 찾지 못했습니다: {model_dir}")
        return None

    def _model_prefix_for_mapping_json(self, model_path):
        """모델 파일명에서 label_mapping JSON prefix를 추정."""
        base = os.path.basename(model_path or "")
        prefix = base.replace("model.pkl", "").replace(".pkl", "")
        for model_name in ["Xgboost", "RandomForest", "LGBM", "LogisticRegression", "LSTM"]:
            if prefix.endswith(model_name):
                prefix = prefix[:-len(model_name)]
                break
        prefix = prefix.rstrip("_")
        if prefix.endswith(".csv"):
            prefix = prefix[:-4]
        return prefix.rstrip("_")

    def _find_mapping_json_for_detect_model(self, model_path=None):
        """Detect에서 CSV 없이도 모델 폴더/상세 폴더 config의 label_mapping을 찾는다."""
        candidates = []
        model_path = model_path or getattr(self, "pklpath", None) or getattr(self, "selected_model_path", None)
        if model_path:
            model_dir = os.path.dirname(os.path.abspath(model_path))
            prefix = self._model_prefix_for_mapping_json(model_path)
            possible_roots = [model_dir]
            if "training_results" in os.path.normpath(model_dir):
                # detail/training_results/run -> detail
                possible_roots.append(os.path.dirname(os.path.dirname(model_dir)))
            possible_roots.append(os.path.dirname(model_dir))
            for root in list(dict.fromkeys([r for r in possible_roots if r and os.path.isdir(r)])):
                config_dir = os.path.join(root, "config")
                if prefix:
                    candidates.append(os.path.join(config_dir, f"{prefix}_label_mapping.json"))
                    candidates.append(os.path.join(root, f"{prefix}_label_mapping.json"))
                candidates.append(os.path.join(config_dir, "label_mapping.json"))
                candidates.append(os.path.join(root, "label_mapping.json"))
                try:
                    candidates.extend(sorted(glob.glob(os.path.join(config_dir, "*_label_mapping.json"))))
                    candidates.extend(sorted(glob.glob(os.path.join(root, "*_label_mapping.json"))))
                except Exception:
                    pass

        if getattr(self, "csv_path", None):
            csv_mapping = self.find_mapping_json(self.csv_path)
            if csv_mapping:
                candidates.append(csv_mapping)

        case_config = self._config_dir() if getattr(self, "case_direc", None) else None
        if case_config:
            candidates.append(os.path.join(case_config, "label_mapping.json"))
        if getattr(self, "mapping_json_path", None) and os.path.exists(self.mapping_json_path):
            candidates.append(self.mapping_json_path)

        seen = set()
        for path in candidates:
            if not path or path in seen:
                continue
            seen.add(path)
            if os.path.exists(path):
                return path
        return None

    def _detect_model_search_roots(self):
        """Detect 모델 자동 선택용 검색 루트. 상세 폴더가 선택되면 그 폴더만 사용."""
        detail_dir = getattr(self, "detect_detail_dir", None)
        if detail_dir and os.path.isdir(detail_dir):
            return [os.path.abspath(detail_dir)]

        csv_path = getattr(self, "csv_path", "")
        if csv_path and os.path.exists(csv_path):
            csv_dir = os.path.dirname(os.path.abspath(csv_path))
            self.detect_detail_dir = csv_dir
            return [csv_dir]

        roots = []
        case = getattr(self, "case_direc", None)
        if case:
            case = os.path.abspath(case)
            roots.append(case)
            roots.extend(self._case_detail_dirs_with_csv(case))
        return list(dict.fromkeys([p for p in roots if p and os.path.isdir(p)]))

    def _detect_paths_in_current_scope(self, *paths):
        """선택된 Detect 상세 폴더/CSV 범위 안의 파일인지 확인."""
        roots = self._detect_model_search_roots()
        if not roots:
            return True
        norm_roots = [os.path.normcase(os.path.abspath(r)) for r in roots]
        for path in paths:
            if not path:
                return False
            try:
                abs_path = os.path.normcase(os.path.abspath(path))
                if not any(os.path.commonpath([abs_path, root]) == root for root in norm_roots):
                    return False
            except Exception:
                return False
        return True

    def _detect_model_display_name(self, model_path):
        """training_results 실행 폴더명 등으로 구분 가능한 모델 표시명."""
        model_path = os.path.normpath(model_path or "")
        parent = os.path.basename(os.path.dirname(model_path))
        if parent and parent != "training_results" and "training_results" in model_path:
            return parent
        return self._detect_model_family_from_path(model_path)

    def _detect_algorithm_token_map(self):
        return {
            "xgboost": "Xgboost",
            "xgb": "Xgboost",
            "randomforest": "RandomForest",
            "randomforestmodel": "RandomForest",
            "rf": "RandomForest",
            "lightgbm": "LightGBM",
            "lgbm": "LightGBM",
            "logisticregression": "LogisticRegression",
            "logistic": "LogisticRegression",
            "lstm": "LSTM",
        }

    def _detect_algorithm_from_compact_token(self, compact):
        """공백/구분자 제거 후 알고리즘명 매칭."""
        compact = re.sub(r"[\s_\-]+", "", str(compact or "").lower())
        if not compact:
            return None
        token_map = self._detect_algorithm_token_map()
        for token, display in sorted(token_map.items(), key=lambda x: -len(x[0])):
            if compact == token or compact.startswith(token) or token in compact:
                return display
        return None

    def _detect_algorithm_from_model_basename(self, base):
        """*_Xgboostmodel.pkl, *_Random forestmodel.pkl 등 레거시 파일명에서 알고리즘 추출."""
        base = os.path.basename(base or "")
        if not base:
            return None
        stem = base
        if stem.lower().endswith("model.pkl"):
            stem = stem[: -len("model.pkl")]
        low = stem.lower()
        if ".csv_" in low:
            algo_part = stem[low.rindex(".csv_") + len(".csv_") :]
        elif "_" in stem:
            algo_part = stem.rsplit("_", 1)[-1]
        else:
            algo_part = stem
        return self._detect_algorithm_from_compact_token(algo_part)

    def _detect_model_family_from_path(self, model_path):
        """model.pkl 위치/파일명에서 사람이 읽을 모델 종류를 추정."""
        base = os.path.basename(model_path or "")
        parent = os.path.basename(os.path.dirname(model_path or ""))
        from_basename = self._detect_algorithm_from_model_basename(base)
        if from_basename:
            return from_basename

        haystack = f"{base} {parent}".lower()
        haystack_compact = re.sub(r"[\s_\-]+", "", haystack)
        families = [
            ("Xgboost", ("xgboost", "xgb")),
            ("RandomForest", ("randomforest", "random_forest", "random forest", "rf")),
            ("LightGBM", ("lightgbm", "lgbm")),
            ("LogisticRegression", ("logisticregression", "logistic_regression", "logistic")),
            ("LSTM", ("lstm",)),
        ]
        for display, tokens in families:
            for token in tokens:
                tok_compact = re.sub(r"[\s_\-]+", "", token)
                if token in haystack or tok_compact in haystack_compact:
                    return display

        if parent and parent != "training_results":
            parent_algo = self._detect_algorithm_from_compact_token(parent.split("_")[0])
            if parent_algo:
                return parent_algo
            if "training_results" in os.path.normpath(model_path or ""):
                parent_algo = self._detect_algorithm_from_compact_token(parent)
                if parent_algo:
                    return parent_algo
                return parent
            if not re.match(r"^\d+([_\-]|$)", parent):
                return parent.split("_")[0]
        return base or "Unknown"

    def _detect_record_model_name(self, model_path):
        """results.csv용 model_name — 실제 pkl 경로 기준 (콤보/aimodel과 무관)."""
        model_path = model_path or ""
        if not model_path:
            return ""
        norm = os.path.normpath(model_path)
        parent = os.path.basename(os.path.dirname(norm))
        if parent and parent != "training_results" and "training_results" in norm:
            return parent
        return self._detect_model_family_from_path(model_path)

    def _detect_model_matches_name(self, path, model_name):
        if not model_name:
            return True
        model_name_l = str(model_name).strip().lower()
        if model_name_l == self._detect_all_models_label().lower():
            return True
        base = os.path.basename(path).lower()
        parent = os.path.basename(os.path.dirname(path)).lower()
        aliases = {
            "lgbm": ("lgbm", "lightgbm"),
            "lightgbm": ("lgbm", "lightgbm"),
            "xgboost": ("xgboost", "xgb"),
            "randomforest": ("randomforest", "random_forest", "rf"),
            "logisticregression": ("logisticregression", "logistic_regression", "logistic"),
        }
        tokens = aliases.get(model_name_l, (model_name_l,))
        return any(token in base or parent.startswith(token) or token in parent for token in tokens)

    def _detect_model_scaler_candidates(self, model_name=None, latest_per_family=False):
        """현재 Detect 범위에서 호환되는 model/scaler 후보를 최신순으로 반환."""
        search_roots = self._detect_model_search_roots()
        model_files = []
        scaler_files = []
        for root in search_roots:
            model_files.extend(glob.glob(os.path.join(root, "*model.pkl")))
            scaler_files.extend(glob.glob(os.path.join(root, "*scaler.pkl")))
            tr_dir = os.path.join(root, "training_results")
            if os.path.isdir(tr_dir):
                for run_name in os.listdir(tr_dir):
                    run_path = os.path.join(tr_dir, run_name)
                    if not os.path.isdir(run_path):
                        continue
                    m = os.path.join(run_path, "model.pkl")
                    s = os.path.join(run_path, "scaler.pkl")
                    if os.path.isfile(m):
                        model_files.append(m)
                    if os.path.isfile(s):
                        scaler_files.append(s)
            try:
                for walk_root, dirnames, _filenames in os.walk(root):
                    if os.path.basename(walk_root) != "training_results":
                        continue
                    for run_name in dirnames:
                        run_path = os.path.join(walk_root, run_name)
                        m = os.path.join(run_path, "model.pkl")
                        s = os.path.join(run_path, "scaler.pkl")
                        if os.path.isfile(m):
                            model_files.append(m)
                        if os.path.isfile(s):
                            scaler_files.append(s)
            except Exception as e:
                print(f"[WARN] Detect 모델 하위 폴더 검색 실패({root}): {e}")

        model_files = [
            p for p in set(model_files)
            if self._detect_model_matches_name(p, model_name)
            and self._detect_paths_in_current_scope(p)
        ]
        scaler_files = [
            p for p in set(scaler_files)
            if self._detect_paths_in_current_scope(p)
        ]
        model_files = sorted(model_files, key=lambda p: os.path.getmtime(p), reverse=True)

        candidates = []
        used_families = set()
        for model_path in model_files:
            scaler_path = self._find_matching_scaler_for_model(model_path, scaler_files)
            if not scaler_path:
                continue
            family = self._detect_model_family_from_path(model_path)
            display_name = self._detect_model_display_name(model_path)
            family_key = family.lower()
            dedupe_key = os.path.normcase(os.path.abspath(model_path))
            if dedupe_key in used_families:
                continue
            if latest_per_family and family_key in used_families:
                continue
            candidates.append({
                "family": family,
                "display_name": display_name,
                "model_path": model_path,
                "scaler_path": scaler_path,
            })
            used_families.add(dedupe_key if not latest_per_family else family_key)
        return candidates

    def _auto_select_detect_model_for_current_csv(self, show_message=False):
        """현재 Detect 모델 콤보 기준으로 최신 model.pkl/scaler.pkl을 자동 선택."""
        csv_path = getattr(self, "csv_path", "")
        model_name = self._detect_model_name()
        search_roots = self._detect_model_search_roots()

        model_files = []
        scaler_files = []
        for root in search_roots:
            model_files.extend(glob.glob(os.path.join(root, "*model.pkl")))
            scaler_files.extend(glob.glob(os.path.join(root, "*scaler.pkl")))
            tr_dir = os.path.join(root, "training_results")
            if os.path.isdir(tr_dir):
                for run_name in os.listdir(tr_dir):
                    run_path = os.path.join(tr_dir, run_name)
                    if not os.path.isdir(run_path):
                        continue
                    m = os.path.join(run_path, "model.pkl")
                    s = os.path.join(run_path, "scaler.pkl")
                    if os.path.isfile(m):
                        model_files.append(m)
                    if os.path.isfile(s):
                        scaler_files.append(s)
            # 상세 폴더가 선택되지 않은 초기 Detect에서도 케이스 하위 training_results를 찾는다.
            if os.path.abspath(root) == os.path.abspath("."):
                continue
            try:
                for walk_root, dirnames, _filenames in os.walk(root):
                    if os.path.basename(walk_root) != "training_results":
                        continue
                    for run_name in dirnames:
                        run_path = os.path.join(walk_root, run_name)
                        m = os.path.join(run_path, "model.pkl")
                        s = os.path.join(run_path, "scaler.pkl")
                        if os.path.isfile(m):
                            model_files.append(m)
                        if os.path.isfile(s):
                            scaler_files.append(s)
            except Exception as e:
                print(f"[WARN] Detect 모델 하위 폴더 검색 실패({root}): {e}")

        def matches_model(path):
            return self._detect_model_matches_name(path, model_name)

        model_files = [p for p in set(model_files) if matches_model(p)]
        scaler_files = list(set(scaler_files))
        model_files = sorted(model_files, key=lambda p: os.path.getmtime(p), reverse=True)
        if not model_files:
            self.selected_model_path = ""
            self.selected_scaler_path = ""
            print(f"[WARN] Detect 자동 모델 선택 실패: {model_name or '모델'} pkl 없음")
            print(f"[DEBUG] Detect 모델 검색 경로: {search_roots}")
            return False

        for model_path in model_files:
            scaler_path = self._find_matching_scaler_for_model(model_path, scaler_files)
            if scaler_path and os.path.exists(scaler_path):
                self.selected_model_path = model_path
                self.selected_scaler_path = scaler_path
                print(f"[INFO] Detect Model 자동 선택: {model_path}")
                print(f"[INFO] Detect Scaler 자동 선택: {scaler_path}")
                try:
                    mj = self._find_mapping_json_for_detect_model(model_path)
                    self.set_mapping_json_path(mj)
                    if mj:
                        print(f"[INFO] Detect 매핑 JSON 자동 선택: {mj}")
                except Exception:
                    pass
                if show_message:
                    self.show_alert(
                        "Detect Model/Scaler 자동 선택 완료:\n"
                        f"Model: {os.path.basename(model_path)}\n"
                        f"Scaler: {os.path.basename(scaler_path)}"
                    )
                return True

        self.selected_model_path = ""
        self.selected_scaler_path = ""
        print("[WARN] Detect 자동 모델 선택 실패: 매칭되는 scaler.pkl 없음")
        return False

    def switch_case_from_menu(self):
        """File 메뉴: 다른 케이스로 전환 (Cases 목록 또는 폴더 직접 선택)."""
        cases_root = self._createtraining_cases_root()
        dlg = QDialog(self)
        dlg.setWindowTitle("케이스 변경")
        dlg.setMinimumWidth(420)
        lay = QVBoxLayout(dlg)
        lay.addWidget(
            QLabel("다른 케이스를 선택하세요. (프로젝트 Cases 폴더 또는 임의 케이스 폴더)")
        )
        lw = QListWidget()
        if os.path.isdir(cases_root):
            try:
                for d in sorted(os.listdir(cases_root)):
                    p = os.path.join(cases_root, d)
                    if os.path.isdir(p):
                        lw.addItem(d)
            except Exception as e:
                print(f"[WARN] Cases 목록 읽기 실패: {e}")
        lay.addWidget(lw)

        try:
            cur = os.path.basename(os.path.normpath(self.case_direc or ""))
            for i in range(lw.count()):
                if lw.item(i).text() == cur:
                    lw.setCurrentRow(i)
                    break
        except Exception:
            pass

        chosen = [None]

        def pick_folder():
            start = cases_root if os.path.isdir(cases_root) else os.getcwd()
            folder = QFileDialog.getExistingDirectory(self, "케이스(작업) 폴더 선택", start)
            if folder:
                chosen[0] = os.path.abspath(folder)
                dlg.accept()

        def accept_list():
            it = lw.currentItem()
            if it:
                chosen[0] = os.path.join(cases_root, it.text())
            dlg.accept()

        btn_row = QHBoxLayout()
        btn_pick = QPushButton("폴더 직접 선택…")
        btn_ok = QPushButton("확인")
        btn_cancel = QPushButton("취소")
        btn_pick.clicked.connect(pick_folder)
        btn_ok.clicked.connect(accept_list)
        btn_cancel.clicked.connect(dlg.reject)
        btn_row.addWidget(btn_pick)
        btn_row.addStretch()
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        lay.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return
        if not chosen[0]:
            self._copyable_msg(
                QMessageBox.Warning,
                "선택 없음",
                "목록에서 케이스를 선택하거나 「폴더 직접 선택」을 사용하세요.",
            )
            return

        new_case, preselected_detail = self._resolve_case_and_detail_folder(chosen[0], cases_root)
        new_case = os.path.abspath(os.path.normpath(new_case))
        if not os.path.isdir(new_case):
            self._copyable_msg(QMessageBox.Warning, "오류", "유효하지 않은 폴더입니다.")
            return

        old_case = os.path.abspath(os.path.normpath(getattr(self, "case_direc", "") or ""))
        same_case = new_case == old_case

        reply = QMessageBox.question(
            self,
            "케이스 변경",
            "케이스/상세 폴더를 바꾸면 데이터셋 트리·states·CSV·Detect 모델이 새 선택 기준으로 맞춰집니다.\n"
            "저장하지 않은 내용은 각 케이스 폴더에 반영되지 않을 수 있습니다.\n\n계속할까요?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        self.case_direc = new_case
        if getattr(self, "detectmode", 0):
            detail_choice = self._choose_detail_folder_for_case(new_case, preselected_detail)
        else:
            detail_choice = preselected_detail
        try:
            os.makedirs(self._config_dir(), exist_ok=True)
        except Exception as e:
            print(f"[WARN] config 폴더 생성 실패: {e}")

        ds = self._read_dataset_dir_from_case_config(new_case)
        if ds:
            self.dataset_direc = ds
        elif getattr(self, "dataset_direc", None) and os.path.isdir(self.dataset_direc):
            pass
        else:
            picked = QFileDialog.getExistingDirectory(
                self,
                "새 케이스에 저장된 데이터셋 경로가 없습니다. 데이터셋 폴더를 선택하세요.",
                self.dataset_direc if self.dataset_direc else os.getcwd(),
            )
            if picked:
                self.dataset_direc = picked
                try:
                    self._write_dataset_dir_to_case_config(new_case, self.dataset_direc)
                except Exception as e:
                    self._copyable_msg(
                        QMessageBox.Warning,
                        "경고",
                        f"데이터셋 경로는 적용했지만 케이스에 저장하지 못했습니다:\n{e}",
                    )

        try:
            self._file_hash_cache.clear()
            self._simhash_cache.clear()
        except Exception:
            pass

        try:
            self.listWidget.clear()
            self.file_paths = []
        except Exception:
            pass

        self.trainclass.case_direc = getattr(self, "case_direc", None) or ""

        self.load_or_initialize_states()
        try:
            self.treeView.setRootIndex(self.dirModel.index(self.dataset_direc))
        except Exception as e:
            print(f"[WARN] treeView 루트 설정 실패: {e}")

        if getattr(self, "detectmode", 0):
            self._apply_detect_detail_choice(detail_choice)
            model_auto_selected = False
            if not getattr(self, "detect_all_detail_folders", False):
                model_auto_selected = self._auto_select_detect_model_for_current_csv(show_message=False)
        else:
            self.detect_all_detail_folders = False
            self._reload_primary_csv_for_current_case(detail_choice)
            model_auto_selected = False

        try:
            self.load_excel_data()
        except Exception:
            pass

        try:
            detail_text = ""
            if getattr(self, "detect_all_detail_folders", False):
                detail_text = f" / {self._detect_all_models_label()}"
            elif detail_choice and detail_choice is not _DETECT_DETAIL_ALL:
                detail_text = f" / {os.path.relpath(detail_choice, new_case)}"
            model_text = ""
            if getattr(self, "detectmode", 0):
                if getattr(self, "detect_all_detail_folders", False):
                    model_text = " / Detect: 전체 상세폴더 모드"
                else:
                    model_text = " / Detect 모델 자동 선택됨" if model_auto_selected else " / Detect 모델 자동 선택 실패"
            prefix = "케이스 상세 갱신됨" if same_case else "케이스 변경됨"
            self.statusBar().showMessage(f"{prefix}: {new_case}{detail_text}{model_text}", 8000)
        except Exception:
            pass
    
    def _select_csv_for_train(self):
        """Train용 CSV 파일을 선택하는 함수 (Detect의 model/scaler 선택과 유사)"""
        csv_files = self._collect_train_csv_candidates()
        if not csv_files:
            self._copyable_msg(QMessageBox.Warning, "오류", "CSV 파일을 찾을 수 없습니다.")
            return None
        
        # 현재 선택된 CSV가 있으면 우선 표시
        current_index = 0
        if self.csv_path and os.path.exists(self.csv_path):
            try:
                current_index = csv_files.index(self.csv_path)
            except ValueError:
                pass
        
        # CSV 파일 선택
        selected_csv_path = None
        if len(csv_files) == 1:
            # 파일이 하나면 자동 선택
            selected_csv_path = csv_files[0]
            print(f"[INFO] Train용 CSV 파일 자동 선택: {os.path.basename(selected_csv_path)}")
        else:
            # 여러 개면 사용자에게 선택 요청
            csv_names = [os.path.basename(f) for f in csv_files]
            csv_name, ok = QInputDialog.getItem(
                self, 
                "Train용 CSV 파일 선택", 
                "학습에 사용할 CSV 파일을 선택하세요:",
                csv_names, 
                current_index, 
                False
            )
            if not ok:
                return None
            
            # 선택한 파일명으로 경로 찾기
            selected_csv_path = next((f for f in csv_files if os.path.basename(f) == csv_name), None)
            if not selected_csv_path:
                self._copyable_msg(QMessageBox.Warning, "오류", f"선택한 CSV 파일을 찾을 수 없습니다: {csv_name}")
                return None
        
        return selected_csv_path

    def _select_csvs_for_train(self):
        """Train용 CSV 파일을 단일/다중으로 선택하는 함수."""
        single_csv = self._select_csv_for_train()
        if not single_csv:
            return []

        csv_files = self._collect_train_csv_candidates()
        if single_csv and os.path.abspath(single_csv) not in [os.path.abspath(p) for p in csv_files]:
            csv_files.insert(0, single_csv)

        if len(csv_files) <= 1:
            # 사용자가 "일괄 팝업이 안 뜬다"를 즉시 확인할 수 있도록 안내 팝업 표시
            self._copyable_msg(
                QMessageBox.Information,
                "CSV 일괄 학습",
                "일괄 대상 CSV가 1개만 발견되어 단일 CSV 학습으로 진행합니다.\n\n"
                f"- 대상: {os.path.basename(single_csv)}"
            )
            return [single_csv]

        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Question)
        msg.setWindowTitle("CSV 일괄 학습")
        msg.setText(
            "현재 케이스에서 찾은 모든 학습 CSV를 한 번에 학습할까요?\n\n"
            f"- 전체 CSV 수: {len(csv_files)}개\n"
            f"- 아니오를 누르면 선택한 CSV 1개만 학습합니다."
        )
        btn_all = msg.addButton("예 (전체 CSV)", QMessageBox.AcceptRole)
        btn_single = msg.addButton("아니오 (선택 1개)", QMessageBox.DestructiveRole)
        btn_cancel = msg.addButton("취소", QMessageBox.RejectRole)
        msg.setDefaultButton(btn_all)
        msg.exec_()

        clicked = msg.clickedButton()
        if clicked == btn_cancel:
            return []
        if clicked == btn_all:
            return csv_files
        return [single_csv]

    def _collect_train_csv_candidates(self):
        """Train 대상 CSV 후보를 수집해 최신순으로 반환."""
        search_dirs = []
        if self.case_direc:
            search_dirs.append(os.path.abspath(self.case_direc))
        if self.csv_path:
            search_dirs.append(os.path.dirname(os.path.abspath(self.csv_path)))
        search_dirs.append(os.path.abspath("."))

        csv_files = []
        seen = set()
        for search_dir in set(search_dirs):
            if not os.path.isdir(search_dir):
                continue

            # 1) 현재 폴더
            flat_pattern = os.path.join(search_dir, "*.csv")
            for f in glob.glob(flat_pattern):
                base = os.path.basename(f)
                if 'feature_importance.csv' in base:
                    continue
                # Detect/보조 데이터 CSV는 기본 제외
                low = base.lower()
                if low in {"labeldata_bin.csv", "labeldata_mul.csv"}:
                    continue
                ap = os.path.abspath(f)
                if ap not in seen:
                    seen.add(ap)
                    csv_files.append(ap)

            # 2) 하위 폴더 재귀 검색(케이스 내부 CSV 누락 방지)
            recursive_pattern = os.path.join(search_dir, "**", "*.csv")
            for f in glob.glob(recursive_pattern, recursive=True):
                base = os.path.basename(f)
                if 'feature_importance.csv' in base:
                    continue
                low = base.lower()
                if low in {"labeldata_bin.csv", "labeldata_mul.csv"}:
                    continue
                ap = os.path.abspath(f)
                if ap not in seen:
                    seen.add(ap)
                    csv_files.append(ap)

        csv_files = sorted(csv_files, key=lambda p: os.path.getmtime(p), reverse=True)
        print(f"[INFO] Train CSV 후보 수집 완료: {len(csv_files)}개")
        return csv_files

    def _collect_untrained_case_csv_targets(self):
        """
        케이스 내부 폴더를 순회해 training_results가 없는 폴더의 CSV만 수집.
        폴더마다 CSV가 여러 개면 최신 파일 1개를 대상으로 사용.
        """
        case_root = os.path.abspath(getattr(self, "case_direc", "") or "")
        if not case_root or not os.path.isdir(case_root):
            print(f"[WARN] 유효한 case_direc가 아닙니다: {case_root}")
            return []

        targets = []
        excluded_dir_names = {
            "config", "detect_results", "clustering_results", "__pycache__",
        }
        excluded_csv_names = {"feature_importance.csv", "labeldata_bin.csv", "labeldata_mul.csv"}

        for root, dirs, files in os.walk(case_root):
            dirs[:] = [d for d in dirs if d not in excluded_dir_names]

            # training_results 하위는 탐색 불필요
            if os.path.basename(root) == "training_results":
                dirs[:] = []
                continue

            # 폴더에 training_results가 있으면 학습 완료로 간주하고 해당 폴더 CSV는 제외
            if "training_results" in dirs:
                dirs[:] = [d for d in dirs if d != "training_results"]
                continue

            csv_candidates = []
            for name in files:
                if not name.lower().endswith(".csv"):
                    continue
                if name in excluded_csv_names:
                    continue
                csv_candidates.append(os.path.join(root, name))

            if not csv_candidates:
                continue

            csv_candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
            targets.append(csv_candidates[0])

        targets = sorted(set(os.path.abspath(p) for p in targets), key=lambda p: os.path.getmtime(p), reverse=True)
        print(f"[INFO] training_results 미존재 폴더 기준 자동 학습 대상: {len(targets)}개")
        return targets

    @staticmethod
    def _accuracy_percent_from_history_record(rec):
        """학습 기록 1건에서 정확도(%)를 추출. 실패 시 None."""
        try:
            tr = rec.get("training_result", {}) if isinstance(rec, dict) else {}
            if isinstance(tr, dict):
                ap = tr.get("accuracy_percent", None)
                if ap is not None:
                    return float(ap)
                a = tr.get("accuracy", None)
                if a is not None:
                    av = float(a)
                    return av * 100.0 if av <= 1.0 else av
        except Exception:
            pass
        try:
            metrics = rec.get("metrics", {}) if isinstance(rec, dict) else {}
            if isinstance(metrics, dict) and metrics.get("accuracy", None) is not None:
                av = float(metrics.get("accuracy"))
                return av * 100.0 if av <= 1.0 else av
        except Exception:
            pass
        return None

    def _feature_combo_from_history_record(self, rec):
        """학습 기록 1건에서 피처 조합 문자열 추출."""
        if not isinstance(rec, dict):
            return ""
        combo = str(rec.get("selected_features", "") or "").strip()
        if combo:
            return combo
        tc = rec.get("training_condition", {}) or {}
        if isinstance(tc, dict):
            combo = str(tc.get("selected_features", "") or "").strip()
            if combo:
                return combo
        ds = rec.get("dataset_summary", {}) or {}
        csv_path = ""
        if isinstance(ds, dict):
            csv_path = str(ds.get("csv_path", "") or "").strip()
        if not csv_path:
            csv_path = str(rec.get("csv_path", "") or "").strip()
        return self._feature_set_from_folder_name(csv_path) if csv_path else ""

    def _iter_case_training_history_records(self, case_root):
        """케이스 하위 학습 히스토리 JSON들을 순회하며 (record, source_path) 반환."""
        if not case_root or not os.path.isdir(case_root):
            return
        targets = []
        for root, _dirs, files in os.walk(case_root):
            for fn in files:
                low = fn.lower()
                if low == "training_history.json" or low == "training_history_all.json":
                    targets.append(os.path.join(root, fn))
        for path in sorted(set(targets)):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception as e:
                print(f"[WARN] 학습 히스토리 로드 실패: {path} ({e})")
                continue
            if isinstance(data, dict):
                data = [data]
            if not isinstance(data, list):
                continue
            for rec in data:
                if isinstance(rec, dict):
                    yield rec, path

    def show_case_feature_combo_summary(self):
        """현재 케이스의 피처 조합별 학습 성능 요약을 표시."""
        case_root = os.path.abspath(getattr(self, "case_direc", "") or "")
        if not case_root or not os.path.isdir(case_root):
            self._copyable_msg(QMessageBox.Warning, "요약 실패", "유효한 케이스 경로가 없습니다.")
            return

        combo_stats = {}
        seen = set()
        total_records = 0

        for rec, src in self._iter_case_training_history_records(case_root):
            acc = self._accuracy_percent_from_history_record(rec)
            if acc is None:
                continue
            total_records += 1
            combo = self._feature_combo_from_history_record(rec) or "(미표기)"
            model = str((rec.get("training_condition", {}) or {}).get("model_name", "") or rec.get("model", "") or "")
            ts = str(rec.get("timestamp", "") or rec.get("datetime", "") or "")
            csv_path = str((rec.get("dataset_summary", {}) or {}).get("csv_path", "") or rec.get("csv_path", "") or "")
            dedup_key = (combo, model, ts, csv_path, round(float(acc), 6))
            if dedup_key in seen:
                continue
            seen.add(dedup_key)

            row = combo_stats.setdefault(combo, {"accs": [], "best_acc": -1.0, "best_model": "", "best_csv": "", "best_src": ""})
            row["accs"].append(float(acc))
            if float(acc) > row["best_acc"]:
                row["best_acc"] = float(acc)
                row["best_model"] = model
                row["best_csv"] = os.path.basename(csv_path) if csv_path else ""
                row["best_src"] = os.path.relpath(src, case_root)

        if not combo_stats:
            self._copyable_msg(
                QMessageBox.Information,
                "피처 조합 성능 요약",
                "요약할 학습 기록이 없습니다.\n(케이스 하위 training_history.json / training_history_all.json 확인)"
            )
            return

        ranked = []
        for combo, row in combo_stats.items():
            accs = row["accs"]
            avg_acc = sum(accs) / len(accs)
            ranked.append((combo, row["best_acc"], avg_acc, len(accs), row["best_model"], row["best_csv"], row["best_src"]))
        ranked.sort(key=lambda x: (x[1], x[2], x[3]), reverse=True)

        best = ranked[0]
        lines = [
            f"케이스: {case_root}",
            f"분석 기록 수: {len(seen)}개 (원본 로드 {total_records}건)",
            "",
            f"[최고 정확도 조합] {best[0]}",
            f"- 최고 정확도: {best[1]:.3f}%",
            f"- 평균 정확도: {best[2]:.3f}% (표본 {best[3]}개)",
            f"- 최고 기록 모델: {best[4] or '(알 수 없음)'}",
            f"- 최고 기록 CSV: {best[5] or '(알 수 없음)'}",
            "",
            "[조합별 순위 TOP 10]",
        ]
        for idx, item in enumerate(ranked[:10], 1):
            combo, best_acc, avg_acc, cnt, model, csv_name, _src = item
            lines.append(
                f"{idx}. {combo} | 최고 {best_acc:.3f}% | 평균 {avg_acc:.3f}% | 표본 {cnt} | 최고모델 {model or '-'} | CSV {csv_name or '-'}"
            )

        self._copyable_msg(QMessageBox.Information, "피처 조합 성능 요약", "\n".join(lines))
    
    def find_mapping_json(self, csv_path):
        """CSV 파일과 연관된 매핑 JSON 파일을 찾는 함수. config 폴더(케이스 루트) 우선."""
        if not csv_path:
            return None
        
        csv_dir = os.path.dirname(os.path.abspath(csv_path))
        csv_name = os.path.splitext(os.path.basename(csv_path))[0]
        # 케이스 루트의 config 폴더 우선
        case_config = None
        if getattr(self, 'case_direc', None) and str(self.case_direc).strip():
            case_config = os.path.join(os.path.abspath(self.case_direc), "config")
        config_dir = os.path.join(csv_dir, "config")
        
        possible_paths = []
        if case_config and os.path.isdir(case_config):
            possible_paths.extend([
                os.path.join(case_config, f"{csv_name}_label_mapping.json"),
                os.path.join(case_config, "label_mapping.json"),
            ])
        possible_paths.extend([
            os.path.join(config_dir, f"{csv_name}_label_mapping.json"),
            os.path.join(config_dir, "label_mapping.json"),
            os.path.join(csv_dir, f"{csv_name}_label_mapping.json"),
            os.path.join(csv_dir, "label_mapping.json"),
            os.path.join(os.path.dirname(csv_dir), f"{csv_name}_label_mapping.json"),
        ])
        
        for path in possible_paths:
            if os.path.exists(path):
                return path
        
        return None

    def set_mapping_json_path(self, path):
        """매핑 JSON 경로를 설정하고 Label List를 갱신"""
        self.mapping_json_path = path
        try:
            self.update_label_list_from_json(path)
        except Exception as e:
            print(f"[WARN] Label List 갱신 실패: {e}")

    def update_label_list_from_json(self, mapping_json_path):
        """Train 탭 tableWidget_train에 매핑 JSON 기준 클래스→표시 라벨만 표시 (labeldata_mul.csv는 Detect 탭 전용)."""
        tw = getattr(self, "tableWidget_train", None)
        if tw is None:
            return

        def _fill_message(col1: str, col2: str):
            tw.blockSignals(True)
            tw.clear()
            tw.setColumnCount(2)
            tw.setHorizontalHeaderLabels(["클래스", "라벨(표시)"])
            tw.setRowCount(1)
            tw.setItem(0, 0, QTableWidgetItem(col1))
            tw.setItem(0, 1, QTableWidgetItem(col2))
            tw.resizeColumnsToContents()
            tw.blockSignals(False)

        if not mapping_json_path or not os.path.exists(mapping_json_path):
            _fill_message("—", "매핑 JSON 없음")
            return
        try:
            import json
            with open(mapping_json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            label_to_group = _mapping_class_key_to_display_name(data)
            if not label_to_group:
                for entry in data.get("groups_detail", []) or data.get("groups", []):
                    try:
                        lbl = str(entry.get("label", "")).strip()
                        grp = str(entry.get("group", "")).strip()
                        if lbl and grp and lbl not in label_to_group:
                            label_to_group[lbl] = grp
                    except Exception:
                        continue
            if not label_to_group and isinstance(data.get("groups_detail"), list):
                for idx, entry in enumerate(data.get("groups_detail")):
                    grp = str(entry.get("group", "")).strip()
                    if grp:
                        label_to_group[str(idx)] = grp
            if not label_to_group and isinstance(data.get("groups"), list):
                for idx, entry in enumerate(data.get("groups")):
                    grp = str(entry.get("group", "")).strip()
                    if grp:
                        label_to_group[str(idx)] = grp

            if not label_to_group:
                _fill_message("—", "매핑 정보가 비어 있습니다.")
                return

            def _sort_key(x):
                s = str(x).strip()
                try:
                    return (0, int(float(s)))
                except Exception:
                    return (1, s)

            keys = sorted(label_to_group.keys(), key=_sort_key)
            tw.blockSignals(True)
            tw.clear()
            tw.setColumnCount(2)
            tw.setHorizontalHeaderLabels(["클래스", "라벨(표시)"])
            tw.setRowCount(len(keys))
            for row, k in enumerate(keys):
                tw.setItem(row, 0, QTableWidgetItem(str(k)))
                tw.setItem(row, 1, QTableWidgetItem(str(label_to_group[k])))
            tw.resizeColumnsToContents()
            tw.blockSignals(False)
        except Exception as e:
            _fill_message("오류", str(e))
    
    def select_csv_file_from_menu(self):
        """File 메뉴에서 CSV 파일을 선택하는 함수"""
        # 기본 경로를 현재 케이스 디렉토리로 설정
        default_path = self.case_direc if hasattr(self, 'case_direc') and self.case_direc else os.getcwd()
        
        # CSV 파일 선택 다이얼로그
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "CSV 파일 선택", 
            default_path, 
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if file_path:
            try:
                self.csv_path = file_path
                self.csv_file = file_path
                
                # CSV 파일과 연관된 매핑 JSON 파일 찾기
                mapping_json_path = self.find_mapping_json(file_path)
                if mapping_json_path:
                    self.set_mapping_json_path(mapping_json_path)
                    print(f"[INFO] 매핑 JSON 파일 자동 로드: {os.path.basename(mapping_json_path)}")
                else:
                    self.set_mapping_json_path(None)
                    print(f"[INFO] 매핑 JSON 파일을 찾을 수 없습니다.")
                
                # CSV 파일을 테이블에 표시
                try:
                    self.open_csv2(file_path, self.tableWidget)
                    try:
                        if hasattr(self, "csv_info_label"):
                            self.csv_info_label.setVisible(True)
                    except Exception:
                        pass
                except Exception as e:
                    print(f"[WARN] CSV 파일 열기 오류: {e}")
                
                # csvlabel 업데이트
                try:
                    if hasattr(self, 'csvlabel'):
                        self.csvlabel.setText(os.path.basename(file_path))
                except Exception:
                    pass

                try:
                    self.load_or_initialize_states()
                except Exception:
                    pass
                try:
                    self.detect_detail_dir = os.path.dirname(os.path.abspath(file_path))
                    self._auto_select_detect_model_for_current_csv(show_message=False)
                except Exception as e:
                    print(f"[WARN] Detect 모델 자동 선택 실패: {e}")
                
                # 매핑 JSON 정보도 메시지에 포함
                mapping_info = f"\n매핑 JSON: {os.path.basename(mapping_json_path)}" if mapping_json_path else "\n매핑 JSON: 없음"
                self.show_alert(f"CSV 파일이 선택되었습니다:\n{os.path.basename(file_path)}{mapping_info}")
                print(f"[INFO] CSV 파일 선택: {file_path}")
                
            except Exception as e:
                self._copyable_msg(QMessageBox.Warning, "오류", f"CSV 파일 선택 중 오류가 발생했습니다:\n{str(e)}")
                print(f"[ERROR] CSV 파일 선택 오류: {e}")

    def load_excel_data(self):
        """labeldata_mul.csv는 Detect 탭에만 표시. Train 탭 tableWidget_train은 매핑 JSON(클래스→표시명) 전용이라 덮어쓰지 않음."""
        labelname = "labeldata_mul.csv"
        labelname = os.path.join(self.case_direc, labelname)
        labelpath = self.resource_path(labelname)
        if not os.path.exists(labelpath):
            print(self, "Warning", "No Excel file found!")
            try:
                self.update_label_list_from_json(getattr(self, "mapping_json_path", None))
            except Exception:
                pass
            return

        df = pd.read_csv(labelpath)
        df.columns = [str(col) for col in df.columns]

        self.display_dataframe(df, widgettype=self.tableWidget_detect)
        try:
            self.update_label_list_from_json(getattr(self, "mapping_json_path", None))
        except Exception:
            pass



    def resource_path(self, relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)


    def show_input_dialog(self, title, label):
        """커스텀 입력창을 표시하고 입력된 텍스트를 반환하는 함수."""
        app = QApplication.instance()  # 이미 실행 중인 QApplication 인스턴스 확인
        if not app:
            app = QApplication([])

        # QDialog 생성 및 스타일 설정
        dialog = QDialog()
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)  # 타이틀 바 제거 및 최상단 설정

        dialog.setStyleSheet("""
            QDialog {
                background-color: #2e2e2e;
                border: 2px solid #444;
                border-radius: 10px;
                padding: 20px;
            }
            QLabel {
                color: #f5f5f5;
                font-size: 16px;
            }
            QLineEdit {
                background-color: #444;
                color: #f5f5f5;
                border: 1px solid #777;
                border-radius: 5px;
                padding: 8px;
                margin-top: 10px;
            }
            QPushButton {
                background-color: #555;
                color: white;
                border-radius: 5px;
                padding: 8px 15px;
                margin-top: 10px;
            }
            QPushButton:hover {
                background-color: #666;
            }
        """)

        # 레이아웃 생성
        layout = QVBoxLayout()

        # 라벨 추가
        message_label = QLabel(label)
        layout.addWidget(message_label)

        # 입력 필드 추가
        line_edit = QLineEdit()
        layout.addWidget(line_edit)

        # 버튼 생성 및 추가
        button_layout = QVBoxLayout()

        ok_button = QPushButton("확인")
        ok_button.clicked.connect(dialog.accept)  # 확인 클릭 시 다이얼로그 닫기
        layout.addWidget(ok_button)

        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(dialog.reject)  # 취소 클릭 시 다이얼로그 닫기
        layout.addWidget(cancel_button)

        # 다이얼로그에 레이아웃 설정
        dialog.setLayout(layout)

        # 다이얼로그 실행 및 결과 처리
        if dialog.exec_() == QDialog.Accepted:
            return line_edit.text(), True
        return "", False

    @staticmethod
    def _feature_column_category_for_record(col_name):
        """training_history 기록용: 컬럼 → 피처 종류. Train_GRUprocess_multi와 동일 규칙."""
        if col_name in ('name', 'label', 'md5'):
            return None
        if col_name == 'sequence':
            return 'seq'
        if isinstance(col_name, str) and (col_name.startswith('SPS_') or col_name.startswith('PPS_') or col_name in ('SPS', 'PPS')):
            return 'sps'
        if col_name == 'GOP':
            return 'gop'
        if col_name == 'GOP compression':
            return 'ratio'
        return 'val'

    @staticmethod
    def _filter_feature_names_by_states(all_feature_names, selected_feature_set):
        """states(선택된 피처)에 걸러진 피처 이름만 반환. selected_feature_set 비어 있으면 전체 반환."""
        if not selected_feature_set or not isinstance(selected_feature_set, str):
            return list(all_feature_names)
        want = set(s.strip().lower() for s in selected_feature_set.split('_') if s.strip())
        if not want:
            return list(all_feature_names)
        keep = []
        for col in all_feature_names:
            cat = createtrainclass._feature_column_category_for_record(col)
            if cat and cat in want:
                keep.append(col)
        return keep

    def save_training_result(self, classmode, aimodel, trainindex, csv_path, selected_features=None, training_duration=None):
        """학습 결과를 JSON 파일로 기록"""
        try:
            # 데이터셋 정보 수집
            dataset_info = {}
            if csv_path and os.path.exists(csv_path):
                try:
                    df = pd.read_csv(csv_path)
                    all_feature_names = [col for col in df.columns if col not in ['name', 'label', 'md5']]
                    # 폴더 기반(없으면 UI 선택) 피처셋 기준으로 기록
                    sf = selected_features
                    if sf is None:
                        sf = self._feature_set_from_folder_name(csv_path) or self.get_train_selected_feature_set_string()
                    feature_names = self._filter_feature_names_by_states(all_feature_names, sf)
                    dataset_info = {
                        'csv_path': csv_path,
                        'total_files': len(df),
                        'total_features': len(feature_names),
                        'feature_names': feature_names
                    }
                    
                    # 레이블 정보
                    if 'label' in df.columns:
                        label_counts = df['label'].value_counts().to_dict()
                        dataset_info['label_distribution'] = {str(k): int(v) for k, v in label_counts.items()}
                        dataset_info['num_labels'] = len(label_counts)
                    else:
                        dataset_info['num_labels'] = 0
                        dataset_info['label_distribution'] = {}
                except Exception as e:
                    print(f"[WARN] 데이터셋 정보 수집 실패: {e}")
                    dataset_info = {'error': str(e)}
            
            # 학습 결과 정보 수집
            training_result = {}
            try:
                # train_obj에서 정확도 정보 가져오기
                if hasattr(self.trainclass, 'accuracy'):
                    acc = self.trainclass.accuracy
                    if acc is not None:
                        # accuracy가 백분율인지 소수인지 확인
                        if isinstance(acc, str):
                            # 문자열에서 숫자 추출 (예: "95.5%" -> 95.5)
                            import re
                            match = re.search(r'(\d+\.?\d*)', str(acc))
                            if match:
                                acc = float(match.group(1))
                                if acc > 1.0:
                                    acc = acc / 100.0  # 백분율을 소수로 변환
                        elif acc > 1.0:
                            acc = acc / 100.0  # 백분율을 소수로 변환
                        training_result['accuracy'] = float(acc)
                        training_result['accuracy_percent'] = float(acc) * 100.0
                    else:
                        training_result['accuracy'] = None
                elif hasattr(self.trainclass, 'model'):
                    # 모델이 있으면 학습은 완료된 것으로 간주
                    training_result['accuracy'] = None
                    training_result['note'] = '정확도 정보 없음 (LSTM 모델일 수 있음)'
                else:
                    training_result['accuracy'] = None
                    training_result['note'] = '정확도 정보를 찾을 수 없음'
            except Exception as e:
                print(f"[WARN] 정확도 정보 수집 실패: {e}")
                training_result['accuracy'] = None
                training_result['error'] = str(e)
            
            # 모델 인덱스에 따른 모델 이름 매핑
            model_names = {
                0: 'XGBoost',
                1: 'LSTM',
                2: 'RandomForest',
                3: 'LightGBM',
                4: 'LogisticRegression'
            }
            model_name = model_names.get(trainindex, f'Unknown({trainindex})')
            
            # 학습 조건 정보
            training_condition = {
                'classification_mode': 'Binary' if classmode == 'bin_' else 'Multi-class',
                'model_name': aimodel if aimodel else model_name,
                'model_index': trainindex,
                'model_type': model_name
            }
            
            # 학습 시간 정보
            time_info = {}
            if training_duration is not None:
                time_info['duration_seconds'] = float(training_duration)
                time_info['duration_minutes'] = float(training_duration / 60.0)
                time_info['duration_formatted'] = f"{int(training_duration // 60)}분 {int(training_duration % 60)}초"
            
            # 학습에 실제 사용된 피처 = 폴더 기반(없으면 UI 선택)
            sf = selected_features
            if sf is None:
                sf = self._feature_set_from_folder_name(csv_path) or self.get_train_selected_feature_set_string()
            if sf:
                training_condition['selected_features'] = sf

            # 최종 기록 데이터
            record = {
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'training_condition': training_condition,
                'dataset_summary': dataset_info,
                'training_result': training_result,
                'training_time': time_info,
                'case_directory': getattr(self, 'case_direc', ''),
                'extension': getattr(self, 'extension', ''),
                'selected_features': sf or ""
            }
            
            # 저장 경로 결정
            save_dir = None
            if csv_path:
                save_dir = os.path.dirname(os.path.abspath(csv_path))
            elif hasattr(self, 'case_direc') and self.case_direc:
                save_dir = os.path.abspath(self.case_direc)
            else:
                save_dir = os.path.abspath(".")
            
            # 학습 기록 파일 경로
            training_log_file = os.path.join(save_dir, 'training_history.json')
            
            # 기존 기록 불러오기
            training_history = []
            if os.path.exists(training_log_file):
                try:
                    with open(training_log_file, 'r', encoding='utf-8') as f:
                        training_history = json.load(f)
                    if not isinstance(training_history, list):
                        training_history = []
                except Exception as e:
                    print(f"[WARN] 기존 학습 기록 읽기 실패: {e}")
                    training_history = []
            
            # 새 기록 추가
            training_history.append(record)
            
            # 파일로 저장
            with open(training_log_file, 'w', encoding='utf-8') as f:
                json.dump(training_history, f, ensure_ascii=False, indent=2)
            
            print(f"[INFO] 학습 결과가 기록되었습니다: {training_log_file}")
            
        except Exception as e:
            print(f"[ERROR] 학습 결과 기록 중 오류: {e}")
            import traceback
            traceback.print_exc()
    
    ##############라벨입력
    def input_label(self):
        if self.binButton.isChecked():
            self.label_datacsv = 'labeldata_bin.csv'
        elif self.mulButton.isChecked():
            self.label_datacsv = 'labeldata_mul.csv'

        # 라벨 데이터 입력 받기
        self.label_data, ok = self.show_input_dialog("입력", "라벨 데이터를 입력하세요.")
        if not ok or not self.label_data:
            return

        try:
            number = float(self.label_data)  # 숫자 입력 여부 확인
        except ValueError:
            self.show_alert("에러", "유효한 숫자를 입력해주세요.")
            return


        # CSV에서 해당 라벨 데이터를 찾기
        try:
            self.label_datacsv = os.path.join(self.case_direc, self.label_datacsv)
            aaa = self.fetch_name_from_csv(self.label_data)
        except Exception as e:
            self.show_alert("binary/multi class 모드를 선택하세요")
            return

        if aaa is None:  # 해당 데이터가 없을 경우 새로 입력
            name, ok = self.show_input_dialog("입력", "매핑되는 속성을 입력하세요.")
            if not ok or not name:
                return

            # 파일이 없으면 생성
            if not os.path.exists(self.label_datacsv):
                with open(self.label_datacsv, mode='w', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    writer.writerow(["Label", "Name"])  # 헤더 작성

            # CSV 파일에 새 데이터 추가
            with open(self.label_datacsv, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow([self.label_data, name])
        else :
            self.show_alert("해당 라벨이 이미 존재합니다!")

    def fetch_name_from_csv(self, max_key):
        filename = self.label_datacsv

        # 파일이 없을 경우 None 반환
        if not os.path.exists(filename):
            return None

        # 파일에서 해당 키에 해당하는 값을 찾기
        with open(filename, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                if row[0] == max_key:
                    return row[1]
        return None



    def open_existcsv(self):
        """기존 CSV 데이터셋에 새로운 데이터를 추가하는 함수
        
        사용 방법:
        1. 파일 목록에 추가할 파일들을 선택
        2. 이 함수를 호출하여 CSV 파일 선택
        3. 'Load' 버튼을 클릭하여 데이터 처리 시작
        """
        # choice를 2로 설정하여 추가 모드 활성화
        self.choice = 2
        self.existval = 1
        
        # 현재 케이스의 CSV 파일 목록 찾기 (_train_이 포함된 파일만)
        csv_files = [file for file in glob.glob(os.path.join(self.case_direc, "*.csv")) 
                     if '_train_' in os.path.basename(file)
                     and 'feature_importance.csv' not in os.path.basename(file)]
        
        # _train_이 포함된 CSV 파일 찾기
        train_csv_files = csv_files
        
        if train_csv_files:
            # 최신 mtime 순으로 정렬해 가장 최근 CSV를 선택
            train_csv_files = sorted(train_csv_files, key=lambda p: os.path.getmtime(p), reverse=True)
            self.csv_path = train_csv_files[0]
            self.csv_file = ''  # 현재 케이스의 CSV 사용
            # 매핑 JSON 자동 찾기
            mapping_json_path = self.find_mapping_json(self.csv_path)
            if mapping_json_path:
                self.set_mapping_json_path(mapping_json_path)
                print(f"[INFO] 매핑 JSON 파일 자동 로드: {os.path.basename(mapping_json_path)}")
            else:
                self.set_mapping_json_path(None)
            print(f"기존 CSV 파일 선택: {self.csv_path}")
            try:
                self.load_or_initialize_states()
            except Exception:
                pass
            
            # 다이얼로그 표시하여 사용자에게 확인
            result = self.show_select_file()
            
            if result:
                # 사용자가 '예'를 선택한 경우 - 현재 CSV 사용 (이미 csv_path 설정됨)
                self.show_alert(f"CSV 파일이 선택되었습니다:\n{os.path.basename(self.csv_path)}\n\n이제 'Load' 버튼을 클릭하여 데이터를 추가하세요.")
            else:
                # 사용자가 취소한 경우
                self.choice = 0  # 초기화
                self.existval = 0
        else:
            # _train CSV 파일이 없으면 파일 선택 다이얼로그 표시
            reply = self._copyable_msg(
                QMessageBox.Information,
                "알림",
                "현재 케이스에 _train CSV 파일이 없습니다.\n다른 CSV 파일을 선택하시겠습니까?",
                buttons=QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                if self.filedialog():
                    self.show_alert(f"CSV 파일이 선택되었습니다:\n{os.path.basename(self.csv_file)}\n\n이제 'Load' 버튼을 클릭하여 데이터를 추가하세요.")
                else:
                    self.choice = 0
                    self.existval = 0
            else:
                self.choice = 0
                self.existval = 0


##################라벨입력
class DataEntryWindow(QWidget):
    def __init__(self, overwrite_labels, direc):
        super().__init__()
        self.setWindowTitle("Data Entry")
        self.label_counter = 0
        self.headers = []
        self.values = []
        labelpath = 'labeldata_mul.csv'
        labelpath = os.path.join(direc, labelpath)
        self.filename = createtrainclass.resource_path(self, labelpath)


        if not overwrite_labels:
            self.load_existing_labels()

        # UI 구성
        layout = QVBoxLayout()

        self.header_label = QLabel(f"Current Label: {self.label_counter}")
        layout.addWidget(self.header_label)

        self.value_input = QLineEdit()
        self.value_input.setPlaceholderText("Enter value...")
        layout.addWidget(self.value_input)

        add_value_button = QPushButton("Add Value")
        add_value_button.clicked.connect(self.add_value)
        layout.addWidget(add_value_button)

        stop_button = QPushButton("Stop and Save")
        stop_button.clicked.connect(self.stop_and_save)
        layout.addWidget(stop_button)

        self.table = QTableWidget()
        layout.addWidget(self.table)

        self.setLayout(layout)
        self.update_display()

    def load_existing_labels(self):
        """기존 엑셀 파일에서 라벨 로드."""
        if os.path.exists(self.filename):
            workbook = load_workbook(self.filename)
            sheet = workbook.active
            if sheet.max_row > 0:
                self.headers = [cell.value for cell in sheet[1]]
                self.label_counter = len(self.headers)
            workbook.close()

    def add_value(self):
        """값을 추가하고 다음 라벨로 이동."""
        value = self.value_input.text()
        if not value:
            copyable_message_box(self, QMessageBox.Critical, "Error", "Value cannot be empty!")
            return

        self.values.append(value)
        self.headers.append(str(self.label_counter))
        self.label_counter += 1
        self.value_input.clear()
        self.update_display()
        self.header_label.setText(f"Current Label: {self.label_counter}")

    def update_display(self):
        """현재까지 입력된 데이터를 테이블에 표시."""
        self.table.setRowCount(1)
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        for col, value in enumerate(self.values):
            self.table.setItem(0, col, QTableWidgetItem(value))

    def stop_and_save(self):
        """엑셀에 데이터를 저장하고 창 닫기."""
        if not self.values:
            copyable_message_box(self, QMessageBox.Warning, "Warning", "No data to save!")
            return

        if os.path.exists(self.filename):
            workbook = load_workbook(self.filename)
            sheet = workbook.active
            if sheet.max_row == 0:
                sheet.append(self.headers)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(self.headers)

        sheet.append(self.values)
        workbook.save(self.filename)
        workbook.close()

        copyable_message_box(self, QMessageBox.Information, "Success", "Data saved successfully!")
        self.close()


class CaseSelectorApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # Initialize case and dataset directories
        self.case_direc = None
        self.dataset_direc = None

        # Set up the main window
        self.setWindowTitle("Select a Case")
        self.setFixedSize(500, 400)

        # Set up menu bar
        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("File")

        # Add NewCase action to the menu
        new_case_action = QAction("NewCase", self)
        new_case_action.triggered.connect(self.create_new_case)
        file_menu.addAction(new_case_action)
        
        # Add Dataset Directory setting action to the menu
        dataset_dir_action = QAction("데이터셋 경로 설정", self)
        dataset_dir_action.triggered.connect(self.set_dataset_directory)
        file_menu.addAction(dataset_dir_action)

        # Set up main layout with a dark theme
        layout = QVBoxLayout()

        # Title label
        label = QLabel("Select a Case:")
        label.setStyleSheet("color: #f5f5f5; font-size: 20px; font-weight: bold;")
        layout.addWidget(label)

        # List widget to display directories
        self.case_list_widget = QListWidget()
        self.case_list_widget.setStyleSheet("""
            QListWidget {
                background-color: #333;
                color: #f5f5f5;
                border: 1px solid #444;
                border-radius: 10px;
                padding: 10px;
            }
            QListWidget::item {
                padding: 8px;
                font-size: 12pt;
            }
            QListWidget::item:selected {
                background-color: #555;
                color: #f5f5f5;
            }
        """)
        layout.addWidget(self.case_list_widget)

        # Load cases from the 'Cases' directory
        self.load_cases()

        # Connect selection change signal to the method
        self.case_list_widget.itemClicked.connect(self.select_case)
        
        # 컨텍스트 메뉴 설정 (오른쪽 클릭 메뉴)
        self.case_list_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.case_list_widget.customContextMenuRequested.connect(self.show_case_context_menu)

        # Confirm button
        confirm_button = QPushButton("Select Case")
        confirm_button.setStyleSheet("""
            QPushButton {
                background-color: #444;
                color: white;
                border: 1px solid #777;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #555;
            }
        """)
        confirm_button.clicked.connect(self.confirm_selection)
        layout.addWidget(confirm_button)

        # Set up the central widget with dark mode styling
        container = QWidget()
        container.setStyleSheet("""
            QWidget {
                background-color: #2e2e2e;
                border: 2px solid #444;
                border-radius: 15px;
                padding: 20px;
            }
        """)
        container.setLayout(layout)
        self.setCentralWidget(container)

    def _cases_root(self):
        """스크립트 위치 기준 Cases 폴더 (getcwd와 무관하게 동일한 케이스 폴더 사용)."""
        try:
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), "Cases")
        except Exception:
            return os.path.join(os.getcwd(), "Cases")

    def load_cases(self):
        """Load directory names from the 'Cases' folder (스크립트 위치 기준)."""
        cases_path = self._cases_root()
        if os.path.exists(cases_path):
            case_dirs = [d for d in os.listdir(cases_path) if os.path.isdir(os.path.join(cases_path, d))]
            self.case_list_widget.addItems(case_dirs)
        else:
            print("No 'Cases' directory found in the current path.")

    def select_case(self, item):
        """Set selected case directory."""
        cases_path = self._cases_root()
        self.case_direc = os.path.join(cases_path, item.text())
        print(f"Selected case directory: {self.case_direc}")

    def _config_dir(self):
        """케이스 기준 config 폴더 경로 (저장/불러올 때 동일)."""
        if not self.case_direc:
            return None
        return os.path.join(self.case_direc, "config")

    def confirm_selection(self):
        """Confirm the selected case, prompt for dataset directory, and start training."""
        # 리스트에서 현재 선택된 항목으로 case_direc 동기화 (키보드 선택 시에도 반영)
        current = self.case_list_widget.currentItem()
        if current:
            self.case_direc = os.path.join(self._cases_root(), current.text())
        if not self.case_direc:
            copyable_message_box(self, QMessageBox.Warning, "선택 필요", "목록에서 케이스를 선택한 뒤 Select Case를 눌러주세요.")
            return
        print(f"Confirmed selection: {self.case_direc}")

        # 1차 모드 선택 (Create / Train / Detect)
        mode_selected = None
        create_mode_selected = None
        mode_items = ["Create (데이터 생성)", "Train (학습)", "Detect"]
        mode_selected, ok = QInputDialog.getItem(
            self,
            "모드 선택",
            "모드를 선택하세요 (데이터 생성/학습/예측):",
            mode_items,
            0,  # 기본 선택: Create
            False
        )
        if not ok:
            return

        # Create 모드일 경우 세부 선택: 새로 만들기 / 기존 데이터에 추가
        if mode_selected.startswith("Create"):
            create_mode_items = ["새로 만들기", "기존 데이터에 추가"]
            create_mode_selected, ok = QInputDialog.getItem(
                self,
                "Create 모드 선택",
                "새로 만들기 또는 기존 데이터에 추가를 선택하세요:",
                create_mode_items,
                0,  # 기본 선택: 새로 만들기
                False
            )
            if not ok:
                return

        # config/base_directory.xml 우선, 없으면 케이스 루트 (후보를 순서대로 시도)
        case_root = os.path.abspath(os.path.normpath(self.case_direc))
        xml_candidates = [
            os.path.join(case_root, "config", "base_directory.xml"),
            os.path.join(case_root, "base_directory.xml"),
        ]
        xml_path = None
        for cand in xml_candidates:
            if os.path.isfile(cand):
                xml_path = cand
                break
        dataset_direc = None

        if xml_path:
            try:
                tree = ET.parse(xml_path)
                root = tree.getroot()
                dataset_direc = root.findtext("dataset_directory")
                if dataset_direc:
                    print(f"Loaded dataset directory from XML: {xml_path}")
                else:
                    print("Dataset directory not found in XML (tag 내용 없음).")
            except ET.ParseError as e:
                print(f"Error parsing base_directory.xml: {e}")
            except Exception as e:
                print(f"base_directory.xml 읽기 오류: {e}")
        else:
            print(f"base_directory.xml not found. case_direc={case_root}")
            print(f"  tried: {xml_candidates}")

        if not dataset_direc:
            dataset_direc, ok = QInputDialog.getText(self, "Dataset Directory",
                                                     "데이터셋 디렉터리를 입력하세요 ex) Y://, Z://")
            if ok and dataset_direc:
                # Save the directory to config/base_directory.xml for future use
                config_d = self._config_dir()
                if config_d:
                    os.makedirs(config_d, exist_ok=True)
                xml_path = os.path.join(self._config_dir(), "base_directory.xml")
                root = ET.Element("settings")
                ET.SubElement(root, "dataset_directory").text = dataset_direc
                tree = ET.ElementTree(root)
                try:
                    with open(xml_path, "wb") as file:
                        tree.write(file, encoding="utf-8", xml_declaration=True)
                    print(f"Dataset directory saved to XML: {dataset_direc}")
                except IOError:
                    copyable_message_box(self, QMessageBox.Warning, "Error", "Failed to save dataset directory to XML.")
            else:
                return  # User canceled the input dialog

        self.dataset_direc = dataset_direc
        print(f"Dataset directory set to: {self.dataset_direc}")

        # Create and show the CreateTrain window (skip 초기 CSV 로드 여부 전달)
        skip_flag = False
        if mode_selected.startswith("Create") and create_mode_selected == "새로 만들기":
            skip_flag = True
        self.create_train_window = createtrainclass(self.case_direc, self.dataset_direc, skip_initial_csv_load=skip_flag)
        # 초기 모드 선택 적용
        try:
            self.apply_initial_mode_selection(mode_selected, create_mode_selected)
        except Exception as e:
            print(f"[WARN] 초기 모드 적용 중 오류 (계속 진행): {e}")
        self.create_train_window.show()
        self.close()

    def apply_initial_mode_selection(self, mode_selected, create_mode_selected=None):
        """케이스 선택 직후 초기 모드(생성/기타) 설정을 적용"""
        if not hasattr(self, "create_train_window") or self.create_train_window is None:
            return

        # 기록용
        try:
            self.create_train_window.initial_mode_selected = mode_selected
            if create_mode_selected:
                self.create_train_window.create_mode_selected = create_mode_selected
        except Exception:
            pass

        if mode_selected.startswith("Create"):
            # 데이터 생성 모드 (CSV 생성)
            self.create_train_window.choice = 2
            self.create_train_window.detectmode = 0
            self.create_train_window.existval = 0
            # Create 새로 만들기 시 초기 CSV 자동 로드 방지 플래그
            self.create_train_window.skip_initial_csv_load = True
            # 새로 만들기 / 기존 데이터 추가 분기
            if create_mode_selected == "새로 만들기":
                # 신규 CSV 생성: csv_file 비우고 최신 규칙으로 경로 준비
                self.create_train_window.csv_file = ''
                try:
                    self.create_train_window.ensure_new_csv_path()
                except Exception as e:
                    print(f"[WARN] 신규 CSV 경로 설정 실패(무시): {e}")
                print("[INFO] 초기 모드: Create > 새로 만들기")
            elif create_mode_selected == "기존 데이터에 추가":
                # 기존 CSV 선택 후 추가 (이미 선택된 CSV가 있으면 그대로 사용)
                csv_file_path = (
                    getattr(self.create_train_window, "csv_file", "") 
                    or getattr(self.create_train_window, "csv_path", "") 
                    or getattr(self, "csv_path", "")
                )
                if not csv_file_path:
                    csv_file_path, _ = QFileDialog.getOpenFileName(
                        self,
                        "기존 CSV 파일 선택",
                        self.case_direc,
                        "CSV Files (*.csv);;All Files (*)"
                    )
                    if not csv_file_path:
                        # 선택 취소 시 생성 모드 초기화
                        try:
                            self.create_train_window.show_alert("CSV 파일이 선택되지 않았습니다.")
                        except Exception:
                            pass
                        self.create_train_window.choice = 0
                        return

                self.create_train_window.csv_file = csv_file_path
                self.create_train_window.csv_path = csv_file_path
                self.create_train_window.existval = 1
                self.create_train_window.skip_initial_csv_load = False

                # 매핑 JSON 로드 (피처는 체크·states.json만 사용)
                try:
                    mapping_json_path = self.create_train_window.find_mapping_json(csv_file_path)
                    self.create_train_window.set_mapping_json_path(mapping_json_path)
                except Exception as e:
                    print(f"[WARN] 매핑 JSON 설정 실패(무시): {e}")
                # UI 표시 업데이트
                try:
                    self.create_train_window.open_csv2(csv_file_path, self.create_train_window.tableWidget)
                    self.create_train_window.current_csv_path = csv_file_path
                except Exception as e:
                    print(f"[WARN] CSV 표시 실패(무시): {e}")
                try:
                    if hasattr(self.create_train_window, 'csvlabel'):
                        self.create_train_window.csvlabel.setText(os.path.basename(csv_file_path))
                except Exception:
                    pass
                try:
                    self.create_train_window.load_or_initialize_states()
                except Exception:
                    pass
                print("[INFO] 초기 모드: Create > 기존 데이터에 추가")

        elif mode_selected.startswith("Train"):
            # 학습 모드: 기존 CSV를 선택해 학습 준비
            self.create_train_window.choice = 0  # 생성 모드 해제
            self.create_train_window.detectmode = 0
            self.create_train_window.existval = 1
            self.create_train_window.skip_initial_csv_load = False

            csv_file_path, _ = QFileDialog.getOpenFileName(
                self,
                "학습용 CSV 파일 선택",
                self.case_direc,
                "CSV Files (*.csv);;All Files (*)"
            )
            if not csv_file_path:
                try:
                    self.create_train_window.show_alert("CSV 파일이 선택되지 않았습니다.")
                except Exception:
                    pass
                self.create_train_window.choice = 0
                return

            self.create_train_window.csv_file = csv_file_path
            self.create_train_window.csv_path = csv_file_path

            # 매핑 JSON 로드 (피처는 체크·states.json만 사용)
            try:
                mapping_json_path = self.create_train_window.find_mapping_json(csv_file_path)
                self.create_train_window.set_mapping_json_path(mapping_json_path)
            except Exception as e:
                print(f"[WARN] 매핑 JSON 설정 실패(무시): {e}")
            # UI 표시 업데이트
            try:
                self.create_train_window.open_csv2(csv_file_path, self.create_train_window.tableWidget)
                self.create_train_window.current_csv_path = csv_file_path
            except Exception as e:
                print(f"[WARN] CSV 표시 실패(무시): {e}")
            try:
                if hasattr(self.create_train_window, 'csvlabel'):
                    self.create_train_window.csvlabel.setText(os.path.basename(csv_file_path))
            except Exception:
                pass
            try:
                self.create_train_window.load_or_initialize_states()
            except Exception:
                pass
            print("[INFO] 초기 모드: Train (학습)")

        else:
            # Detect 모드: 케이스 변경과 동일하게 상세 폴더/CSV/모델을 확정
            self.create_train_window.choice = 0
            self.create_train_window.detectmode = 1
            self.create_train_window.skip_initial_csv_load = False
            try:
                self.create_train_window._ensure_detect_all_models_option()
                detail_choice = self.create_train_window._choose_detail_folder_for_case(
                    self.create_train_window.case_direc
                )
                self.create_train_window._apply_detect_detail_choice(detail_choice)
                model_auto_selected = False
                if not getattr(self.create_train_window, "detect_all_detail_folders", False):
                    model_auto_selected = self.create_train_window._auto_select_detect_model_for_current_csv(
                        show_message=False
                    )
                detail_text = ""
                if getattr(self.create_train_window, "detect_all_detail_folders", False):
                    detail_text = f" / {self.create_train_window._detect_all_models_label()}"
                elif detail_choice and detail_choice is not _DETECT_DETAIL_ALL:
                    detail_text = f" / {os.path.relpath(detail_choice, self.create_train_window.case_direc)}"
                if getattr(self.create_train_window, "detect_all_detail_folders", False):
                    model_text = " / Detect: 전체 상세폴더 모드"
                else:
                    model_text = " / Detect 모델 자동 선택됨" if model_auto_selected else " / Detect 모델 자동 선택 실패"
                try:
                    self.create_train_window.statusBar().showMessage(
                        f"Detect 준비 완료: {self.create_train_window.case_direc}{detail_text}{model_text}",
                        8000
                    )
                except Exception:
                    pass
            except Exception as e:
                print(f"[WARN] 초기 Detect 상세 폴더/모델 자동 설정 실패: {e}")
            print("[INFO] 초기 모드: Detect (예측)")

    def create_new_case(self):
        """Prompt for a new case name and create the case directory."""
        case_name, ok = QInputDialog.getText(self, "New Case", "Enter the name of the new case:")
        if ok and case_name:
            # Create the new case directory (스크립트 위치 기준)
            cases_path = self._cases_root()
            new_case_path = os.path.join(cases_path, case_name)

            os.makedirs(cases_path, exist_ok=True)

            # Check if the case already exists to avoid duplicates
            if not os.path.exists(new_case_path):
                os.mkdir(new_case_path)
                print(f"Created new case directory: {new_case_path}")

                # Refresh the case list and select the new case
                self.case_list_widget.addItem(case_name)
                self.case_direc = new_case_path
                # 새 케이스 생성 시에는 초기 CSV 자동 로드 건너뜀
                self.create_train_window = createtrainclass(self.case_direc, self.dataset_direc, skip_initial_csv_load=True)
                self.create_train_window.show()
            else:
                print(f"Case '{case_name}' already exists.")

    def set_dataset_directory(self):
        """데이터셋 경로를 설정하는 함수"""
        # 현재 선택된 케이스가 있는지 확인
        if not self.case_direc:
            # 케이스가 선택되지 않았으면 리스트에서 선택된 항목 확인
            selected_items = self.case_list_widget.selectedItems()
            if selected_items:
                cases_path = self._cases_root()
                self.case_direc = os.path.join(cases_path, selected_items[0].text())
            else:
                copyable_message_box(self, QMessageBox.Warning, "경고", "먼저 케이스를 선택해주세요.")
                return
        
        # 폴더 선택 다이얼로그 사용
        dataset_direc = QFileDialog.getExistingDirectory(
            self, 
            "데이터셋 디렉터리 선택",
            self.dataset_direc if self.dataset_direc else os.getcwd()
        )
        
        if dataset_direc:
            # config/base_directory.xml에 저장
            config_d = self._config_dir()
            if config_d:
                os.makedirs(config_d, exist_ok=True)
            xml_path = os.path.join(self._config_dir(), "base_directory.xml")
            root = ET.Element("settings")
            ET.SubElement(root, "dataset_directory").text = dataset_direc
            tree = ET.ElementTree(root)
            try:
                with open(xml_path, "wb") as file:
                    tree.write(file, encoding="utf-8", xml_declaration=True)
                self.dataset_direc = dataset_direc
                copyable_message_box(self, QMessageBox.Information, "성공", f"데이터셋 경로가 설정되었습니다:\n{dataset_direc}")
                print(f"Dataset directory saved to XML: {dataset_direc}")
            except IOError as e:
                self._copyable_msg(QMessageBox.Warning, "오류", f"데이터셋 경로 저장 실패: {str(e)}")

    def show_case_context_menu(self, position):
        """케이스 리스트에서 오른쪽 클릭 시 컨텍스트 메뉴 표시"""
        item = self.case_list_widget.itemAt(position)
        if item is None:
            return
        
        # 컨텍스트 메뉴 생성
        context_menu = QMenu(self)
        
        # Windows Explorer에서 열기
        open_explorer_action = QAction("Windows Explorer에서 열기", self)
        open_explorer_action.triggered.connect(lambda: self.open_case_in_explorer(item))
        context_menu.addAction(open_explorer_action)
        
        # 구분선 추가
        context_menu.addSeparator()
        
        # 삭제
        delete_action = QAction("삭제", self)
        delete_action.triggered.connect(lambda: self.delete_case(item))
        context_menu.addAction(delete_action)
        
        # 메뉴 표시
        context_menu.exec_(self.case_list_widget.mapToGlobal(position))

    def open_case_in_explorer(self, item):
        """Windows Explorer에서 케이스 폴더 열기"""
        case_name = item.text()
        cases_path = self._cases_root()
        case_path = os.path.join(cases_path, case_name)
        
        if os.path.exists(case_path):
            try:
                # Windows에서 폴더를 탐색기로 열기
                os.startfile(case_path)
                print(f"Windows Explorer에서 케이스 폴더 열기: {case_path}")
            except Exception as e:
                copyable_message_box(self, QMessageBox.Warning, "오류", f"폴더를 열 수 없습니다:\n{str(e)}")
                print(f"폴더 열기 오류: {e}")
        else:
            copyable_message_box(self, QMessageBox.Warning, "오류", "케이스 폴더를 찾을 수 없습니다.")

    def delete_case(self, item):
        """선택된 케이스 폴더를 삭제하는 함수"""
        case_name = item.text()
        
        # 확인 메시지
        reply = QMessageBox.question(
            self, 
            "케이스 삭제 확인",
            f"케이스 '{case_name}'을(를) 삭제하시겠습니까?\n\n이 작업은 되돌릴 수 없습니다.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            cases_path = self._cases_root()
            case_path = os.path.join(cases_path, case_name)
            
            try:
                # 폴더와 모든 내용 삭제
                if os.path.exists(case_path):
                    shutil.rmtree(case_path)
                    print(f"케이스 폴더 삭제 완료: {case_path}")
                    
                    # 리스트에서도 제거
                    row = self.case_list_widget.row(item)
                    self.case_list_widget.takeItem(row)
                    
                    # 현재 선택된 케이스가 삭제된 케이스면 초기화
                    if self.case_direc == case_path:
                        self.case_direc = None
                    
                    copyable_message_box(self, QMessageBox.Information, "완료", f"케이스 '{case_name}'이(가) 삭제되었습니다.")
                else:
                    copyable_message_box(self, QMessageBox.Warning, "오류", "케이스 폴더를 찾을 수 없습니다.")
            except Exception as e:
                copyable_message_box(self, QMessageBox.Critical, "오류", f"케이스 삭제 중 오류가 발생했습니다:\n{str(e)}")
                print(f"케이스 삭제 오류: {e}")



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = CaseSelectorApp()
    ex.show()
    app.exec_()
