import csv
import hashlib
import json
import math
import os
import pickle
import re
import struct
import sys
import glob
import xml.etree.ElementTree as ET
from datetime import datetime
from tkinter import simpledialog, messagebox
from frame_compression import process_videos_in_folder
import joblib
import numpy as np
import pandas as pd
from PyQt5.QtCore import QDir, Qt
from PyQt5.QtWidgets import QApplication, QWidget, QFileSystemModel, QMainWindow, QProgressBar, QDialog, QLabel, \
    QVBoxLayout, QTableWidgetItem, QMessageBox, QLineEdit, QPushButton, QTableWidget, QInputDialog, QFileDialog, \
    QListWidget, QAction
from PyQt5 import uic, QtWidgets
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from simhash import Simhash

from Train_GRUprocess import twoTrainClass
from clustering1 import trainClustering
from Train_GRUprocess_multi import TrainClass
from extractframe_single import extractGOP
from extract_sps import parse_sps
from pps import analyzesps

os.environ["CUDA_VISIBLE_DEVICES"]="0"

class ProgressWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("파일 처리 진행 상황")

        self.progress_bar = QProgressBar(self)
        self.progress_bar.setGeometry(100, 50, 300, 20)

        self.label = QLabel("파일 처리 중...", self)

        layout = QVBoxLayout()
        layout.addWidget(self.label)
        layout.addWidget(self.progress_bar)
        self.setLayout(layout)

    def set_progress(self, value):
        self.progress_bar.setValue(value)

    def set_label_text(self, text):
        self.label.setText(text)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

form_path = resource_path("new.ui")
form_class = uic.loadUiType(form_path)[0]

class createtrainclass(QMainWindow, form_class):
    def __init__(self, case_direc, dataset_direc):
        super(createtrainclass, self).__init__()
        self.choice = 0
        self.file_paths = []


        self.setupUi(self) # UI 요소 초기화
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.clustering = trainClustering()
        self.trainclass = twoTrainClass()
        self.existval = 0
        # 확장자 필터
        self.extension_list = ["확장자", ".mp4",  ".mov"]
        self.comboBox.addItems(self.extension_list)
        self.csv_file = ''
        self.tempcsv_file = ''
        self.comboBox.currentIndexChanged.connect(
            lambda index: self.filter_files_by_extension(self.comboBox.itemText(index)))

        self.dirModel = QFileSystemModel()
        self.dirModel.setRootPath(QDir.rootPath())

        self.exist_csv_but.clicked.connect(self.open_existcsv)
        self.treeView.setModel(self.dirModel)

        initialcode = 0
        self.detectmode = 0

        if initialcode == 0:
            print("=====================================")
            try:
                self.case_direc = case_direc
                self.dataset_direc = dataset_direc
                # 케이스 이름이 공백 또는 비어 있는 경우
                ''' if not self.case_direc or self.case_direc.strip() == "":
                    print("기본 default_case를 생성합니다.")
                    self.case_direc = 'default_case'
                    raise ValueError("케이스명이 공백이라 default_case로 설정")
                
                else:
                    # 공백이 아닌 경우
                    if os.path.exists(self.case_direc):
                        print(f"{self.case_direc} 케이스가 이미 존재합니다.")
                    else:
                        print(f"{self.case_direc} 케이스가 존재하지 않아 생성합니다.")
                        os.makedirs(self.case_direc) # 경로 생성
                        raise ValueError("경로가 존재하지 않아 생성")'''

                # 데이터셋 경로 유효성 검사 (존재하는 경로인지 확인)
                if not os.path.isdir(self.dataset_direc):
                    raise ValueError(f"유효하지 않은 데이터셋 경로: {self.dataset_direc}")

                print("유효한 데이터셋 경로와 케이스 경로가 설정되었습니다.")

            except Exception as e:
                # 경로가 유효하지 않으면 기본 경로를 설정
                self.dataset_direc = 'y:\\'
                print("데이터셋 경로가 Y:\\로 설정되었습니다.")

            print("=====================================")
            try:
                print("케이스 이름: [", self.case_direc, "]")
                print("데이터셋 경로: [", self.dataset_direc, "]")
                print("=====================================")
            except:
                self.ask_input()
            initialcode = 1

        self.default_states = {
            "structure_seq_state": 0,
            "frame_gop_state": 0,
            "frame_ratio_state": 0,
            "frame_sps_state": 0,
            "structure_val_state": 0
        }
        self.treeView.setRootIndex(self.dirModel.index(self.dataset_direc))
        self.treeView.clicked.connect(self.file_selected)
        self.load_or_initialize_states()
        # 케이스 디렉토리에서 .csv 찾아서 csv_files 리스트로 반환
        all_files = glob.glob(os.path.join(self.case_direc, "*.csv"))
        print("All CSV files:", all_files)
        csv_files = [file for file in all_files if 'feature_importance.csv' not in os.path.basename(file)]
        self.csv_path = ''

        # .csv 파일이 하나 이상 있을 때 일단은 첫 번째 파일을 열기
        if csv_files:
            for csv_file in csv_files:
               if '_train' in csv_file or '241' in csv_file:
                    self.csv_path = csv_file  # 첫 번째 CSV 파일 경로 선택
                    self.open_csv2(self.csv_path, self.tableWidget)

        try :
            file_name = os.path.basename(self.csv_path)
            self.csvlabel.setText(file_name)
        except Exception as e:
            pass
        # 헤더 설정
        header = self.treeView.header()
        header.setSectionResizeMode(0, header.Interactive)
        header.resizeSection(0, 400)
        self.create_value2.clicked.connect(lambda: setattr(self, 'choice', 2))

        self.structure_val_state = False
        self.structure_seq_state = False
        self.frame_sps_state = False
        self.frame_gop_state = False
        self.frame_ratio_state = False

        self.tabWidget.setCurrentIndex(0)

        self.structure_val_but.stateChanged.connect(self.on_structure_val_changed)
        self.structure_seq_but.stateChanged.connect(self.on_structure_seq_changed)
        self.frame_sps_but.stateChanged.connect(self.on_frame_sps_changed)
        self.frame_gop_but.stateChanged.connect(self.on_frame_gop_changed)
        self.frame_ratio_but.stateChanged.connect(self.on_frame_ratio_changed)
        self.stateButton.clicked.connect(self.save_states)

        self.LoadButton.clicked.connect(self.main) # Load 버튼 클릭 시 self.main() 호출
        self.cluster_train.clicked.connect(self.clustermain)
        self.class_train.clicked.connect(self.classmain)

        self.class_detect.clicked.connect(self.load_file_for_prediction)
        # 파일 목록에서 아이템을 더블 클릭할 때 호출되는 슬롯을 연결합니다.
        self.listWidget.itemDoubleClicked.connect(self.remove_selected_file)
        self.list_del.clicked.connect(self.remove_all_file)

        self.label_info.clicked.connect(self.open_data_entry_window)
        if self.binButton.isChecked():
            self.label_datacsv = 'labeldata_bin.csv'
        elif self.mulButton.isChecked():
            self.label_datacsv = 'labeldata_mul.csv'
        self.labelinfofile = ""

        try :
            pass
            self.load_excel_data()
        except:
            pass
        self.label_input_but.clicked.connect(self.input_label)
        self.model_combo.currentTextChanged.connect(self.on_modelcombobox_select)
        self.trainindex = self.comboBox.currentIndex()



    # 상태값을 불러오거나 초기화하는 함수
    def load_or_initialize_states(self):
        # 파일이 있으면 불러오기, 없으면 초기화
        statepath = self.resource_path(os.path.join(self.case_direc, "states.json"))
        if os.path.exists(statepath):
            with open(statepath, "r") as file:
                self.states = json.load(file)
                print("기존 상태값을 불러왔습니다:", self.states)
        else:
            self.states = self.default_states.copy()

            print("파일이 없어 기본 상태값으로 초기화하고 저장했습니다:", self.states)

        try:
            self.structure_seq_state = self.states["structure_seq_state"]
            self.frame_gop_state = self.states["frame_gop_state"]
            self.frame_ratio_state = self.states["frame_ratio_state"]
            self.frame_sps_state = self.states["frame_sps_state"]
            self.structure_val_state = self.states["structure_val_state"]
        except Exception as e:
            pass


    def set_state(self, state_name, value):
        self.states[state_name] = value

    def save_states(self):
        statepath = self.resource_path(os.path.join(self.case_direc, "states.json"))
        with open(statepath, "w") as file:
            json.dump(self.states, file)
        print("상태값이 JSON 파일에 저장되었습니다.")

    def on_combobox_select(self, index):
        self.trainclass.index = index
    def on_modelcombobox_select(self):
        self.aimodel = self.model_combo.currentText()

    def clustermain(self):

        self.clustering.gotrain(self.csv_path)

    def showFileDialog(self):
        # 파일 다이얼로그를 띄워서 파일 선택
        self.csv_path, _ = QFileDialog.getOpenFileName(self, '파일 선택', '',
                                                   '모든 파일 (*);;텍스트 파일 (*.csv)')

        # 선택한 파일 경로를 라벨에 표시
        if self.csv_path:
            print(f'선택된 파일 경로: {self.csv_path}')
        else:
            print('파일이 선택되지 않았습니다.')

    def classmain(self):
        binstat = self.binButton_3.isChecked()
        mulstat = self.mulButton_3.isChecked()
        states = self.load_or_initialize_states()
        if binstat:
            self.trainclass = twoTrainClass() ##### 이진으로 설정
            self.classmode = 'bin_'
        elif mulstat:
            self.trainclass = TrainClass() ##### 다중으로 설정
            self.classmode = 'mul_'
        else :
            messagebox.showerror("에러", "바이너리/멀티 모드를 선택")
        self.trainindex = self.model_combo.currentIndex()
        self.aimodel = 'Xgboost'
        self.model_combo.activated.connect(self.on_combobox_select)
        self.aimodel = self.model_combo.currentText()
        self.trainclass.csv_path = self.csv_path # 객체 csv 경로 설정
        self.trainclass.comboBox = self.model_combo_2
        try:
            self.trainclass.gotrain(self.classmode, self.aimodel, self.trainindex, self.csv_path)
        except Exception as e:
            pass


    def classdetect(self):
        self.detectclass.predict(file_path=self.file_paths[0])

    def filter_files_by_extension(self, extension):  # 선택된 확장자에 따라 필터링
        if extension and extension != "확장자":  # Ensure it's not the placeholder text
            self.extension = extension  # Directly assign the selected extension
            self.dirModel.setNameFilters([f"*{extension}"])
            self.dirModel.setNameFilterDisables(False)
        else:
            self.dirModel.setNameFilters([])

    def file_selected(self, index): # 파일 또는 디렉토리 선택 시 호출
        file_info = self.dirModel.fileInfo(index)
        try:
            if file_info.isDir():  # If a directory is selected
                if self.extension :
                    self.select_all_files_in_directory(file_info.absoluteFilePath())
            else:
                file_path = file_info.absoluteFilePath()
                extension = os.path.splitext(file_path)[1]
                self.filter_files_by_extension(extension)
                if file_path not in self.file_paths:
                    if extension.lower() == self.extension.lower():
                        self.listWidget.addItem(file_path)
                        self.file_paths.append(file_path)
                if extension == '.csv':
                    self.csv_path = file_path

        except Exception as e:
            self.show_alert(str(e))

    def ask_input(self):
            try:
                self.dataset_direc = input("데이터셋 경로를 입력하세요: ")
            except:
                pass


    def display_dataframe(self, df, widgettype):
        try:
            df = self.move_label_to_second_column(df)
        except:
            pass
        widgettype.setRowCount(df.shape[0])
        widgettype.setColumnCount(df.shape[1])
        widgettype.setHorizontalHeaderLabels(df.columns)

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                if i<100:
                    item = QTableWidgetItem(str(df.iat[i, j]))
                    widgettype.setItem(i, j, item)
                else :
                    return


    def open_csv2(self, csvfile, widgett):
        file_name = csvfile
        if file_name:
            try:
                sample_df = pd.read_csv(file_name, nrows=1, header=None)
                tempvalue = sample_df.iloc[0, 0]
                # 첫 번째 행의 첫 번째 값이 'name'이 아닌 경우 두 번째 행을 헤더로 설정
                if tempvalue != 'name':
                    # 첫 번째 행에 컬럼 이름이 없으면 두 번째 행을 헤더로 설정하여 다시 읽어옵니다
                    df = pd.read_csv(file_name, header=1)
                else:
                    # 첫 번째 행이 컬럼 이름이면 기본적으로 읽어옵니다
                    df = pd.read_csv(file_name)
                self.display_dataframe(df, widgettype=widgett)
            except Exception as e:
                self.show_alert("CSV 파일을 읽는 중 오류가 발생했습니다: " + str(e))

        return


    def select_all_files_in_directory(self, directory_path):
        try:
            for root, _, files in os.walk(directory_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    extension = os.path.splitext(file_path)[1][0:]

                    #같은 확장자만 담아야한다면 조건추가
                    if extension.lower() == self.extension:
                        if file_path not in self.file_paths:
                            self.listWidget.addItem(file_path)
                            self.file_paths.append(file_path)
        except Exception as e:
            self.show_alert(str(e))


    def remove_selected_file(self, item): # 선택한 파일을 목록에서 제거
        # 더블 클릭한 파일 아이템을 목록에서 제거합니다.
        file_path = item.text()
        self.listWidget.takeItem(self.listWidget.row(item))
        self.file_paths.remove(file_path)

    def remove_all_file(self):
        self.listWidget.clear()
        self.file_paths = []

    def lcs(self, X, Y):
        m = len(X)
        n = len(Y)

        # memoization table 초기화
        L = [[0] * (n + 1) for _ in range(m + 1)]

        # X[0..m-1]와 Y[0..n-1] LCS 계산
        for i in range(m + 1):
            for j in range(n + 1):
                if i == 0 or j == 0:
                    L[i][j] = 0
                elif X[i - 1] == Y[j - 1]:
                    L[i][j] = L[i - 1][j - 1] + 1
                else:
                    L[i][j] = max(L[i - 1][j], L[i][j - 1])

        # LCS
        index = L[m][n]
        lcs_list = [None] * index

        i = m
        j = n
        while i > 0 and j > 0:
            if X[i - 1] == Y[j - 1]:
                lcs_list[index - 1] = X[i - 1]
                i -= 1
                j -= 1
                index -= 1
            elif L[i - 1][j] > L[i][j - 1]:
                i -= 1
            else:
                j -= 1

        return lcs_list

    def lcs_multiple_lists(self, lists): # 가장 긴 공통 서브시퀀스 찾기
        if len(lists) < 2: # 리스트가 2개 이상 있어야 LCS 찾기 가능
            raise ValueError("At least two lists are required")

        current_lcs = self.lcs(lists[0], lists[1]) # 처음 두 리스트의 LCS 계산

        for lst in lists[2:]: # 나머지 리스트와의 LCS 반복 계산
            current_lcs = self.lcs(current_lcs, lst)
            if not current_lcs:
                return []

        return current_lcs

    def process_files(self):
        progress_window = ProgressWindow()
        progress_window.setModal(True)
        progress_window.show()

        total_files = len(self.file_paths)

        for i, fname in enumerate(self.file_paths):
            full_path = fname
            if os.path.isfile(full_path):
                self.extract_value(full_path)
                self.all_result.append(self.reres)

            # 진행 상황 업데이트
            progress_percentage = (i + 1) / total_files * 100
            self.progress_bar.setValue(progress_percentage)
            QApplication.processEvents()

        progress_window.set_label_text("작업 완료")
        progress_window.exec_()

    def get_files_value(self): # self.file_paths의 모든 파일 처리하고 self.all_result 저장 및 반환
        self.all_result = []
        self.count = -1
        # 폴더 내 모든 파일에 대해 수행
        for i, fname in enumerate(self.file_paths):
            self.count += 1
            full_path = fname

            if os.path.isfile(full_path):
                self.extract_value(full_path)
                self.all_result.append(self.reres)

            # 진행 상황 업데이트
            progress_percentage = (i + 1) / len(self.file_paths) * 100
            self.progress_bar.setValue(progress_percentage)
            QApplication.processEvents()

        # 파일 처리가 완료되면 최장 리스트를 찾음
        longest_list = max(self.all_result, key=len)

        # 모든 리스트를 가장 긴 리스트와 비교하며 첫 번째 요소가 없으면 추가하고 두 번째 요소는 0으로 설정
        for lst in self.all_result:
            for i in range(len(longest_list)):
                if len(lst) <= i or lst[i][0] != longest_list[i][0]:
                    # 첫 번째 요소가 없으면 추가하고 두 번째 요소를 0으로 설정
                    if len(lst) <= i:
                        lst.insert(i, [longest_list[i][0], '0'])
                    else:
                        lst[i] = [longest_list[i][0], '0']

        return self.all_result

    def merge_lists2(self, ngram): # LCS 받아서 연속되거나 유사한 n-gram 병합하여 하나의 긴 패턴으로 생성
        count, count2, onecount = 0,0,0
        new_list = []
        merged_list=[]
        one_merged_list = []
        previous_gram = ''
        for onegram1 in ngram :

            if previous_gram == onegram1 and onegram1 != '00000000': # 00000000 이면 제외
                one_merged_list.append(onegram1)
                onecount += 1
                pass

            else : # 현재 n-gram과 이전 n-gram 마지막 부분이 일치하는지 확인
                if count == 0:
                    previous_gram = onegram1
                    count += 1
                else :
                    lengh = (len(onegram1) - 1)
                    previous_gram_2=previous_gram[-lengh:]
                    onegram_2 = onegram1[:-1]

                    if previous_gram_2 == onegram_2 and onegram1 != '00000000':
                        previous_gram = previous_gram+onegram1[-1]
                        #count2 += 1
                    else :
                        previous_gram_2 = ''
                        if count2 != 0 and onegram1 != '00000000':
                            one_merged_list.append(previous_gram)
                            count = 0
                        else :
                            if previous_gram != '00000000':
                                one_merged_list.append(previous_gram)

                        previous_gram = onegram1
                if onecount == len(ngram)-1: # 리스트의 마지막 n-gram 처리
                    if previous_gram != '00000000':
                        one_merged_list.append(previous_gram)
                    break

                onecount += 1

        merged_list.append(one_merged_list)
        return merged_list

    def add_numbers_to_duplicates(self, input_list):
        counts = {}  # 요소별로 카운트를 저장할 딕셔너리

        for i in range(len(input_list[0])):
            item = input_list[0][i]

            # 이미 등장한 요소인 경우
            if item in counts:
                counts[item] += 1
                input_list[0][i] = f"{item}_{counts[item]}"
            else:
                counts[item] = 0

        return input_list

    def extract_value(self, fpath): # 파일에서 n-gram 추출하고 병합된 n-gram 리스트와 비교하여 일치하는 패턴 찾아 res 리스트에 저장
        file_type = ""
        res = []
        self.reres = []

        # 파일 경로에서 파일 이름 추출
        file_name = os.path.basename(fpath)

        for file_info in self.ngrams_list:
            _, ngrams = file_info
            _ = os.path.basename(_)
            if file_name == _:  # 파일 이름과 fpath의 파일 이름을 비교
                res.append(('name', os.path.basename(fpath)))

                with open(fpath, 'rb') as fp:
                    check_opcode = self.mergelist
                    check_opcode = self.add_numbers_to_duplicates(check_opcode)

                    # 데이터 추출
                    count = 0
                    tempvalue = ''
                    mvalue = 0
                    for i in range(len(check_opcode[0]) + 1):
                        for j in range(mvalue, len(ngrams)):
                            try:
                                nowvalue = ngrams[j]

                                headerfeat = check_opcode[0][i]

                                if '_' in headerfeat :
                                    headerfeatemp = headerfeat[:-2]
                                else :
                                    headerfeatemp = headerfeat
                                k = 8
                                m = j
                                if nowvalue in headerfeatemp:

                                    while len(nowvalue) < len(headerfeatemp) and (j + k) < len(ngrams):
                                        temppp = ngrams[m + k]
                                        nowvalue += temppp[0]
                                        k += 1

                                if nowvalue == headerfeatemp:
                                    count += 1
                                elif count != 0:
                                    lennowvalue = len(nowvalue)
                                    testvalue = j + lennowvalue
                                    for kn in range(testvalue, testvalue + (lennowvalue * 2), 8):
                                        temppppp = ngrams[kn]  # 수정된 부분
                                        tempvalue += temppppp

                                    res.append((headerfeat, tempvalue))
                                    tempvalue = ''
                                    count = 0
                                    mvalue = j + 1
                                    break

                                if j == len(ngrams) - 1 and tempvalue == '':
                                    res.append((headerfeat, '0'))

                                    break

                            except Exception as e:
                                pass

        self.reres = res
        return res

    def extract_rengram(self, result): # self.ngrams_list와 result 간 교집합을 계산해 각 파일별로 공통된 n-gram들을 찾아냄
        self.intersection_lists = []

        result_set = set(result)
        for name, ngram in self.ngrams_list:
            intersection_list = [onegram for onegram in ngram if onegram in result_set]
            self.intersection_lists.append(intersection_list)

        return self.intersection_lists

    def find_duplicates_count(self): # ngrams_list에서 공통적으로 출현하는 요소 찾기

        self.data_list = []
        self.newlist = []
        duplicates = []

        element_count = {} # n-gram 요소 출현 횟수 저장 딕셔너리, 키: 요소, 값: 출현 횟수

        # 모든 리스트에서 요소의 출현 횟수를 카운트
        for k in range(len(self.ngrams_list)):
            for lst in self.ngrams_list[k][1]:
                    if lst in element_count:
                        element_count[lst] += 1
                    else:
                        element_count[lst] = 1

        # 출현 횟수가 2번 이상인 요소만 self.newlist에 저장
        basenum = int(len(self.ngrams_list)*0.7)
        self.newlist = [key for key, value in element_count.items() if value >= basenum]

        #중복이 없는 교집합 리스트를 commonlist.pkl에 저장
        commonlistpkl = str(self.extension + '\\' + "commonlist.pkl")
        with open(commonlistpkl, "wb") as fw: #
            pickle.dump(self.newlist, fw)

    def add_string_if_not_exists(self, filename, target_string):
        try:
            with open(filename, 'r', encoding='utf-8') as file:
                content = file.read()

            # 없으면 추가
            if target_string not in content:
                with open(filename, 'a', encoding='utf-8') as file:
                    file.write(target_string)

        except Exception as e :
            with open(filename, 'a', encoding='utf-8') as file:
                file.write(target_string)

    def feature_dictionary(self, hexa):

        array10 = []

        # 엑셀 파일에서 데이터 읽어오기 (엑셀 파일 경로를 설정해 주세요)
        excel_file = str(self.extension + '\\' + '_dict.xlsx')  # 엑셀 파일 경로
        self.resource_path(excel_file)
        df = pd.read_excel(excel_file)  # 엑셀 파일 읽기

        # 엑셀 데이터를 딕셔너리로 변환 (엑셀 파일의 첫 번째 열을 key로, 두 번째 열을 value로)
        newdict = dict(zip(df.iloc[:, 0], df.iloc[:, 1]))

        result = hexa

        # newdict의 value가 result[1]에 있으면 key를 array10에 추가
        for key, value in newdict.items():
            if value in result[1]:
                array10.append(str(key))

        # array10을 ","로 구분된 문자열로 만들어서 simhash 계산
        sequencedem = ", ".join(array10)
        sequencedem = self.simhash(sequencedem)

        # sequencedem과 hexa[0]을 함께 저장
        self.sequencedem.append((hexa[0], sequencedem))

    def save_lists_of_10_to_csv(self, data_list, file_name): # 더 긴 패턴(mergelist 리스트) 주어진 데이터를 CSV로 저장
        with open(file_name, 'w', newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)

            row = [j for j in range(1, len(data_list[0])+1)] # 헤더 작성
            csv_writer.writerow(row)

            for row in data_list:
                csv_writer.writerow(row)

    # value로 key찾기
    def find_key_by_value(self, dictionary, value):
        for key, val in dictionary.items():
            if val == value:
                return key
        return None  # 해당 값과 일치하는 키가 없을 경우 None을 반환

    def make_features(self, input_str):
        length = 3
        input_str = input_str.lower()
        out_str = re.sub(r'[^\w]+', '', input_str)
        return [out_str[i:i + length] for i in range(max(len(out_str) - length + 1, 1))]
    def simhash(self, input_str):
        features = self.make_features(input_str)
        return Simhash(features).value
    def headersimhash(self, input_list):
        string_result = ''.join(map(str, input_list))
        self.simhash(string_result)
    def makearray(self, featurelist, newdict):
        newlist2 = []
        newlist = []
        # 피쳐를 딕셔너리 사전의 10진수값에 매핑
        for item in featurelist:
            a = newdict.values()
            if item in newdict.values():
                newlist.append(self.find_key_by_value(newdict, item))

        newlist2 =[self.file_paths]
        newlist2.append(newlist)

        pklname = str(self.extension + '\\' + "vectordb.pkl")
        with open(pklname, "wb") as fw:
            pickle.dump(newlist2, fw)

    # 헤더를 csv에 저장
    def save_list_of_indivi_to_csv(self, data_list, file_name):

        with open(file_name, 'w', newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            sub_strings = data_list.split(',')
            csv_writer.writerow(sub_strings)
    def file_exists(self, folder_path, filename):
        file_path = os.path.join(folder_path, filename)
        return os.path.isfile(file_path)

    def center_window(self, root, width=300, height=200):
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()

        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)

        root.geometry('%dx%d+%d+%d' % (width, height, x, y))

    def merge_and_save_pkl(self, data, pkl_path):
        if os.path.exists(pkl_path) and os.path.getsize(pkl_path) > 0:
            with open(pkl_path, 'rb') as f:
                existing_data = pickle.load(f)
        else:
            existing_data = []

        # 기존 데이터와 새로운 데이터 병합
        existing_file_names = {item[0] for item in existing_data}
        new_items = [item for item in data if item[0] not in existing_file_names]
        combined_data = existing_data + new_items

        with open(pkl_path, 'wb') as f:
            pickle.dump(combined_data, f)

        return combined_data

    def get_fast_file_hash(self, filepath, hash_type='md5', chunk_size=8192, sample_size=1024):
        hash_func = getattr(hashlib, hash_type)()

        with open(filepath, 'rb') as f:
            # 파일의 처음 부분에서 sample_size 만큼 읽기
            start_chunk = f.read(sample_size)
            hash_func.update(start_chunk)

            # 파일의 마지막 부분에서 sample_size 만큼 읽기
            f.seek(0, 2)  # 파일 끝으로 이동
            file_size = f.tell()
            if file_size > sample_size:
                f.seek(-sample_size, 2)  # 파일 끝에서 sample_size 전으로 이동
                end_chunk = f.read(sample_size)
                hash_func.update(end_chunk)

        return hash_func.hexdigest()

    def extract_box_feature(self, file_paths):
        #excel_file = str((self.extension[1:]).lower() + '\\' + '_dict.xlsx')  # 엑셀 파일 경로
        all_results = []  # 모든 파일 데이터 저장할 리스트

        excel_file = str(('mp4').lower() + '\\' + '_dict.xlsx')  # 엑셀 파일 경로  # 엑셀 파일 경로
        excel_file = self.resource_path(excel_file)
        df = pd.read_excel(excel_file)  # 엑셀 파일 읽기

        # 엑셀 데이터를 딕셔너리로 변환 (엑셀 파일의 첫 번째 열을 key로, 두 번째 열을 value로)
        self.seqdict = dict(zip(df.iloc[:, 0], df.iloc[:, 1]))

        filecount = 0
        if isinstance(file_paths, str):
            file_paths = [file_paths]

        for file_path in file_paths:
            filecount +=1
            results = []
            results.append(('name', os.path.basename(file_path)))  # 파일명 열 추가
            nameprintformat = f"{filecount}/{len(file_paths)}_{results}"
            print(nameprintformat)

            hashf = self.get_fast_file_hash(file_path)
            results.append(('md5', hashf))

            if self.structure_val_state == True or self.structure_seq_state == True:

                self.set_state("structure_val_state", 1)

                # 각 파일 데이터 저장 리스트
                onesequence = []
                # 파일 내 Box 파싱
                def parse_box(f, end_position, depth=0, max_depth=300):
                    if depth > max_depth:
                        print("최대 재귀 깊이 도달 에러")
                        return

                    while f.tell() < end_position:
                        box_header = f.read(8)  # 첫 8Bytes Box 헤더
                        if len(box_header) < 8:
                            break

                        box_size, box_type = struct.unpack(">I4s", box_header)  # size 4Bytes, type 4Bytes 추출
                        try:
                            lenbox_type = len(box_type)

                            box_type = box_type.decode("utf-8")
                            if len(box_type) != 4 or "\\x" in repr(box_type):
                                raise ValueError(f"잘못된 박스 타입 감지: {box_type}")

                        except Exception as e:
                            f.seek(0,1)
                            continue

                        if box_size == 0:  # 파일의 끝까지 Box가 확장됨을 의미
                            break
                        elif box_size == 1:  # 실제 크기는 다음 8Bytes에 저장됨
                            large_size = f.read(8)
                            actual_box_size = struct.unpack(">Q", large_size)[0]
                        else:
                            actual_box_size = box_size

                        box_end_position = f.tell() + (actual_box_size - 8 if box_size == 1 else box_size - 8)

                        if box_type in ('moov', 'trak', 'mdia', 'minf', 'stbl', 'udta', 'edts', 'moof', 'traf'):
                            # 컨테이너 Box 처리
                            parse_box(f, box_end_position, depth + 1, max_depth)
                        else:  # 컨테이너가 아닌 Box 처리
                            if box_type == 'mdat':
                                f.seek(box_end_position)
                                continue
                            box_data = f.read(actual_box_size - 8)

                            box_data_hex = box_data.hex()

                            if self.structure_seq_state == True :
                                self.set_state("structure_seq_state", 1)
                                if box_type in self.seqdict :
                                    onesequence.append(str(self.seqdict[box_type]))

                            # 각 Box의 속성을 구체적으로 추출
                            if box_type == 'ftyp':
                                major_brand = box_data[0:4].decode("utf-8")
                                minor_version = struct.unpack(">I", box_data[4:8])[0]
                                compatible_brands = [box_data[i:i + 4].decode("utf-8") for i in range(8, len(box_data), 4)]
                                results.append((box_type,
                                                f"Major Brand: {major_brand}, Minor Version: {minor_version}, Compatible Brands: {', '.join(compatible_brands)}"))

                            elif box_type == 'mvhd':
                                version = box_data[0]
                                if version == 0:
                                    create_time, modify_time, timescale, duration = struct.unpack(">IIII", box_data[4:20])
                                else:
                                    create_time, modify_time = struct.unpack(">QQ", box_data[4:20])
                                    timescale, duration = struct.unpack(">II", box_data[20:28])
                                preferred_rate = struct.unpack(">I", box_data[28:32])[0]
                                preferred_volume = struct.unpack(">H", box_data[32:34])[0]
                                box104_108 = box_data[96:100]
                                next_track_id = struct.unpack(">I", box_data[96:100])[0]
                                results.append((box_type,
                                                f"Create Time: {create_time}, Modify Time: {modify_time}, Timescale: {timescale}, Duration: {duration}, Preferred Rate: {preferred_rate}, Preferred Volume: {preferred_volume}, Next Track ID: {next_track_id}"))

                            elif box_type == 'tkhd':
                                version = box_data[0]
                                if version == 0:
                                    create_time, modify_time, track_id, duration = struct.unpack(">IIII", box_data[4:20])
                                else:
                                    create_time, modify_time = struct.unpack(">QQ", box_data[4:20])
                                    track_id, duration = struct.unpack(">II", box_data[20:28])
                                width, height = struct.unpack(">II", box_data[76:84])
                                results.append((box_type,
                                                f"Track ID: {track_id}, Create Time: {create_time}, Modify Time: {modify_time}, Duration: {duration}, Width: {width}, Height: {height}"))

                            elif box_type == 'mdhd':
                                version = box_data[0]
                                if version == 0:
                                    create_time, modify_time, timescale, duration = struct.unpack(">IIII", box_data[4:20])
                                else:
                                    create_time, modify_time = struct.unpack(">QQ", box_data[4:20])
                                    timescale, duration = struct.unpack(">II", box_data[20:28])
                                language_code = struct.unpack(">H", box_data[20:22])[0]
                                results.append((box_type,
                                                f"Create Time: {create_time}, Modify Time: {modify_time}, Timescale: {timescale}, Duration: {duration}, Language Code: {language_code}"))

                            elif box_type == 'elst':
                                version = box_data[0]
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                entries = []
                                offset = 8
                                for _ in range(entry_count):
                                    if version == 1:
                                        segment_duration, media_time, media_rate = struct.unpack(">QqI", box_data[
                                                                                                         offset:offset + 16])
                                        entries.append(
                                            f"Duration: {segment_duration}, Media Time: {media_time}, Rate: {media_rate}")
                                        offset += 16
                                    else:
                                        segment_duration, media_time, media_rate = struct.unpack(">Iii", box_data[
                                                                                                         offset:offset + 12])
                                        entries.append(
                                            f"Duration: {segment_duration}, Media Time: {media_time}, Rate: {media_rate}")
                                        offset += 12
                                results.append((box_type, f"Entry Count: {entry_count}, Entries: {entries}"))

                            elif box_type == 'stsd':
                                version = box_data[0]
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                            elif box_type == 'stts':
                                version = box_data[0]
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                            elif box_type == 'stsc':
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                            elif box_type == 'stsz':
                                sample_size = struct.unpack(">I", box_data[4:8])[0]
                                sample_count = struct.unpack(">I", box_data[8:12])[0]
                                results.append((box_type, f"Sample Size: {sample_size}, Sample Count: {sample_count}"))

                            elif box_type == 'co64':
                                entry_count = struct.unpack(">I", box_data[4:8])[0]
                                results.append((box_type, f"Entry Count: {entry_count}"))

                            else:
                                results.append((box_type, box_data_hex[:5000]))  # Default for other box types

                        # 다음 Box로 이동
                            f.seek(box_end_position)

                with open(file_path, 'rb') as f:
                    file_size = f.seek(0, 2)  # 파일 끝으로 커서 옮겨서 파일 크기 계산
                    f.seek(0)  # 커서를 파일 시작 위치로 이동
                    parse_box(f, file_size)  # 재귀

            if self.frame_gop_state == 1:
                self.set_state("frame_gop_state", 1)
                onesequence = extractGOP(file_path)
                results.append(('GOP', onesequence))

            if self.frame_ratio_state == 1:
                self.set_state("frame_ratio_state", 1)
                ratio = process_videos_in_folder(file_path)
                results.append(('GOP compression', ratio))


            if self.frame_sps_state == 1:
                self.set_state("frame_sps_state", 1)
                try :
                    parse_sps(file_path)
                    file_name = os.path.basename(file_path)
                    file_name +=".264"
                    sps_filepath = file_name

                    spsresult = analyzesps(sps_filepath)
                    results.append(('SPS', spsresult))
                finally:
                    if os.path.exists(sps_filepath):
                        os.remove(sps_filepath)
                        print(f"{sps_filepath} 파일이 삭제되었습니다.")
                    else:
                        print(f"{sps_filepath} 파일이 존재하지 않습니다.")


            if self.structure_seq_state == 1:
                self.set_state("structure_seq_state", 1)
                try:
                    try:
                        try:
                            onesequence = Simhash(onesequence).value
                        except:
                            onesequence = onesequence
                    except:
                        onesequence = np.uint16(onesequence)
                except:
                        onesequence = 0

                results.append(('sequence', onesequence))


            all_results.append(results)
            # 각 파일의 결과를 전체 리스트에 추가

        if self.detectmode == 0:
            self.save_to_csv(all_results)
        return results

    def on_structure_val_changed(self, state):
        if state == Qt.Checked:
            print('structure_val Box is checked')
            self.structure_val_state = 1
            self.set_state("structure_val_state", 1)
            if '_val' not in self.tempcsv_file:
                self.tempcsv_file += '_val'
        else:
            print('structure_val Box is unchecked')
            self.structure_val_state = 1
            self.set_state("structure_val_state", 0)
            self.tempcsv_file = self.tempcsv_file.replace('_val', '')

    def on_structure_seq_changed(self, state):
        if state == Qt.Checked:
            print('structure_seq Box is checked')
            self.structure_seq_state = 1
            self.set_state("structure_seq_state", 1)
            if '_seq' not in self.tempcsv_file:
                self.tempcsv_file += '_seq'
        else:
            print('structure_seq Box is unchecked')
            self.structure_seq_state = 0
            self.set_state("structure_seq_state", 0)
            self.tempcsv_file = self.tempcsv_file.replace('_seq', '')

    def on_frame_sps_changed(self, state):
        if state == Qt.Checked:
            print('frame_sps Box is checked')
            self.set_state("frame_sps_state", 1)
            self.frame_sps_state = 1
            if '_sps' not in self.tempcsv_file:
                self.tempcsv_file += '_sps'
        else:
            print('frame_sps Box is unchecked')
            self.frame_sps_state = 0
            self.set_state("frame_sps_state", 0)
            self.tempcsv_file = self.tempcsv_file.replace('_sps', '')

    def on_frame_gop_changed(self, state):
        if state == Qt.Checked:
            print('frame_gop Box is checked')
            self.frame_gop_state = 1
            self.set_state("frame_gop_state", 1)
            if '_gop' not in self.tempcsv_file:
                self.tempcsv_file += '_gop'
        else:
            print('frame_gop Box is unchecked')
            self.set_state("frame_gop_state", 0)
            self.frame_gop_state = 0
            self.tempcsv_file = self.tempcsv_file.replace('_gop', '')

    def on_frame_ratio_changed(self, state):
        if state == Qt.Checked:
            print('frame_ratio Box is checked')
            self.set_state("frame_ratio_state", 1)
            self.frame_ratio_state = 1
            if '_ratio' not in self.tempcsv_file:
                self.tempcsv_file += '_ratio'
        else:
            print('frame_ratio Box is unchecked')
            self.set_state("frame_ratio_state", 0)
            self.frame_ratio_state = 0
            self.tempcsv_file = self.tempcsv_file.replace('_ratio', '')

    # 결과를 CSV로 저장
    def save_to_csv(self, all_data):
        if self.csv_path == '':
            csv_file = self.csv_path
            timestamp = datetime.now().strftime("%y%m%d%H%M")

            # 파일 이름에 타임스탬프 추가
            filename = f"{csv_file}_train_{timestamp}.csv"
            self.csv_path = os.path.join(self.case_direc, filename)
        csv_path = self.csv_path

        if self.csv_file!='':
            csv_path = self.csv_file

        # 기존 파일에서 데이터와 헤더 불러오기
        existing_data = []
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                existing_fieldnames = reader.fieldnames if reader.fieldnames else []
                for row in reader:
                    existing_data.append(row)
        else:
            existing_fieldnames = []

        # 필드명 추출 - 모든 파일의 필드를 확인하여 중복 필드 처리
        fieldnames = existing_fieldnames.copy()
        key_count_global = {}  # 전체 파일에서의 중복 key 카운트 딕셔너리

        # 모든 파일의 데이터를 순회하여 필드 추출
        for file_data in all_data:
            key_count_local = {}  # 각 파일 내에서의 중복 key 카운트
            for key, value in file_data:
                if key!='GOP':
                    # 중복 필드 처리 (필드 이름 중복 시 숫자를 붙임)
                    if key in key_count_local:
                        key_count_local[key] += 1
                        key_with_count = f"{key}({key_count_local[key]})"
                    else:
                        key_count_local[key] = 1
                        key_with_count = key

                    # 콜론이 있는 경우 자식 속성 분리
                    if isinstance(value, str) and ":" in value:
                        attributes = [attr.strip() for attr in value.split(",")]
                        for attr in attributes:
                            if ":" in attr:
                                attr_name = f"{key_with_count}_{attr.split(':')[0].strip()}"
                                attr_value = attr.split(":")[1].strip()

                                # 필드가 fieldnames에 없고 값이 있는 경우만 필드를 추가
                                if attr_name not in fieldnames and attr_value:
                                    fieldnames.append(attr_name)
                    else:
                        if key_with_count not in fieldnames:
                            fieldnames.append(key_with_count)

                #print('keycount (local):', key_count_local)
                #print('new_fieldnames:', fieldnames)
                else:
                    fieldnames.append(key)

        ##1025 레이블 추가
        if 'label' not in fieldnames:
            fieldnames.append('label')

        # GOP 처리
        # for onedata in all_data:
        #     for key, value in onedata:
        #         if key == 'GOP':
        #             if isinstance(value, str) and ":" in value:
        #                 # 자식 속성 있는 경우 속성 분할(예: “생성 시간: 1234, 수정 시간: 5678”).
        #                 attributes = [attr.strip() for attr in value.split(",")]
        #                 for attr in attributes:
        #                     if ":" in attr:
        #                         attr_name = f"{key}_{attr.split(':')[0].strip()}"
        #                         if attr_name not in fieldnames:
        #                             fieldnames.append(attr_name)
        #             else:
        #                 if key not in fieldnames:
        #                     fieldnames.append(key)

        print('최종 필드명 확인: ', fieldnames)

        # CSV에 기존 데이터와 함께 쓰기
        try:
            all_data = self.move_label_to_second_column(all_data)
        except Exception as e:
            strtem = f'{e}, " -- label 컬럼이 존재하지 않습니다."'
            self.show_alert(strtem)

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            # 기존 데이터 쓰기
            for row in existing_data:
                writer.writerow(row)

            # 새로운 데이터 쓰기
            for file_data in all_data:
                row_data = {}
                key_count_local = {}

                for key, value in file_data:
                    if key in key_count_local:
                        key_count_local[key] += 1
                        key_with_count = f"{key}({key_count_local[key]})"
                    else:
                        key_count_local[key] = 1
                        key_with_count = key

                    if isinstance(value, str):
                        attributes = [attr.strip() for attr in value.split(",")]
                        for attr in attributes:
                            if ":" in attr:
                                attr_name, attr_value = attr.split(":", 1)
                                if f"{key_with_count}_{attr_name.strip()}" in fieldnames:
                                    row_data[f"{key_with_count}_{attr_name.strip()}"] = attr_value.strip()
                            else:
                                if key_with_count in fieldnames:
                                    row_data[key_with_count] = value
                    else:
                        if key_with_count in fieldnames:
                            row_data[key_with_count] = value

                try:
                    if hasattr(self, 'label_data') and self.label_data:
                        row_data['label'] = self.label_data
                    else:
                        print("Warning: 'label_data' is not set or is empty.")
                except Exception as e:
                    pass

                writer.writerow({key: row_data.get(key, "") for key in fieldnames})

        print(f"Results saved to {csv_path}")
        savemassage = f"학습데이터셋이 파일 {csv_path} 에 저장되었습니다."
        self.show_file_alert(csv_path, savemassage, self.tableWidget_Create)

        self.save_states()
        self.post_process_csv(csv_path)

#preprocessing
    def get_fieldnames(self, all_data, existing_data):
        fieldnames = set()
        for row in existing_data:
            fieldnames.update(row.keys())
        for file_data in all_data:
            fieldnames.update(file_data.keys())
        if 'label' not in fieldnames:
            fieldnames.add('label')
        return list(fieldnames)

    # Helper function to prepare a row for writing
    def prepare_row_data(self, file_data, fieldnames):
        row_data = {}
        for key, value in file_data.items():
            row_data[key] = value
        if 'label' in fieldnames and self.label_data:
            row_data['label'] = self.label_data
        return row_data

    # Function to post-process the saved CSV
    def post_process_csv(self, csv_path):
        result_path = csv_path.replace(".csv", "_processed.csv")
        df = pd.read_csv(csv_path)

        # Apply transformations for specific columns
        self.adjust_time_columns(df)
        self.adjust_duration_columns(df)
        self.adjust_dimensions(df)

        # Save processed data
        df.to_csv(result_path, index=False)
        print(f"Processed data saved to {result_path}")

    # Functions for column-specific adjustments
    def adjust_time_columns(self, df):
        pattern = re.compile(r'.*(Create Time|Modify Time)', re.IGNORECASE)
        for col in df.columns:
            if pattern.search(col):
                df[col] = df[col].apply(self.transform_time)

    def transform_time(self, value):
        if pd.notna(value) and len(str(value)) >= 4:
            if str(value).startswith('1'):
                return 1
            elif str(value).startswith('3'):
                return 3
        return value

    def adjust_duration_columns(self, df):
        pattern = re.compile(r'.*(duration|Entry|Entries)\b', re.IGNORECASE)
        for col in df.columns:
            if pattern.search(col):
                df[col] = df[col].apply(self.transform_duration)

    def transform_duration(self, value):
        if pd.notna(value):
            return 1
        return -1 if pd.isna(value) or value == '' else value

    def adjust_dimensions(self, df):
        pattern = re.compile(r'.*(width|height)\b$', re.IGNORECASE)
        for col in df.columns:
            if pattern.search(col):
                df[col] = df[col].apply(self.transform_dimension)

    def transform_dimension(self, value):
        if pd.notna(value):
            if value > 1:
                return 1
            elif value == 0:
                return 0
        return -1 if pd.isna(value) or value == '' else value




    @staticmethod
    def calculate_simhash_lib(value):
        try:
            if value in [0, None, ""] or (isinstance(value, float) and math.isnan(value)):
                return -111111111
        except Exception as e:
            pass
        try:
            try:
                simval = Simhash(str(value)).value
            except:
                simval = Simhash(str(value[:100])).value
        except Exception as e:
            print(e)
            simval = 0
        return simval

    def apply_simhash(self, df):
        """Simhash 적용"""
        # df.columns = df.columns.astype(str)
        # columns_to_process = [col for col in df.columns if col not in ['name', 'label']]
        # for column in columns_to_process:
        #     df[column] = df[column].apply(self.calculate_simhash_lib)
        # return df

        """Simhash 적용"""
        df.columns = df.columns.astype(str)
        columns_to_process = [col for col in df.columns if col not in ['name', 'label']]
        for column in columns_to_process:
            df[column] = df[column].apply(self.calculate_simhash_lib)
        # return df

        # df.columns = df.columns.astype(str)
        # columns_to_process = [col for col in df.columns if col not in ['name', 'label']]
        #
        # def safe_hex_to_int(value):
        #     try:
        #         # 1. 문자열 값 확인
        #         try:
        #             value = int(value)
        #         except:
        #             pass
        #
        #         if isinstance(value, str):
        #             # 과학적 표기법 확인 및 처리
        #             if "E" in value.upper():
        #                 # 과학적 표기법 값을 정수로 변환
        #                 try:
        #                     changeint =  int(float(value))
        #                 except :
        #                     changeint =  int(float(value[:100]))
        #                 return changeint
        #             # 일반 문자열을 16진수로 변환
        #             else:
        #                 return value
        #         # 2. 이미 숫자인 경우
        #         elif isinstance(value, (int, float)):
        #             return int(value)
        #     except Exception as e :
        #         print(e)
        #         return float('nan')
        #
        # for column in columns_to_process:
        #     df[column] = df[column].apply(safe_hex_to_int)

        return df

    def load_file_for_prediction(self):
        """Open a dialog to select a file for prediction."""
        self.detectmode = 1
        file_path= self.file_paths[0]
        if file_path:
            self.predict_on_file(file_path)

    def move_label_to_second_column(self, df):
        # 'label' 컬럼을 분리합니다.
        label_column = df.pop('label')

        # 'label' 컬럼을 두 번째 위치로 삽입합니다.
        df.insert(1, 'label', label_column)

        return df

    def predict_on_file(self, file_path):
        """Predict the label for a new file while handling missing and extra features."""
        try:
            # Extract and flatten features from the file

            feature_data = self.extract_features_from_file(file_path)

            structured_data = self.flatten_features_for_prediction(feature_data)

            # Load model and scaler
            self.load_model_and_scaler()

            # Predict and show results
            predicted_df = self.predict_data1(structured_data)
            print(predicted_df)
            self.detectmode = 0

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Prediction failed for {file_path}: {str(e)}")

    def list_to_dict(self, data):
        """Convert a list of strings into a dictionary by identifying 'key: value' pairs."""
        data_dict = {}
        for index, item in enumerate(data):
            # If the item is a string with colon-separated pairs
            if isinstance(item, str) and ":" in item:
                attributes = [attr.strip() for attr in item.split(",")]
                for attr in attributes:
                    if ":" in attr:
                        key, value = attr.split(":", 1)
                        data_dict[key.strip()] = value.strip()
                    else:
                        # If parsing fails, add the entire item as a single entry
                        data_dict[key] = item
            else:
                # Add non-string items or items without ':' as is
                data_dict[key] = item
        return data_dict

    def flatten_features_for_prediction(self, data):
        # Initialize an empty list for the flattened rows
        flattened_rows = []

        # To store unique field names
        global_fieldnames = set()

        # Flatten each file data
        for file_data in data:
            flattened = {}
            key_count_local = {}  # Local key duplication count for each file

            for key, value in file_data:
                if key != 'GOP':
                    # Handle key duplication within the current file
                    if key in key_count_local:
                        key_count_local[key] += 1
                        key_with_count = f"{key}({key_count_local[key]})"
                    else:
                        key_count_local[key] = 1
                        key_with_count = key

                    # Flatten the attributes in the value if it's a string with nested attributes
                    if isinstance(value, str) and ":" in value:
                        attributes = [attr.strip() for attr in value.split(",")]
                        for attr in attributes:
                            if ":" in attr:
                                attr_name, attr_value = attr.split(":", 1)
                                attr_name = f"{key_with_count}_{attr_name.strip()}"
                                flattened[attr_name] = attr_value.strip()
                                global_fieldnames.add(attr_name)
                            else:
                                flattened[key_with_count] = value
                                global_fieldnames.add(key_with_count)
                    elif isinstance(value, list):
                        for item in value:
                            if ":" in item:
                                attr_name, attr_value = item.split(":", 1)
                                attr_name = f"{key_with_count}_{attr_name.strip()}"
                                flattened[attr_name] = attr_value.strip()
                                global_fieldnames.add(attr_name)
                            else:
                                flattened[key_with_count] = item
                                global_fieldnames.add(key_with_count)
                    else:
                        flattened[key_with_count] = value
                        global_fieldnames.add(key_with_count)
                else:
                    # Handle the 'GOP' key separately
                    gop_value = value.split(':')[0]
                    flattened[key] = gop_value
                    global_fieldnames.add(key)

            flattened_rows.append(flattened)

        # Convert flattened rows into a DataFrame
        df = pd.DataFrame(flattened_rows)

        # Apply transformation functions
        self.adjust_time_columns(df)
        self.adjust_duration_columns(df)
        self.adjust_dimensions(df)

        # Fill missing fields with None
        for field in global_fieldnames:
            if field not in df.columns:
                df[field] = None

        return df

        # Ensure all flattened rows have the same fields, filling missing ones with None
        global_fieldnames = list(global_fieldnames)  # Convert to a sorted list for consistency
        for row in flattened_rows:
            for field in global_fieldnames:
                if field not in row:
                    row[field] = None  # Fill missing fields with None

        return flattened_rows

    def transform_in_flattened(self, flattened, pattern, transform_func):
        """Applies transformation to the flattened data based on the column name pattern."""
        for key, value in flattened.items():
            if pattern.search(key):  # Apply transformation to matching columns
                flattened[key] = transform_func(value)
        return flattened

    def transform_duration_entry(self, x):
        """Transformation logic for duration, entry, and entries columns."""
        if pd.notna(x):  # If value is not NaN, set it to 1
            return 1
        elif pd.isna(x) or x == '':  # If value is NaN or empty, set it to -1
            return -1
        return x  # Otherwise, keep the original value

    def transform_time_value(self, x):
        """Transformation logic for time-related columns (Create Time, Modify Time)."""
        if pd.notna(x) and len(str(x)) >= 4:
            x_str = str(x)
            if x_str.startswith('1'):
                return 1
            elif x_str.startswith('3'):
                return 3
            else:
                return 0
        return x  # Keep original value if it doesn't match

    def transform_width_height(self, x):
        """Transformation logic for width and height columns."""
        x = int(x)
        if pd.notna(x):
            if x > 1:
                return 1  # If width/height is greater than 1, return 1
            elif x == 0:
                return 0  # If width/height is 0, return 0
        elif pd.isna(x) or x == '':
            return -1  # If value is NaN or empty, return -1
        return x  # Otherwise, keep the original value


    def extract_features_from_file(self, file_path):
        """Extract features from the selected file, using the same state-based logic."""
        results = []
        self.load_or_initialize_states()

        if self.structure_val_state:
            box_features = self.extract_box_feature(file_path)
            results.append((box_features))
        elif self.frame_gop_state:
            gop_features = extractGOP(file_path)
            results.append(('GOP', gop_features))
        elif self.frame_sps_state:
            sps_result = self.extract_sps_features(file_path)
            results.append(('SPS', sps_result))
        elif self.structure_seq_state:
            ######1031
            results = [f[1] for f in results if f and len(f) > 1 and f[0] != 'name']
            sequence_feature = Simhash([f for f in results if isinstance(f, str) and 'name' not in f]).value
            results.append(('sequence', sequence_feature))

        return results

    def load_model_and_scaler(self):
        """Load the trained model and scaler from disk."""


        binstat = self.binButton_2.isChecked()
        mulstat = self.mulButton_2.isChecked()
        self.aimodel = 'Xgboost'
        self.model_combo_2.activated.connect(self.on_combobox_select)
        self.aimodel = self.model_combo_2.currentText()
        if binstat:
            self.classmode = 'bin_'
        elif mulstat:
            self.classmode = 'mul_'
        else :
            messagebox.showerror("에러", "바이너리/멀티 모드를 선x택하세요")


        self.pklname = str(self.csv_path+"_" + self.aimodel + "model.pkl")

        self.pklpath = self.resource_path(self.pklname)

        self.scalername = str(self.csv_path+"_" + self.aimodel + "scaler.pkl")
        self.scalerpath = self.resource_path(self.scalername)
        if os.path.exists(self.pklpath) and os.path.exists(self.scalerpath):
            self.model = joblib.load(self.pklpath)
            self.scaler = joblib.load(self.scalerpath)
        else:
            raise FileNotFoundError("Model or scaler file not found.")

    def predict_data1(self, structured_data):
        """Scale the features and predict the label."""
        df = pd.DataFrame(structured_data)
  # Drop the first row and reset the index

        # Load model features from feature.json

        jsonpath = os.path.join(os.path.dirname(self.csv_path), "feature.json")
        with open(jsonpath, 'r') as f:
            model_features = json.load(f)

        # Add missing features with default value 0
        for feature in model_features:
            if feature not in df.columns:
                df[feature] = 0  # Add missing features with 0

        # Keep only the relevant features and ensure the order matches
        df = df[model_features]
        try:
            df = df.drop(columns='md5')
        except Exception as e:
            pass
        # Drop unnecessary columns, e.g., 'name'
        df = df.drop(columns=[col for col in df.columns if col == 'name'], errors='ignore')

        # Apply Simhash transformation (assuming apply_simhash is defined)
        df = self.apply_simhash(df)

        # Align df with the scaler's features by reordering
        scaler_features = self.scaler.feature_names_in_
        df = df.reindex(columns=scaler_features)  # Fill missing features with 0

        # Scale features and predict
        X_new_scaled = self.scaler.transform(df)
        y_pred = self.model.predict(X_new_scaled)
        y_pred_probs = self.model.predict_proba(X_new_scaled)
        predicted_class_probs = y_pred_probs[np.arange(len(y_pred)), y_pred]

        # Add predictions to DataFrame
        df['predicted_label'] = y_pred

        # Load label information from Excel
        labelpath = "labeldata_mul.csv"
        labelpath = os.path.join(labelpath)
        labelpath = self.resource_path(labelpath)
        labeltransferdf = pd.read_csv(labelpath)

        try:
            # Convert prediction to integer for column access
            temppred = int(y_pred[0])

            # Ensure label columns are integers
            labeltransferdf.columns = [int(col) for col in labeltransferdf.columns]

            # Filter the relevant label
            filtered_df = labeltransferdf[temppred]
        except KeyError:
            message = f"해당 라벨({temppred})이 존재하지 않습니다. 라벨을 업데이트하세요."
            self.show_alert(message)
        except Exception as e:
            self.show_alert(str(e))

        try:
            # Format probability and show message
            fileaccuracy = "{:.3f}".format(predicted_class_probs[0][1] * 100)
            message = f"{fileaccuracy}% 확률로 {filtered_df[0]}({y_pred}) 입니다"
            self.show_alert(message)
        except Exception as e:
            fileaccuracy = "{:.3f}".format(predicted_class_probs[0] * 100)
            message = f"{fileaccuracy}% 확률로({y_pred}) 입니다"
            self.show_alert(message)

        return df

    def show_select_file(self):
        """파일 경로를 받아 사용자에게 알림을 표시하고 파일을 여는 함수."""
        app = QApplication.instance()  # 이미 실행 중인 QApplication 인스턴스 확인
        if not app:
            app = QApplication(sys.argv)

        # QDialog를 사용해 타이틀 없는 커스텀 알림창 생성
        dialog = QDialog()
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)  # 타이틀 바 제거 및 최상단 설정

        # 다크 모드 스타일 적용
        dialog.setStyleSheet("""
                    QDialog {
                        background-color: #2e2e2e;
                        border: 2px solid #444;
                        border-radius: 15px;
                        padding: 20px;
                        font: bold 10pt "Playfair Display";
                    }
                    QLabel {
                        color: #f5f5f5;
                        font-size: 20px;
                        font-weight: bold;
                        margin-bottom: 10px;
                        font: bold 10pt "Playfair Display";
                    }
                    QPushButton {
                        background-color: #444;
                        color: white;
                        border: 1px solid #777;
                        border-radius: 5px;
                        padding: 8px 15px;
                        margin-top: 10px;
                        font: bold 10pt "Playfair Display";
                    }
                    QPushButton:hover {
                        background-color: #555;
                    }
                """)

        # 레이아웃 생성 및 위젯 추가
        layout = QVBoxLayout()
        messages = "현재 케이스의 csv에 추가하겠습니까?(No 선택시 다른 케이스의 csv를 직접 선택)"
        message_label = QLabel(messages)
        message_label.setWordWrap(True)
        layout.addWidget(message_label)

        # '확인' 버튼 추가
        open_button = QPushButton("확인") # 파일 열기 함수 호출
        open_button.clicked.connect(dialog.accept)
        layout.addWidget(open_button)

        # '취소' 버튼 추가
        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(self.filedialog)
        cancel_button.clicked.connect(dialog.accept) # 창 닫기
        layout.addWidget(cancel_button)
        dialog.setFixedSize(400, 200)
        dialog.setLayout(layout)

        # 창 크기 조정 및 화면 중앙 배치
        dialog.adjustSize()

        # 알림 창 표시
        dialog.exec_()
        return


    def filedialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "파일 선택", "", "All Files (*);;Text Files (*.txt)")
        self.csv_file = file_path
        return file_path

    def show_file_alert(self, file_path, messagea, widgett):
        """파일 경로를 받아 사용자에게 알림을 표시하고 파일을 여는 함수."""
        app = QApplication.instance()  # 이미 실행 중인 QApplication 인스턴스 확인
        if not app:
            app = QApplication(sys.argv)

        # QDialog를 사용해 타이틀 없는 커스텀 알림창 생성
        dialog = QDialog()
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)  # 타이틀 바 제거 및 최상단 설정

        # 다크 모드 스타일 적용
        dialog.setStyleSheet("""
                    QDialog {
                        background-color: #2e2e2e;
                        border: 2px solid #444;
                        border-radius: 15px;
                        padding: 20px;
                        font: bold 10pt "Playfair Display";
                    }
                    QLabel {
                        color: #f5f5f5;
                        font-size: 20px;
                        font-weight: bold;
                        margin-bottom: 10px;
                        font: bold 10pt "Playfair Display";
                    }
                    QPushButton {
                        background-color: #444;
                        color: white;
                        border: 1px solid #777;
                        border-radius: 5px;
                        padding: 8px 15px;
                        margin-top: 10px;
                        font: bold 10pt "Playfair Display";
                    }
                    QPushButton:hover {
                        background-color: #555;
                    }
                """)

        # 레이아웃 생성 및 위젯 추가
        layout = QVBoxLayout()
        messages = messagea + "바로 확인하시겠습니까?"
        message_label = QLabel(messages)
        message_label.setWordWrap(True)
        layout.addWidget(message_label)

        # '확인' 버튼 추가
        open_button = QPushButton("확인")
        open_button.clicked.connect(lambda: self.open_csv2(file_path, widgett))  # 파일 열기 함수 호출
        open_button.clicked.connect(dialog.accept)
        layout.addWidget(open_button)

        # '취소' 버튼 추가
        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(dialog.reject)  # 창 닫기
        layout.addWidget(cancel_button)
        dialog.setFixedSize(400, 200)
        dialog.setLayout(layout)

        # 창 크기 조정 및 화면 중앙 배치
        dialog.adjustSize()

        # 알림 창 표시
        dialog.exec_()
        return


    def show_alert(self, message):
        title = "알림"
        app = QApplication.instance()  # 이미 실행 중인 QApplication 인스턴스 확인
        if not app:
            app = QApplication(sys.argv)

        # QDialog를 사용해 타이틀 없는 커스텀 알림창 생성
        dialog = QDialog()
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)  # 타이틀 바 제거 및 최상단 설정

        # 다크 모드 스타일 적용
        dialog.setStyleSheet("""
            QDialog {
                background-color: #2e2e2e;
                border: 2px solid #444;
                border-radius: 15px;
                padding: 20px;
                font: bold 10pt "Playfair Display";
            }
            QLabel {
                color: #f5f5f5;
                font-size: 20px;
                font-weight: bold;
                margin-bottom: 10px;
                font: bold 10pt "Playfair Display";
            }
            QPushButton {
                background-color: #444;
                color: white;
                border: 1px solid #777;
                border-radius: 5px;
                padding: 8px 15px;
                margin-top: 10px;
                font: bold 10pt "Playfair Display";
            }
            QPushButton:hover {
                background-color: #555;
            }
        """)

        # 레이아웃 생성 및 위젯 추가
        layout = QVBoxLayout()
        title_label = QLabel(title)
        message_label = QLabel(message)
        layout.addWidget(title_label)
        layout.addWidget(message_label)

        # 확인 버튼 추가
        button = QPushButton("확인")
        button.clicked.connect(dialog.accept)  # 버튼 클릭 시 창 닫기
        layout.addWidget(button)

        dialog.setLayout(layout)

        # 창 크기 조정 및 화면 중앙 배치
        dialog.adjustSize()
        # screen_center = QApplication.primaryScreen().geometry().center()
        # dialog.move(screen_center - dialog.rect().center())

        # 알림 창 표시
        dialog.exec_()


    def extract_sps_features(self, file_path):
        """Extract SPS features."""
        try:
            parse_sps(file_path)
            file_name = os.path.basename(file_path) + ".264"
            sps_result = analyzesps(file_name)
            return sps_result
        finally:
            if os.path.exists(file_name):
                os.remove(file_name)




    def main(self):
        self.ngrams = []

        while True:
            choice = self.choice

            folder_path = os.getcwd()  # 폴더 경로
            filename = 'lcsdata.pkl'  # 확인하고 싶은 파일 이름

            a = self.file_exists(folder_path, filename) # lcsdata.pkl 확인

            if choice == 1: #기준 피처를 만들기 위함, 10개 이내의 파일로 파일형식의 피처 생성
                print("1클릭")
                self.extension = (self.file_paths[0].split('.'))[1]

                #헤더딕셔너리(기존딕셔너리에 없으면 추가하기 위함)
                """header = self.extract_value_tocsv(choice) # 헤더 추출해서 문자열로 반환
                headersave = header.replace('name,', '') # header에서 name 문자열 제거한 결과 저장
                filename = str(self.extension+ 'header.txt') # 헤더 정보 저장할 파일 경로, 이름
                self.add_string_if_not_exists(filename, headersave)
                messagebox.showinfo("Notification", "Learning data extraction has been completed")"""

                break

            elif choice == 2:


                print("2클릭", )
                print("선택한 파일", self.file_paths)
                if self.csv_file != '' or self.existval == 1:
                    file_names = [os.path.basename(path) for path in self.file_paths]
                    if self.csv_file != '':
                        csv_file_path = self.csv_file
                    else :
                        csv_file_path = self.csv_path
                    df = pd.read_csv(csv_file_path)
                    file_info = {self.get_fast_file_hash(path): path for path in self.file_paths}

                    existing_hashes = set(df['md5'])

                    # CSV의 hash 컬럼과 file_info 해시 값을 비교합니다.
                    # 해시가 같으면서 name 컬럼의 파일명이 다른 경우 name 값을 업데이트합니다.
                    for i, row in df.iterrows():
                        file_hash = row['md5']
                        if file_hash in file_info:
                            csv_name = row['name']
                            file_name = os.path.basename(file_info[file_hash])
                            if csv_name != file_name:
                                df.at[i, 'name'] = file_name
                    df.to_csv(csv_file_path, index=False)

                    # 기존 CSV에 없는 새로운 해시만 추출하여 new_entries에 저장합니다.
                    new_entries = [file_info[hash_val] for hash_val in file_info if hash_val not in existing_hashes]

                    self.extract_box_feature(new_entries)

                else :
                    self.extract_box_feature(self.file_paths)



                break

            elif choice == 3:
                self.file_paths = self.listWidget
                #self.extension = (self.file_paths[0].split('.'))[1]
                #self.folderpath("inputfile2")
                # 파일을 순서딕셔너리와 비교하여 1-2-3-4, 2-3-1-4 등 순서 리스트를 만들고, 이를 심해시화
                self.extension = 'mp4'

                self.sequencedem = []
                '''                hexlist = str(self.extension + '\\' +  "hexlist.pkl")
                                with open(hexlist, 'rb') as f:
                                    hexvalues = pickle.load(f)
                                for h in range(len(hexvalues)):
                                    self.feature_dictionary(hexvalues[h])'''


                self.feature_dictionary()

                filename_to_sequence = {}
                for path, value in self.sequencedem:
                    filename = os.path.basename(path)
                    filename_to_sequence[filename] = value

                extractvalue = str(self.extension + '\\' +  "extractvalues.csv")
                df = pd.read_csv(extractvalue)

                for index, row in df.iterrows():
                    if row['name'] in filename_to_sequence:
                        df.at[index, 'sequence'] = filename_to_sequence[row['name']]

                df.to_csv(extractvalue, index=False)
                messagebox.showinfo("Notification", "Learning data extraction has been completed")
                break

            else :
                self.show_alert("학습버튼을 선택하세요")
                break


    def ask_overwrite_labels(self):
        """덮어쓰기 여부를 묻는 팝업."""
        reply = QMessageBox.question(
            self, "Overwrite Labels", "Do you want to overwrite existing labels?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        return reply == QMessageBox.Yes

    def open_data_entry_window(self):
        """데이터 입력 창 열기."""
        overwrite_labels = self.ask_overwrite_labels()
        self.data_entry_window = DataEntryWindow(overwrite_labels, self.case_direc)
        self.data_entry_window.show()
        self.load_excel_data()

    def load_excel_data(self):
        """엑셀 데이터를 DataFrame으로 불러와 테이블에 표시."""
        labelname = "labeldata_mul.csv"
        labelname = os.path.join(self.case_direc, labelname)
        labelpath = self.resource_path(labelname)
        if not os.path.exists(labelpath):
            print(self, "Warning", "No Excel file found!")

        df = pd.read_csv(labelpath)  # 엑셀 파일을 DataFrame으로 로드
        df.columns = [str(col) for col in df.columns]

        self.display_dataframe(df, widgettype=self.tableWidget_train)
        self.display_dataframe(df, widgettype=self.tableWidget_detect)



    def resource_path(self, relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)


    def show_input_dialog(self, title, label):
        """커스텀 입력창을 표시하고 입력된 텍스트를 반환하는 함수."""
        app = QApplication.instance()  # 이미 실행 중인 QApplication 인스턴스 확인
        if not app:
            app = QApplication([])

        # QDialog 생성 및 스타일 설정
        dialog = QDialog()
        dialog.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)  # 타이틀 바 제거 및 최상단 설정

        dialog.setStyleSheet("""
            QDialog {
                background-color: #2e2e2e;
                border: 2px solid #444;
                border-radius: 10px;
                padding: 20px;
            }
            QLabel {
                color: #f5f5f5;
                font-size: 16px;
            }
            QLineEdit {
                background-color: #444;
                color: #f5f5f5;
                border: 1px solid #777;
                border-radius: 5px;
                padding: 8px;
                margin-top: 10px;
            }
            QPushButton {
                background-color: #555;
                color: white;
                border-radius: 5px;
                padding: 8px 15px;
                margin-top: 10px;
            }
            QPushButton:hover {
                background-color: #666;
            }
        """)

        # 레이아웃 생성
        layout = QVBoxLayout()

        # 라벨 추가
        message_label = QLabel(label)
        layout.addWidget(message_label)

        # 입력 필드 추가
        line_edit = QLineEdit()
        layout.addWidget(line_edit)

        # 버튼 생성 및 추가
        button_layout = QVBoxLayout()

        ok_button = QPushButton("확인")
        ok_button.clicked.connect(dialog.accept)  # 확인 클릭 시 다이얼로그 닫기
        layout.addWidget(ok_button)

        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(dialog.reject)  # 취소 클릭 시 다이얼로그 닫기
        layout.addWidget(cancel_button)

        # 다이얼로그에 레이아웃 설정
        dialog.setLayout(layout)

        # 다이얼로그 실행 및 결과 처리
        if dialog.exec_() == QDialog.Accepted:
            return line_edit.text(), True
        return "", False

    ##############라벨입력
    def input_label(self):
        if self.binButton.isChecked():
            self.label_datacsv = 'labeldata_bin.csv'
        elif self.mulButton.isChecked():
            self.label_datacsv = 'labeldata_mul.csv'

        # 라벨 데이터 입력 받기
        self.label_data, ok = self.show_input_dialog("입력", "라벨 데이터를 입력하세요.")
        if not ok or not self.label_data:
            return

        try:
            number = float(self.label_data)  # 숫자 입력 여부 확인
        except ValueError:
            self.show_alert("에러", "유효한 숫자를 입력해주세요.")
            return


        # CSV에서 해당 라벨 데이터를 찾기
        try:
            self.label_datacsv = os.path.join(self.case_direc, self.label_datacsv)
            aaa = self.fetch_name_from_csv(self.label_data)
        except Exception as e:
            self.show_alert("binary/multi class 모드를 선택하세요")
            return

        if aaa is None:  # 해당 데이터가 없을 경우 새로 입력
            name, ok = self.show_input_dialog("입력", "매핑되는 속성을 입력하세요.")
            if not ok or not name:
                return

            # 파일이 없으면 생성
            if not os.path.exists(self.label_datacsv):
                with open(self.label_datacsv, mode='w', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    writer.writerow(["Label", "Name"])  # 헤더 작성

            # CSV 파일에 새 데이터 추가
            with open(self.label_datacsv, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow([self.label_data, name])
        else :
            self.show_alert("해당 라벨이 이미 존재합니다!")

    def fetch_name_from_csv(self, max_key):
        filename = self.label_datacsv

        # 파일이 없을 경우 None 반환
        if not os.path.exists(filename):
            return None

        # 파일에서 해당 키에 해당하는 값을 찾기
        with open(filename, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                if row[0] == max_key:
                    return row[1]
        return None



    def open_existcsv(self):
        self.existval = 1
        self.choice == 2
        self.show_select_file()
        csv_files = [file for file in glob.glob(os.path.join(self.case_direc, "*.csv")) if
                     'feature_importance.csv' not in os.path.basename(file)]

        # .csv 파일이 하나 이상 있을 때 일단은 첫 번째 파일을 열기

        for csv_file in csv_files:
            if '_train' in csv_file or '241' in csv_file:
                self.csv_path = csv_file  # 첫 번째 CSV 파일 경로 선택


##################라벨입력
class DataEntryWindow(QWidget):
    def __init__(self, overwrite_labels, direc):
        super().__init__()
        self.setWindowTitle("Data Entry")
        self.label_counter = 0
        self.headers = []
        self.values = []
        labelpath = 'labeldata_mul.csv'
        labelpath = os.path.join(direc, labelpath)
        self.filename = createtrainclass.resource_path(self, labelpath)


        if not overwrite_labels:
            self.load_existing_labels()

        # UI 구성
        layout = QVBoxLayout()

        self.header_label = QLabel(f"Current Label: {self.label_counter}")
        layout.addWidget(self.header_label)

        self.value_input = QLineEdit()
        self.value_input.setPlaceholderText("Enter value...")
        layout.addWidget(self.value_input)

        add_value_button = QPushButton("Add Value")
        add_value_button.clicked.connect(self.add_value)
        layout.addWidget(add_value_button)

        stop_button = QPushButton("Stop and Save")
        stop_button.clicked.connect(self.stop_and_save)
        layout.addWidget(stop_button)

        self.table = QTableWidget()
        layout.addWidget(self.table)

        self.setLayout(layout)
        self.update_display()

    def load_existing_labels(self):
        """기존 엑셀 파일에서 라벨 로드."""
        if os.path.exists(self.filename):
            workbook = load_workbook(self.filename)
            sheet = workbook.active
            if sheet.max_row > 0:
                self.headers = [cell.value for cell in sheet[1]]
                self.label_counter = len(self.headers)
            workbook.close()

    def add_value(self):
        """값을 추가하고 다음 라벨로 이동."""
        value = self.value_input.text()
        if not value:
            QMessageBox.critical(self, "Error", "Value cannot be empty!")
            return

        self.values.append(value)
        self.headers.append(str(self.label_counter))
        self.label_counter += 1
        self.value_input.clear()
        self.update_display()
        self.header_label.setText(f"Current Label: {self.label_counter}")

    def update_display(self):
        """현재까지 입력된 데이터를 테이블에 표시."""
        self.table.setRowCount(1)
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        for col, value in enumerate(self.values):
            self.table.setItem(0, col, QTableWidgetItem(value))

    def stop_and_save(self):
        """엑셀에 데이터를 저장하고 창 닫기."""
        if not self.values:
            QMessageBox.warning(self, "Warning", "No data to save!")
            return

        if os.path.exists(self.filename):
            workbook = load_workbook(self.filename)
            sheet = workbook.active
            if sheet.max_row == 0:
                sheet.append(self.headers)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(self.headers)

        sheet.append(self.values)
        workbook.save(self.filename)
        workbook.close()

        QMessageBox.information(self, "Success", "Data saved successfully!")
        self.close()


class CaseSelectorApp(QMainWindow):
    def __init__(self):
        super().__init__()

        # Initialize case and dataset directories
        self.case_direc = None
        self.dataset_direc = None

        # Set up the main window
        self.setWindowTitle("Select a Case")
        self.setFixedSize(500, 400)

        # Set up menu bar
        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("File")

        # Add NewCase action to the menu
        new_case_action = QAction("NewCase", self)
        new_case_action.triggered.connect(self.create_new_case)
        file_menu.addAction(new_case_action)

        # Set up main layout with a dark theme
        layout = QVBoxLayout()

        # Title label
        label = QLabel("Select a Case:")
        label.setStyleSheet("color: #f5f5f5; font-size: 20px; font-weight: bold;")
        layout.addWidget(label)

        # List widget to display directories
        self.case_list_widget = QListWidget()
        self.case_list_widget.setStyleSheet("""
            QListWidget {
                background-color: #333;
                color: #f5f5f5;
                border: 1px solid #444;
                border-radius: 10px;
                padding: 10px;
            }
            QListWidget::item {
                padding: 8px;
                font-size: 12pt;
            }
            QListWidget::item:selected {
                background-color: #555;
                color: #f5f5f5;
            }
        """)
        layout.addWidget(self.case_list_widget)

        # Load cases from the 'Cases' directory
        self.load_cases()

        # Connect selection change signal to the method
        self.case_list_widget.itemClicked.connect(self.select_case)

        # Confirm button
        confirm_button = QPushButton("Select Case")
        confirm_button.setStyleSheet("""
            QPushButton {
                background-color: #444;
                color: white;
                border: 1px solid #777;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #555;
            }
        """)
        confirm_button.clicked.connect(self.confirm_selection)
        layout.addWidget(confirm_button)

        # Set up the central widget with dark mode styling
        container = QWidget()
        container.setStyleSheet("""
            QWidget {
                background-color: #2e2e2e;
                border: 2px solid #444;
                border-radius: 15px;
                padding: 20px;
            }
        """)
        container.setLayout(layout)
        self.setCentralWidget(container)

    def load_cases(self):
        """Load directory names from the 'Cases' folder in the current directory."""
        cases_path = os.path.join(os.getcwd(), "Cases")
        if os.path.exists(cases_path):
            case_dirs = [d for d in os.listdir(cases_path) if os.path.isdir(os.path.join(cases_path, d))]
            self.case_list_widget.addItems(case_dirs)
        else:
            print("No 'Cases' directory found in the current path.")

    def select_case(self, item):
        """Set selected case directory."""
        cases_path = os.path.join(os.getcwd(), "Cases")
        self.case_direc = os.path.join(cases_path, item.text())
        print(f"Selected case directory: {self.case_direc}")

    def confirm_selection(self):
        """Confirm the selected case, prompt for dataset directory, and start training."""
        if self.case_direc:
            print(f"Confirmed selection: {self.case_direc}")

            # Check for base_directory.xml and load path if available
            xml_path = "base_directory.xml"
            xml_path = os.path.join(self.case_direc, xml_path)
            dataset_direc = None

            if os.path.exists(xml_path):
                try:
                    tree = ET.parse(xml_path)
                    root = tree.getroot()
                    dataset_direc = root.findtext("dataset_directory")
                    if dataset_direc:
                        print(f"Loaded dataset directory from XML: {dataset_direc}")
                    else:
                        print("Dataset directory not found in XML. Requesting input.")
                except ET.ParseError:
                    print("Error parsing base_directory.xml. Requesting input.")
            else:
                print("base_directory.xml not found. Requesting input.")

            # If no valid directory is found, prompt user for input
            if not dataset_direc:
                dataset_direc, ok = QInputDialog.getText(self, "Dataset Directory",
                                                         "데이터셋 디렉터리를 입력하세요 ex) Y://, Z://")
                if ok and dataset_direc:
                    # Save the directory to base_directory.xml for future use
                    root = ET.Element("settings")
                    ET.SubElement(root, "dataset_directory").text = dataset_direc
                    tree = ET.ElementTree(root)
                    try:
                        with open(xml_path, "wb") as file:
                            tree.write(file, encoding="utf-8", xml_declaration=True)
                        print(f"Dataset directory saved to XML: {dataset_direc}")
                    except IOError:
                        QMessageBox.warning(self, "Error", "Failed to save dataset directory to XML.")
                else:
                    return  # User canceled the input dialog

            self.dataset_direc = dataset_direc
            print(f"Dataset directory set to: {self.dataset_direc}")

            # Create and show the CreateTrain window
            self.create_train_window = createtrainclass(self.case_direc, self.dataset_direc)
            self.create_train_window.show()  # Display the CreateTrain window
            self.close()  # Close CaseSelectorApp after confirmation # Close CaseSelectorApp after confirmation

    def create_new_case(self):
        """Prompt for a new case name and create the case directory."""
        case_name, ok = QInputDialog.getText(self, "New Case", "Enter the name of the new case:")
        if ok and case_name:
            # Create the new case directory
            cases_path = os.path.join(os.getcwd(), "Cases")
            new_case_path = os.path.join(cases_path, case_name)

            os.makedirs(cases_path, exist_ok=True)

            # Check if the case already exists to avoid duplicates
            if not os.path.exists(new_case_path):
                os.mkdir(new_case_path)
                print(f"Created new case directory: {new_case_path}")

                # Refresh the case list and select the new case
                self.case_list_widget.addItem(case_name)
                self.case_direc = new_case_path
                self.create_train_window = createtrainclass(self.case_direc, self.dataset_direc)
                self.create_train_window.show()
            else:
                print(f"Case '{case_name}' already exists.")



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = CaseSelectorApp()
    ex.show()
    app.exec_()
